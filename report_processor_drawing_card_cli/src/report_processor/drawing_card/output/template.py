"""Template analysis and default contract checks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def analyze_template(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheet = workbook["Лист1"] if "Лист1" in workbook.sheetnames else workbook.active
        merges = sorted(str(item) for item in sheet.merged_cells.ranges)
        widths = {
            key: value.width
            for key, value in sheet.column_dimensions.items()
            if value.width is not None
        }
        heights = {
            str(key): value.height
            for key, value in sheet.row_dimensions.items()
            if value.height is not None
        }
        return {
            "path": str(path),
            "sheets": workbook.sheetnames,
            "active_sheet": sheet.title,
            "dimensions": sheet.calculate_dimension(),
            "merged_ranges": merges,
            "column_widths": widths,
            "row_heights": heights,
            "headers": {
                "B2": sheet["B2"].value,
                "E2": sheet["E2"].value,
                "B3:F3": [sheet.cell(3, column).value for column in range(2, 7)],
            },
        }
    finally:
        workbook.close()

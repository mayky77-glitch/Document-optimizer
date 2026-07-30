import json
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from report_processor.drawing_card.review.io import (
    export_manual_review,
    import_review_approvals,
    review_approvals_payload,
)


def test_empty_manual_review_has_no_invalid_empty_validation_container(tmp_path: Path) -> None:
    path = tmp_path / "manual_review.xlsx"
    assert export_manual_review(path, [], []) == 0
    with ZipFile(path) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
        ElementTree.fromstring(sheet_xml)
        assert b"dataValidations" not in sheet_xml
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        assert workbook["Review"].max_row == 1
        assert workbook["Review"]["A1"].value == "Review ID"
    finally:
        workbook.close()


def test_import_requires_valid_category_for_approved_actions(tmp_path: Path) -> None:
    path = tmp_path / "manual_review.xlsx"
    assert export_manual_review(path, [], []) == 0
    workbook = load_workbook(path)
    try:
        sheet = workbook["Review"]
        sheet.append(["row-1", *([None] * 15), "approve", None])
        workbook.save(path)
    finally:
        workbook.close()
    with pytest.raises(ValueError, match="requires a valid category"):
        import_review_approvals(path)


@pytest.mark.parametrize(
    ("action", "category"),
    [
        ("approve", "concrete_works"),
        ("reject", None),
        ("skip", None),
        ("quantity_only", "concrete_works"),
        ("cost_only", "concrete_works"),
        ("change_category", "concrete_works"),
    ],
)
def test_review_decision_json_round_trip(action: str, category: str | None, tmp_path: Path) -> None:
    source = tmp_path / "manual_review.xlsx"
    assert export_manual_review(source, [], []) == 0
    workbook = load_workbook(source)
    try:
        values = [None] * 18
        values[0] = "row-1"
        values[10] = category
        values[16] = action
        workbook["Review"].append(values)
        workbook.save(source)
    finally:
        workbook.close()

    imported = import_review_approvals(source)
    decisions = tmp_path / "review_decisions.json"
    decisions.write_text(
        json.dumps(review_approvals_payload(imported)),
        encoding="utf-8",
    )

    assert import_review_approvals(decisions) == imported


def test_review_decision_json_rejects_invalid_category(tmp_path: Path) -> None:
    decisions = tmp_path / "review_decisions.json"
    decisions.write_text(
        '{"row-1": {"row_id": "row-1", "action": "approve", "category": "unknown"}}',
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="Unsupported review category"):
        import_review_approvals(decisions)

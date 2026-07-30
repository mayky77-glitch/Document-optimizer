import hashlib
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

from report_processor.drawing_card.models import WorkflowRequest, WorkflowResult
from report_processor.drawing_card.output.validator import validate_card
from report_processor.drawing_card.workflow import _publication_blockers, run_workflow


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_demo_create_is_valid_and_sources_unchanged(project_root: Path, tmp_path: Path) -> None:
    source = project_root / "examples" / "0906_demo_input.xlsx"
    template = project_root / "templates" / "default_drawing_card_template.xlsx"
    output = tmp_path / "result.xlsx"
    before = (_sha(source), _sha(template))
    result = run_workflow(
        WorkflowRequest(
            inputs=(source,),
            template=template,
            output=output,
            mode="create",
            rag_mode="off",
            strict=True,
            work_dir=tmp_path / "work",
        )
    )
    assert result.status == "OK"
    assert output.exists()
    assert validate_card(output)["status"] == "OK"
    assert before == (_sha(source), _sha(template))
    assert (result.work_dir / "processing_summary.json").exists()
    assert (result.work_dir / "write_operations.jsonl").exists()

    workbook = load_workbook(output, data_only=True)
    try:
        sheet = workbook["Лист1"]
        assert sheet["B2"].value == "Индекс объекта: 0906"
        assert sheet["E2"].value == "Остаток работ по договору"
        assert "E2:F2" in {str(item) for item in sheet.merged_cells.ranges}
    finally:
        workbook.close()


def test_demo_dry_run_writes_audit_but_not_excel(project_root: Path, tmp_path: Path) -> None:
    output = tmp_path / "should_not_exist.xlsx"
    result = run_workflow(
        WorkflowRequest(
            inputs=(project_root / "examples" / "0906_demo_input.xlsx",),
            template=project_root / "templates" / "default_drawing_card_template.xlsx",
            output=output,
            mode="create",
            rag_mode="off",
            strict=True,
            dry_run=True,
            work_dir=tmp_path / "work",
        )
    )
    assert result.status == "OK"
    assert not output.exists()
    assert (result.work_dir / "planned_write_operations.json").exists()
    assert (result.work_dir / "manual_review.xlsx").exists()


def test_invalid_output_is_blocked_with_audit(project_root: Path, tmp_path: Path) -> None:
    result = run_workflow(
        WorkflowRequest(
            inputs=(project_root / "examples" / "0906_demo_input.xlsx",),
            template=project_root / "templates" / "default_drawing_card_template.xlsx",
            output=tmp_path,
            mode="create",
            rag_mode="off",
            strict=True,
            work_dir=tmp_path / "work",
        )
    )
    assert result.status == "BLOCKED"
    assert any(warning.startswith("REQUEST_VALIDATION_FAILED:") for warning in result.warnings)
    assert (result.work_dir / "error.json").exists()
    assert (result.work_dir / "processing_summary.json").exists()


def test_output_cannot_overwrite_source_and_still_records_audit(
    project_root: Path, tmp_path: Path
) -> None:
    source = project_root / "examples" / "0906_demo_input.xlsx"
    result = run_workflow(
        WorkflowRequest(
            inputs=(source,),
            template=project_root / "templates" / "default_drawing_card_template.xlsx",
            output=source,
            mode="create",
            rag_mode="off",
            strict=True,
            work_dir=tmp_path / "work",
        )
    )
    assert result.status == "BLOCKED"
    assert any("must not overwrite an input" in warning for warning in result.warnings)
    assert (result.work_dir / "processing_summary.json").exists()


@pytest.mark.parametrize("base_kind", ["template", "existing_card"])
def test_output_cannot_overwrite_publication_base_before_inspection(
    base_kind: str, project_root: Path, tmp_path: Path
) -> None:
    template = project_root / "templates" / "default_drawing_card_template.xlsx"
    existing_card = template if base_kind == "existing_card" else None
    result = run_workflow(
        WorkflowRequest(
            inputs=(project_root / "examples" / "0906_demo_input.xlsx",),
            template=template,
            existing_card=existing_card,
            output=template,
            mode="update" if existing_card else "create",
            rag_mode="off",
            strict=True,
            work_dir=tmp_path / "work",
        )
    )

    assert result.status == "BLOCKED"
    assert not result.manifest
    assert any(
        "must not overwrite the template or existing card" in warning for warning in result.warnings
    )
    assert (result.work_dir / "error.json").exists()
    assert (result.work_dir / "processing_summary.json").exists()


@pytest.mark.parametrize(
    "warning",
    [
        "INVALID_NUMBER:bad",
        "EXCEL_ERROR:#VALUE!",
        "DRAWING_CODE_NOT_FOUND",
        "FORMULA_WITHOUT_CACHED_VALUE",
    ],
)
def test_strict_blockers_include_invalid_source_values(warning: str, tmp_path: Path) -> None:
    result = WorkflowResult(run_id="test", status="OK", work_dir=tmp_path, warnings=[warning])

    assert _publication_blockers(result) == [warning]


def test_formula_backend_limitation_is_not_a_strict_blocker(tmp_path: Path) -> None:
    warning = "FORMULA_NOT_AVAILABLE_FOR_BACKEND"
    result = WorkflowResult(run_id="test", status="OK", work_dir=tmp_path, warnings=[warning])

    assert _publication_blockers(result) == []


def test_invalid_review_json_is_blocked_with_error_and_summary(
    project_root: Path, tmp_path: Path
) -> None:
    decisions = tmp_path / "review_decisions.json"
    decisions.write_text('{"row-1": {"action": "approve", "category": "invalid"}}')
    result = run_workflow(
        WorkflowRequest(
            inputs=(project_root / "examples" / "0906_demo_input.xlsx",),
            template=project_root / "templates" / "default_drawing_card_template.xlsx",
            output=tmp_path / "result.xlsx",
            review_decisions=decisions,
            rag_mode="off",
            work_dir=tmp_path / "work",
        )
    )

    assert result.status == "BLOCKED"
    assert any(warning.startswith("REVIEW_DECISIONS_INVALID:") for warning in result.warnings)
    assert (result.work_dir / "error.json").exists()
    assert (result.work_dir / "processing_summary.json").exists()


def test_strict_invalid_number_blocks_publication_after_complete_audit(
    monkeypatch: pytest.MonkeyPatch, project_root: Path, tmp_path: Path
) -> None:
    import report_processor.drawing_card.workflow as workflow_module

    original_extract_rows = workflow_module.extract_rows

    def extract_with_invalid_number(*args, **kwargs):
        for index, row in enumerate(original_extract_rows(*args, **kwargs)):
            yield (
                replace(row, warnings=(*row.warnings, "INVALID_NUMBER:bad")) if index == 0 else row
            )

    monkeypatch.setattr(workflow_module, "extract_rows", extract_with_invalid_number)
    output = tmp_path / "result.xlsx"
    result = run_workflow(
        WorkflowRequest(
            inputs=(project_root / "examples" / "0906_demo_input.xlsx",),
            template=project_root / "templates" / "default_drawing_card_template.xlsx",
            output=output,
            rag_mode="off",
            strict=True,
            work_dir=tmp_path / "work",
        )
    )

    assert result.status == "BLOCKED"
    assert result.extracted_row_count > 1
    assert result.card_rows
    assert not output.exists()
    assert (result.work_dir / "extracted_rows.jsonl").exists()
    assert (result.work_dir / "source_hashes_after.json").exists()
    assert (result.work_dir / "processing_summary.json").exists()


def test_validator_rejects_unit_marker_as_drawing_code(project_root: Path, tmp_path: Path) -> None:
    source = project_root / "examples" / "0906_demo_input.xlsx"
    template = project_root / "templates" / "default_drawing_card_template.xlsx"
    output = tmp_path / "result.xlsx"
    run_workflow(
        WorkflowRequest(
            inputs=(source,),
            template=template,
            output=output,
            mode="create",
            rag_mode="off",
            strict=True,
            work_dir=tmp_path / "work",
        )
    )
    workbook = load_workbook(output)
    try:
        workbook["Лист1"]["B4"] = "м"
        workbook.save(output)
    finally:
        workbook.close()
    result = validate_card(output)
    assert result["status"] != "OK"
    assert any(item.startswith("INVALID_DRAWING_CODE:") for item in result["errors"])

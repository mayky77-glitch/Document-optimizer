"""Единая точка выбора движка чтения книги."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from ..models import CachedWorksheet
from .xlsb_reader import cache_xlsb_workbook_sheet, open_xlsb_workbook
from .xlsx_reader import cache_read_only_worksheet


@dataclass(slots=True)
class WorkbookSource:
    workbook: Any
    sheet_names: list[str]
    source_reader: str
    is_xlsb: bool

    def cache_sheet(self, sheet_name: str) -> CachedWorksheet:
        if self.is_xlsb:
            return cache_xlsb_workbook_sheet(self.workbook, sheet_name)
        return cache_read_only_worksheet(self.workbook[sheet_name])

    def close(self) -> None:
        self.workbook.close()


def open_source_workbook(path: Path) -> WorkbookSource:
    """Открывает поддерживаемую книгу соответствующим движком."""
    if path.suffix.lower() == ".xlsb":
        workbook, sheet_names = open_xlsb_workbook(path)
        return WorkbookSource(workbook, sheet_names, "pyxlsb", True)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Для чтения .xlsx/.xlsm не установлен openpyxl. Выполните:\n"
            "python3 -m pip install --upgrade openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    return WorkbookSource(
        workbook=workbook,
        sheet_names=list(workbook.sheetnames),
        source_reader="openpyxl",
        is_xlsb=False,
    )

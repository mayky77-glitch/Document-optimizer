"""Потоковое чтение XLSX/XLSM через openpyxl."""

from __future__ import annotations

import sys
from typing import Any

from ..models import CachedWorksheet
from ..normalization import clean_text

def cache_read_only_worksheet(worksheet: Any) -> CachedWorksheet:
    """
    Читает XML текущего листа один раз.

    Используется потоковый парсер openpyxl, но пустые отформатированные ячейки
    не помещаются в кэш. Это особенно важно для книг, где Excel ошибочно
    считает используемыми сотни или тысячи пустых столбцов.
    """
    rows: dict[int, dict[int, Any]] = {}
    actual_max_row = 0
    actual_max_column = 0

    try:
        from openpyxl.worksheet._reader import WorkSheetParser

        with worksheet._get_source() as source:
            parser = WorkSheetParser(
                source,
                worksheet._shared_strings,
                data_only=worksheet.parent.data_only,
                epoch=worksheet.parent.epoch,
                date_formats=worksheet.parent._date_formats,
                timedelta_formats=worksheet.parent._timedelta_formats,
            )

            for row_number, cells in parser.parse():
                sparse: dict[int, Any] = {}

                for cell in cells:
                    value = cell.get("value")
                    if clean_text(value) is None:
                        continue

                    column_number = int(cell["column"])
                    sparse[column_number] = value
                    if column_number > actual_max_column:
                        actual_max_column = column_number

                if sparse:
                    rows[int(row_number)] = sparse
                    actual_max_row = int(row_number)

    except Exception as internal_error:
        # Совместимый запасной путь на случай изменения внутренних API openpyxl.
        # Он всё равно читает лист только один раз, но может быть тяжелее для
        # файлов с искусственно раздутым используемым диапазоном.
        print(
            f"  Быстрый XML-кэш недоступен ({type(internal_error).__name__}); "
            "использую совместимый потоковый режим.",
            file=sys.stderr,
        )

        declared_max_column = min(int(worksheet.max_column or 0), 1024)
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=int(worksheet.max_row or 0),
                min_col=1,
                max_col=declared_max_column,
                values_only=True,
            ),
            start=1,
        ):
            sparse = {
                column_number: value
                for column_number, value in enumerate(row, start=1)
                if clean_text(value) is not None
            }
            if sparse:
                rows[row_number] = sparse
                actual_max_row = row_number
                actual_max_column = max(actual_max_column, max(sparse))

    return CachedWorksheet(
        title=str(worksheet.title),
        rows=rows,
        max_row=actual_max_row,
        max_column=actual_max_column,
    )

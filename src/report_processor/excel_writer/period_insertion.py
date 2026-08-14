"""Fail-closed direct OOXML insertion of reconciliation reporting-period columns."""
# ruff: noqa: E501

from __future__ import annotations

import hashlib
import os
import posixpath
import re
import tempfile
import zipfile
from contextlib import suppress
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, range_boundaries

from report_processor.admin_panel.reconciliation_period import (
    PreparedReconciliationTarget,
    ReconciliationPeriodError,
    ReconciliationPeriodInsertionPlan,
    ReconciliationSheetAnchor,
    ReportingPeriod,
)
from report_processor.admin_panel.reconciliation_target_measure import (
    ReconciliationTargetMeasureError,
    bounded_header_windows,
    calendar_identities,
    discover_historical_target_measures,
    discover_target_measures,
)
from report_processor.target_report.ooxml import worksheet_parts

_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_Q = lambda name: f"{{{_MAIN}}}{name}"  # noqa: E731
# Deliberately only used by the forward scanner *after* quoted strings have
# been removed.  Applying a cell-reference regexp to a whole formula is a
# correctness bug: ``="N4"`` is text, not a dependency.
_A1_TOKEN = re.compile(r"\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?")
_A1_RANGE = re.compile(r"\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?")
_UNSUPPORTED_FORMULA = re.compile(r"\[|\]|\b(?:INDIRECT|ADDRESS)\s*\(", re.IGNORECASE)
_UNSUPPORTED_PART = re.compile(
    r"(?:^|/)(?:tables|pivotTables|pivotCache|slicers|externalLinks|embeddings|controls|charts)/",
    re.IGNORECASE,
)
_PERMITTED_DEFINED_NAMES = {
    "_xlnm.Print_Area",
    "_xlnm.Print_Titles",
    "_xlnm._FilterDatabase",
}
_XMLNS = re.compile(rb"\s(xmlns(?::[A-Za-z_][\w.-]*)?=\"[^\"]+\")")
_MAX_ROWS = 1_048_576
_MAX_COLUMNS = 16_384


def _raw_merge_inventory(source: Path, sheet_names) -> dict[str, tuple[str, ...]]:
    """Validate unnormalised OOXML merge topology before opening a workbook."""

    parts = worksheet_parts(source)
    requested = tuple(sheet_names)
    if any(name not in parts for name in requested):
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
    try:
        with zipfile.ZipFile(source) as archive:
            return {name: _raw_sheet_merges(archive.read(parts[name])) for name in requested}
    except ReconciliationPeriodError:
        raise
    except Exception as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID") from error


def _raw_sheet_merges(payload: bytes) -> tuple[str, ...]:
    try:
        root = ET.fromstring(payload)
    except ET.ParseError as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID") from error
    containers = root.findall(_Q("mergeCells"))
    if not containers:
        return ()
    if len(containers) != 1:
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
    container = containers[0]
    count = container.attrib.get("count")
    if count is None or not count.isdecimal():
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
    references: list[tuple[str, tuple[int, int, int, int]]] = []
    for child in container:
        if child.tag != _Q("mergeCell") or set(child.attrib) != {"ref"}:
            raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
        reference = child.attrib["ref"]
        try:
            left, top, right, bottom = range_boundaries(reference)
        except ValueError as error:
            raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID") from error
        canonical = f"{get_column_letter(left)}{top}:{get_column_letter(right)}{bottom}"
        if (
            reference != canonical
            or (left == right and top == bottom)
            or not 1 <= left <= right <= _MAX_COLUMNS
            or not 1 <= top <= bottom <= _MAX_ROWS
        ):
            raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
        if any(
            not (
                right < existing_left
                or existing_right < left
                or bottom < existing_top
                or existing_bottom < top
            )
            for _existing, (
                existing_left,
                existing_top,
                existing_right,
                existing_bottom,
            ) in references
        ):
            raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
        references.append((reference, (left, top, right, bottom)))
    if int(count) != len(references):
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
    return tuple(reference for reference, _bounds in references)


def _validated_header_merges(
    source: Path,
    sheet_names,
    supplied: dict[str, tuple[str, ...]] | None,
) -> dict[str, tuple[str, ...]]:
    raw = _raw_merge_inventory(source, sheet_names)
    if supplied is not None and any(tuple(supplied.get(name, ())) != raw[name] for name in raw):
        raise ReconciliationPeriodError("PERIOD_INSERTION_ANCHOR_INVALID")
    return raw


def build_period_insertion_plan(
    source_path: Path,
    period: ReportingPeriod | str,
    first_detail_rows: dict[str, int],
    merged_ranges_by_sheet: dict[str, tuple[str, ...]] | None = None,
) -> ReconciliationPeriodInsertionPlan:
    """Freeze the only allowed structural insertion before touching a ZIP."""

    reporting_period = (
        period if isinstance(period, ReportingPeriod) else ReportingPeriod.parse(period)
    )
    source = Path(source_path)
    _reject_package(source)
    raw_merges = _validated_header_merges(source, first_detail_rows, merged_ranges_by_sheet)
    parts = worksheet_parts(source)
    sheet_ids = _sheet_id_map(source)
    with source.open("rb") as stream:
        digest = hashlib.sha256(stream.read()).hexdigest()
    workbook = load_workbook(source, read_only=False, data_only=False, keep_links=True)
    try:
        header_windows = bounded_header_windows(workbook, first_detail_rows, raw_merges)
        current: set[str] = set()
        missing: set[str] = set()
        for sheet_name, detail_row in first_detail_rows.items():
            try:
                pair = discover_target_measures(
                    workbook,
                    {sheet_name: detail_row},
                    raw_merges,
                    {sheet_name: header_windows[sheet_name]},
                )
            except ReconciliationTargetMeasureError as error:
                if str(error) != "TARGET_CURRENT_PERIOD_PAIR_MISSING":
                    raise ReconciliationPeriodError("PERIOD_INSERTION_ANCHOR_INVALID") from None
                missing.add(sheet_name)
            else:
                if _pair_period_conflict(pair[0], reporting_period):
                    raise ReconciliationPeriodError("REPORTING_PERIOD_CONFLICT")
                current.add(sheet_name)
        if current and missing:
            raise ReconciliationPeriodError("PERIOD_INSERTION_MIXED_STATE")
        if current:
            return ReconciliationPeriodInsertionPlan(
                "ReconciliationPeriodInsertion-1.0",
                digest,
                reporting_period,
                (),
                tuple(parts.items()),
                _affected_parts(source, ()),
                tuple(sorted(first_detail_rows.items())),
                True,
            )
        historical = discover_historical_target_measures(
            workbook, first_detail_rows, raw_merges, header_windows
        )
        anchors = tuple(
            ReconciliationSheetAnchor(
                pair.sheet_name,
                sheet_ids[pair.sheet_name],
                parts[pair.sheet_name],
                pair.quantity_column,
                pair.cost_column,
                first_detail_rows[pair.sheet_name],
                pair.parent_span,
                pair.historical_parent_label,
                pair.quantity_leaf_row,
                pair.cost_leaf_row,
                pair.quantity_leaf_label,
                pair.cost_leaf_label,
                pair.suffix_nonempty_count,
                pair.suffix_first_coordinate,
                pair.suffix_last_coordinate,
                pair.suffix_rightmost_coordinate,
                pair.suffix_coordinate_sha256,
            )
            for pair in historical
        )
        _preflight_worksheets(source, anchors)
        return ReconciliationPeriodInsertionPlan(
            "ReconciliationPeriodInsertion-1.0",
            digest,
            reporting_period,
            anchors,
            tuple(parts.items()),
            _affected_parts(source, anchors),
            tuple(sorted(first_detail_rows.items())),
        )
    except ReconciliationPeriodError:
        raise
    except ReconciliationTargetMeasureError as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_ANCHOR_INVALID") from error
    finally:
        workbook.close()


def prepare_period_insertion(
    source_path: Path,
    output_path: Path,
    plan: ReconciliationPeriodInsertionPlan,
) -> PreparedReconciliationTarget:
    """Transform one private copy, independently verify it, then no-clobber publish."""

    source, output = Path(source_path), Path(output_path)
    _assert_digest(source, plan.source_sha256)
    _validate_plan(source, plan)
    _raw_merge_inventory(source, (anchor.sheet_name for anchor in plan.anchors))
    if plan.idempotent:
        raise ReconciliationPeriodError("PERIOD_INSERTION_IDEMPOTENT")
    try:
        descriptor, name = tempfile.mkstemp(
            prefix="period-insertion-", suffix=".xlsx", dir=output.parent
        )
        os.close(descriptor)
        temporary = Path(name)
        _transform_package(source, temporary, plan)
        verify_period_insertion(source, temporary, plan)
        _assert_digest(source, plan.source_sha256)
        # Finish every fallible proof before publication.  Reopen once more to
        # reject a transient/truncated temporary package before link(2).
        published_hash = _sha256(temporary)
        with zipfile.ZipFile(temporary) as archive:
            if archive.testzip() is not None:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        # link(2) is the publication transition: it is no-clobber and, unlike
        # stat/unlink based publication, cannot delete a concurrently created
        # user file.  Everything which may reject the workbook happens above.
        os.link(temporary, output)
        # The link is already durable enough to be a valid private result;
        # critically, never inspect or unlink ``output`` here.
        with suppress(OSError):
            _fsync_directory(output.parent)
        with suppress(OSError):
            temporary.unlink()
        return PreparedReconciliationTarget(str(output), published_hash, plan)
    except FileExistsError as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_OUTPUT_EXISTS") from error
    except ReconciliationPeriodError:
        raise
    except Exception as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_FAILED") from error
    finally:
        if "temporary" in locals():
            # Once the link exists, output is not ours to inspect or remove.
            # This best-effort cleanup only addresses our unique temp path.
            with suppress(OSError):
                temporary.unlink(missing_ok=True)


def verify_period_insertion(
    source_path: Path, output_path: Path, plan: ReconciliationPeriodInsertionPlan
) -> None:
    """Independent package and inverse-coordinate verification of a frozen plan."""

    source, output = Path(source_path), Path(output_path)
    _assert_digest(source, plan.source_sha256)
    _validate_plan(source, plan)
    _raw_merge_inventory(source, (anchor.sheet_name for anchor in plan.anchors))
    _raw_merge_inventory(output, (anchor.sheet_name for anchor in plan.anchors))
    anchors = {item.sheet_name: item for item in plan.anchors}
    source_parts = dict(plan.worksheet_parts)
    try:
        with zipfile.ZipFile(source) as before, zipfile.ZipFile(output) as after:
            if tuple(before.namelist()) != tuple(after.namelist()) or after.testzip() is not None:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            if before.comment != after.comment:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            changed = set(plan.affected_parts)
            for info, candidate_info in zip(before.infolist(), after.infolist(), strict=True):
                # CRC and sizes necessarily change in edited members.  The
                # remaining ZipInfo identity is part of the frozen artifact.
                if (
                    info.filename != candidate_info.filename
                    or info.date_time != candidate_info.date_time
                    or info.compress_type != candidate_info.compress_type
                    or info.external_attr != candidate_info.external_attr
                    or info.create_system != candidate_info.create_system
                    or info.create_version != candidate_info.create_version
                    or info.extract_version != candidate_info.extract_version
                    or info.reserved != candidate_info.reserved
                    or info.flag_bits != candidate_info.flag_bits
                    or info.volume != candidate_info.volume
                    or info.internal_attr != candidate_info.internal_attr
                    or info.extra != candidate_info.extra
                    or info.comment != candidate_info.comment
                    or info.orig_filename != candidate_info.orig_filename
                ):
                    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
                if info.filename not in changed and before.read(info.filename) != after.read(
                    info.filename
                ):
                    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            for sheet_name, anchor in anchors.items():
                _verify_sheet_delta(
                    before.read(source_parts[sheet_name]),
                    after.read(source_parts[sheet_name]),
                    anchor,
                    plan.period,
                )
            _verify_other_changed_parts(before, after, plan)
        workbook = load_workbook(output, read_only=False, data_only=False, keep_links=True)
        try:
            rows = {anchor.sheet_name: anchor.first_detail_row for anchor in plan.anchors}
            header_windows = bounded_header_windows(workbook, rows)
            pairs = discover_target_measures(workbook, rows, header_windows=header_windows)
            if {(pair.sheet_name, pair.quantity_column, pair.cost_column) for pair in pairs} != {
                (anchor.sheet_name, anchor.cost_column + 1, anchor.cost_column + 2)
                for anchor in plan.anchors
            }:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DETECTOR_INVALID")
        finally:
            workbook.close()
    except ReconciliationPeriodError:
        raise
    except Exception as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID") from error


def _transform_package(
    source: Path, temporary: Path, plan: ReconciliationPeriodInsertionPlan
) -> None:
    anchors = {item.sheet_name: item for item in plan.anchors}
    parts = dict(plan.worksheet_parts)
    try:
        with (
            zipfile.ZipFile(source) as before,
            zipfile.ZipFile(temporary, "w", allowZip64=True) as after,
        ):
            after.comment = before.comment
            for info in before.infolist():
                payload = before.read(info.filename)
                sheet_name = next(
                    (name for name, part in parts.items() if part == info.filename), None
                )
                if sheet_name in anchors:
                    payload = _transform_sheet(payload, anchors[sheet_name], plan.period)
                elif info.filename == "xl/workbook.xml" and info.filename in plan.affected_parts:
                    payload = _transform_defined_names(payload, anchors)
                elif info.filename == "xl/calcChain.xml" and info.filename in plan.affected_parts:
                    payload = _transform_calc_chain(payload, _sheet_name_by_id(source), anchors)
                elif info.filename in plan.affected_parts:
                    drawing_anchor = _drawing_owner(source, info.filename, anchors)
                    if drawing_anchor is None:
                        raise ReconciliationPeriodError("PERIOD_INSERTION_TRANSFORM_INVALID")
                    payload = _transform_drawing(payload, drawing_anchor.insertion_after_column)
                after.writestr(info, payload)
        _fsync_file(temporary)
    except ReconciliationPeriodError:
        raise
    except Exception as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_TRANSFORM_INVALID") from error


def _transform_sheet(
    payload: bytes, anchor: ReconciliationSheetAnchor, period: ReportingPeriod
) -> bytes:
    root = ET.fromstring(payload)
    boundary = anchor.insertion_after_column
    _reject_sheet_features(root, boundary)
    parent_row = _historical_parent_xml_row(root, anchor)
    auto_filter = root.find(_Q("autoFilter"))
    original_filter_ref = auto_filter.attrib.get("ref") if auto_filter is not None else None
    for node in root.iter():
        if node.tag == _Q("c") and "r" in node.attrib:
            node.attrib["r"] = _map_coordinate(node.attrib["r"], boundary)
        elif node.tag == _Q("f") and node.attrib.get("t") != "shared" and node.text:
            node.text = _translate_formula(node.text, boundary)
        elif (
            (node.tag == _Q("mergeCell") and "ref" in node.attrib)
            or (node.tag == _Q("autoFilter") and "ref" in node.attrib)
            or (node.tag == _Q("dimension") and "ref" in node.attrib)
        ):
            node.attrib["ref"] = _map_range(node.attrib["ref"], boundary)
        elif node.tag == _Q("conditionalFormatting") and "sqref" in node.attrib:
            node.attrib["sqref"] = _map_sqref(node.attrib["sqref"], boundary)
        elif node.tag == _Q("filterColumn") and "colId" in node.attrib:
            node.attrib["colId"] = str(
                _map_filter_colid(int(node.attrib["colId"]), boundary, original_filter_ref)
            )
        elif node.tag in {_Q("pane"), _Q("sheetView")} and "topLeftCell" in node.attrib:
            node.attrib["topLeftCell"] = _map_coordinate(node.attrib["topLeftCell"], boundary)
        elif node.tag == _Q("selection"):
            if "activeCell" in node.attrib:
                node.attrib["activeCell"] = _map_coordinate(node.attrib["activeCell"], boundary)
            if "sqref" in node.attrib:
                node.attrib["sqref"] = _map_range(node.attrib["sqref"], boundary)
        elif node.tag == _Q("col"):
            # handled as a set below: a spanning column definition needs a
            # three-way split, not an expanded ``max``.
            pass
    _translate_cols(root, anchor)
    sheet_data = root.find(_Q("sheetData"))
    if sheet_data is None:
        raise ReconciliationPeriodError("PERIOD_INSERTION_TRANSFORM_INVALID")
    for row in sheet_data.findall(_Q("row")):
        number = int(row.attrib.get("r", "0"))
        old_cells = {
            _unmap_coordinate(cell.attrib["r"], boundary): cell for cell in row.findall(_Q("c"))
        }
        if not old_cells:
            continue
        inserted = []
        for original_column, new_column in (
            (anchor.quantity_column, boundary + 1),
            (anchor.cost_column, boundary + 2),
        ):
            source_cell = old_cells.get(f"{get_column_letter(original_column)}{number}")
            cell = ET.Element(_Q("c"), {"r": f"{get_column_letter(new_column)}{number}"})
            if source_cell is not None and "s" in source_cell.attrib:
                cell.attrib["s"] = source_cell.attrib["s"]
            if number == parent_row:
                _inline_string(
                    cell,
                    f"{period.label} "
                    f"{anchor.quantity_leaf_label if new_column == boundary + 1 else anchor.cost_leaf_label}",
                )
            inserted.append(cell)
        for cell in inserted:
            row.append(cell)
        row[:] = sorted(row, key=lambda cell: _cell_column(cell.attrib.get("r", "A1")))
        _translate_row_spans(row, boundary)
    return _serialize_preserving_ignorable_namespaces(root, payload)


def _verify_sheet_delta(
    before: bytes, after: bytes, anchor: ReconciliationSheetAnchor, period: ReportingPeriod
) -> None:
    old, new = ET.fromstring(before), ET.fromstring(after)
    boundary = anchor.insertion_after_column
    if _verify_shared_formula_topology(old, boundary) != _verify_shared_formula_topology(
        new, boundary
    ):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
    parent = _historical_parent_xml_row(old, anchor)
    _verify_inserted_cells(old, new, anchor, parent, period)
    _verify_cols_delta(old, new, anchor)
    _inverse_sheet_tree(new, boundary)
    # Cols were independently checked; restore the exact original tree before
    # semantic equality so every unrelated attr/node/text/tail is covered.
    new_cols, old_cols = new.find(_Q("cols")), old.find(_Q("cols"))
    if new_cols is not None:
        parent_node = next((node for node in new.iter() if new_cols in list(node)), None)
        if parent_node is None:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        position = list(parent_node).index(new_cols)
        parent_node.remove(new_cols)
        if old_cols is not None:
            parent_node.insert(position, ET.fromstring(ET.tostring(old_cols)))
    if _semantic_xml(new) != _semantic_xml(old):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")


def _verify_inserted_cells(
    old: ET.Element,
    new: ET.Element,
    anchor: ReconciliationSheetAnchor,
    parent_row: int,
    period: ReportingPeriod,
) -> None:
    boundary = anchor.insertion_after_column
    old_rows = {row.attrib.get("r"): row for row in old.findall(f".//{_Q('row')}")}
    new_rows = {row.attrib.get("r"): row for row in new.findall(f".//{_Q('row')}")}
    if set(old_rows) != set(new_rows):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
    for number, old_row in old_rows.items():
        new_row = new_rows[number]
        old_cells = old_row.findall(_Q("c"))
        inserted = [
            cell
            for cell in new_row.findall(_Q("c"))
            if _cell_column(cell.attrib.get("r", "A1")) in {boundary + 1, boundary + 2}
        ]
        if not old_cells:
            if inserted:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            continue
        if len(inserted) != 2:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        expected_columns = [boundary + 1, boundary + 2]
        if [_cell_column(item.attrib.get("r", "A1")) for item in inserted] != expected_columns:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        for cell, source_column, metric in zip(
            inserted,
            (anchor.quantity_column, anchor.cost_column),
            (anchor.quantity_leaf_label, anchor.cost_leaf_label),
            strict=True,
        ):
            source = next(
                (
                    item
                    for item in old_cells
                    if _cell_column(item.attrib.get("r", "A1")) == source_column
                ),
                None,
            )
            expected_style = source.attrib.get("s") if source is not None else None
            if cell.attrib.get("s") != expected_style:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            if int(number or "0") == parent_row:
                if (
                    cell.attrib.get("t") != "inlineStr"
                    or "".join(cell.itertext()) != f"{period.label} {metric}"
                ):
                    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            elif set(cell.attrib) - {"r", "s"} or list(cell) or cell.text:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")


def _inverse_sheet_tree(root: ET.Element, boundary: int) -> None:
    for parent in root.iter():
        for child in list(parent):
            if child.tag == _Q("c") and _cell_column(child.attrib.get("r", "A1")) in {
                boundary + 1,
                boundary + 2,
            }:
                parent.remove(child)
    for node in root.iter():
        if node.tag == _Q("c") and "r" in node.attrib:
            node.attrib["r"] = _verify_inverse_coordinate(node.attrib["r"], boundary)
        elif node.tag == _Q("f") and node.text:
            node.text = _verify_inverse_formula(node.text, boundary)
        elif (
            node.tag in {_Q("mergeCell"), _Q("autoFilter"), _Q("dimension")}
            and "ref" in node.attrib
        ):
            node.attrib["ref"] = _verify_inverse_range(node.attrib["ref"], boundary)
        elif node.tag == _Q("conditionalFormatting") and "sqref" in node.attrib:
            node.attrib["sqref"] = " ".join(
                _verify_inverse_range(item, boundary) for item in node.attrib["sqref"].split()
            )
        elif node.tag in {_Q("pane"), _Q("sheetView")} and "topLeftCell" in node.attrib:
            node.attrib["topLeftCell"] = _verify_inverse_coordinate(
                node.attrib["topLeftCell"], boundary
            )
        elif node.tag == _Q("selection"):
            for attribute in ("activeCell", "sqref"):
                if attribute in node.attrib:
                    node.attrib[attribute] = (
                        _verify_inverse_coordinate(node.attrib[attribute], boundary)
                        if attribute == "activeCell"
                        else " ".join(
                            _verify_inverse_range(item, boundary)
                            for item in node.attrib[attribute].split()
                        )
                    )
        elif node.tag == _Q("row"):
            _inverse_row_spans(node, boundary)
    auto = root.find(_Q("autoFilter"))
    if auto is not None and "ref" in auto.attrib:
        left, _top, _right, _bottom = range_boundaries(auto.attrib["ref"])
        for node in auto.findall(_Q("filterColumn")):
            if "colId" in node.attrib and left + int(node.attrib["colId"]) > boundary:
                node.attrib["colId"] = str(int(node.attrib["colId"]) - 2)


def _verify_inverse_range(value: str, boundary: int) -> str:
    if ":" not in value:
        return _verify_inverse_coordinate(value, boundary)
    left, top, right, bottom = range_boundaries(value.replace("$", ""))
    if left > boundary + 2:
        left -= 2
        right -= 2
    elif right > boundary + 2:
        right -= 2
    return f"{get_column_letter(left)}{top}:{get_column_letter(right)}{bottom}"


def _inverse_row_spans(row: ET.Element, boundary: int) -> None:
    if "spans" not in row.attrib:
        return
    restored = []
    for item in row.attrib["spans"].split():
        left, right = (int(value) for value in item.split(":", 1))
        if left > boundary + 2:
            left -= 2
            right -= 2
        elif right > boundary + 2:
            right -= 2
        restored.append(f"{left}:{right}")
    row.attrib["spans"] = " ".join(restored)


def _semantic_xml(root: ET.Element) -> bytes:
    return ET.tostring(root, encoding="utf-8")


def _verify_cols_delta(old: ET.Element, new: ET.Element, anchor: ReconciliationSheetAnchor) -> None:
    """Compute the expected expanded `<cols>` independently from source XML."""
    old_cols, new_cols = old.find(_Q("cols")), new.find(_Q("cols"))
    if old_cols is None:
        if new_cols is not None:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        return
    if new_cols is None:
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
    boundary = anchor.insertion_after_column
    expected: list[ET.Element] = []
    quantity: ET.Element | None = None
    cost: ET.Element | None = None
    for source in old_cols.findall(_Q("col")):
        minimum, maximum = int(source.attrib["min"]), int(source.attrib["max"])
        if minimum <= anchor.quantity_column <= maximum:
            quantity = source
        if minimum <= anchor.cost_column <= maximum:
            cost = source
        fragments = ((minimum, min(maximum, boundary), 0), (max(minimum, boundary + 1), maximum, 2))
        for left, right, shift in fragments:
            if left <= right:
                node = ET.Element(_Q("col"), dict(source.attrib))
                node.attrib.update({"min": str(left + shift), "max": str(right + shift)})
                expected.append(node)
    for source, target in ((quantity, boundary + 1), (cost, boundary + 2)):
        if source is not None:
            node = ET.Element(_Q("col"), dict(source.attrib))
            node.attrib.update({"min": str(target), "max": str(target)})
            expected.append(node)
    expected.sort(key=lambda item: int(item.attrib["min"]))
    actual = list(new_cols.findall(_Q("col")))
    if len(actual) != len(expected) or any(
        _semantic_xml(left) != _semantic_xml(right)
        for left, right in zip(actual, expected, strict=True)
    ):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")


def _verify_forward_coordinate(coordinate: str, boundary: int) -> str:
    """Verifier-local coordinate function: intentionally not a forward helper."""
    column, row = coordinate_from_string(coordinate.replace("$", ""))
    index = column_index_from_string(column)
    col_absolute = "$" if coordinate.startswith("$") else ""
    rest = coordinate[len(col_absolute) + len(column) :]
    row_absolute = "$" if rest.startswith("$") else ""
    return f"{col_absolute}{get_column_letter(index + 2 if index > boundary else index)}{row_absolute}{row}"


def _canonical_cell(node: ET.Element, boundary: int, inverse: bool = False) -> bytes:
    """Compare all pre-existing cell state after inverse coordinate recovery."""
    clone = ET.fromstring(ET.tostring(node))
    coordinate = clone.attrib.get("r")
    if inverse and coordinate:
        col, row = coordinate_from_string(coordinate.replace("$", ""))
        index = column_index_from_string(col)
        if index > boundary + 2:
            clone.attrib["r"] = f"{get_column_letter(index - 2)}{row}"
    if inverse:
        formula = clone.find(_Q("f"))
        if formula is not None and formula.text:
            formula.text = _verify_inverse_formula(formula.text, boundary)
    return ET.tostring(clone)


def _verify_inverse_formula(value: str, boundary: int) -> str:
    """Verifier-local token walk; never reuse the forward formula translator."""
    output: list[str] = []
    index = 0
    while index < len(value):
        if value[index] == '"':
            end = index + 1
            while end < len(value):
                if value[end] == '"':
                    if end + 1 < len(value) and value[end + 1] == '"':
                        end += 2
                        continue
                    end += 1
                    break
                else:
                    end += 1
            if end > len(value) or (end == len(value) and value[end - 1] != '"'):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            output.append(value[index:end])
            index = end
            continue
        match = _A1_TOKEN.match(value, index)
        if match and _formula_token_boundary(value, match.start(), match.end()):
            token = match.group(0)
            output.append(
                ":".join(_verify_inverse_coordinate(item, boundary) for item in token.split(":"))
            )
            index = match.end()
            continue
        output.append(value[index])
        index += 1
    return "".join(output)


def _verify_other_changed_parts(
    before: zipfile.ZipFile, after: zipfile.ZipFile, plan: ReconciliationPeriodInsertionPlan
) -> None:
    """Independent structural checks for non-worksheet members.

    These routines intentionally do not call any forward transformation
    function.  They prove that only a permitted coordinate delta occurred.
    """
    anchors = {anchor.sheet_name: anchor for anchor in plan.anchors}
    for part in plan.affected_parts:
        if part in {anchor.worksheet_part for anchor in plan.anchors}:
            continue
        old, new = before.read(part), after.read(part)
        if part == "xl/workbook.xml":
            old_root, new_root = ET.fromstring(old), ET.fromstring(new)
            old_names = list(old_root.iter(_Q("definedName")))
            new_names = list(new_root.iter(_Q("definedName")))
            if len(old_names) != len(new_names):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            for left, right in zip(old_names, new_names, strict=True):
                if left.attrib != right.attrib:
                    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
                if left.text == right.text:
                    continue
                if left.attrib.get("name") not in _PERMITTED_DEFINED_NAMES:
                    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
                inverse = _inverse_defined_name_text(right.text or "", anchors)
                if inverse != (left.text or ""):
                    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
                right.text = inverse
            if _semantic_xml(new_root) != _semantic_xml(old_root):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        elif part == "xl/calcChain.xml":
            _verify_calc_chain_delta(
                old,
                new,
                {anchor.sheet_id: anchor.sheet_name for anchor in plan.anchors},
                anchors,
            )
        else:
            _verify_drawing_delta(old, new, before, part, anchors)


def _inverse_defined_name_text(value: str, anchors: dict[str, ReconciliationSheetAnchor]) -> str:
    """Invert only the narrow, already-whitelisted built-in name grammar."""
    terms: list[str] = []
    for term in value.split(","):
        if "!" not in term:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        sheet, reference = term.rsplit("!", 1)
        name = (
            sheet[1:-1].replace("''", "'")
            if sheet.startswith("'") and sheet.endswith("'")
            else sheet
        )
        anchor = anchors.get(name)
        terms.append(
            term
            if anchor is None
            else f"{sheet}!{_verify_inverse_reference(reference, anchor.insertion_after_column)}"
        )
    return ",".join(terms)


def _verify_inverse_reference(value: str, boundary: int) -> str:
    if re.fullmatch(r"\$?\d+:\$?\d+", value):
        return value
    if re.fullmatch(r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}", value):
        left, right = value.split(":", 1)
        return f"{_verify_inverse_column(left, boundary)}:{_verify_inverse_column(right, boundary)}"
    if re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+", value):
        return ":".join(_verify_inverse_coordinate(item, boundary) for item in value.split(":"))
    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")


def _verify_inverse_column(value: str, boundary: int) -> str:
    absolute = "$" if value.startswith("$") else ""
    index = column_index_from_string(value.lstrip("$"))
    return absolute + get_column_letter(index - 2 if index > boundary + 2 else index)


def _verify_inverse_coordinate(value: str, boundary: int) -> str:
    column, row = coordinate_from_string(value.replace("$", ""))
    index = column_index_from_string(column)
    col_absolute = "$" if value.startswith("$") else ""
    rest = value[len(col_absolute) + len(column) :]
    row_absolute = "$" if rest.startswith("$") else ""
    return f"{col_absolute}{get_column_letter(index - 2 if index > boundary + 2 else index)}{row_absolute}{row}"


def _verify_calc_chain_delta(
    old: bytes,
    new: bytes,
    sheet_names: dict[int, str],
    anchors: dict[str, ReconciliationSheetAnchor],
) -> None:
    old_root, new_root = ET.fromstring(old), ET.fromstring(new)
    left, right = list(old_root.iter(_Q("c"))), list(new_root.iter(_Q("c")))
    if len(left) != len(right):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
    sheet_index: int | None = None
    for prior, later in zip(left, right, strict=True):
        if prior.attrib.get("i") != later.attrib.get("i"):
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        if "i" in prior.attrib:
            sheet_index = int(prior.attrib["i"])
        if sheet_index is None or sheet_index not in sheet_names:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        anchor = anchors.get(sheet_names[sheet_index])
        expected = prior.attrib.get("r")
        actual = later.attrib.get("r")
        if not expected or not actual:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        if anchor is not None and _cell_column(expected) > anchor.insertion_after_column:
            if _verify_inverse_coordinate(actual, anchor.insertion_after_column) != expected:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            later.attrib["r"] = expected
        elif expected != actual:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
    if _semantic_xml(new_root) != _semantic_xml(old_root):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")


def _verify_drawing_delta(
    old: bytes,
    new: bytes,
    archive: zipfile.ZipFile,
    part: str,
    anchors: dict[str, ReconciliationSheetAnchor],
) -> None:
    """Inverse only drawing marker columns owned by the frozen sheet anchor."""
    anchor = _drawing_owner_from_archive(archive, part, anchors)
    if anchor is None:
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
    old_root, new_root = ET.fromstring(old), ET.fromstring(new)
    left, right = list(old_root.iter(_DX("col"))), list(new_root.iter(_DX("col")))
    if len(left) != len(right):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
    boundary = anchor.insertion_after_column - 1
    for prior, later in zip(left, right, strict=True):
        if (
            prior.text is None
            or later.text is None
            or not prior.text.isdigit()
            or not later.text.isdigit()
        ):
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
        if int(prior.text) > boundary:
            if int(later.text) - 2 != int(prior.text):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            later.text = prior.text
        elif prior.text != later.text:
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
    if _semantic_xml(new_root) != _semantic_xml(old_root):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")


def _reject_package(source: Path) -> None:
    try:
        with zipfile.ZipFile(source) as archive:
            names = tuple(archive.namelist())
            if len(names) != len(set(names)) or archive.testzip() is not None:
                raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
            if any(_UNSUPPORTED_PART.search(name) for name in names):
                raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    except ReconciliationPeriodError:
        raise
    except Exception as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID") from error


def _sheet_id_map(source: Path) -> dict[str, int]:
    with zipfile.ZipFile(source) as archive:
        root = ET.fromstring(archive.read("xl/workbook.xml"))
    result = {}
    for node in root.iter(_Q("sheet")):
        name, sheet_id = node.attrib.get("name"), node.attrib.get("sheetId")
        if not name or not sheet_id or not sheet_id.isdigit():
            raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
        result[name] = int(sheet_id)
    return result


def _sheet_name_by_id(source: Path) -> dict[int, str]:
    return {sheet_id: name for name, sheet_id in _sheet_id_map(source).items()}


def _affected_parts(
    source: Path, anchors: tuple[ReconciliationSheetAnchor, ...]
) -> tuple[str, ...]:
    if not anchors:
        return ()
    with zipfile.ZipFile(source) as archive:
        names = set(archive.namelist())
        affected = {anchor.worksheet_part for anchor in anchors}
        boundaries = {anchor.sheet_name: anchor.insertion_after_column for anchor in anchors}
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        if any(
            _defined_name_is_affected(node, boundaries) for node in workbook.iter(_Q("definedName"))
        ):
            affected.add("xl/workbook.xml")
        if "xl/calcChain.xml" in names and _calc_chain_is_affected(
            archive.read("xl/calcChain.xml"), _sheet_name_by_id(source), boundaries
        ):
            affected.add("xl/calcChain.xml")
        for anchor in anchors:
            for _relation_id, target, relation_type, target_mode in _part_relationships(
                archive, anchor.worksheet_part
            ):
                if target_mode == "External":
                    # Validated against the worksheet hyperlink reference in
                    # preflight.  It is never a ZIP member and never rewritten.
                    continue
                if relation_type.endswith("/drawing"):
                    if target is None or target not in names:
                        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
                    drawing = archive.read(target)
                    _preflight_drawing(drawing, anchor.insertion_after_column)
                    if _drawing_is_affected(drawing, anchor.insertion_after_column):
                        affected.add(target)
                elif relation_type.endswith(("/hyperlink", "/comments", "/vmlDrawing")):
                    continue
                else:
                    raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    return tuple(sorted(affected))


def _validate_plan(source: Path, plan: ReconciliationPeriodInsertionPlan) -> None:
    """Reject fabricated plans before a temporary output identity exists."""

    if plan.plan_digest != hashlib.sha256(plan.canonical_bytes()).hexdigest():
        raise ReconciliationPeriodError("PERIOD_INSERTION_PLAN_INVALID")
    expected = build_period_insertion_plan(source, plan.period, dict(plan.selected_detail_rows))
    if expected.canonical_bytes() != plan.canonical_bytes():
        raise ReconciliationPeriodError("PERIOD_INSERTION_PLAN_INVALID")
    parts = worksheet_parts(source)
    sheet_ids = _sheet_id_map(source)
    if tuple(parts.items()) != plan.worksheet_parts:
        raise ReconciliationPeriodError("PERIOD_INSERTION_PLAN_INVALID")
    if _affected_parts(source, plan.anchors) != plan.affected_parts:
        raise ReconciliationPeriodError("PERIOD_INSERTION_PLAN_INVALID")
    if len({anchor.sheet_name for anchor in plan.anchors}) != len(plan.anchors):
        raise ReconciliationPeriodError("PERIOD_INSERTION_PLAN_INVALID")
    for anchor in plan.anchors:
        if (
            parts.get(anchor.sheet_name) != anchor.worksheet_part
            or sheet_ids.get(anchor.sheet_name) != anchor.sheet_id
            or anchor.cost_column != anchor.quantity_column + 1
            or anchor.parent_span[1] != anchor.quantity_column
            or anchor.parent_span[3] != anchor.cost_column
            or not anchor.historical_parent_label
            or anchor.suffix_nonempty_count < 1
            or not anchor.suffix_first_coordinate
            or not anchor.suffix_last_coordinate
            or not anchor.suffix_rightmost_coordinate
            or len(anchor.suffix_coordinate_sha256) != 64
        ):
            raise ReconciliationPeriodError("PERIOD_INSERTION_PLAN_INVALID")


def _part_relationships(
    archive: zipfile.ZipFile, part: str
) -> tuple[tuple[str, str | None, str, str | None], ...]:
    """Return a validated relationship inventory without path-normalising URLs."""
    rel = posixpath.join(posixpath.dirname(part), "_rels", posixpath.basename(part) + ".rels")
    if rel not in archive.namelist():
        return ()
    result: list[tuple[str, str | None, str, str | None]] = []
    for node in ET.fromstring(archive.read(rel)):
        identifier = node.attrib.get("Id")
        target, relation_type, target_mode = (
            node.attrib.get("Target"),
            node.attrib.get("Type", ""),
            node.attrib.get("TargetMode"),
        )
        if not identifier or not target:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        if target_mode == "External":
            if not relation_type.endswith("/hyperlink"):
                raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
            result.append((identifier, target, relation_type, target_mode))
            continue
        if target_mode is not None:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        result.append(
            (
                identifier,
                posixpath.normpath(posixpath.join(posixpath.dirname(part), target)).lstrip("/"),
                relation_type,
                None,
            )
        )
    return tuple(result)


def _drawing_owner(
    source: Path, drawing_part: str, anchors: dict[str, ReconciliationSheetAnchor]
) -> ReconciliationSheetAnchor | None:
    with zipfile.ZipFile(source) as archive:
        return _drawing_owner_from_archive(archive, drawing_part, anchors)


def _drawing_owner_from_archive(
    archive: zipfile.ZipFile,
    drawing_part: str,
    anchors: dict[str, ReconciliationSheetAnchor],
) -> ReconciliationSheetAnchor | None:
    owners = [
        anchor
        for anchor in anchors.values()
        if any(
            target == drawing_part and kind.endswith("/drawing") and mode is None
            for _identifier, target, kind, mode in _part_relationships(
                archive, anchor.worksheet_part
            )
        )
    ]
    return owners[0] if len(owners) == 1 else None


_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DX = lambda name: f"{{{_XDR}}}{name}"  # noqa: E731


def _drawing_columns(payload: bytes) -> tuple[tuple[int, int], ...]:
    root = ET.fromstring(payload)
    result = []
    for anchor in root:
        if anchor.tag not in {_DX("twoCellAnchor"), _DX("oneCellAnchor")}:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        markers = [node.find(_DX("col")) for node in anchor if node.tag in {_DX("from"), _DX("to")}]
        if not markers or any(
            marker is None or marker.text is None or not marker.text.isdigit() for marker in markers
        ):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        result.append(tuple(int(marker.text) for marker in markers))
    return tuple(result)


def _preflight_drawing(payload: bytes, boundary: int) -> None:
    for columns in _drawing_columns(payload):
        # Drawing marker columns are zero based.
        left, right = min(columns), max(columns)
        if left <= boundary - 1 < right:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")


def _drawing_is_affected(payload: bytes, boundary: int) -> bool:
    return any(max(columns) > boundary - 1 for columns in _drawing_columns(payload))


def _transform_drawing(payload: bytes, boundary: int) -> bytes:
    root = ET.fromstring(payload)
    for marker in root.iter(_DX("col")):
        if marker.text is None or not marker.text.isdigit():
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        if int(marker.text) > boundary - 1:
            marker.text = str(int(marker.text) + 2)
    return _serialize_preserving_ignorable_namespaces(root, payload)


def _preflight_worksheets(source: Path, anchors: tuple[ReconciliationSheetAnchor, ...]) -> None:
    """Parse every changing sheet and reject unsupported coordinates before temp creation."""

    try:
        with zipfile.ZipFile(source) as archive:
            for anchor in anchors:
                root = ET.fromstring(archive.read(anchor.worksheet_part))
                _reject_sheet_features(root, anchor.insertion_after_column)
                _require_insertible_rows(root, anchor)
                _validate_sheet_relationships(archive, anchor, root)
            _preflight_workbook(archive, anchors)
    except ReconciliationPeriodError:
        raise
    except Exception as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID") from error


def _reject_sheet_features(root: ET.Element, boundary: int) -> None:
    rejected = {_Q(name) for name in ("dataValidations", "tableParts", "extLst", "legacyDrawingHF")}
    if any(node.tag in rejected for node in root.iter()):
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    for node in root.iter(_Q("mergeCell")):
        if "ref" in node.attrib and _crosses_boundary(node.attrib["ref"], boundary):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    for node in root.iter(_Q("conditionalFormatting")):
        if any(child.tag == _Q("formula") for child in node.iter()):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        if "sqref" not in node.attrib:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    for node in root.iter(_Q("hyperlink")):
        reference = node.attrib.get("ref", "")
        if not reference or _range_touches_or_right(reference, boundary):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    auto_filter = root.find(_Q("autoFilter"))
    if auto_filter is not None:
        if any(child.tag not in {_Q("filterColumn")} for child in auto_filter):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        if not auto_filter.attrib.get("ref"):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    _shared_formula_topology(root, boundary, "PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    for node in root.iter(_Q("c")):
        formulas = node.findall(_Q("f"))
        if len(formulas) > 1:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        formula = formulas[0] if formulas else None
        if formula is not None and formula.attrib.get("t") not in {None, "shared"}:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        if formula is not None and formula.text and formula.attrib.get("t") != "shared":
            _translate_formula(formula.text, boundary)


def _shared_formula_topology(
    root: ET.Element, boundary: int, error_code: str
) -> tuple[tuple[str, str, tuple[tuple[str, tuple[tuple[str, str], ...], str | None], ...]], ...]:
    """Strictly parse unchanged shared groups and retain their exact formula topology."""

    try:
        formula_nodes = list(root.iter(_Q("f")))
        members_by_si: dict[int, list[tuple[str, int, int, ET.Element]]] = {}
        physical_cell_counts: dict[tuple[int, int], int] = {}
        direct_formula_count = 0
        for cell in root.iter(_Q("c")):
            coordinate = cell.attrib.get("r")
            if coordinate is not None:
                physical_column, physical_row = coordinate_from_string(coordinate)
                physical_key = (column_index_from_string(physical_column), physical_row)
                physical_cell_counts[physical_key] = physical_cell_counts.get(physical_key, 0) + 1
            formulas = cell.findall(_Q("f"))
            direct_formula_count += len(formulas)
            if len(formulas) > 1:
                raise ReconciliationPeriodError(error_code)
            if not formulas:
                continue
            formula = formulas[0]
            if list(formula):
                raise ReconciliationPeriodError(error_code)
            formula_type = formula.attrib.get("t")
            if formula_type != "shared":
                if "si" in formula.attrib or "ref" in formula.attrib:
                    raise ReconciliationPeriodError(error_code)
                continue
            if set(formula.attrib) - {"t", "si", "ref"}:
                raise ReconciliationPeriodError(error_code)
            si = formula.attrib.get("si")
            if si is None or not si.isascii() or not si.isdecimal() or coordinate is None:
                raise ReconciliationPeriodError(error_code)
            canonical_si = si.lstrip("0") or "0"
            if len(canonical_si) > 10 or (len(canonical_si) == 10 and canonical_si > "4294967295"):
                raise ReconciliationPeriodError(error_code)
            column, row = coordinate_from_string(coordinate)
            members_by_si.setdefault(int(canonical_si), []).append(
                (coordinate, column_index_from_string(column), row, formula)
            )
        if direct_formula_count != len(formula_nodes):
            raise ReconciliationPeriodError(error_code)

        topology = []
        covered: set[tuple[int, int]] = set()
        for si, members in members_by_si.items():
            anchors = [member for member in members if "ref" in member[3].attrib]
            if len(anchors) != 1:
                raise ReconciliationPeriodError(error_code)
            anchor = anchors[0]
            anchor_formula = anchor[3]
            if (
                set(anchor_formula.attrib) != {"t", "si", "ref"}
                or anchor_formula.text is None
                or not anchor_formula.text.strip()
            ):
                raise ReconciliationPeriodError(error_code)
            anchor_si = anchor_formula.attrib["si"]
            for member in members:
                formula = member[3]
                if member is anchor:
                    continue
                if (
                    set(formula.attrib) != {"t", "si"}
                    or formula.attrib["si"] != anchor_si
                    or formula.text is not None
                ):
                    raise ReconciliationPeriodError(error_code)

            reference = anchor_formula.attrib["ref"]
            if not _A1_RANGE.fullmatch(reference):
                raise ReconciliationPeriodError(error_code)
            start, end = [*reference.split(":", 1), reference][:2]
            start_column, top = coordinate_from_string(start.replace("$", ""))
            end_column, bottom = coordinate_from_string(end.replace("$", ""))
            left, right = (
                column_index_from_string(start_column),
                column_index_from_string(end_column),
            )
            if left > right or top > bottom:
                raise ReconciliationPeriodError(error_code)
            if right > boundary or any(
                column > boundary for _coordinate, column, _row, _f in members
            ):
                raise ReconciliationPeriodError(error_code)
            actual_members = {(column, row) for _coordinate, column, row, _f in members}
            expected_count = (right - left + 1) * (bottom - top + 1)
            if (
                anchor[1:3] != (left, top)
                or len(members) != len(actual_members)
                or len(members) != expected_count
                or any(physical_cell_counts.get(member) != 1 for member in actual_members)
                or any(
                    column < left or column > right or row < top or row > bottom
                    for column, row in actual_members
                )
            ):
                raise ReconciliationPeriodError(error_code)
            if covered & actual_members:
                raise ReconciliationPeriodError(error_code)
            covered.update(actual_members)
            if _translate_formula(anchor_formula.text, boundary) != anchor_formula.text:
                raise ReconciliationPeriodError(error_code)
            topology.append(
                (
                    str(si),
                    reference,
                    tuple(
                        sorted(
                            (
                                coordinate,
                                tuple(sorted(formula.attrib.items())),
                                formula.text,
                            )
                            for coordinate, _column, _row, formula in members
                        )
                    ),
                )
            )
        return tuple(sorted(topology))
    except ReconciliationPeriodError:
        raise
    except (TypeError, ValueError):
        raise ReconciliationPeriodError(error_code) from None


def _verify_shared_formula_topology(
    root: ET.Element, boundary: int
) -> tuple[tuple[str, str, tuple[tuple[str, tuple[tuple[str, str], ...], str | None], ...]], ...]:
    """Verifier-local shared-formula parser; never call the forward preflight helper."""

    try:
        groups: dict[int, list[tuple[str, int, int, ET.Element]]] = {}
        physical_cell_counts: dict[tuple[int, int], int] = {}
        direct_formula_count = 0
        for cell in root.iter(_Q("c")):
            coordinate = cell.attrib.get("r")
            if coordinate is not None:
                physical_column, physical_row = coordinate_from_string(coordinate)
                physical_key = (column_index_from_string(physical_column), physical_row)
                physical_cell_counts[physical_key] = physical_cell_counts.get(physical_key, 0) + 1
            formulas = cell.findall(_Q("f"))
            direct_formula_count += len(formulas)
            if len(formulas) > 1:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            if not formulas:
                continue
            formula = formulas[0]
            if list(formula):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            formula_type = formula.attrib.get("t")
            if formula_type not in {None, "shared"}:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            if formula_type is None:
                if "si" in formula.attrib or "ref" in formula.attrib:
                    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
                continue
            si = formula.attrib.get("si")
            if (
                set(formula.attrib) - {"t", "si", "ref"}
                or si is None
                or not si.isascii()
                or not si.isdecimal()
                or coordinate is None
            ):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            canonical_si = si.lstrip("0") or "0"
            if len(canonical_si) > 10 or (len(canonical_si) == 10 and canonical_si > "4294967295"):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            column, row = coordinate_from_string(coordinate)
            groups.setdefault(int(canonical_si), []).append(
                (coordinate, column_index_from_string(column), row, formula)
            )
        if direct_formula_count != sum(1 for _formula in root.iter(_Q("f"))):
            raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")

        topology = []
        occupied: set[tuple[int, int]] = set()
        for numeric_si, members in groups.items():
            anchors = tuple(member for member in members if member[3].attrib.get("ref") is not None)
            if len(anchors) != 1:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            _anchor_coordinate, anchor_column, anchor_row, anchor_formula = anchors[0]
            anchor_si = anchor_formula.attrib.get("si")
            anchor_text = anchor_formula.text
            if (
                set(anchor_formula.attrib) != {"t", "si", "ref"}
                or anchor_si is None
                or anchor_text is None
                or not anchor_text.strip()
            ):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            for _coordinate, _column, _row, formula in members:
                if formula is anchor_formula:
                    continue
                if (
                    set(formula.attrib) != {"t", "si"}
                    or formula.attrib.get("si") != anchor_si
                    or formula.text is not None
                ):
                    raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")

            reference = anchor_formula.attrib["ref"]
            if not _A1_RANGE.fullmatch(reference):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            if ":" in reference:
                first, last = reference.split(":", 1)
            else:
                first = last = reference
            first_column, first_row = coordinate_from_string(first.replace("$", ""))
            last_column, last_row = coordinate_from_string(last.replace("$", ""))
            minimum_column = column_index_from_string(first_column)
            maximum_column = column_index_from_string(last_column)
            if minimum_column > maximum_column or first_row > last_row:
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            coordinates = {(column, row) for _coord, column, row, _formula in members}
            area = (maximum_column - minimum_column + 1) * (last_row - first_row + 1)
            if (
                (anchor_column, anchor_row) != (minimum_column, first_row)
                or len(coordinates) != len(members)
                or len(coordinates) != area
                or any(physical_cell_counts.get(member) != 1 for member in coordinates)
                or any(
                    column < minimum_column
                    or column > maximum_column
                    or row < first_row
                    or row > last_row
                    for column, row in coordinates
                )
                or maximum_column > boundary
                or any(column > boundary for column, _row in coordinates)
                or occupied.intersection(coordinates)
                or not _verify_shared_formula_is_noop(anchor_text, boundary)
            ):
                raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID")
            occupied.update(coordinates)
            topology.append(
                (
                    str(numeric_si),
                    reference,
                    tuple(
                        sorted(
                            (
                                coordinate,
                                tuple(sorted(formula.attrib.items())),
                                formula.text,
                            )
                            for coordinate, _column, _row, formula in members
                        )
                    ),
                )
            )
        return tuple(sorted(topology))
    except ReconciliationPeriodError:
        raise
    except (TypeError, ValueError):
        raise ReconciliationPeriodError("PERIOD_INSERTION_DELTA_INVALID") from None


def _verify_shared_formula_is_noop(value: str, boundary: int) -> bool:
    """Verifier-local operand scan for an unchanged wholly-left shared anchor."""

    if not value.strip() or "!" in value or "#" in value or "@" in value:
        return False
    unquoted = _strip_quoted(value)
    if (
        _UNSUPPORTED_FORMULA.search(unquoted)
        or re.search(
            r"(?<![A-Za-z0-9_.])\$?[A-Z]{1,3}:\$?[A-Z]{1,3}(?![A-Za-z0-9_.])", unquoted, re.I
        )
        or re.search(r"(?<![A-Za-z0-9_.])\$?\d+:\$?\d+(?![A-Za-z0-9_.])", unquoted)
    ):
        return False
    index = 0
    while index < len(value):
        if value[index] == '"':
            index += 1
            while index < len(value):
                if value[index] == '"':
                    index += 2 if index + 1 < len(value) and value[index + 1] == '"' else 1
                    break
                index += 1
            else:
                return False
            continue
        match = _A1_TOKEN.match(value, index)
        if match and _formula_token_boundary(value, match.start(), match.end()):
            for operand in match.group(0).split(":"):
                column, _row = coordinate_from_string(operand.replace("$", ""))
                if column_index_from_string(column) > boundary:
                    return False
            index = match.end()
            continue
        index += 1
    without_operands = _A1_TOKEN.sub("", unquoted)
    for token in re.finditer(
        r"(?<![A-Z0-9_])[A-Z_][A-Z0-9_.]*(?![A-Z0-9_])", without_operands, re.I
    ):
        if not without_operands[token.end() :].lstrip().startswith("(") and not re.fullmatch(
            r"[A-Z]{1,3}", token.group(), re.I
        ):
            return False
    return True


def _validate_sheet_relationships(
    archive: zipfile.ZipFile, anchor: ReconciliationSheetAnchor, root: ET.Element
) -> None:
    """Permit only wholly-left comments/VML and external hyperlinks unchanged."""
    rels = _part_relationships(archive, anchor.worksheet_part)
    relationship_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    hyperlinks = {node.attrib.get(relationship_ns): node for node in root.iter(_Q("hyperlink"))}
    for identifier, _target, kind, mode in rels:
        if mode == "External":
            hyperlink = hyperlinks.get(identifier)
            if hyperlink is None or _range_touches_or_right(
                hyperlink.attrib.get("ref", ""), anchor.insertion_after_column
            ):
                raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        elif kind.endswith("/hyperlink"):
            # Internal hyperlinks are deliberately unsupported: their target
            # semantics are not a package part and they can encode local refs.
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    legacy = list(root.iter(_Q("legacyDrawing")))
    comment_rels = [
        (identifier, target)
        for identifier, target, kind, mode in rels
        if kind.endswith("/comments") and mode is None
    ]
    vml_rels = [
        (identifier, target)
        for identifier, target, kind, mode in rels
        if kind.endswith("/vmlDrawing") and mode is None
    ]
    if not legacy and not comment_rels and not vml_rels:
        return
    if len(legacy) != 1 or len(comment_rels) != 1 or len(vml_rels) != 1:
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    legacy_id = legacy[0].attrib.get(relationship_ns)
    if legacy_id != vml_rels[0][0] or comment_rels[0][1] is None or vml_rels[0][1] is None:
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    comments_part, vml_part = comment_rels[0][1], vml_rels[0][1]
    if comments_part not in archive.namelist() or vml_part not in archive.namelist():
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID")
    try:
        comments = ET.fromstring(archive.read(comments_part))
    except ET.ParseError as error:
        raise ReconciliationPeriodError("PERIOD_INSERTION_PACKAGE_INVALID") from error
    for comment in comments.iter(_Q("comment")):
        reference = comment.attrib.get("ref")
        if not reference or _range_touches_or_right(reference, anchor.insertion_after_column):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")


def _require_insertible_rows(root: ET.Element, anchor: ReconciliationSheetAnchor) -> None:
    """Every materialised row receives two blank styled cells; no ghost rows."""
    rows = root.find(_Q("sheetData"))
    if rows is None:
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    required = {anchor.quantity_leaf_row, anchor.cost_leaf_row, anchor.first_detail_row}
    present = {int(row.attrib.get("r", "0")) for row in rows.findall(_Q("row"))}
    if not required <= present:
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")


def _preflight_workbook(
    archive: zipfile.ZipFile, anchors: tuple[ReconciliationSheetAnchor, ...]
) -> None:
    boundaries = {anchor.sheet_name: anchor.insertion_after_column for anchor in anchors}
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    for name in workbook.iter(_Q("definedName")):
        if _defined_name_is_affected(name, boundaries):
            if name.attrib.get("name") not in _PERMITTED_DEFINED_NAMES:
                raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
            _translate_defined_name(name.text or "", boundaries)
    if "xl/calcChain.xml" in archive.namelist():
        _validate_calc_chain(
            archive.read("xl/calcChain.xml"), _sheet_name_by_id(Path(archive.filename)), boundaries
        )


def _historical_parent_row(sheet, anchor: ReconciliationSheetAnchor) -> int:
    matches = []
    for merged in sheet.merged_cells.ranges:
        if merged.min_col == anchor.quantity_column and merged.max_col == anchor.cost_column:
            value = sheet.cell(merged.min_row, merged.min_col).value
            if value and any(
                token in str(value).casefold()
                for token in ("истор", "документ", "накоп", "прошл", "весь период")
            ):
                matches.append(merged.min_row)
    if len(matches) != 1:
        raise ReconciliationPeriodError("PERIOD_INSERTION_ANCHOR_INVALID")
    return matches[0]


def _historical_parent_xml_row(root: ET.Element, anchor: ReconciliationSheetAnchor) -> int:
    parent = anchor.parent_span
    for node in root.iter(_Q("mergeCell")):
        ref = node.attrib.get("ref", "")
        try:
            left, top, right, bottom = range_boundaries(ref)
        except ValueError:
            continue
        if (top, left, bottom, right) == parent:
            return top
    if parent[1] != anchor.quantity_column or parent[3] != anchor.cost_column:
        raise ReconciliationPeriodError("PERIOD_INSERTION_ANCHOR_INVALID")
    raise ReconciliationPeriodError("PERIOD_INSERTION_ANCHOR_INVALID")


def _pair_period_conflict(pair, period: ReportingPeriod) -> bool:
    quantity = calendar_identities(pair.quantity_header)
    cost = calendar_identities(pair.cost_header)
    evidence = quantity | cost
    return any(month != period.month or year not in {None, period.year} for month, year in evidence)


def _map_coordinate(coordinate: str, boundary: int) -> str:
    column, row = coordinate_from_string(coordinate.replace("$", ""))
    index = column_index_from_string(column)
    prefix = "$" if coordinate.startswith("$") else ""
    row_prefix = "$" if "$" in coordinate[len(prefix) + len(column) :] else ""
    return f"{prefix}{get_column_letter(index + 2 if index > boundary else index)}{row_prefix}{row}"


def _unmap_coordinate(coordinate: str, boundary: int) -> str:
    column, row = coordinate_from_string(coordinate.replace("$", ""))
    index = column_index_from_string(column)
    return f"{get_column_letter(index - 2 if index > boundary + 2 else index)}{row}"


def _map_range(value: str, boundary: int) -> str:
    if " " in value:
        return " ".join(_map_range(item, boundary) for item in value.split())
    if ":" not in value:
        return _map_coordinate(value, boundary)
    left, top, right, bottom = range_boundaries(value.replace("$", ""))
    if left <= boundary < right:
        right += 2
    elif left > boundary:
        left += 2
        right += 2
    return f"{get_column_letter(left)}{top}:{get_column_letter(right)}{bottom}"


def _map_sqref(value: str, boundary: int) -> str:
    items = value.split()
    if not items:
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    return " ".join(_map_range(item, boundary) for item in items)


def _range_touches_or_right(value: str, boundary: int) -> bool:
    left, _top, right, _bottom = range_boundaries(value.replace("$", ""))
    return right > boundary or left > boundary


def _map_filter_colid(value: int, boundary: int, original_ref: str | None) -> int:
    if original_ref is None:
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    left, _top, _right, _bottom = range_boundaries(original_ref.replace("$", ""))
    absolute = left + value
    return value + 2 if absolute > boundary else value


def _crosses_boundary(value: str, boundary: int) -> bool:
    left, _top, right, _bottom = range_boundaries(value.replace("$", ""))
    return left <= boundary < right


def _translate_formula(formula: str, boundary: int) -> str:
    if _UNSUPPORTED_FORMULA.search(formula) or "!" in formula or "#" in formula or "@" in formula:
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    unquoted_source = _strip_quoted(formula)
    if re.search(
        r"(?<![A-Za-z0-9_.])\$?[A-Z]{1,3}:\$?[A-Z]{1,3}(?![A-Za-z0-9_.])",
        unquoted_source,
        re.I,
    ) or re.search(r"(?<![A-Za-z0-9_.])\$?\d+:\$?\d+(?![A-Za-z0-9_.])", unquoted_source):
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    result: list[str] = []
    index = 0
    while index < len(formula):
        if formula[index] == '"':
            end = index + 1
            while end < len(formula):
                if formula[end] == '"':
                    if end + 1 < len(formula) and formula[end + 1] == '"':
                        end += 2
                        continue
                    end += 1
                    break
                else:
                    end += 1
            if end > len(formula) or (end == len(formula) and formula[end - 1] != '"'):
                raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
            result.append(formula[index:end])
            index = end
            continue
        match = _A1_TOKEN.match(formula, index)
        if match and _formula_token_boundary(formula, match.start(), match.end()):
            token = match.group(0)
            result.append(_translate_a1_operand(token, boundary))
            index = match.end()
            continue
        result.append(formula[index])
        index += 1
    # A bare identifier not followed by '(' is a named range.  We do not have
    # a safe name resolver in this structural layer.
    unquoted = _A1_TOKEN.sub("", _strip_quoted("".join(result)))
    for token in re.finditer(r"(?<![A-Z0-9_])[A-Z_][A-Z0-9_.]*(?![A-Z0-9_])", unquoted, re.I):
        tail = unquoted[token.end() :].lstrip()
        if not tail.startswith("(") and not re.fullmatch(r"[A-Z]{1,3}", token.group(), re.I):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    return "".join(result)


def _strip_quoted(formula: str) -> str:
    return re.sub(r'"(?:[^"]|"")*"', '""', formula)


def _formula_token_boundary(value: str, start: int, end: int) -> bool:
    before = value[start - 1] if start else ""
    after = value[end] if end < len(value) else ""
    return (
        not before
        or before not in "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_."
    ) and (
        not after or after not in "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_."
    )


def _translate_a1_operand(value: str, boundary: int) -> str:
    if ":" not in value:
        return _map_coordinate(value, boundary)
    left, right = value.split(":", 1)
    # Preserve the independent absolute markers rather than constructing a
    # normalised reference from range_boundaries.
    return f"{_map_coordinate(left, boundary)}:{_map_coordinate(right, boundary)}"


def _map_col(node: ET.Element, boundary: int) -> None:
    minimum, maximum = int(node.attrib["min"]), int(node.attrib["max"])
    if minimum <= boundary < maximum:
        node.attrib["max"] = str(maximum + 2)
    elif minimum > boundary:
        node.attrib["min"], node.attrib["max"] = str(minimum + 2), str(maximum + 2)


def _translate_cols(root: ET.Element, anchor: ReconciliationSheetAnchor) -> None:
    cols = root.find(_Q("cols"))
    if cols is None:
        return
    boundary = anchor.insertion_after_column
    original = list(cols.findall(_Q("col")))
    replacement: list[ET.Element] = []
    quantity_definition: ET.Element | None = None
    cost_definition: ET.Element | None = None
    for node in original:
        try:
            minimum, maximum = int(node.attrib["min"]), int(node.attrib["max"])
        except (KeyError, ValueError) as error:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE") from error
        if minimum > maximum:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        if minimum <= anchor.quantity_column <= maximum:
            quantity_definition = node
        if minimum <= anchor.cost_column <= maximum:
            cost_definition = node
        # Disjoint fragments maintain exact attributes.  A definition which
        # spans the insertion is split so new columns do not inherit a broad
        # definition accidentally.
        for low, high, offset in (
            (minimum, min(maximum, boundary), 0),
            (max(minimum, boundary + 1), maximum, 2),
        ):
            if low <= high:
                clone = ET.Element(_Q("col"), dict(node.attrib))
                clone.attrib["min"], clone.attrib["max"] = str(low + offset), str(high + offset)
                replacement.append(clone)
    if quantity_definition is not None:
        replacement.append(_column_clone(quantity_definition, boundary + 1))
    if cost_definition is not None:
        replacement.append(_column_clone(cost_definition, boundary + 2))
    if not _nonoverlapping_cols(replacement):
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    cols[:] = sorted(replacement, key=lambda item: int(item.attrib["min"]))


def _column_clone(source: ET.Element, target: int) -> ET.Element:
    node = ET.Element(_Q("col"), dict(source.attrib))
    node.attrib["min"] = node.attrib["max"] = str(target)
    return node


def _nonoverlapping_cols(nodes: list[ET.Element]) -> bool:
    previous = 0
    for node in sorted(nodes, key=lambda item: int(item.attrib["min"])):
        minimum, maximum = int(node.attrib["min"]), int(node.attrib["max"])
        if minimum <= previous or minimum > maximum:
            return False
        previous = maximum
    return True


def _translate_row_spans(row: ET.Element, boundary: int) -> None:
    spans = row.attrib.get("spans")
    if spans is None:
        return
    pieces: list[str] = []
    for item in spans.split():
        try:
            left, right = (int(value) for value in item.split(":", 1))
        except (TypeError, ValueError):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE") from None
        if left <= boundary < right:
            right += 2
        elif left > boundary:
            left += 2
            right += 2
        pieces.append(f"{left}:{right}")
    row.attrib["spans"] = " ".join(pieces)


def _inline_string(cell: ET.Element, text: str) -> None:
    cell.attrib["t"] = "inlineStr"
    inline = ET.SubElement(cell, _Q("is"))
    ET.SubElement(inline, _Q("t")).text = text


def _cell_column(coordinate: str) -> int:
    return column_index_from_string(coordinate_from_string(coordinate)[0])


def _transform_defined_names(
    payload: bytes, anchors: dict[str, ReconciliationSheetAnchor]
) -> bytes:
    root = ET.fromstring(payload)
    boundaries = {name: anchor.insertion_after_column for name, anchor in anchors.items()}
    for node in root.iter(_Q("definedName")):
        if _defined_name_is_affected(node, boundaries):
            node.text = _translate_defined_name(node.text or "", boundaries)
    return _serialize_preserving_ignorable_namespaces(root, payload)


def _transform_calc_chain(
    payload: bytes,
    sheet_names: dict[int, str],
    anchors: dict[str, ReconciliationSheetAnchor],
) -> bytes:
    root = ET.fromstring(payload)
    active_index: int | None = None
    for node in root.iter(_Q("c")):
        if "i" in node.attrib:
            try:
                active_index = int(node.attrib["i"])
            except ValueError as error:
                raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE") from error
        if active_index is None or active_index not in sheet_names:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        anchor = anchors.get(sheet_names[active_index])
        if anchor is not None and "r" in node.attrib:
            node.attrib["r"] = _map_coordinate(node.attrib["r"], anchor.insertion_after_column)
    return _serialize_preserving_ignorable_namespaces(root, payload)


def _defined_name_is_affected(node: ET.Element, boundaries: dict[str, int]) -> bool:
    text = node.text or ""
    for sheet, boundary in boundaries.items():
        marker = _quote_sheet(sheet) + "!"
        if marker in text:
            reference = text.split(marker, 1)[1]
            try:
                if _defined_reference_touches(reference.split(",", 1)[0], boundary):
                    return True
            except ValueError:
                return True
    return False


def _defined_reference_touches(reference: str, boundary: int) -> bool:
    if re.fullmatch(r"\$?\d+:\$?\d+", reference):
        return False
    if re.fullmatch(r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}", reference):
        left, right = (column_index_from_string(item.lstrip("$")) for item in reference.split(":"))
        return right > boundary or left > boundary
    return _range_touches_or_right(reference, boundary)


def _quote_sheet(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def _translate_defined_name(text: str, boundaries: dict[str, int]) -> str:
    if not text or "!" not in text:
        raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
    # The permitted built-ins are straightforward union lists.  Parse each
    # quoted-sheet term exactly; any expression syntax is intentionally out.
    terms = text.split(",")
    translated: list[str] = []
    for term in terms:
        if "!" not in term:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        sheet, reference = term.rsplit("!", 1)
        unquoted = (
            sheet[1:-1].replace("''", "'")
            if sheet.startswith("'") and sheet.endswith("'")
            else sheet
        )
        boundary = boundaries.get(unquoted)
        if boundary is None:
            translated.append(term)
            continue
        if not re.fullmatch(
            r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}|\$?\d+:\$?\d+|\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+",
            reference,
        ):
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        if re.fullmatch(r"\$?\d+:\$?\d+", reference):
            translated.append(term)
        elif re.fullmatch(r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}", reference):
            left, right = reference.split(":", 1)
            translated.append(
                f"{sheet}!{_map_column_token(left, boundary)}:{_map_column_token(right, boundary)}"
            )
        else:
            translated.append(f"{sheet}!{_translate_a1_operand(reference, boundary)}")
    return ",".join(translated)


def _map_column_token(value: str, boundary: int) -> str:
    absolute = "$" if value.startswith("$") else ""
    index = column_index_from_string(value.lstrip("$"))
    return absolute + get_column_letter(index + 2 if index > boundary else index)


def _validate_calc_chain(
    payload: bytes, sheet_names: dict[int, str], boundaries: dict[str, int]
) -> None:
    root = ET.fromstring(payload)
    current: int | None = None
    for node in root.iter(_Q("c")):
        if "i" in node.attrib:
            try:
                current = int(node.attrib["i"])
            except ValueError as error:
                raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE") from error
        if current is None or current not in sheet_names or "r" not in node.attrib:
            raise ReconciliationPeriodError("PERIOD_INSERTION_UNSUPPORTED_FEATURE")
        if sheet_names[current] in boundaries:
            _map_coordinate(node.attrib["r"], boundaries[sheet_names[current]])


def _calc_chain_is_affected(
    payload: bytes, sheet_names: dict[int, str], boundaries: dict[str, int]
) -> bool:
    root = ET.fromstring(payload)
    current: int | None = None
    for node in root.iter(_Q("c")):
        if "i" in node.attrib:
            current = int(node.attrib["i"])
        if current is None or current not in sheet_names:
            return True
        boundary = boundaries.get(sheet_names[current])
        if boundary is not None and _cell_column(node.attrib.get("r", "A1")) > boundary:
            return True
    return False


def _assert_digest(path: Path, expected: str) -> None:
    if _sha256(path) != expected:
        raise ReconciliationPeriodError("PERIOD_INSERTION_SOURCE_CHANGED")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _serialize_preserving_ignorable_namespaces(root: ET.Element, source: bytes) -> bytes:
    """Keep namespace bindings referenced only by ``mc:Ignorable`` valid.

    ElementTree discards declarations that are only lexical tokens in
    ``mc:Ignorable``.  Reinsert original declarations into the root start tag
    when serialization did not retain them; this is a namespace-validity guard,
    not a semantic rewrite of any workbook node.
    """
    serialized = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    root_open = serialized.find(b"<", serialized.find(b"?>") + 2)
    start = serialized.find(b">", root_open)
    source_open = source.find(b"<", source.find(b"?>") + 2)
    source_end = source.find(b">", source_open)
    if start < 0 or source_end < 0:
        raise ReconciliationPeriodError("PERIOD_INSERTION_TRANSFORM_INVALID")
    original = _XMLNS.findall(source[source_open : source_end + 1])
    missing = [
        declaration
        for declaration in original
        if declaration not in serialized[root_open : start + 1]
    ]
    if missing:
        serialized = (
            serialized[:start] + b"".join(b" " + item for item in missing) + serialized[start:]
        )
    return serialized


def _fsync_file(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _fsync_directory(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY | getattr(os, "O_DIRECTORY", 0))
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)

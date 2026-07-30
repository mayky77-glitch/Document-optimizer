from __future__ import annotations

from dataclasses import replace

from openpyxl.utils.cell import get_column_letter

from report_processor.excel import DualWorkbookSession
from report_processor.schema import LogicalColumn, SheetType, WorksheetSchema

from .config import ExtractionConfig
from .exceptions import ExtractionSchemaError
from .models import ExtractionPlan

_REQUIRED_BY_TYPE: dict[SheetType, tuple[LogicalColumn, ...]] = {
    SheetType.KS2: (LogicalColumn.WORK_NAME,),
    SheetType.KS6A: (LogicalColumn.WORK_NAME,),
    SheetType.SVVR: (
        LogicalColumn.WORK_NAME,
        LogicalColumn.CURRENT_PERIOD_QUANTITY,
    ),
}

_OPTIONAL_BY_TYPE: dict[SheetType, tuple[LogicalColumn, ...]] = {
    SheetType.KS2: (
        LogicalColumn.POSITION_CODE,
        LogicalColumn.UNIT,
        LogicalColumn.CURRENT_PERIOD_QUANTITY,
        LogicalColumn.UNIT_PRICE,
        LogicalColumn.CURRENT_PERIOD_COST,
    ),
    SheetType.KS6A: (
        LogicalColumn.UNIT,
        LogicalColumn.CURRENT_PERIOD_QUANTITY,
        LogicalColumn.CUMULATIVE_QUANTITY,
        LogicalColumn.CURRENT_PERIOD_COST,
        LogicalColumn.CUMULATIVE_COST,
    ),
    SheetType.SVVR: (
        LogicalColumn.OBJECT_CODE,
        LogicalColumn.SUBOBJECT_CODE,
        LogicalColumn.POSITION_CODE,
        LogicalColumn.UNIT,
    ),
}


def _has_required_column(
    present: set[LogicalColumn],
    required: LogicalColumn,
) -> bool:
    return required in present


def build_extraction_plan(
    session: DualWorkbookSession,
    schema: WorksheetSchema,
    config: ExtractionConfig,
) -> ExtractionPlan:
    if schema.sheet_name not in session.formula_workbook.sheetnames:
        raise ExtractionSchemaError(f"Лист отсутствует в workbook: {schema.sheet_name}")
    if schema.sheet_name not in session.values_workbook.sheetnames:
        raise ExtractionSchemaError(f"Лист отсутствует в data_only workbook: {schema.sheet_name}")
    if schema.data_start_row < 1:
        raise ExtractionSchemaError("data_start_row должен быть >= 1")
    if not schema.columns:
        raise ExtractionSchemaError("Схема листа не содержит разрешённых столбцов")
    if schema.status.upper() == "SCHEMA_INVALID":
        raise ExtractionSchemaError("WorksheetSchema имеет статус SCHEMA_INVALID")

    warnings: list[str] = []
    if schema.status.upper() != "OK":
        warnings.append(f"SCHEMA_STATUS:{schema.status}")
    seen_columns: set[int] = set()
    seen_logical: set[LogicalColumn] = set()
    normalized_columns = []
    for column in schema.columns:
        if column.column_index is None or not 1 <= column.column_index <= 16_384:
            raise ExtractionSchemaError(
                f"Недопустимый физический столбец {column.column_index} "
                f"для {column.logical_column.value}"
            )
        if column.column_index in seen_columns:
            warnings.append(f"DUPLICATE_PHYSICAL_COLUMN:{column.column_index}")
        if column.logical_column in seen_logical:
            warnings.append(f"DUPLICATE_LOGICAL_COLUMN:{column.logical_column.value}")
        seen_columns.add(column.column_index)
        seen_logical.add(column.logical_column)
        expected_letter = get_column_letter(column.column_index)
        if column.column_letter.upper() != expected_letter:
            warnings.append(f"COLUMN_LETTER_CORRECTED:{column.column_letter}->{expected_letter}")
            column = replace(column, column_letter=expected_letter)
        if column.status.upper() != "OK":
            warnings.append(f"COLUMN_STATUS:{column.logical_column.value}:{column.status}")
        warnings.extend(
            f"COLUMN_WARNING:{column.logical_column.value}:{item}" for item in column.warnings
        )
        normalized_columns.append(column)

    required = _REQUIRED_BY_TYPE.get(schema.sheet_type, ())
    optional = _OPTIONAL_BY_TYPE.get(schema.sheet_type, ())
    missing = tuple(item for item in required if not _has_required_column(seen_logical, item))
    if missing:
        names = ",".join(item.value for item in missing)
        raise ExtractionSchemaError(f"REQUIRED_COLUMNS_MISSING:{names}")

    worksheet = session.formula_workbook[schema.sheet_name]
    reported_max_raw = getattr(worksheet, "max_row", None)
    reported_max_column = getattr(worksheet, "max_column", None)
    if reported_max_column is not None:
        for column in normalized_columns:
            if column.column_index > int(reported_max_column):
                warnings.append(
                    "COLUMN_OUTSIDE_REPORTED_RANGE:"
                    f"{column.logical_column.value}:{column.column_index}>"
                    f"{reported_max_column}"
                )
    hard_limit_end = schema.data_start_row + config.max_rows - 1
    if reported_max_raw is None:
        max_end_row = hard_limit_end
        warnings.append("REPORTED_MAX_ROW_UNKNOWN")
    else:
        reported_max = int(reported_max_raw)
        max_end_row = min(max(reported_max, schema.data_start_row), hard_limit_end)
        if reported_max > hard_limit_end:
            warnings.append(f"REPORTED_MAX_ROW_CAPPED:{reported_max}->{hard_limit_end}")

    return ExtractionPlan(
        sheet_name=schema.sheet_name,
        sheet_type=schema.sheet_type,
        data_start_row=schema.data_start_row,
        max_end_row=max_end_row,
        columns=tuple(normalized_columns),
        required_columns=tuple(required),
        optional_columns=tuple(optional),
        warnings=tuple((*schema.warnings, *warnings)),
    )

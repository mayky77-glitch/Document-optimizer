"""Merged-cell metadata without mutating or expanding worksheet cells."""

from __future__ import annotations

import xml.etree.ElementTree as ET
from collections.abc import Mapping
from typing import Any

from openpyxl.utils.cell import get_column_letter, range_boundaries

from report_processor.schema.models import MergedRangeInfo


def _range_info(ref: str, anchor_value: Any = None) -> MergedRangeInfo:
    min_column, min_row, max_column, max_row = range_boundaries(ref)
    anchor = f"{get_column_letter(min_column)}{min_row}"
    return MergedRangeInfo(
        range_string=ref,
        min_row=min_row,
        max_row=max_row,
        min_column=min_column,
        max_column=max_column,
        anchor_coordinate=anchor,
        anchor_value=anchor_value,
    )


def _read_readonly_merge_refs(worksheet: Any) -> tuple[str, ...]:
    workbook = worksheet.parent
    archive = getattr(workbook, "_archive", None)
    worksheet_path = getattr(worksheet, "_worksheet_path", None)
    if archive is None or worksheet_path is None:
        return ()
    refs: list[str] = []
    with archive.open(worksheet_path) as stream:
        for _event, element in ET.iterparse(stream, events=("end",)):
            if element.tag.endswith("mergeCell"):
                ref = element.attrib.get("ref")
                if ref:
                    refs.append(ref)
            if element.tag.endswith("mergeCells"):
                element.clear()
                break
            element.clear()
    return tuple(refs)


def collect_merged_range_geometries(
    worksheet: Any,
    *,
    max_row: int,
    max_column: int,
) -> tuple[MergedRangeInfo, ...]:
    if hasattr(worksheet, "merged_cells"):
        refs = tuple(str(item) for item in worksheet.merged_cells.ranges)
    else:
        refs = _read_readonly_merge_refs(worksheet)
    result: list[MergedRangeInfo] = []
    for ref in refs:
        info = _range_info(ref)
        intersects = (
            info.min_row <= max_row
            and info.min_column <= max_column
            and info.max_row >= 1
            and info.max_column >= 1
        )
        if intersects:
            result.append(info)
    return tuple(result)


def attach_merged_anchor_values(
    ranges: tuple[MergedRangeInfo, ...],
    anchor_values: Mapping[str, Any],
) -> tuple[MergedRangeInfo, ...]:
    return tuple(
        MergedRangeInfo(
            range_string=item.range_string,
            min_row=item.min_row,
            max_row=item.max_row,
            min_column=item.min_column,
            max_column=item.max_column,
            anchor_coordinate=item.anchor_coordinate,
            anchor_value=anchor_values.get(item.anchor_coordinate),
        )
        for item in ranges
    )


def collect_relevant_merged_ranges(
    worksheet: Any,
    *,
    max_row: int,
    max_column: int,
) -> tuple[MergedRangeInfo, ...]:
    geometries = collect_merged_range_geometries(
        worksheet,
        max_row=max_row,
        max_column=max_column,
    )
    values = {
        item.anchor_coordinate: worksheet[item.anchor_coordinate].value for item in geometries
    }
    return attach_merged_anchor_values(geometries, values)


def merged_range_for_cell(
    ranges: tuple[MergedRangeInfo, ...],
    row: int,
    column: int,
) -> MergedRangeInfo | None:
    return next(
        (
            item
            for item in ranges
            if item.min_row <= row <= item.max_row and item.min_column <= column <= item.max_column
        ),
        None,
    )

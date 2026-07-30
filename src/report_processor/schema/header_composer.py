"""Compose logical column headers from bounded rows and virtual merged parents."""

from __future__ import annotations

from openpyxl.utils.cell import get_column_letter

from report_processor.schema.merged_cells import merged_range_for_cell
from report_processor.schema.models import ComposedHeader, MergedRangeInfo, WorksheetScanWindow
from report_processor.schema.text_normalization import clean_display_text, normalize_header_text


def _usable_header_part(value: object) -> str:
    text = clean_display_text(value)
    if not text:
        return ""
    if text.replace(".", "", 1).isdigit():
        return ""
    return text


def compose_logical_headers(
    scan: WorksheetScanWindow,
    merged_ranges: tuple[MergedRangeInfo, ...],
    *,
    start_row: int,
    end_row: int,
) -> tuple[ComposedHeader, ...]:
    direct = {(cell.row, cell.column): cell for cell in scan.cells}
    candidate_columns = {cell.column for cell in scan.cells if start_row <= cell.row <= end_row}
    for merged in merged_ranges:
        if merged.min_row <= end_row and merged.max_row >= start_row:
            upper = min(merged.max_column, scan.max_scanned_column)
            candidate_columns.update(range(merged.min_column, upper + 1))
    if not candidate_columns:
        return ()

    headers: list[ComposedHeader] = []
    for column in range(min(candidate_columns), max(candidate_columns) + 1):
        parts: list[str] = []
        normalized_parts: set[str] = set()
        sources: list[str] = []
        merged_sources: list[str] = []
        for row in range(start_row, end_row + 1):
            cell = direct.get((row, column))
            value = cell.raw_value if cell is not None else None
            source_coordinate = cell.coordinate if cell is not None else None
            merged = merged_range_for_cell(merged_ranges, row, column)
            if (value is None or not _usable_header_part(value)) and merged is not None:
                value = merged.anchor_value
                source_coordinate = merged.anchor_coordinate
                merged_sources.append(merged.range_string)
            part = _usable_header_part(value)
            normalized = normalize_header_text(part)
            if not part or not normalized or normalized in normalized_parts:
                continue
            parts.append(part)
            normalized_parts.add(normalized)
            if source_coordinate:
                sources.append(source_coordinate)
        raw_text = " ".join(parts)
        headers.append(
            ComposedHeader(
                column_index=column,
                column_letter=get_column_letter(column),
                parts=tuple(parts),
                raw_text=raw_text,
                normalized_text=normalize_header_text(raw_text),
                is_empty=not bool(parts),
                source_coordinates=tuple(dict.fromkeys(sources)),
                merged_sources=tuple(dict.fromkeys(merged_sources)),
            )
        )
    return tuple(headers)

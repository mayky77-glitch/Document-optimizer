"""Generate and score one- to four-row header candidates."""

from __future__ import annotations

from datetime import date, datetime

from openpyxl.utils.cell import get_column_letter, range_boundaries

from report_processor.schema.config import HeaderDetectionConfig
from report_processor.schema.header_composer import compose_logical_headers
from report_processor.schema.models import (
    HeaderCandidate,
    MergedRangeInfo,
    SheetClassification,
    WorksheetScanWindow,
)

_ALIAS_PHRASES = (
    "наименование",
    "количество",
    "единица измерения",
    "стоимость",
    "цена",
    "объект",
    "подобъект",
    "позиция",
    "этап",
    "индекс",
    "обоснование",
    "шифр чертежа",
)
_TITLE_MARKERS = (
    "организация",
    "заказчик",
    "подрядчик",
    "утверждаю",
    "согласовано",
    "доверенность",
    "акт о приемке выполненных работ",
    "справка о стоимости выполненных работ",
)


def _merged_infos_from_scan(scan: WorksheetScanWindow) -> tuple[MergedRangeInfo, ...]:
    anchors = {cell.coordinate: cell.raw_value for cell in scan.cells if cell.is_merged_anchor}
    result: list[MergedRangeInfo] = []
    for ref in scan.merged_ranges:
        min_column, min_row, max_column, max_row = range_boundaries(ref)
        anchor = f"{get_column_letter(min_column)}{min_row}"
        result.append(
            MergedRangeInfo(
                ref, min_row, max_row, min_column, max_column, anchor, anchors.get(anchor)
            )
        )
    return tuple(result)


def _is_numeric(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _next_row_data_score(scan: WorksheetScanWindow, end_row: int) -> float:
    for row in range(end_row + 1, min(end_row + 4, scan.max_scanned_row) + 1):
        cells = [cell for cell in scan.cells if cell.row == row and not cell.is_empty]
        if not cells:
            continue
        numeric = sum(_is_numeric(cell.raw_value) for cell in cells)
        textual = sum(isinstance(cell.raw_value, str) and not cell.is_formula for cell in cells)
        if numeric and textual:
            return 1.0
        if numeric >= 2:
            return 0.65
        if textual >= 2:
            return 0.35
        return 0.0
    return 0.0


def _candidate_for_range(
    scan: WorksheetScanWindow,
    merged_ranges: tuple[MergedRangeInfo, ...],
    start_row: int,
    end_row: int,
    config: HeaderDetectionConfig,
) -> HeaderCandidate | None:
    range_cells = [
        cell for cell in scan.cells if start_row <= cell.row <= end_row and not cell.is_empty
    ]
    if not any(cell.row == start_row for cell in range_cells):
        return None
    if not any(cell.row == end_row for cell in range_cells):
        return None
    columns = {cell.column for cell in range_cells}
    headers = compose_logical_headers(
        scan,
        merged_ranges,
        start_row=start_row,
        end_row=end_row,
    )
    nonempty_headers = [header for header in headers if not header.is_empty]
    if len(nonempty_headers) < config.min_distinct_columns:
        return None

    text_count = sum(
        isinstance(cell.raw_value, str) and not cell.is_formula for cell in range_cells
    )
    numeric_count = sum(_is_numeric(cell.raw_value) for cell in range_cells)
    total = max(text_count + numeric_count, 1)
    text_ratio = text_count / total
    corpus = " | ".join(header.normalized_text for header in nonempty_headers)
    aliases = tuple(alias for alias in _ALIAS_PHRASES if alias in corpus)
    penalties: list[str] = []
    reasons: list[str] = []

    score = 0.08
    score += min(len(nonempty_headers) / 10, 0.24)
    score += min(len(aliases) / 7, 0.36)
    score += min(text_ratio, 1.0) * 0.16
    next_score = _next_row_data_score(scan, end_row)
    score += next_score * 0.12
    merged_hits = sum(
        item.min_row <= end_row and item.max_row >= start_row for item in merged_ranges
    )
    score += min(merged_hits, 2) * 0.035

    if text_ratio < config.min_text_ratio:
        score -= 0.18
        penalties.append("LOW_TEXT_RATIO")
    if numeric_count and not text_count:
        score -= 0.35
        penalties.append("NUMERIC_ONLY")
    if len(nonempty_headers) == 1 and len(nonempty_headers[0].normalized_text) > 45:
        score -= 0.28
        penalties.append("SINGLE_LONG_PHRASE")
    title_hits = tuple(marker for marker in _TITLE_MARKERS if marker in corpus)
    if title_hits and len(aliases) < 3:
        score -= min(0.12 * len(title_hits), 0.32)
        penalties.append("TITLE_LIKE")
    if "итого" in corpus and len(aliases) < 3:
        score -= 0.18
        penalties.append("TOTAL_LIKE")
    if any(isinstance(cell.raw_value, (date, datetime)) for cell in range_cells):
        score -= 0.08
        penalties.append("DATE_PRESENT")

    reasons.extend(
        (
            f"nonempty_columns={len(nonempty_headers)}",
            f"alias_hits={len(aliases)}",
            f"text_ratio={text_ratio:.3f}",
            f"next_data_score={next_score:.3f}",
        )
    )
    return HeaderCandidate(
        start_row=start_row,
        end_row=end_row,
        score=round(min(max(score, 0.0), 1.0), 4),
        nonempty_columns=len(columns),
        text_cell_count=text_count,
        numeric_cell_count=numeric_count,
        matched_aliases=aliases,
        penalties=tuple(penalties),
        reasons=tuple(reasons),
    )


def find_header_candidates(
    scan: WorksheetScanWindow,
    classification: SheetClassification,
    config: HeaderDetectionConfig,
) -> tuple[HeaderCandidate, ...]:
    del classification  # reserved for type-specific scoring without coupling algorithms
    merged_ranges = _merged_infos_from_scan(scan)
    max_start = min(config.max_header_start_row, scan.max_scanned_row)
    candidates: list[HeaderCandidate] = []
    for start_row in range(1, max_start + 1):
        for depth in range(1, config.max_header_depth + 1):
            end_row = start_row + depth - 1
            if end_row > scan.max_scanned_row:
                break
            candidate = _candidate_for_range(
                scan,
                merged_ranges,
                start_row,
                end_row,
                config,
            )
            if candidate is not None and candidate.score >= 0.34:
                candidates.append(candidate)
    candidates.sort(
        key=lambda item: (
            -item.score,
            -len(item.matched_aliases),
            item.start_row,
            item.end_row,
        )
    )
    return tuple(candidates[:20])

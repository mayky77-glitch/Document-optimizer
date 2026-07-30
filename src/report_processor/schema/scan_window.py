"""Bounded worksheet scanning that never materializes a full two-dimensional sheet."""

from __future__ import annotations

from openpyxl.utils.cell import get_column_letter

from report_processor.excel import DualWorkbookSession
from report_processor.schema.config import SheetScanConfig
from report_processor.schema.exceptions import WorksheetScanError
from report_processor.schema.merged_cells import (
    attach_merged_anchor_values,
    collect_merged_range_geometries,
)
from report_processor.schema.models import MergedRangeInfo, ScannedCell, WorksheetScanWindow
from report_processor.schema.text_normalization import normalize_header_text

_HEADER_TERMS = {
    "наименование",
    "количество",
    "стоимость",
    "единица",
    "объект",
    "позиция",
    "этап",
    "индекс",
}


def _is_formula(cell: object) -> bool:
    data_type = getattr(cell, "data_type", None)
    value = getattr(cell, "value", None)
    return data_type == "f" or (isinstance(value, str) and value.startswith("="))


def _row_is_structural(nonempty_values: list[object]) -> bool:
    if len(nonempty_values) >= 3:
        return True
    normalized = " ".join(normalize_header_text(value) for value in nonempty_values)
    return len(_HEADER_TERMS.intersection(normalized.split())) >= 2


def scan_worksheet_window(
    session: DualWorkbookSession,
    sheet_name: str,
    config: SheetScanConfig,
) -> WorksheetScanWindow:
    if session.closed:
        raise WorksheetScanError("Workbook-сессия уже закрыта")
    if sheet_name not in session.sheet_names:
        raise WorksheetScanError(f"Лист не найден: {sheet_name}")

    worksheet = session.formula_workbook[sheet_name]
    row_limit = min(config.max_scan_rows, max(int(worksheet.max_row or 1), 1))
    column_limit = min(config.max_scan_columns, max(int(worksheet.max_column or 1), 1))
    geometries = collect_merged_range_geometries(
        worksheet,
        max_row=row_limit,
        max_column=column_limit,
    )
    anchors = {item.anchor_coordinate: item.range_string for item in geometries}

    cells: list[ScannedCell] = []
    anchor_values: dict[str, object] = {}
    nonempty_count = 0
    warnings: list[str] = []
    empty_streak = 0
    structural_area_seen = False
    stopped_early = False
    max_scanned_row = 0

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=row_limit,
            min_col=1,
            max_col=column_limit,
        ),
        start=1,
    ):
        max_scanned_row = row_number
        row_nonempty: list[object] = []
        for column_number, cell in enumerate(row, start=1):
            value = cell.value
            coordinate = f"{get_column_letter(column_number)}{row_number}"
            merged_range = anchors.get(coordinate)
            if merged_range is not None:
                anchor_values[coordinate] = value
            if value is None and merged_range is None:
                continue
            is_formula = _is_formula(cell)
            normalized = None
            if value is not None and not is_formula:
                normalized = normalize_header_text(value) or None
            scanned = ScannedCell(
                row=row_number,
                column=column_number,
                coordinate=coordinate,
                raw_value=value,
                normalized_text=normalized,
                is_formula=is_formula,
                is_empty=value is None,
                is_merged_anchor=merged_range is not None,
                merged_range=merged_range,
            )
            cells.append(scanned)
            if value is not None:
                nonempty_count += 1
                row_nonempty.append(value)
            if nonempty_count >= config.max_nonempty_cells:
                warnings.append("SCAN_CELL_LIMIT_REACHED")
                stopped_early = True
                break
        if stopped_early:
            break
        if row_nonempty:
            empty_streak = 0
            structural_area_seen = structural_area_seen or _row_is_structural(row_nonempty)
        elif structural_area_seen:
            empty_streak += 1
            if empty_streak >= config.stop_after_empty_rows:
                warnings.append("SCAN_STOPPED_AFTER_EMPTY_ROWS")
                stopped_early = True
                break

    merged_ranges = attach_merged_anchor_values(geometries, anchor_values)
    session.structure_cache[f"merged:{sheet_name}"] = merged_ranges
    session.structure_cache[f"scan:{sheet_name}"] = tuple(cells)
    if geometries and not hasattr(worksheet, "merged_cells"):
        warnings.append("MERGED_RANGES_STREAMED_FROM_XML")
    return WorksheetScanWindow(
        sheet_name=sheet_name,
        max_scanned_row=max_scanned_row,
        max_scanned_column=column_limit,
        nonempty_cell_count=nonempty_count,
        cells=tuple(cells),
        merged_ranges=tuple(item.range_string for item in merged_ranges),
        stopped_early=stopped_early,
        warnings=tuple(dict.fromkeys(warnings)),
    )


def get_cached_merged_ranges(
    session: DualWorkbookSession,
    sheet_name: str,
) -> tuple[MergedRangeInfo, ...]:
    value = session.structure_cache.get(f"merged:{sheet_name}", ())
    return tuple(value)

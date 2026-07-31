"""Bounded, read-only period discovery for uploaded drawing-card workbooks."""

from __future__ import annotations

import os
import re
import tempfile
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .sources.readers import open_reader

MAX_PERIOD_SCAN_CELLS = 1_000_000
_MIN_YEAR = 2000
_MAX_YEAR = 2100
_FULL_DATE_RE = re.compile(
    r"(?<!\d)(?:0?[1-9]|[12]\d|3[01])[._/-](0?[1-9]|1[0-2])[._/-](20\d{2})(?!\d)"
)
_YEAR_MONTH_RE = re.compile(r"(?<!\d)(20\d{2})[._/-](0?[1-9]|1[0-2])(?!\d)")
_MONTH_YEAR_RE = re.compile(r"(?<!\d)(0?[1-9]|1[0-2])[._/-](20\d{2})(?!\d)")
_NAMED_MONTH_YEAR_RE = re.compile(
    r"(?<![а-я])"
    r"(янв|фев|мар|апр|май|июн|июл|авг|сен|окт|ноя|дек)"
    r"[а-я]*[\s._/-]+(20\d{2})(?!\d)"
)
_MONTH_PREFIXES = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}


def discover_workbook_periods(
    sources: Iterable[tuple[str, bytes]],
    *,
    temporary_root: Path,
) -> tuple[str, ...]:
    """Return every detected period without retaining uploads or workbook objects."""
    source_entries = tuple(sources)
    filename_periods: set[str] = set()
    for name, _content in source_entries:
        _add_filename_periods(filename_periods, name)
    periods = set(filename_periods)
    reference_years = frozenset(int(value[:4]) for value in filename_periods)
    with tempfile.TemporaryDirectory(prefix="period-scan-", dir=temporary_root) as directory:
        root = Path(directory)
        os.chmod(root, 0o700)
        for index, (name, content) in enumerate(source_entries, 1):
            path = root / f"{index:02d}{Path(name).suffix.casefold()}"
            path.write_bytes(content)
            os.chmod(path, 0o600)
            _scan_workbook(path, periods, reference_years)
    return tuple(sorted(periods))


def _scan_workbook(
    path: Path,
    periods: set[str],
    reference_years: frozenset[int],
) -> None:
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        _scan_openxml_workbook(path, periods, reference_years)
        return
    reader = None
    scanned = 0
    try:
        reader = open_reader(path)
        for sheet_name in reader.list_sheets():
            for _formula_row, cached_row in reader.iter_rows(sheet_name):
                for value in cached_row:
                    scanned += 1
                    if scanned > MAX_PERIOD_SCAN_CELLS:
                        return
                    _add_periods(periods, value, reference_years)
    except (KeyError, OSError, TypeError, ValueError):
        return
    finally:
        if reader is not None:
            reader.close()


def _scan_openxml_workbook(
    path: Path,
    periods: set[str],
    reference_years: frozenset[int],
) -> None:
    workbook = None
    scanned = 0
    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    scanned += 1
                    if scanned > MAX_PERIOD_SCAN_CELLS:
                        return
                    _add_periods(periods, value, reference_years)
    except (KeyError, OSError, TypeError, ValueError):
        return
    finally:
        if workbook is not None:
            workbook.close()


def _add_periods(
    periods: set[str],
    value: object,
    reference_years: frozenset[int],
) -> None:
    if isinstance(value, (date, datetime)):
        return
    if not isinstance(value, str) or not value.strip():
        return
    normalized = value.casefold().replace("ё", "е")
    without_full_dates = _FULL_DATE_RE.sub(" ", normalized)
    for year, month in _YEAR_MONTH_RE.findall(without_full_dates):
        _add_cell_period(periods, int(year), int(month), reference_years)
    for month, year in _MONTH_YEAR_RE.findall(without_full_dates):
        _add_cell_period(periods, int(year), int(month), reference_years)
    for prefix, year in _NAMED_MONTH_YEAR_RE.findall(without_full_dates):
        _add_cell_period(periods, int(year), _MONTH_PREFIXES[prefix], reference_years)


def _add_filename_periods(periods: set[str], name: str) -> None:
    normalized = name.casefold().replace("ё", "е")
    full_date = _FULL_DATE_RE.search(normalized)
    if full_date:
        _add_period(periods, int(full_date.group(2)), int(full_date.group(1)))
    year_month = _YEAR_MONTH_RE.search(normalized)
    if year_month:
        _add_period(periods, int(year_month.group(1)), int(year_month.group(2)))
    month_year = _MONTH_YEAR_RE.search(normalized)
    if month_year:
        _add_period(periods, int(month_year.group(2)), int(month_year.group(1)))
    year = re.search(r"20\d{2}", normalized)
    if year:
        for prefix, month in _MONTH_PREFIXES.items():
            if prefix in normalized:
                _add_period(periods, int(year.group()), month)


def _add_period(periods: set[str], year: int, month: int) -> None:
    if _MIN_YEAR <= year <= _MAX_YEAR and 1 <= month <= 12:
        periods.add(f"{year:04d}-{month:02d}")


def _add_cell_period(
    periods: set[str],
    year: int,
    month: int,
    reference_years: frozenset[int],
) -> None:
    if reference_years and not any(abs(year - reference) <= 1 for reference in reference_years):
        return
    if not reference_years and year > date.today().year + 5:
        return
    _add_period(periods, year, month)

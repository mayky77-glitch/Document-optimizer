"""XLSX rendering for contract-cost discrepancies."""

from __future__ import annotations

from decimal import Decimal

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..contract_check import ContractCostViolation
from ..models import CATEGORY_DISPLAY_NAMES, CATEGORY_ORDER, ObjectBlockLayout, TargetWorkCategory
from ..statuses import Status
from .contract import COST_FORMAT, FRACTIONAL_QUANTITY_FORMAT, cost_to_million_rubles

DISCREPANCY_SHEET_NAME = "Расхождения и ошибки"
CONTRACT_COST_COLUMN_OFFSET = 6
_CONTRACT_COST_FILL = PatternFill("solid", fgColor="FFC7CE")
_EMPTY_FILL = PatternFill()
_HEADER_FILL = PatternFill("solid", fgColor="F4CCCC")
_BORDER = Border(
    left=Side(style="thin", color="7F8C7A"),
    right=Side(style="thin", color="7F8C7A"),
    top=Side(style="thin", color="7F8C7A"),
    bottom=Side(style="thin", color="7F8C7A"),
)
_HEADERS = (
    "Индекс объекта",
    "Шифр чертежа",
    "Этап / категория",
    "Ед. изм.",
    "По договору — объём",
    "По договору — стоимость, млн руб.",
    "Выполнено за весь период — объём",
    "Выполнено за весь период — стоимость, млн руб.",
    "Причина ошибки",
    "Разница стоимости, млн руб.",
    "Ссылка на договорную стоимость",
)

CONTRACT_COST_ERROR_REASON = (
    "Стоимость выполненных работ превышает договорную стоимость более чем на допустимые 1 000 руб."
)

_ISSUE_LABELS = {
    "REMAINING_QUANTITY_REPAIRED_FROM_DIMENSIONAL_FORMULA": (
        "В исходной книге формула остаточного объёма ссылалась на стоимость. "
        "Объём пересчитан по договорному количеству и вычитаемым объёмам."
    ),
    Status.POSSIBLE_DUPLICATE.value: "В исходных данных найдены возможные дубли этой работы.",
    Status.UNIT_MISMATCH.value: "Для объединяемых строк обнаружены разные единицы измерения.",
    Status.INVALID_NUMBER.value: "Числовое значение заполнено некорректно.",
    Status.EXCEL_ERROR.value: "В исходной ячейке содержится ошибка Excel.",
    Status.FORMULA_WITHOUT_CACHED_VALUE.value: (
        "Формула не содержит сохранённого результата. Пересчитайте и сохраните исходный файл."
    ),
    Status.FORMULA_NOT_AVAILABLE_FOR_BACKEND.value: (
        "Формулу не удалось безопасно прочитать используемым модулем."
    ),
    Status.VALUE_NOT_FOUND.value: "Не найдено значение для заполнения показателя.",
    Status.CONFLICT_REQUIRES_REVIEW.value: "Новое значение расходится с существующей карточкой.",
    Status.UNCONFIRMED_CLASSIFICATION.value: "Категория работы не подтверждена.",
}


def report_issue_text(warnings: tuple[str, ...], status: str) -> str:
    """Return readable report text without leaking row IDs or internal evidence hashes."""

    values = warnings or (status,)
    reasons: list[str] = []
    for value in values:
        code = str(value).partition(":")[0]
        reason = _ISSUE_LABELS.get(
            code,
            f"Обнаружено предупреждение обработки ({code.replace('_', ' ').lower()}).",
        )
        if reason not in reasons:
            reasons.append(reason)
    return "\n".join(reasons)


def contract_cost_coordinate(
    layout: ObjectBlockLayout, *, drawing_code: str, category_index: int
) -> str:
    """Find the main-card contract-cost cell for one rendered category row."""

    block = next(
        block for block in layout.drawing_code_blocks if block.drawing_code == drawing_code
    )
    column = get_column_letter(layout.start_column + CONTRACT_COST_COLUMN_OFFSET)
    return f"{column}{block.start_row + category_index}"


def excel_internal_location(sheet_name: str, coordinate: str) -> str:
    """Quote sheet names so spaces, Cyrillic characters and apostrophes remain valid."""

    escaped_sheet_name = sheet_name.replace("'", "''")
    return f"#'{escaped_sheet_name}'!{coordinate}"


def apply_contract_cost_highlights(
    workbook,
    layouts: list[ObjectBlockLayout],
    violations: list[ContractCostViolation],
) -> dict[tuple[str, str, object], tuple[str, str]]:
    """Fill precisely the violating contract-cost cells and return their locations."""

    layout_by_object = {layout.object_index: layout for layout in layouts}
    locations: dict[tuple[str, str, TargetWorkCategory], tuple[str, str]] = {}
    for violation in violations:
        row = violation.row
        layout = layout_by_object[row.object_index]
        coordinate = contract_cost_coordinate(
            layout,
            drawing_code=row.drawing_code.raw,
            category_index=CATEGORY_ORDER.index(row.category),
        )
        workbook[layout.sheet_name][coordinate].fill = _CONTRACT_COST_FILL
        locations[(row.object_index, row.drawing_code.raw, row.category)] = (
            layout.sheet_name,
            coordinate,
        )
    return locations


def clear_contract_cost_highlights(workbook, layouts: list[ObjectBlockLayout]) -> None:
    """Remove only discrepancy fills left by an earlier drawing-card render."""

    for layout in layouts:
        sheet = workbook[layout.sheet_name]
        column = layout.start_column + CONTRACT_COST_COLUMN_OFFSET
        for block in layout.drawing_code_blocks:
            for category_offset, _category in enumerate(CATEGORY_ORDER):
                cell = sheet.cell(block.start_row + category_offset, column)
                if _is_contract_cost_fill(cell.fill):
                    cell.fill = _EMPTY_FILL


def _is_contract_cost_fill(fill) -> bool:
    color = fill.fgColor
    return fill.fill_type == "solid" and color.type == "rgb" and color.rgb.endswith("FFC7CE")


def add_discrepancy_sheet(
    workbook,
    violations: list[ContractCostViolation],
    locations: dict[tuple[str, str, TargetWorkCategory], tuple[str, str]],
    *,
    cost_scale: int,
) -> dict[tuple[str, str], Decimal]:
    """Create the optional discrepancy registry with durable internal links."""

    if not violations:
        return {}
    if DISCREPANCY_SHEET_NAME in workbook.sheetnames:
        del workbook[DISCREPANCY_SHEET_NAME]
    sheet = workbook.create_sheet(DISCREPANCY_SHEET_NAME)
    exact_numeric_cells: dict[tuple[str, str], Decimal] = {}
    for column, header in enumerate(_HEADERS, start=1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(name="Arial", bold=True)
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_number, violation in enumerate(violations, start=2):
        row = violation.row
        values = (
            row.object_index,
            row.drawing_code.raw,
            CATEGORY_DISPLAY_NAMES[row.category],
            row.result_unit,
            row.contract_quantity,
            cost_to_million_rubles(row.contract_total_cost, cost_scale),
            row.performed_quantity,
            cost_to_million_rubles(row.performed_total_cost, cost_scale),
            CONTRACT_COST_ERROR_REASON,
            cost_to_million_rubles(violation.difference_rub, cost_scale),
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in {5, 7}:
                cell.number_format = FRACTIONAL_QUANTITY_FORMAT
            elif column in {6, 8, 10}:
                cell.number_format = COST_FORMAT
            if column in {5, 6, 7, 8, 10} and isinstance(value, Decimal):
                exact_numeric_cells[(sheet.title, cell.coordinate)] = value
        target_sheet, target_coordinate = locations[
            (row.object_index, row.drawing_code.raw, row.category)
        ]
        link = sheet.cell(row_number, 11, "Перейти к ячейке")
        link.hyperlink = excel_internal_location(target_sheet, target_coordinate)
        link.style = "Hyperlink"
        link.border = _BORDER
    widths = (16, 24, 42, 12, 20, 24, 28, 32, 52, 24, 28)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{sheet.max_row}"
    return exact_numeric_cells

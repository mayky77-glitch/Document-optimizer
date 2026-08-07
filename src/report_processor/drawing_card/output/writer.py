"""Template-based atomic create/update writer without business matching logic."""

from __future__ import annotations

import os
import tempfile
from dataclasses import replace
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from ..contract_check import find_contract_cost_violations
from ..models import (
    CATEGORY_DISPLAY_NAMES,
    CATEGORY_ORDER,
    DrawingCardResultRow,
    ObjectBlockLayout,
    WriteOperation,
)
from ..sources.normalization import build_drawing_code, normalize_text
from ..statuses import Status
from .contract import (
    CARD_BLOCK_COLUMN_SPAN,
    CARD_HEADERS,
    COST_FORMAT,
    DISPLAY_COST_SCALE,
    FRACTIONAL_QUANTITY_FORMAT,
    INTEGER_QUANTITY_FORMAT,
    MAIN_CARD_SHEET_NAME,
    SUMMARY_SHEET_NAME,
    cost_to_million_rubles,
)
from .discrepancies import (
    DISCREPANCY_SHEET_NAME,
    add_discrepancy_sheet,
    apply_contract_cost_highlights,
    clear_contract_cost_highlights,
    report_issue_text,
)
from .planner import plan_write_operations
from .styles import apply_card_style, clone_block_columns, clone_row_style
from .summary import add_summary_sheet
from .validator import validate_card
from .xlsx_xml import rewrite_exact_numeric_cells


def _decimal_or_value(value):
    """Keep Decimal values exact when serializing through openpyxl."""
    return value


def _existing_values(
    workbook, *, cost_scale: int = 1
) -> dict[tuple[str, str, str], tuple[object, object, object]]:
    result: dict[tuple[str, str, str], tuple[object, object, object]] = {}
    for sheet in workbook.worksheets:
        for start_column in range(1, sheet.max_column + 1):
            header = sheet.cell(2, start_column).value
            if not isinstance(header, str) or "индекс объекта" not in normalize_text(header):
                continue
            object_index = header.split(":", 1)[-1].strip()
            current_drawing: str | None = None
            for row in range(4, sheet.max_row + 1):
                drawing = sheet.cell(row, start_column).value
                if drawing not in (None, ""):
                    current_drawing = str(drawing).strip()
                category = sheet.cell(row, start_column + 1).value
                if not current_drawing or not category:
                    continue
                if normalize_text(str(category)) not in {
                    normalize_text(value) for value in CATEGORY_DISPLAY_NAMES.values()
                }:
                    continue
                cost = sheet.cell(row, start_column + 4).value
                cost_header = sheet.cell(3, start_column + 4).value
                if (
                    cost is not None
                    and isinstance(cost_header, str)
                    and "млн" in normalize_text(cost_header)
                ):
                    cost = Decimal(str(cost)) * DISPLAY_COST_SCALE * Decimal(cost_scale)
                result[(object_index, current_drawing, normalize_text(str(category)))] = (
                    sheet.cell(row, start_column + 2).value,
                    sheet.cell(row, start_column + 3).value,
                    cost,
                )
    return result


def merge_update_rows(
    new_rows: list[DrawingCardResultRow],
    existing: dict[tuple[str, str, str], tuple[object, object, object]],
    policy: str,
) -> tuple[list[DrawingCardResultRow], list[str]]:
    warnings: list[str] = []
    merged: list[DrawingCardResultRow] = []
    for row in new_rows:
        key = (row.object_index, row.drawing_code.raw, normalize_text(row.display_name))
        old = existing.get(key)
        if old is None:
            merged.append(row)
            continue
        old_unit, old_quantity, old_cost = old
        unit = row.result_unit
        quantity = row.remaining_quantity
        cost = row.remaining_total_cost
        if policy == "keep_existing" or policy == "fill_empty_only":
            unit = old_unit or unit
            quantity = old_quantity if old_quantity is not None else quantity
            cost = old_cost if old_cost is not None else cost
        elif policy == "overwrite":
            unit = unit or old_unit
            quantity = quantity if quantity is not None else old_quantity
            cost = cost if cost is not None else old_cost
        elif policy == "conflicts_to_review":
            for metric, old_value, new_value in (
                ("unit", old_unit, unit),
                ("quantity", old_quantity, quantity),
                ("total_cost", old_cost, cost),
            ):
                if (
                    old_value not in (None, "")
                    and new_value not in (None, "")
                    and str(old_value) != str(new_value)
                ):
                    warnings.append(
                        f"{Status.CONFLICT_REQUIRES_REVIEW}:{key}:{metric}:{old_value}:{new_value}"
                    )
            unit = old_unit or unit
            quantity = old_quantity if old_quantity is not None else quantity
            cost = old_cost if old_cost is not None else cost
        merged.append(
            replace(
                row,
                result_unit=None if unit is None else str(unit),
                remaining_quantity=(
                    quantity
                    if isinstance(quantity, Decimal) or quantity is None
                    else Decimal(str(quantity))
                ),
                remaining_total_cost=(
                    cost if isinstance(cost, Decimal) or cost is None else Decimal(str(cost))
                ),
                warnings=row.warnings + tuple(item for item in warnings[-3:] if str(key) in item),
            )
        )
    known = {
        (row.object_index, row.drawing_code.raw, normalize_text(row.display_name))
        for row in new_rows
    }
    for (object_index, drawing_raw, category_name), (unit, quantity, cost) in existing.items():
        if (object_index, drawing_raw, category_name) in known:
            continue
        category = next(
            item
            for item in CATEGORY_ORDER
            if normalize_text(CATEGORY_DISPLAY_NAMES[item]) == category_name
        )
        drawing = build_drawing_code(drawing_raw)
        merged.append(
            DrawingCardResultRow(
                object_index=object_index,
                drawing_code=drawing,
                category=category,
                display_name=CATEGORY_DISPLAY_NAMES[category],
                result_unit=None if unit is None else str(unit),
                remaining_quantity=None if quantity is None else Decimal(str(quantity)),
                remaining_total_cost=None if cost is None else Decimal(str(cost)),
                quantity_source_rows=(),
                cost_source_rows=(),
                quantity_rule_id="existing-card",
                cost_rule_id="existing-card",
                quantity_confidence=None,
                cost_confidence=None,
                requires_manual_review=False,
                status=Status.OK,
                warnings=("PRESERVED_FROM_EXISTING_CARD",),
                quantity_matching_strategies=("existing_card",),
                cost_matching_strategies=("existing_card",),
            )
        )
    return merged, warnings


def _ensure_sheets(workbook, layouts: list[ObjectBlockLayout]) -> None:
    first = (
        workbook[MAIN_CARD_SHEET_NAME]
        if MAIN_CARD_SHEET_NAME in workbook.sheetnames
        else workbook["Лист1"]
        if "Лист1" in workbook.sheetnames
        else next(sheet for sheet in workbook.worksheets if sheet.title != SUMMARY_SHEET_NAME)
    )
    required = tuple(dict.fromkeys(layout.sheet_name for layout in layouts))
    if required and first.title != required[0]:
        first.title = required[0]
    for name in required[1:]:
        if name not in workbook.sheetnames:
            clone = workbook.copy_worksheet(first)
            clone.title = name


def _clear_template_values(sheet, layouts: list[ObjectBlockLayout]) -> None:
    """Clear only the template-managed card slots before rendering new values.

    The shipped template has four legacy six-column slots, whereas new cards use
    ten-column slots.  Cover both geometries so legacy labels cannot survive in
    new spacer columns, while leaving cells outside the controlled card region
    untouched for update-mode users.
    """

    max_row = max(
        19,
        max(
            (block.end_row for layout in layouts for block in layout.drawing_code_blocks),
            default=0,
        ),
    )
    legacy_starts = range(2, min(sheet.max_column, 25), 6)
    current_starts = {2 + slot * CARD_BLOCK_COLUMN_SPAN for slot in range(4)} | {
        layout.start_column for layout in layouts
    }
    for start_column in set(legacy_starts) | current_starts:
        for row in range(2, max_row + 1):
            for column in range(start_column, start_column + len(CARD_HEADERS)):
                cell = sheet.cell(row, column)
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.comment = None


def _prepare_block(sheet, layout: ObjectBlockLayout) -> None:
    start = layout.start_column
    if start > 2:
        clone_block_columns(
            sheet,
            source_start=2,
            target_start=start,
            width=len(CARD_HEADERS),
        )
    for existing in tuple(sheet.merged_cells.ranges):
        if existing.min_row <= 2 <= existing.max_row and not (
            existing.max_col < start or existing.min_col > layout.end_column
        ):
            sheet.unmerge_cells(str(existing))
    index_merge_range = f"{get_column_letter(start)}2:{get_column_letter(start + 2)}2"
    merge_range = f"{get_column_letter(start + 3)}2:{get_column_letter(layout.end_column)}2"
    if index_merge_range not in {str(item) for item in sheet.merged_cells.ranges}:
        sheet.merge_cells(index_merge_range)
    if merge_range not in {str(item) for item in sheet.merged_cells.ranges}:
        sheet.merge_cells(merge_range)
    sheet.cell(2, start).value = f"Индекс объекта: {layout.object_index}"
    sheet.cell(2, start + 3).value = "Показатели объёма и стоимости"
    for offset, header in enumerate(CARD_HEADERS):
        sheet.cell(3, start + offset).value = header


def _ensure_data_styles(sheet, layout: ObjectBlockLayout) -> None:
    max_row = max((block.end_row for block in layout.drawing_code_blocks), default=11)
    for row in range(4, max_row + 1):
        source_row = 4 + ((row - 4) % 8)
        clone_row_style(
            sheet,
            source_row=source_row,
            target_row=row,
            start_column=layout.start_column,
            end_column=layout.end_column,
        )


def _trim_unused_right_template_slots(workbook, layouts: list[ObjectBlockLayout]) -> None:
    """Remove only unoccupied right-side template slots on the final card sheet."""

    if not layouts:
        return
    final_sheet_name = layouts[-1].sheet_name
    final_layouts = [layout for layout in layouts if layout.sheet_name == final_sheet_name]
    sheet = workbook[final_sheet_name]
    last_occupied_column = max(layout.end_column for layout in final_layouts)
    trim_start = last_occupied_column + 2
    if trim_start <= sheet.max_column:
        for existing in tuple(sheet.merged_cells.ranges):
            if existing.max_col >= trim_start:
                sheet.unmerge_cells(str(existing))
        sheet.delete_cols(trim_start, sheet.max_column - trim_start + 1)


def write_card(
    *,
    base_path: Path,
    output_path: Path,
    rows: list[DrawingCardResultRow],
    layouts: list[ObjectBlockLayout],
    run_id: str,
    cost_scale: int,
) -> list[WriteOperation]:
    if base_path.resolve() == output_path.resolve():
        raise ValueError("Input template/existing card and output path must differ")
    if output_path.exists() and output_path.is_dir():
        raise IsADirectoryError(f"Output must be an .xlsx file, not a directory: {output_path}")
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Output path must end with .xlsx: {output_path}")
    if output_path.parent.exists() and not output_path.parent.is_dir():
        raise NotADirectoryError(output_path.parent)
    workbook = load_workbook(base_path)
    operations = plan_write_operations(
        rows=rows,
        layouts=layouts,
        run_id=run_id,
        cost_scale=cost_scale,
    )
    operations_by_cell = {
        (operation.output_sheet, operation.output_cell): index
        for index, operation in enumerate(operations)
    }
    exact_numeric_cells: dict[tuple[str, str], Decimal] = {}
    try:
        _ensure_sheets(workbook, layouts)
        previous_values = {
            (operation.output_sheet, operation.output_cell): workbook[operation.output_sheet][
                operation.output_cell
            ].value
            for operation in operations
        }
        for sheet_name in dict.fromkeys(layout.sheet_name for layout in layouts):
            _clear_template_values(
                workbook[sheet_name],
                [layout for layout in layouts if layout.sheet_name == sheet_name],
            )
        rows_by_key = {(row.object_index, row.drawing_code.raw, row.category): row for row in rows}
        for layout in layouts:
            sheet = workbook[layout.sheet_name]
            _prepare_block(sheet, layout)
            _ensure_data_styles(sheet, layout)
            for block in layout.drawing_code_blocks:
                for offset, category in enumerate(CATEGORY_ORDER):
                    row_number = block.start_row + offset
                    result = rows_by_key[(layout.object_index, block.drawing_code, category)]
                    values = (
                        block.drawing_code if offset == 0 else None,
                        result.display_name,
                        result.result_unit,
                        result.remaining_quantity,
                        cost_to_million_rubles(result.remaining_total_cost, cost_scale),
                        result.contract_quantity,
                        cost_to_million_rubles(result.contract_total_cost, cost_scale),
                        result.performed_quantity,
                        cost_to_million_rubles(result.performed_total_cost, cost_scale),
                    )
                    for column_offset, value in enumerate(values):
                        cell = sheet.cell(row_number, layout.start_column + column_offset)
                        cell.value = _decimal_or_value(value)
                        if isinstance(value, Decimal) and column_offset in {3, 4, 5, 6, 7, 8}:
                            exact_numeric_cells[(sheet.title, cell.coordinate)] = value
                        if column_offset in {3, 5, 7}:
                            cell.number_format = (
                                INTEGER_QUANTITY_FORMAT
                                if isinstance(value, Decimal) and value == value.to_integral()
                                else FRACTIONAL_QUANTITY_FORMAT
                            )
                        elif column_offset in {4, 6, 8}:
                            cell.number_format = COST_FORMAT
                        if result.requires_manual_review or result.status not in {
                            Status.OK,
                            Status.VALUE_NOT_FOUND,
                            Status.UNIT_FROM_TEMPLATE,
                        }:
                            cell.comment = Comment(
                                report_issue_text(result.warnings, result.status),
                                "report_processor",
                            )
                        metric = (
                            "drawing_code",
                            "category",
                            "unit",
                            "quantity",
                            "total_cost",
                            "contract_quantity",
                            "contract_total_cost",
                            "performed_quantity",
                            "performed_total_cost",
                        )[column_offset]
                        if metric not in {"drawing_code", "category"}:
                            old_value = previous_values[(sheet.title, cell.coordinate)]
                            operation_index = operations_by_cell[(sheet.title, cell.coordinate)]
                            operations[operation_index] = replace(
                                operations[operation_index],
                                old_value=old_value,
                                new_value=cell.value,
                            )
            apply_card_style(sheet, layout)
        _trim_unused_right_template_slots(workbook, layouts)
        if SUMMARY_SHEET_NAME in workbook.sheetnames:
            del workbook[SUMMARY_SHEET_NAME]
        if DISCREPANCY_SHEET_NAME in workbook.sheetnames:
            del workbook[DISCREPANCY_SHEET_NAME]
        clear_contract_cost_highlights(workbook, layouts)
        violations = find_contract_cost_violations(rows)
        locations = apply_contract_cost_highlights(workbook, layouts, violations)
        exact_numeric_cells.update(
            add_summary_sheet(
                workbook,
                layouts,
                rows,
                cost_scale=cost_scale,
            )
        )
        exact_numeric_cells.update(
            add_discrepancy_sheet(
                workbook,
                violations,
                locations,
                cost_scale=cost_scale,
            )
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            prefix=output_path.stem + ".", suffix=".tmp.xlsx", dir=output_path.parent, delete=False
        ) as handle:
            temp_path = Path(handle.name)
        try:
            workbook.save(temp_path)
            rewrite_exact_numeric_cells(temp_path, exact_numeric_cells)
            validation = validate_card(temp_path, layouts)
            if validation["status"] != Status.OK.value:
                details = "; ".join(validation["errors"][:10])
                raise ValueError(f"Temporary output validation failed: {details}")
            os.replace(temp_path, output_path)
        except BaseException:
            temp_path.unlink(missing_ok=True)
            raise
    finally:
        workbook.close()
    return operations


def load_existing_values(
    path: Path, *, cost_scale: int = 1
) -> dict[tuple[str, str, str], tuple[object, object, object]]:
    workbook = load_workbook(path, data_only=True)
    try:
        return _existing_values(workbook, cost_scale=cost_scale)
    finally:
        workbook.close()

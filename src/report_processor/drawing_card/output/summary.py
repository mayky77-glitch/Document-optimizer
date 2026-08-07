"""Literal-value card-style summary sheet for published drawing-card workbooks."""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import CATEGORY_DISPLAY_NAMES, CATEGORY_ORDER, DrawingCardResultRow, ObjectBlockLayout
from ..sources.normalization import normalize_text
from .contract import (
    COST_FORMAT,
    FRACTIONAL_QUANTITY_FORMAT,
    SUMMARY_BLOCK_COLUMN_SPAN,
    SUMMARY_BLOCK_ROW_SPAN,
    SUMMARY_BLOCKS_PER_ROW,
    SUMMARY_HEADERS,
    SUMMARY_SHEET_NAME,
    cost_to_million_rubles,
)

_TITLE_FILL = PatternFill("solid", fgColor="548235")
_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_BORDER = Border(
    left=Side(style="thin", color="7F8C7A"),
    right=Side(style="thin", color="7F8C7A"),
    top=Side(style="thin", color="7F8C7A"),
    bottom=Side(style="thin", color="7F8C7A"),
)


def summary_block_position(block_number: int) -> tuple[int, int]:
    """Return the top-left cell for a compact summary card block."""

    return (
        1 + (block_number // SUMMARY_BLOCKS_PER_ROW) * SUMMARY_BLOCK_ROW_SPAN,
        1 + (block_number % SUMMARY_BLOCKS_PER_ROW) * SUMMARY_BLOCK_COLUMN_SPAN,
    )


def summary_row_count(layouts: list[ObjectBlockLayout]) -> int:
    """Return the summary sheet height used by all index cards and the grand total."""

    total_block_number = len(layouts)
    total_start_row, _ = summary_block_position(total_block_number)
    return total_start_row + len(CATEGORY_ORDER) + 1


def _units_by_layout(
    rows: list[DrawingCardResultRow],
) -> dict[tuple[str, object], tuple[str | None, ...]]:
    units: dict[tuple[str, object], list[str | None]] = defaultdict(list)
    for row in rows:
        unit = row.result_unit.strip() if isinstance(row.result_unit, str) else None
        units[(row.object_index, row.category)].append(unit or None)
    return {key: tuple(value) for key, value in units.items()}


def _single_nonempty_unit(units: tuple[str | None, ...]) -> str | None:
    if not units or any(unit is None for unit in units):
        return None
    normalized = {normalize_text(unit) for unit in units if unit is not None}
    return units[0] if len(normalized) == 1 else None


def _sum_quantity(rows: list[DrawingCardResultRow], metric: str) -> Decimal:
    return sum(
        (
            (row.remaining_quantity if metric == "remaining_quantity" else getattr(row, metric))
            or Decimal(0)
            for row in rows
        ),
        Decimal(0),
    )


def _sum_cost(rows: list[DrawingCardResultRow], metric: str, cost_scale: int) -> Decimal:
    """Sum the published row values so summaries reconcile with the main card."""

    return sum(
        (
            cost_to_million_rubles(
                row.remaining_total_cost
                if metric == "remaining_total_cost"
                else getattr(row, metric),
                cost_scale,
            )
            or Decimal(0)
            for row in rows
        ),
        Decimal(0),
    )


def _style_block(sheet, *, start_row: int, start_column: int, title: str) -> None:
    end_column = start_column + len(SUMMARY_HEADERS) - 1
    sheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=start_row,
        end_column=end_column,
    )
    title_cell = sheet.cell(start_row, start_column, title)
    title_cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = _BORDER
    for column, header in enumerate(SUMMARY_HEADERS, start=start_column):
        cell = sheet.cell(start_row + 1, column, header)
        cell.font = Font(name="Arial", bold=True, size=9, color="1F1F1F")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    for row in range(start_row + 2, start_row + 2 + len(CATEGORY_ORDER)):
        for column in range(start_column, end_column + 1):
            cell = sheet.cell(row, column)
            cell.font = Font(name="Arial", size=10, color="1F1F1F")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _BORDER


def _write_summary_metrics(
    sheet,
    *,
    row: int,
    start_column: int,
    rows: list[DrawingCardResultRow],
    unit: str | None,
    cost_scale: int,
    exact_numeric_cells: dict[tuple[str, str], Decimal],
) -> None:
    if unit is not None:
        sheet.cell(row, start_column + 1, unit)
    metrics = (
        ("remaining_quantity", False),
        ("remaining_total_cost", True),
        ("contract_quantity", False),
        ("contract_total_cost", True),
        ("performed_quantity", False),
        ("performed_total_cost", True),
    )
    for offset, (metric, is_cost) in enumerate(metrics, start=2):
        if not is_cost and unit is None:
            continue
        value = _sum_cost(rows, metric, cost_scale) if is_cost else _sum_quantity(rows, metric)
        cell = sheet.cell(row, start_column + offset, value)
        cell.number_format = COST_FORMAT if is_cost else FRACTIONAL_QUANTITY_FORMAT
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        exact_numeric_cells[(sheet.title, cell.coordinate)] = value


def add_summary_sheet(
    workbook,
    layouts: list[ObjectBlockLayout],
    rows: list[DrawingCardResultRow],
    *,
    cost_scale: int,
) -> dict[tuple[str, str], Decimal]:
    """Create horizontal index cards and a separate literal-value grand-total card."""

    if SUMMARY_SHEET_NAME in workbook.sheetnames:
        raise ValueError(f"Summary sheet already exists: {SUMMARY_SHEET_NAME}")
    sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    units = _units_by_layout(rows)
    exact_numeric_cells: dict[tuple[str, str], Decimal] = {}
    rows_by_object_category: dict[tuple[str, object], list[DrawingCardResultRow]] = defaultdict(
        list
    )
    rows_by_category: dict[object, list[DrawingCardResultRow]] = defaultdict(list)
    for result in rows:
        rows_by_object_category[(result.object_index, result.category)].append(result)
        rows_by_category[result.category].append(result)

    for block_number, layout in enumerate(layouts):
        start_row, start_column = summary_block_position(block_number)
        _style_block(
            sheet,
            start_row=start_row,
            start_column=start_column,
            title=f"Индекс объекта: {layout.object_index}",
        )
        for offset, category in enumerate(CATEGORY_ORDER):
            row = start_row + 2 + offset
            category_cell = sheet.cell(row, start_column, CATEGORY_DISPLAY_NAMES[category])
            index_rows = rows_by_object_category[(layout.object_index, category)]
            unit = _single_nonempty_unit(units.get((layout.object_index, category), ()))
            _write_summary_metrics(
                sheet,
                row=row,
                start_column=start_column,
                rows=index_rows,
                unit=unit,
                cost_scale=cost_scale,
                exact_numeric_cells=exact_numeric_cells,
            )
            category_cell.alignment = Alignment(vertical="top", wrap_text=True)

    total_start_row, total_start_column = summary_block_position(len(layouts))
    _style_block(
        sheet,
        start_row=total_start_row,
        start_column=total_start_column,
        title="Все индексы",
    )
    for offset, category in enumerate(CATEGORY_ORDER):
        row = total_start_row + 2 + offset
        sheet.cell(row, total_start_column, CATEGORY_DISPLAY_NAMES[category])
        category_units = tuple(
            _single_nonempty_unit(units.get((layout.object_index, category), ()))
            for layout in layouts
        )
        unit = _single_nonempty_unit(category_units)
        _write_summary_metrics(
            sheet,
            row=row,
            start_column=total_start_column,
            rows=rows_by_category[category],
            unit=unit,
            cost_scale=cost_scale,
            exact_numeric_cells=exact_numeric_cells,
        )

    for column in range(1, sheet.max_column + 1):
        offset = (column - 1) % SUMMARY_BLOCK_COLUMN_SPAN
        sheet.column_dimensions[get_column_letter(column)].width = (
            34,
            10,
            15,
            19,
            15,
            19,
            15,
            19,
            2.5,
            2.5,
        )[offset]
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 22 if row % SUMMARY_BLOCK_ROW_SPAN < 2 else 30
    sheet.freeze_panes = "A3"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_options.horizontalCentered = True
    sheet.print_area = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    return exact_numeric_cells

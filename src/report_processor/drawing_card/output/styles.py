"""Template-derived style cloning helpers."""

from __future__ import annotations

from copy import copy

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models import ObjectBlockLayout

_TITLE_FILL = PatternFill("solid", fgColor="548235")
_HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_GRID_BORDER = Border(
    left=Side(style="thin", color="A6A6A6"),
    right=Side(style="thin", color="A6A6A6"),
    top=Side(style="thin", color="A6A6A6"),
    bottom=Side(style="thin", color="A6A6A6"),
)
_CARD_COLUMN_WIDTHS = (42, 34, 10, 15, 19, 15, 19, 15, 19)


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def clone_row_style(
    worksheet: Worksheet,
    *,
    source_row: int,
    target_row: int,
    start_column: int,
    end_column: int,
) -> None:
    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    for column in range(start_column, end_column + 1):
        copy_cell_style(worksheet.cell(source_row, column), worksheet.cell(target_row, column))


def clone_block_columns(
    worksheet: Worksheet,
    *,
    source_start: int,
    target_start: int,
    width: int = 5,
) -> None:
    from openpyxl.utils import get_column_letter

    for offset in range(width):
        source_letter = get_column_letter(source_start + offset)
        target_letter = get_column_letter(target_start + offset)
        source_dimension = worksheet.column_dimensions[source_letter]
        target_dimension = worksheet.column_dimensions[target_letter]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        for row in range(2, max(12, worksheet.max_row) + 1):
            copy_cell_style(
                worksheet.cell(row, source_start + offset),
                worksheet.cell(row, target_start + offset),
            )


def apply_card_style(worksheet: Worksheet, layout: ObjectBlockLayout) -> None:
    """Apply the unified, template-independent visual contract to one card."""

    start = layout.start_column
    end = layout.end_column
    for column in range(start, end + 1):
        title = worksheet.cell(2, column)
        title.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        title.fill = _TITLE_FILL
        title.border = _GRID_BORDER
        title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header = worksheet.cell(3, column)
        header.font = Font(name="Arial", size=9, bold=True, color="1F1F1F")
        header.fill = _HEADER_FILL
        header.border = _GRID_BORDER
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 28
    worksheet.row_dimensions[3].height = 42
    for offset, width in enumerate(_CARD_COLUMN_WIDTHS):
        worksheet.column_dimensions[get_column_letter(start + offset)].width = width
    worksheet.column_dimensions[get_column_letter(end + 1)].width = 2.5

    max_row = max((block.end_row for block in layout.drawing_code_blocks), default=3)
    for row in range(4, max_row + 1):
        worksheet.row_dimensions[row].height = 28
        for offset in range(end - start + 1):
            cell = worksheet.cell(row, start + offset)
            cell.font = Font(
                name="Arial",
                size=9.5 if offset < 3 else 10,
                bold=offset == 1,
                color="1F1F1F",
            )
            cell.fill = _HEADER_FILL if offset in {1, 2} else _WHITE_FILL
            cell.border = _GRID_BORDER
            cell.alignment = Alignment(
                horizontal="left" if offset in {0, 1} else "center" if offset == 2 else "right",
                vertical="top" if offset == 0 else "center",
                wrap_text=True,
            )

    if start == 2:
        worksheet.freeze_panes = "B4"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_options.horizontalCentered = True

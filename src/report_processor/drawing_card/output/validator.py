"""Post-write validation of drawing-card workbooks."""

from __future__ import annotations

import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..models import CATEGORY_DISPLAY_NAMES, CATEGORY_ORDER, ObjectBlockLayout
from ..sources.normalization import is_plausible_drawing_code, normalize_text
from ..statuses import Status
from .contract import (
    CARD_BLOCK_COLUMN_SPAN,
    CARD_HEADERS,
    COST_FORMAT,
    FRACTIONAL_QUANTITY_FORMAT,
    INTEGER_QUANTITY_FORMAT,
    SUMMARY_HEADERS,
    SUMMARY_SHEET_NAME,
)
from .discrepancies import DISCREPANCY_SHEET_NAME
from .summary import summary_block_position, summary_row_count
from .xlsx_xml import find_binary_tail_cells

_EXPECTED_CATEGORIES = tuple(
    normalize_text(CATEGORY_DISPLAY_NAMES[category]) for category in CATEGORY_ORDER
)
_FORMULA_ERRORS = {"#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"}


def _decimal(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _equal_nonzero(left: Decimal | None, right: Decimal | None) -> bool:
    if left in (None, Decimal(0)) or right in (None, Decimal(0)):
        return False
    tolerance = max(Decimal("0.000001"), abs(right) * Decimal("1e-9"))
    return abs(left - right) <= tolerance


def _equal_values(left: object, right: object) -> bool:
    """Compare literal workbook numbers without requiring binary-float identity."""

    left_decimal = _decimal(left)
    right_decimal = _decimal(right)
    if left_decimal is None or right_decimal is None:
        return left_decimal is right_decimal
    return abs(left_decimal - right_decimal) <= Decimal("0.000001")


def _summary_source_values(
    workbook,
    layout: ObjectBlockLayout,
    category,
    column_offset: int,
) -> list[object]:
    sheet = workbook[layout.sheet_name]
    category_offset = CATEGORY_ORDER.index(category)
    return [
        sheet.cell(block.start_row + category_offset, layout.start_column + column_offset).value
        for block in layout.drawing_code_blocks
    ]


def _sum_literal_values(values: list[object]) -> Decimal:
    return sum((_decimal(value) or Decimal(0) for value in values), Decimal(0))


def _validate_archive(path: Path, errors: list[str]) -> bool:
    try:
        with zipfile.ZipFile(path) as archive:
            broken_member = archive.testzip()
            if broken_member:
                errors.append(f"ZIP_CRC_ERROR:{broken_member}")
            for member in archive.namelist():
                if member.endswith(".xml"):
                    ET.fromstring(archive.read(member))
        return not errors
    except (ET.ParseError, OSError, zipfile.BadZipFile) as error:
        errors.append(f"INVALID_XLSX_ARCHIVE:{error}")
        return False


def _validate_categories(
    *, sheet_name: str, drawing: str, categories: list[str], errors: list[str]
) -> None:
    if len(categories) != len(_EXPECTED_CATEGORIES):
        errors.append(f"DRAWING_CATEGORY_COUNT:{sheet_name}:{drawing}:{len(categories)}")
        return
    if len(set(categories)) != len(categories):
        errors.append(f"DRAWING_CATEGORY_DUPLICATE:{sheet_name}:{drawing}")
    if tuple(categories) != _EXPECTED_CATEGORIES:
        errors.append(f"DRAWING_CATEGORY_ORDER:{sheet_name}:{drawing}")


def _validate_layout_contract(
    workbook, layouts: list[ObjectBlockLayout], errors: list[str]
) -> None:
    for layout in layouts:
        if layout.sheet_name not in workbook.sheetnames:
            errors.append(f"MISSING_SHEET:{layout.sheet_name}")
            continue
        sheet = workbook[layout.sheet_name]
        start = layout.start_column
        merge = f"{get_column_letter(start + 3)}2:{get_column_letter(layout.end_column)}2"
        if merge not in {str(item) for item in sheet.merged_cells.ranges}:
            errors.append(f"MISSING_MERGE:{sheet.title}:{merge}")
        if sheet.cell(2, start).value != f"Индекс объекта: {layout.object_index}":
            errors.append(f"INVALID_OBJECT_HEADER:{sheet.title}:{start}")
        for offset, expected in enumerate(CARD_HEADERS):
            if sheet.cell(3, start + offset).value != expected:
                coordinate = sheet.cell(3, start + offset).coordinate
                errors.append(f"INVALID_COLUMN_HEADER:{sheet.title}:{coordinate}")
        for block in layout.drawing_code_blocks:
            for offset in range(len(CATEGORY_ORDER)):
                row = block.start_row + offset
                quantity = sheet.cell(row, start + 3)
                cost = sheet.cell(row, start + 4)
                contract_quantity = sheet.cell(row, start + 5)
                contract_cost = sheet.cell(row, start + 6)
                performed_quantity = sheet.cell(row, start + 7)
                performed_cost = sheet.cell(row, start + 8)
                expected_quantity_format = (
                    INTEGER_QUANTITY_FORMAT
                    if isinstance(quantity.value, (int, Decimal))
                    and Decimal(str(quantity.value)) == Decimal(str(quantity.value)).to_integral()
                    else FRACTIONAL_QUANTITY_FORMAT
                )
                if quantity.number_format != expected_quantity_format:
                    errors.append(f"INVALID_QUANTITY_FORMAT:{sheet.title}:{quantity.coordinate}")
                if cost.number_format != COST_FORMAT:
                    errors.append(f"INVALID_COST_FORMAT:{sheet.title}:{cost.coordinate}")
                for quantity_cell in (contract_quantity, performed_quantity):
                    if quantity_cell.number_format != FRACTIONAL_QUANTITY_FORMAT:
                        errors.append(
                            f"INVALID_QUANTITY_FORMAT:{sheet.title}:{quantity_cell.coordinate}"
                        )
                for cost_cell in (contract_cost, performed_cost):
                    if cost_cell.number_format != COST_FORMAT:
                        errors.append(f"INVALID_COST_FORMAT:{sheet.title}:{cost_cell.coordinate}")
                if not all(
                    sheet.cell(row, column).has_style
                    for column in range(start, layout.end_column + 1)
                ):
                    errors.append(f"MISSING_DATA_STYLE:{sheet.title}:{row}")
                if sheet.row_dimensions[row].height is None:
                    errors.append(f"MISSING_ROW_DIMENSION:{sheet.title}:{row}")


def _validate_trimmed_template_slots(
    workbook,
    layouts: list[ObjectBlockLayout],
    errors: list[str],
) -> None:
    if not layouts:
        return
    final_sheet_name = layouts[-1].sheet_name
    final_layouts = [layout for layout in layouts if layout.sheet_name == final_sheet_name]
    last_occupied_column = max(layout.end_column for layout in final_layouts)
    sheet = workbook[final_sheet_name]
    for start_column in tuple(2 + slot * CARD_BLOCK_COLUMN_SPAN for slot in range(1, 4)):
        if start_column <= last_occupied_column:
            continue
        if sheet.cell(4, start_column).has_style:
            errors.append(f"VACANT_TEMPLATE_SLOT:{sheet.title}:{get_column_letter(start_column)}")


def _validate_summary_contract(
    workbook,
    layouts: list[ObjectBlockLayout],
    errors: list[str],
) -> None:
    if SUMMARY_SHEET_NAME not in workbook.sheetnames:
        errors.append(f"MISSING_SUMMARY_SHEET:{SUMMARY_SHEET_NAME}")
        return
    sheet = workbook[SUMMARY_SHEET_NAME]
    expected_last_row = summary_row_count(layouts)
    if sheet.max_row != expected_last_row:
        errors.append(f"SUMMARY_ROW_COUNT:{sheet.max_row}!={expected_last_row}")
        return
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                errors.append(f"SUMMARY_FORMULA:{cell.coordinate}")
    category_unit_status: dict[tuple[str, object], str] = {}
    for block_number, layout in enumerate(layouts):
        start_row, start_column = summary_block_position(block_number)
        if sheet.cell(start_row, start_column).value != f"Индекс объекта: {layout.object_index}":
            errors.append(f"SUMMARY_INDEX_TITLE:{start_row}:{start_column}")
        if (
            tuple(
                sheet.cell(start_row + 1, start_column + offset).value
                for offset in range(len(SUMMARY_HEADERS))
            )
            != SUMMARY_HEADERS
        ):
            errors.append(f"INVALID_SUMMARY_HEADERS:{start_row}:{start_column}")
        for category in CATEGORY_ORDER:
            row_number = start_row + 2 + CATEGORY_ORDER.index(category)
            expected_name = CATEGORY_DISPLAY_NAMES[category]
            if sheet.cell(row_number, start_column).value != expected_name:
                errors.append(f"SUMMARY_CATEGORY:{row_number}:{start_column}")
            quantity = sheet.cell(row_number, start_column + 2)
            cost = sheet.cell(row_number, start_column + 3)
            source_units = [
                value.strip() if isinstance(value, str) else None
                for value in _summary_source_values(workbook, layout, category, 2)
            ]
            if not source_units or any(unit is None for unit in source_units):
                category_unit_status[(layout.object_index, category)] = "missing"
            elif len({normalize_text(unit) for unit in source_units if unit is not None}) != 1:
                category_unit_status[(layout.object_index, category)] = "mixed"
            else:
                category_unit_status[(layout.object_index, category)] = "valid"
            expected_cost = _sum_literal_values(
                _summary_source_values(workbook, layout, category, 4)
            )
            if not _equal_values(cost.value, expected_cost):
                errors.append(f"SUMMARY_COST_VALUE:{cost.coordinate}")
            if category_unit_status[(layout.object_index, category)] == "valid":
                expected_quantity = _sum_literal_values(
                    _summary_source_values(workbook, layout, category, 3)
                )
                if not _equal_values(quantity.value, expected_quantity):
                    errors.append(f"SUMMARY_QUANTITY_VALUE:{quantity.coordinate}")
            elif quantity.value is not None:
                errors.append(f"SUMMARY_INVALID_UNIT_QUANTITY:{quantity.coordinate}")
            if quantity.number_format != FRACTIONAL_QUANTITY_FORMAT:
                errors.append(f"SUMMARY_QUANTITY_FORMAT:{quantity.coordinate}")
            if cost.number_format != COST_FORMAT:
                errors.append(f"SUMMARY_COST_FORMAT:{cost.coordinate}")
    total_start_row, total_start_column = summary_block_position(len(layouts))
    if sheet.cell(total_start_row, total_start_column).value != "Все индексы":
        errors.append(f"SUMMARY_ALL_INDICES_TITLE:{total_start_row}:{total_start_column}")
    if (
        tuple(
            sheet.cell(total_start_row + 1, total_start_column + offset).value
            for offset in range(len(SUMMARY_HEADERS))
        )
        != SUMMARY_HEADERS
    ):
        errors.append(f"INVALID_SUMMARY_HEADERS:{total_start_row}:{total_start_column}")
    for category in CATEGORY_ORDER:
        row_number = total_start_row + 2 + CATEGORY_ORDER.index(category)
        expected_name = CATEGORY_DISPLAY_NAMES[category]
        if sheet.cell(row_number, total_start_column).value != expected_name:
            errors.append(f"SUMMARY_ALL_INDICES_CATEGORY:{row_number}:{total_start_column}")
        unit_statuses = [
            category_unit_status[(layout.object_index, category)] for layout in layouts
        ]
        category_units: set[str] = set()
        for block_number in range(len(layouts)):
            block_row, block_column = summary_block_position(block_number)
            unit_value = sheet.cell(
                block_row + 2 + CATEGORY_ORDER.index(category), block_column + 1
            ).value
            if isinstance(unit_value, str) and unit_value.strip():
                category_units.add(normalize_text(unit_value))
        quantity = sheet.cell(row_number, total_start_column + 2)
        if "missing" in unit_statuses:
            if quantity.value is not None:
                errors.append(f"SUMMARY_MISSING_UNIT_QUANTITY:{expected_name}")
            errors.append(f"SUMMARY_MISSING_UNIT:{category.value}")
        elif "mixed" in unit_statuses or len(category_units) != 1:
            if quantity.value is not None:
                errors.append(f"SUMMARY_MIXED_UNIT_QUANTITY:{expected_name}")
            errors.append(f"SUMMARY_MIXED_UNIT:{category.value}")
        else:
            expected_quantity = sum(
                (
                    _sum_literal_values(_summary_source_values(workbook, layout, category, 3))
                    for layout in layouts
                ),
                Decimal(0),
            )
            if not _equal_values(quantity.value, expected_quantity):
                errors.append(f"SUMMARY_ALL_INDICES_QUANTITY_VALUE:{quantity.coordinate}")
        cost = sheet.cell(row_number, total_start_column + 3)
        expected_cost = sum(
            (
                _sum_literal_values(_summary_source_values(workbook, layout, category, 4))
                for layout in layouts
            ),
            Decimal(0),
        )
        if not _equal_values(cost.value, expected_cost):
            errors.append(f"SUMMARY_ALL_INDICES_COST_VALUE:{cost.coordinate}")


def _validate_sheet_data(
    sheet, errors: list[str], numeric_targets: dict[str, set[str]]
) -> tuple[int, int]:
    object_count = 0
    drawing_count = 0
    expected_set = set(_EXPECTED_CATEGORIES)
    for start_column in range(2, sheet.max_column + 1, CARD_BLOCK_COLUMN_SPAN):
        value = sheet.cell(2, start_column).value
        if not isinstance(value, str) or "индекс объекта" not in normalize_text(value):
            continue
        object_count += 1
        current_drawing: str | None = None
        categories: list[str] = []
        dual_nonzero = 0
        equal_nonzero = 0
        for row in range(4, sheet.max_row + 1):
            drawing = sheet.cell(row, start_column).value
            category = sheet.cell(row, start_column + 1).value
            if drawing not in (None, ""):
                if not is_plausible_drawing_code(str(drawing)):
                    coordinate = sheet.cell(row, start_column).coordinate
                    errors.append(f"INVALID_DRAWING_CODE:{sheet.title}:{coordinate}:{drawing}")
                if current_drawing is not None:
                    _validate_categories(
                        sheet_name=sheet.title,
                        drawing=current_drawing,
                        categories=categories,
                        errors=errors,
                    )
                current_drawing = str(drawing)
                drawing_count += 1
                categories = []
            if current_drawing is None or category in (None, ""):
                continue
            normalized = normalize_text(str(category))
            if normalized not in expected_set:
                coordinate = sheet.cell(row, start_column + 1).coordinate
                errors.append(f"UNKNOWN_CATEGORY:{sheet.title}:{coordinate}:{category}")
                continue
            categories.append(normalized)
            quantity_cell = sheet.cell(row, start_column + 3)
            cost_cell = sheet.cell(row, start_column + 4)
            contract_quantity_cell = sheet.cell(row, start_column + 5)
            contract_cost_cell = sheet.cell(row, start_column + 6)
            performed_quantity_cell = sheet.cell(row, start_column + 7)
            performed_cost_cell = sheet.cell(row, start_column + 8)
            numeric_targets.setdefault(sheet.title, set()).update(
                {
                    quantity_cell.coordinate,
                    cost_cell.coordinate,
                    contract_quantity_cell.coordinate,
                    contract_cost_cell.coordinate,
                    performed_quantity_cell.coordinate,
                    performed_cost_cell.coordinate,
                }
            )
            quantity = _decimal(quantity_cell.value)
            cost = _decimal(cost_cell.value)
            if quantity not in (None, Decimal(0)) and cost not in (None, Decimal(0)):
                dual_nonzero += 1
                if _equal_nonzero(quantity, cost):
                    equal_nonzero += 1
        if current_drawing is not None:
            _validate_categories(
                sheet_name=sheet.title,
                drawing=current_drawing,
                categories=categories,
                errors=errors,
            )
        if dual_nonzero >= 3 and equal_nonzero / dual_nonzero >= 0.45:
            errors.append(
                f"SUSPICIOUS_IDENTICAL_QUANTITY_COST:{sheet.title}:{equal_nonzero}/{dual_nonzero}"
            )
    return object_count, drawing_count


def validate_card(path: Path, layouts: list[ObjectBlockLayout] | None = None) -> dict[str, object]:
    """Validate a workbook and return machine-readable output publication evidence."""

    errors: list[str] = []
    warnings: list[str] = []
    if not _validate_archive(path, errors):
        return {
            "status": Status.OUTPUT_VALIDATION_FAILED.value,
            "path": str(path),
            "sheets": [],
            "objects": 0,
            "drawings": 0,
            "errors": errors,
            "warnings": warnings,
        }
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
    except (OSError, ValueError, zipfile.BadZipFile) as error:
        errors.append(f"WORKBOOK_REOPEN_FAILED:{error}")
        return {
            "status": Status.OUTPUT_VALIDATION_FAILED.value,
            "path": str(path),
            "sheets": [],
            "objects": 0,
            "drawings": 0,
            "errors": errors,
            "warnings": warnings,
        }
    numeric_targets: dict[str, set[str]] = {}
    try:
        object_count = drawing_count = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "e" or str(cell.value).upper() in _FORMULA_ERRORS:
                        errors.append(f"FORMULA_ERROR:{sheet.title}:{cell.coordinate}:{cell.value}")
            if sheet.title in {SUMMARY_SHEET_NAME, DISCREPANCY_SHEET_NAME}:
                continue
            objects, drawings = _validate_sheet_data(sheet, errors, numeric_targets)
            object_count += objects
            drawing_count += drawings
        if layouts:
            _validate_layout_contract(workbook, layouts, errors)
            _validate_trimmed_template_slots(workbook, layouts, errors)
            _validate_summary_contract(workbook, layouts, errors)
            expected_objects = len({layout.object_index for layout in layouts})
            if object_count != expected_objects:
                errors.append(f"OBJECT_COUNT:{object_count}!={expected_objects}")
            expected_drawings = sum(len(layout.drawing_code_blocks) for layout in layouts)
            if drawing_count != expected_drawings:
                errors.append(f"DRAWING_COUNT:{drawing_count}!={expected_drawings}")
        for sheet_name, coordinate, raw, canonical in find_binary_tail_cells(path, numeric_targets):
            errors.append(
                f"BINARY_FLOAT_SERIALIZATION:{sheet_name}:{coordinate}:{raw}->{canonical}"
            )
        return {
            "status": Status.OK.value if not errors else Status.OUTPUT_VALIDATION_FAILED.value,
            "path": str(path),
            "sheets": workbook.sheetnames,
            "objects": object_count,
            "drawings": drawing_count,
            "errors": errors,
            "warnings": warnings,
        }
    finally:
        workbook.close()

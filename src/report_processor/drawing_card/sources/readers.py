"""Workbook-reader abstraction for OpenXML and XLSB files."""

from __future__ import annotations

import re
import shutil
import tempfile
import zipfile
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from typing import Any, Protocol
from xml.etree import ElementTree as ET

from openpyxl.formula.translate import Translator

try:
    from pyxlsb import open_workbook
except ImportError:  # Optional until the host application enables XLSB support.
    open_workbook = None

from ..models import ManifestEntry
from ..statuses import Status

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CELL_RE = re.compile(r"([A-Z]+)(\d+)")


class WorkbookReader(Protocol):
    def list_sheets(self) -> tuple[str, ...]: ...

    def iter_rows(
        self,
        sheet_name: str,
        *,
        min_row: int = 1,
        max_row: int | None = None,
        max_col: int | None = None,
        selected_columns: Sequence[int] | None = None,
    ) -> Iterator[tuple[tuple[Any, ...], tuple[Any, ...]]]: ...

    def close(self) -> None: ...


@dataclass(slots=True)
class MaterializedSource:
    path: Path
    temporary_dir: tempfile.TemporaryDirectory[str] | None

    def close(self) -> None:
        if self.temporary_dir is not None:
            self.temporary_dir.cleanup()


def _decode_zip_member_name(name: str) -> str:
    try:
        return name.encode("cp437").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


def materialize_entry(entry: ManifestEntry, max_size: int = 2 * 1024**3) -> MaterializedSource:
    container = Path(entry.container_path)
    if entry.source_kind != "zip":
        if entry.source_kind == "directory":
            return MaterializedSource(container / PurePosixPath(entry.logical_path), None)
        return MaterializedSource(container, None)
    path = PurePosixPath(entry.logical_path)
    if path.is_absolute() or ".." in path.parts:
        raise ValueError(Status.UNSAFE_ARCHIVE_PATH.value)
    if entry.size > max_size:
        raise ValueError(Status.VERY_LARGE_ARCHIVE_ENTRY.value)
    temp_dir = tempfile.TemporaryDirectory(prefix="drawing-card-")
    target = Path(temp_dir.name) / path.name
    try:
        with zipfile.ZipFile(container) as archive:
            try:
                info = archive.getinfo(entry.logical_path)
            except KeyError:
                info = next(
                    (
                        item
                        for item in archive.infolist()
                        if _decode_zip_member_name(item.filename) == entry.logical_path
                    ),
                    None,
                )
                if info is None:
                    raise KeyError(f"Archive member not found: {entry.logical_path}") from None
            with archive.open(info) as source, target.open("wb") as output:
                shutil.copyfileobj(source, output, length=1024 * 1024)
    except BaseException:
        temp_dir.cleanup()
        raise
    return MaterializedSource(target, temp_dir)


def _column_number(reference: str) -> int:
    match = _CELL_RE.fullmatch(reference)
    if not match:
        return 0
    value = 0
    for char in match.group(1):
        value = value * 26 + ord(char) - 64
    return value


def _numeric_value(raw: str | None) -> Any:
    if raw is None or raw == "":
        return None
    try:
        value = Decimal(raw)
    except InvalidOperation:
        return raw
    if value == value.to_integral_value():
        return int(value)
    return value


class OpenXmlWorkbookReader:
    """Streams selected OpenXML cells and exposes formulas plus cached values."""

    def __init__(self, path: Path) -> None:
        self.path = path
        self._archive = zipfile.ZipFile(path)
        self._sheet_paths = self._read_sheet_paths()
        self._shared_strings: tuple[str, ...] | None = None

    def _read_sheet_paths(self) -> dict[str, str]:
        workbook = ET.fromstring(self._archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(self._archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            item.attrib["Id"]: item.attrib["Target"]
            for item in relationships.findall(f"{{{_PACKAGE_REL_NS}}}Relationship")
        }
        result: dict[str, str] = {}
        for sheet in workbook.findall(f".//{{{_MAIN_NS}}}sheet"):
            relationship_id = sheet.attrib[f"{{{_REL_NS}}}id"]
            target = targets[relationship_id].lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            result[sheet.attrib["name"]] = target
        return result

    def _load_shared_strings(self) -> tuple[str, ...]:
        if self._shared_strings is not None:
            return self._shared_strings
        if "xl/sharedStrings.xml" not in self._archive.namelist():
            self._shared_strings = ()
            return self._shared_strings
        root = ET.fromstring(self._archive.read("xl/sharedStrings.xml"))
        values: list[str] = []
        for item in root.findall(f"{{{_MAIN_NS}}}si"):
            values.append("".join(node.text or "" for node in item.iter(f"{{{_MAIN_NS}}}t")))
        self._shared_strings = tuple(values)
        return self._shared_strings

    def list_sheets(self) -> tuple[str, ...]:
        return tuple(self._sheet_paths)

    def _cell_cached_value(self, cell: ET.Element) -> Any:
        cell_type = cell.attrib.get("t")
        if cell_type == "inlineStr":
            return "".join(node.text or "" for node in cell.iter(f"{{{_MAIN_NS}}}t"))
        value_node = cell.find(f"{{{_MAIN_NS}}}v")
        raw = None if value_node is None else value_node.text
        if raw is None:
            return None
        if cell_type == "s":
            strings = self._load_shared_strings()
            index = int(raw)
            return strings[index] if 0 <= index < len(strings) else None
        if cell_type in {"str", "e"}:
            return raw
        if cell_type == "b":
            return raw == "1"
        return _numeric_value(raw)

    def iter_rows(
        self,
        sheet_name: str,
        *,
        min_row: int = 1,
        max_row: int | None = None,
        max_col: int | None = None,
        selected_columns: Sequence[int] | None = None,
    ) -> Iterator[tuple[tuple[Any, ...], tuple[Any, ...]]]:
        sheet_path = self._sheet_paths[sheet_name]
        selected = set(selected_columns or ())
        width = max_col or (max(selected) if selected else 0)
        last_row = min_row - 1
        shared_formulas: dict[str, tuple[str, str]] = {}
        with self._archive.open(sheet_path) as stream:
            for _event, element in ET.iterparse(stream, events=("end",)):
                if element.tag != f"{{{_MAIN_NS}}}row":
                    continue
                row_number = int(element.attrib.get("r", last_row + 1))
                if row_number < min_row:
                    element.clear()
                    continue
                if max_row is not None and row_number > max_row:
                    element.clear()
                    break
                while last_row + 1 < row_number:
                    last_row += 1
                    blank = tuple(None for _ in range(width))
                    yield blank, blank
                formula_values = [None] * width
                cached_values = [None] * width
                for cell in element.findall(f"{{{_MAIN_NS}}}c"):
                    column = _column_number(cell.attrib.get("r", ""))
                    if column < 1 or (max_col is not None and column > max_col):
                        continue
                    if selected and column not in selected:
                        continue
                    if column > width:
                        extension = column - width
                        formula_values.extend([None] * extension)
                        cached_values.extend([None] * extension)
                        width = column
                    cached = self._cell_cached_value(cell)
                    formula_node = cell.find(f"{{{_MAIN_NS}}}f")
                    formula = cached
                    if formula_node is not None:
                        formula_text = formula_node.text or ""
                        formula_type = formula_node.attrib.get("t")
                        if formula_type == "shared":
                            shared_index = formula_node.attrib.get("si")
                            cell_reference = cell.attrib.get("r", "")
                            if shared_index and formula_text:
                                shared_formulas[shared_index] = (cell_reference, formula_text)
                            elif shared_index in shared_formulas:
                                origin, base_formula = shared_formulas[shared_index]
                                try:
                                    formula_text = (
                                        Translator("=" + base_formula, origin=origin)
                                        .translate_formula(cell_reference)
                                        .lstrip("=")
                                    )
                                except (TypeError, ValueError):
                                    formula_text = base_formula
                        formula = "=" + formula_text
                    formula_values[column - 1] = formula
                    cached_values[column - 1] = cached
                last_row = row_number
                yield tuple(formula_values), tuple(cached_values)
                element.clear()

    def close(self) -> None:
        self._archive.close()


# Schema inspection already uses the same low-overhead XML reader.
OpenXmlSchemaReader = OpenXmlWorkbookReader


class XlsbWorkbookReader:
    """Reads XLSB values through pyxlsb; formulas are not exposed by this backend."""

    def __init__(self, path: Path) -> None:
        if open_workbook is None:
            raise ValueError("XLSB support is unavailable")
        self._workbook = open_workbook(path)

    def list_sheets(self) -> tuple[str, ...]:
        return tuple(self._workbook.sheets)

    def iter_rows(
        self,
        sheet_name: str,
        *,
        min_row: int = 1,
        max_row: int | None = None,
        max_col: int | None = None,
        selected_columns: Sequence[int] | None = None,
    ) -> Iterator[tuple[tuple[Any, ...], tuple[Any, ...]]]:
        selected = set(selected_columns or ())
        width = max_col or (max(selected) if selected else 0)
        with self._workbook.get_sheet(sheet_name) as sheet:
            for row_number, row in enumerate(sheet.rows(), 1):
                if row_number < min_row:
                    continue
                if max_row is not None and row_number > max_row:
                    break
                values = [None] * width
                for column, cell in enumerate(row, 1):
                    if max_col is not None and column > max_col:
                        break
                    if selected and column not in selected:
                        continue
                    if column > len(values):
                        values.extend([None] * (column - len(values)))
                    values[column - 1] = cell.v
                row_values = tuple(values)
                yield row_values, row_values

    def close(self) -> None:
        self._workbook.close()


def open_schema_reader(path: Path) -> WorkbookReader:
    return open_reader(path)


def open_reader(path: Path) -> WorkbookReader:
    if path.suffix.lower() == ".xlsb":
        return XlsbWorkbookReader(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return OpenXmlWorkbookReader(path)
    raise ValueError(f"Unsupported workbook extension: {path.suffix}")


def value_at(row: Sequence[Any], column: int | None) -> Any:
    if column is None or column < 1 or column > len(row):
        return None
    return row[column - 1]

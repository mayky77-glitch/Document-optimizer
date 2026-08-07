"""Streaming extraction of canonical drawing source rows."""

from __future__ import annotations

import re
from collections.abc import Iterator
from decimal import Decimal
from itertools import islice

from openpyxl.utils import get_column_letter

from report_processor.identifiers import extract_document_index

from ..models import DrawingSourceLocation, DrawingSourceRow, ManifestEntry, SourceSchema
from ..statuses import Status
from .normalization import is_plausible_drawing_code, parse_decimal, stable_id
from .readers import WorkbookReader, value_at


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\u00a0", " ").strip()
    return text or None


def _position_text(value: object) -> str | None:
    # XLSB readers commonly expose an integer position such as 1 as 1.0.
    # Keep true dotted string codes unchanged, but normalize numeric integer cells.
    if isinstance(value, float):
        if not value.is_integer():
            # A numeric fraction cannot be distinguished from a calculated
            # metric (for example 0.108299999...).  It is therefore unsafe as
            # an automatic hierarchy key; real multi-segment positions should
            # be stored as text in the source workbook.
            return None
        value = int(value)
    return _text(value)


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _decimal_or_zero(value: object) -> tuple[Decimal, tuple[str, ...]]:
    parsed, warnings = parse_decimal(value)
    return (parsed if parsed is not None else Decimal(0)), warnings


_SIMPLE_SUBTRACTION_FORMULA_RE = re.compile(r"^=\$?[A-Z]{1,3}\$?\d+(?:-\$?[A-Z]{1,3}\$?\d+)+$")
_FORMULA_REFERENCE_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")


def _correct_dimensionally_invalid_quantity(
    formula: object,
    cached_row: tuple[object, ...],
    columns: dict[str, int],
    *,
    row_number: int,
) -> tuple[Decimal | None, tuple[str, ...]] | None:
    """Evaluate the schema-approved quantity subtraction with quantity operands."""

    base_column = columns.get("remaining_quantity_base")
    invalid_base_column = columns.get("remaining_quantity_invalid_base")
    if base_column is None or invalid_base_column is None or not _is_formula(formula):
        return None
    normalized_formula = str(formula).replace(" ", "").upper()
    if not _SIMPLE_SUBTRACTION_FORMULA_RE.fullmatch(normalized_formula):
        return None, ("DIMENSIONAL_QUANTITY_REPAIR_SIGNATURE_MISMATCH",)
    parsed_references = _FORMULA_REFERENCE_RE.findall(normalized_formula)
    references = [column for column, _row in parsed_references]
    if not references or any(
        int(reference_row) != row_number for _column, reference_row in parsed_references
    ):
        return None, ("DIMENSIONAL_QUANTITY_REPAIR_SIGNATURE_MISMATCH",)
    if references[0] == get_column_letter(base_column):
        # This row already uses the dimensionally correct quantity base. Keep
        # its cached Excel result instead of applying the workbook-level repair.
        return None
    if references[0] != get_column_letter(invalid_base_column):
        return None, ("DIMENSIONAL_QUANTITY_REPAIR_SIGNATURE_MISMATCH",)
    subtrahend_keys = sorted(
        (key for key in columns if key.startswith("remaining_quantity_subtrahend_")),
        key=lambda key: int(key.rsplit("_", 1)[-1]),
    )
    expected_references = [
        get_column_letter(invalid_base_column),
        *(get_column_letter(columns[key]) for key in subtrahend_keys),
    ]
    if references != expected_references:
        return None, ("DIMENSIONAL_QUANTITY_REPAIR_SIGNATURE_MISMATCH",)
    base, warnings = parse_decimal(value_at(cached_row, base_column))
    if base is None:
        return None, (*warnings, "DIMENSIONAL_QUANTITY_REPAIR_BASE_MISSING")
    result = base
    collected_warnings = list(warnings)
    for key in subtrahend_keys:
        value, value_warnings = _decimal_or_zero(value_at(cached_row, columns[key]))
        result -= value
        collected_warnings.extend(value_warnings)
    collected_warnings.append("REMAINING_QUANTITY_REPAIRED_FROM_DIMENSIONAL_FORMULA")
    return result, tuple(collected_warnings)


def extract_rows(
    reader: WorkbookReader,
    entry: ManifestEntry,
    schema: SourceSchema,
    object_index: str | None,
    *,
    max_rows: int | None = None,
    empty_row_limit: int = 200,
    stats: dict[str, int] | None = None,
) -> Iterator[DrawingSourceRow]:
    columns = schema.columns
    max_col = max(columns.values())
    current_drawing: str | None = None
    current_document_index: str | None = None
    filename_extraction = extract_document_index(entry.logical_path)
    filename_document_index = (
        filename_extraction.value.normalized if filename_extraction.value is not None else None
    )
    rows = reader.iter_rows(
        schema.sheet_name,
        min_row=schema.data_start_row,
        max_col=max_col,
        selected_columns=tuple(sorted(set(columns.values()))),
    )
    if max_rows is not None:
        rows = islice(rows, max_rows)
    empty_streak = 0
    for offset, (formula_row, cached_row) in enumerate(rows):
        row_number = schema.data_start_row + offset
        raw_drawing = _text(value_at(cached_row, columns.get("drawing_code")))
        position_code = _position_text(value_at(cached_row, columns.get("position_code")))
        raw_document_index = _text(value_at(cached_row, columns.get("document_index")))
        cost_type_code = _text(value_at(cached_row, columns.get("cost_type_code")))
        work_name = _text(value_at(cached_row, columns.get("work_name")))
        unit = _text(value_at(cached_row, columns.get("unit")))
        formula_drawing = _text(value_at(formula_row, columns.get("drawing_code")))
        quantity_formula = value_at(formula_row, columns.get("remaining_quantity"))
        quantity_cached = value_at(cached_row, columns.get("remaining_quantity"))
        cost_formula = value_at(formula_row, columns.get("remaining_total_cost"))
        cost_cached = value_at(cached_row, columns.get("remaining_total_cost"))
        contract_quantity_formula = value_at(formula_row, columns.get("contract_quantity"))
        contract_quantity_cached = value_at(cached_row, columns.get("contract_quantity"))
        contract_cost_formula = value_at(formula_row, columns.get("contract_total_cost"))
        contract_cost_cached = value_at(cached_row, columns.get("contract_total_cost"))
        performed_quantity_formula = value_at(formula_row, columns.get("performed_quantity"))
        performed_quantity_cached = value_at(cached_row, columns.get("performed_quantity"))
        performed_cost_formula = value_at(formula_row, columns.get("performed_total_cost"))
        performed_cost_cached = value_at(cached_row, columns.get("performed_total_cost"))
        observed = (
            raw_drawing,
            work_name,
            unit,
            formula_drawing,
            quantity_formula,
            quantity_cached,
            cost_formula,
            cost_cached,
            contract_quantity_formula,
            contract_quantity_cached,
            contract_cost_formula,
            contract_cost_cached,
            performed_quantity_formula,
            performed_quantity_cached,
            performed_cost_formula,
            performed_cost_cached,
        )
        if all(value in (None, "") for value in observed):
            empty_streak += 1
            if empty_streak >= empty_row_limit:
                break
            continue
        if empty_streak and stats is not None:
            stats["skipped_empty_rows"] = stats.get("skipped_empty_rows", 0) + empty_streak
        empty_streak = 0
        if raw_drawing is None and formula_drawing and not _is_formula(formula_drawing):
            raw_drawing = formula_drawing
        extraction_warnings: list[str] = []
        if raw_document_index is not None:
            if raw_document_index.casefold() in {"-", "—", "нет"}:
                current_document_index = None
            else:
                extracted_index = extract_document_index(raw_document_index)
                if extracted_index.value is None:
                    current_document_index = None
                    extraction_warnings.append("INVALID_DOCUMENT_INDEX")
                else:
                    current_document_index = extracted_index.value.normalized
        if raw_drawing and not is_plausible_drawing_code(raw_drawing):
            extraction_warnings.append(f"IGNORED_NON_DRAWING_CELL:{raw_drawing}")
            raw_drawing = None
        if raw_drawing and not work_name:
            current_drawing = raw_drawing
            if stats is not None:
                stats["skipped_header_rows"] = stats.get("skipped_header_rows", 0) + 1
            continue
        if raw_drawing:
            current_drawing = raw_drawing
        drawing = raw_drawing or current_drawing
        if not work_name:
            if stats is not None:
                stats["skipped_header_rows"] = stats.get("skipped_header_rows", 0) + 1
            continue
        quantity, quantity_warnings = parse_decimal(quantity_cached)
        corrected_quantity = _correct_dimensionally_invalid_quantity(
            quantity_formula,
            cached_row,
            columns,
            row_number=row_number,
        )
        if corrected_quantity is not None:
            quantity, quantity_warnings = corrected_quantity
        cost, cost_warnings = parse_decimal(cost_cached)
        contract_quantity, contract_quantity_warnings = _decimal_or_zero(contract_quantity_cached)
        contract_cost, contract_cost_warnings = _decimal_or_zero(contract_cost_cached)
        performed_quantity, performed_quantity_warnings = _decimal_or_zero(
            performed_quantity_cached
        )
        performed_cost, performed_cost_warnings = _decimal_or_zero(performed_cost_cached)
        # A cumulative contract value is comparable only when both semantic
        # roles are explicit: whole-period performed plus residual.  Do not use
        # a month/intermediate performed block as a substitute.
        if (
            "contract_quantity" not in columns
            or (
                "PERIOD_ROLES_AUTHORITATIVE:contract_quantity" in schema.warnings
                and "remaining_quantity_base" not in columns
            )
        ) and {
            "performed_quantity",
            "remaining_quantity",
        }.issubset(columns):
            contract_quantity = performed_quantity + (quantity or Decimal(0))
            contract_quantity_warnings += ("CONTRACT_QUANTITY_DERIVED_FROM_PERFORMED_AND_RESIDUAL",)
        if (
            "contract_total_cost" not in columns
            or (
                "PERIOD_ROLES_AUTHORITATIVE:contract_total_cost" in schema.warnings
                and "remaining_quantity_base" not in columns
            )
        ) and {
            "performed_total_cost",
            "remaining_total_cost",
        }.issubset(columns):
            contract_cost = performed_cost + (cost or Decimal(0))
            contract_cost_warnings += ("CONTRACT_TOTAL_COST_DERIVED_FROM_PERFORMED_AND_RESIDUAL",)
        warnings = extraction_warnings + list(
            quantity_warnings
            + cost_warnings
            + contract_quantity_warnings
            + contract_cost_warnings
            + performed_quantity_warnings
            + performed_cost_warnings
        )
        metric_values = (
            (quantity_formula, quantity_cached),
            (cost_formula, cost_cached),
            (contract_quantity_formula, contract_quantity_cached),
            (contract_cost_formula, contract_cost_cached),
            (performed_quantity_formula, performed_quantity_cached),
            (performed_cost_formula, performed_cost_cached),
        )
        if any(_is_formula(formula) and cached is None for formula, cached in metric_values):
            warnings.append(Status.FORMULA_WITHOUT_CACHED_VALUE)
        if not drawing:
            warnings.append(Status.DRAWING_CODE_NOT_FOUND)
        coordinate_columns = tuple(sorted(set(columns.values())))
        coordinates = tuple(
            f"{get_column_letter(column)}{row_number}" for column in coordinate_columns
        )
        row_id = stable_id(entry.file_id, schema.sheet_name, row_number)
        status = Status.OK.value if not warnings else Status.WARNING.value
        yield DrawingSourceRow(
            row_id=row_id,
            location=DrawingSourceLocation(
                file_id=entry.file_id,
                filename=entry.filename,
                sheet_name=schema.sheet_name,
                row_number=row_number,
                coordinates=coordinates,
                document_index=current_document_index or filename_document_index,
            ),
            object_index_raw=object_index,
            position_code_raw=position_code,
            cost_type_code_raw=cost_type_code,
            drawing_code_raw=drawing,
            work_name_raw=work_name,
            unit_raw=unit,
            remaining_quantity=quantity,
            remaining_total_cost=cost,
            formula_values=tuple(formula for formula, _cached in metric_values),
            cached_values=tuple(cached for _formula, cached in metric_values),
            source_document_type=entry.document_type,
            source_period=entry.period,
            source_revision=entry.revision,
            status=status,
            warnings=tuple(dict.fromkeys(str(item) for item in warnings)),
            contract_quantity=contract_quantity,
            contract_total_cost=contract_cost,
            performed_quantity=performed_quantity,
            performed_total_cost=performed_cost,
        )

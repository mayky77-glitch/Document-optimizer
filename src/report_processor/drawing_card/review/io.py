"""Manual-review workbook export and import."""

from __future__ import annotations

import json
import os
import tempfile
import zipfile
from decimal import Decimal
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..matching.matcher import ReviewApproval
from ..models import DrawingSourceRow, MatchDecision, TargetWorkCategory
from ..sources.normalization import normalize_text

_HEADERS = (
    "Review ID",
    "Индекс объекта",
    "Исходный файл",
    "Лист",
    "Строка",
    "Шифр чертежа",
    "Исходное наименование",
    "Ед. изм.",
    "Остаток количества",
    "Остаток стоимости",
    "Предлагаемая категория",
    "Решение по количеству",
    "Решение по стоимости",
    "Уверенность",
    "Подтверждённые похожие примеры",
    "Причина",
    "Решение пользователя",
    "Комментарий пользователя",
)
_ACTIONS = ("approve", "reject", "change_category", "quantity_only", "cost_only", "skip")
_CATEGORY_ACTIONS = frozenset({"approve", "change_category", "quantity_only", "cost_only"})


def _safe_excel_value(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _verify_xlsx(path: Path) -> None:
    """Verify that the generated package contains parseable XML and reopens."""
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if name.endswith(".xml"):
                ElementTree.fromstring(archive.read(name))
    check = load_workbook(path, read_only=True, data_only=False)
    check.close()


def _atomic_save_review(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix=path.stem + ".",
        suffix=".tmp.xlsx",
        dir=path.parent,
        delete=False,
    ) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        _verify_xlsx(temporary)
        os.replace(temporary, path)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise


def export_manual_review(
    path: Path,
    rows: list[DrawingSourceRow],
    decisions: list[MatchDecision],
) -> int:
    row_map = {row.row_id: row for row in rows}
    review_decisions = [item for item in decisions if item.requires_manual_review]
    count = len(review_decisions)
    write_only = count > 5000
    workbook = Workbook(write_only=write_only)
    sheet = workbook.create_sheet("Review") if write_only else workbook.active
    if not write_only:
        sheet.title = "Review"
        sheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="B7C9D6")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if write_only:
        header_cells = []
        for value in _HEADERS:
            cell = WriteOnlyCell(sheet, value=value)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_cells.append(cell)
        sheet.append(header_cells)
    else:
        sheet.append(_HEADERS)
        sheet.row_dimensions[1].height = 54
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for decision in review_decisions:
        row = row_map.get(decision.row_id)
        if row is None:
            continue
        confidence_values = [
            value
            for value in (decision.quantity_confidence, decision.cost_confidence)
            if value is not None
        ]
        values = (
            row.row_id,
            row.object_index_raw,
            row.location.filename,
            row.location.sheet_name,
            row.location.row_number,
            row.drawing_code_raw,
            row.work_name_raw,
            row.unit_raw,
            row.remaining_quantity,
            row.remaining_total_cost,
            decision.category.value if decision.category else None,
            decision.quantity_decision,
            decision.cost_decision,
            min(confidence_values) if confidence_values else None,
            ", ".join(decision.evidence_ids),
            decision.reason,
            None,
            None,
        )
        sheet.append(tuple(_safe_excel_value(value) for value in values))

    widths = [52, 13, 34, 18, 9, 44, 50, 11, 17, 18, 25, 17, 17, 12, 34, 44, 22, 34]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    if not write_only:
        sheet.freeze_panes = "A2"
        if count:
            decision_validation = DataValidation(
                type="list",
                formula1='"' + ",".join(_ACTIONS) + '"',
                allow_blank=True,
            )
            category_validation = DataValidation(
                type="list",
                formula1='"' + ",".join(item.value for item in TargetWorkCategory) + '"',
                allow_blank=True,
            )
            sheet.add_data_validation(decision_validation)
            sheet.add_data_validation(category_validation)
            decision_validation.add(f"Q2:Q{count + 1}")
            category_validation.add(f"K2:K{count + 1}")
            sheet.auto_filter.ref = f"A1:R{count + 1}"
        for row_values in sheet.iter_rows(min_row=2, max_row=count + 1):
            for cell in row_values:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if count:
            for index, cell in enumerate(sheet[f"I2:I{count + 1}"], start=0):
                quantity = review_decisions[index]
                row = row_map.get(quantity.row_id)
                value = row.remaining_quantity if row is not None else None
                cell[0].number_format = (
                    "0" if isinstance(value, Decimal) and value == value.to_integral() else "0.###"
                )
            for cell in sheet[f"J2:J{count + 1}"]:
                cell[0].number_format = "#,##0.00"
            for cell in sheet[f"N2:N{count + 1}"]:
                cell[0].number_format = "0.00"

    try:
        _atomic_save_review(workbook, path)
    finally:
        workbook.close()
    return count


def _review_approval(row_id: object, action: object, category: object) -> ReviewApproval | None:
    if not row_id or not action:
        return None
    row_id_text = str(row_id)
    action_text = normalize_text(str(action))
    if action_text not in _ACTIONS:
        raise ValueError(f"Unsupported review action for {row_id}: {action}")
    parsed_category = None
    if category:
        try:
            parsed_category = TargetWorkCategory(str(category))
        except ValueError as error:
            raise ValueError(f"Unsupported review category for {row_id}") from error
    if action_text in _CATEGORY_ACTIONS and parsed_category is None:
        raise ValueError(f"Review action {action_text} requires a valid category for {row_id}")
    return ReviewApproval(row_id_text, action_text, parsed_category)


def _import_json_review_approvals(path: Path) -> dict[str, ReviewApproval]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ValueError(f"Review decisions JSON is invalid: {error}") from error
    if not isinstance(payload, dict):
        raise ValueError("Review decisions JSON must be an object keyed by review ID")
    approvals: dict[str, ReviewApproval] = {}
    for key, item in payload.items():
        if not isinstance(item, dict):
            raise ValueError(f"Review decision for {key} must be an object")
        row_id = item.get("row_id", key)
        if str(row_id) != str(key):
            raise ValueError(f"Review decision key does not match row_id: {key}")
        approval = _review_approval(row_id, item.get("action"), item.get("category"))
        if approval is None:
            raise ValueError(f"Review decision for {key} must include row_id and action")
        approvals[approval.row_id] = approval
    return approvals


def import_review_approvals(path: Path | None) -> dict[str, ReviewApproval]:
    if path is None:
        return {}
    if not path.exists():
        raise FileNotFoundError(path)
    if path.suffix.lower() == ".json":
        return _import_json_review_approvals(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active
        headers = {normalize_text(cell.value): index for index, cell in enumerate(sheet[1], 1)}
        id_col = headers.get(normalize_text("Review ID"))
        action_col = headers.get(normalize_text("Решение пользователя"))
        category_col = headers.get(normalize_text("Предлагаемая категория"))
        if not id_col or not action_col:
            raise ValueError("Review workbook has no required decision columns")
        approvals: dict[str, ReviewApproval] = {}
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row_id = values[id_col - 1]
            action = values[action_col - 1]
            approval = _review_approval(
                row_id,
                action,
                values[category_col - 1] if category_col else None,
            )
            if approval is not None:
                approvals[approval.row_id] = approval
        return approvals
    finally:
        workbook.close()


def review_approvals_payload(
    approvals: dict[str, ReviewApproval],
) -> dict[str, dict[str, str | None]]:
    """Build the validated JSON contract consumed by ``--review-decisions``."""
    return {
        row_id: {
            "row_id": approval.row_id,
            "action": approval.action,
            "category": approval.category.value if approval.category else None,
        }
        for row_id, approval in approvals.items()
    }


def append_approved_examples(
    review_path: Path,
    examples_path: Path,
    *,
    confirmed_by: str,
    rule_version: str,
) -> int:
    workbook = load_workbook(review_path, read_only=True, data_only=True)
    created: list[dict[str, object]] = []
    try:
        sheet = workbook["Review"] if "Review" in workbook.sheetnames else workbook.active
        headers = {str(cell.value): index for index, cell in enumerate(sheet[1], 1)}
        for values in sheet.iter_rows(min_row=2, values_only=True):
            action = values[headers["Решение пользователя"] - 1]
            if action not in {"approve", "change_category", "quantity_only", "cost_only"}:
                continue
            category = values[headers["Предлагаемая категория"] - 1]
            if not category:
                continue
            source_text = str(values[headers["Исходное наименование"] - 1] or "")
            review_id = str(values[headers["Review ID"] - 1])
            created.append(
                {
                    "example_id": f"review-{review_id[:16]}",
                    "source_text": source_text,
                    "normalized_text": normalize_text(source_text),
                    "category": str(category),
                    "quantity_decision": "include"
                    if action in {"approve", "change_category", "quantity_only"}
                    else "exclude",
                    "cost_decision": "include"
                    if action in {"approve", "change_category", "cost_only"}
                    else "exclude",
                    "unit": values[headers["Ед. изм."] - 1],
                    "source_type": None,
                    "confirmed": True,
                    "confirmed_by": confirmed_by,
                    "rule_version": rule_version,
                }
            )
    finally:
        workbook.close()
    if created:
        examples_path.parent.mkdir(parents=True, exist_ok=True)
        with examples_path.open("a", encoding="utf-8") as stream:
            for item in created:
                stream.write(json.dumps(item, ensure_ascii=False) + "\n")
    return len(created)

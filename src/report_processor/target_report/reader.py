"""Read a target workbook without changing its OOXML package or openpyxl views."""

from __future__ import annotations

import hashlib
from collections.abc import Iterable
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl.utils import get_column_letter

from report_processor.excel import DualWorkbookSession
from report_processor.excel.workbook_session import validate_dual_workbook_session
from report_processor.schema import LogicalColumn, SheetType, WorkbookSchema, WorksheetSchema

from .models import (
    PackageSanitizationPlan,
    StructuralMutationPlan,
    TargetCellSnapshot,
    TargetColumnBinding,
    TargetDiagnostic,
    TargetFormulaSnapshot,
    TargetNumericCell,
    TargetObjectBlock,
    TargetPeriodIdentity,
    TargetReportOverride,
    TargetReportReadRequest,
    TargetReportResult,
    TargetReportRow,
    TargetReportSchema,
    TargetSourceFingerprint,
    TargetWorksheetSnapshot,
    WritableCellPlan,
)
from .ooxml import (
    RawCellLexemes,
    formula_caches_trusted,
    package_entries,
    read_sheet_comments,
    read_sheet_lexemes,
    read_sheet_structure,
)

_VERSION = "TargetReportSchema-9.0"
_ROW_VERSION = "TargetReportRow-9.0"
_WRITE_VERSION = "WritableCellPlan-1.0"
_STRUCTURAL_VERSION = "StructuralMutationPlan-1.0"
_SANITIZATION_VERSION = "PackageSanitizationPlan-1.0"


def _fingerprint(path: Path, source_file_id: str) -> TargetSourceFingerprint:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1_048_576), b""):
            digest.update(chunk)
    return TargetSourceFingerprint(
        "sha256", digest.hexdigest(), path.stat().st_size, source_file_id
    )


def _diagnostic(code: str, message: str, *, sheet_name: str | None = None) -> TargetDiagnostic:
    return TargetDiagnostic(code, "error", message, sheet_name)


def _selected_schemas(
    workbook_schema: WorkbookSchema,
    request: TargetReportReadRequest,
    fingerprint: TargetSourceFingerprint,
) -> tuple[tuple[WorksheetSchema, ...], tuple[TargetDiagnostic, ...]]:
    override = request.override
    if override is not None and override.source_fingerprint != fingerprint.value:
        return (), (
            _diagnostic("OVERRIDE_SOURCE_FINGERPRINT_MISMATCH", "Override belongs to other bytes"),
        )
    requested = request.sheet_names or (
        (override.sheet_name,) if override and override.sheet_name else None
    )
    by_name = {item.sheet_name: item for item in workbook_schema.worksheets}
    if requested:
        missing = tuple(name for name in requested if name not in by_name)
        if missing:
            return (), tuple(
                _diagnostic("TARGET_SHEET_NOT_IN_SCHEMA", name, sheet_name=name) for name in missing
            )
        return tuple(by_name[name] for name in requested), ()
    # UNKNOWN is deliberately not promoted from a generic semantic hint.
    candidates = tuple(
        item for item in workbook_schema.worksheets if item.sheet_type != SheetType.UNKNOWN
    )
    if not candidates:
        return (), (
            _diagnostic(
                "AMBIGUOUS_TARGET_SHEET_REQUIRES_OVERRIDE",
                "No typed target sheet; provide an override with an exact fingerprint",
            ),
        )
    return candidates, ()


def _bindings(
    schema: WorksheetSchema, override: TargetReportOverride | None
) -> tuple[TargetColumnBinding, ...]:
    if override and override.column_bindings:
        return override.column_bindings
    result: list[TargetColumnBinding] = []
    for column in schema.columns:
        if column.column_index is None or column.status != "OK":
            continue
        result.append(
            TargetColumnBinding(
                column.logical_column,
                column.column_index,
                column.column_letter or get_column_letter(column.column_index),
                column.header_text,
                "WORKBOOK_SCHEMA",
            )
        )
    return tuple(result)


def _semantic_recovery(
    session: DualWorkbookSession, schema: WorksheetSchema
) -> tuple[tuple[TargetColumnBinding, ...], int] | None:
    """Recognize only the documented additional-report table, never a generic UNKNOWN."""

    if schema.sheet_type != SheetType.UNKNOWN:
        return None
    sheet = session.formula_workbook[schema.sheet_name]
    for row_number in range(1, min(int(sheet.max_row or 0), 30) + 1):
        values = [
            str(getattr(sheet.cell(row_number, column), "value", "") or "").casefold()
            for column in range(1, 12)
        ]
        if not ("№ п/п" in values and any("наименование этапа" in item for item in values)):
            continue
        if not any("оперативная отчетность" in item for item in values):
            continue
        if not any("документальная отчетность" in item for item in values):
            continue
        return (
            (
                TargetColumnBinding(
                    LogicalColumn.ROW_NUMBER, 1, "A", values[0], "SEMANTIC_RECOVERY"
                ),
                TargetColumnBinding(
                    LogicalColumn.WORK_NAME, 2, "B", values[1], "SEMANTIC_RECOVERY"
                ),
                TargetColumnBinding(LogicalColumn.UNIT, 3, "C", values[2], "SEMANTIC_RECOVERY"),
                TargetColumnBinding(
                    LogicalColumn.CURRENT_PERIOD_QUANTITY, 4, "D", values[3], "SEMANTIC_RECOVERY"
                ),
                TargetColumnBinding(
                    LogicalColumn.CUMULATIVE_QUANTITY, 5, "E", values[4], "SEMANTIC_RECOVERY"
                ),
                TargetColumnBinding(
                    LogicalColumn.CUMULATIVE_COST, 6, "F", values[5], "SEMANTIC_RECOVERY"
                ),
            ),
            row_number + 2,
        )
    return None


def _period(
    bindings: Iterable[TargetColumnBinding], override: TargetReportOverride | None
) -> TargetPeriodIdentity:
    if override and override.period_identity:
        return override.period_identity
    columns = {item.logical_column: item for item in bindings}
    current = columns.get(LogicalColumn.CURRENT_PERIOD_QUANTITY) or columns.get(
        LogicalColumn.CURRENT_PERIOD_COST
    )
    cumulative = columns.get(LogicalColumn.CUMULATIVE_QUANTITY) or columns.get(
        LogicalColumn.CUMULATIVE_COST
    )
    return TargetPeriodIdentity(
        current_period=current.header_text if current else None,
        cumulative_period=cumulative.header_text if cumulative else None,
        status="OK" if current and cumulative else "PERIOD_UNRESOLVED",
    )


def _decimal(lexeme: str | None, cell_type: str | None) -> tuple[Decimal | None, str]:
    if lexeme is None or cell_type not in {None, "n", "inlineStr", "str"}:
        return None, "NOT_NUMERIC_OOXML_CELL"
    try:
        result = Decimal(lexeme)
    except InvalidOperation:
        return None, "INVALID_NUMERIC_LEXEME"
    return (result, "OK") if result.is_finite() else (None, "NON_FINITE_NUMERIC_LEXEME")


def _cell_snapshot(
    coordinate: str,
    formula_cell,
    value_cell,
    lexemes: RawCellLexemes | None,
    comment_text: str | None,
    caches_trusted: bool,
) -> TargetCellSnapshot:
    formula_value = getattr(formula_cell, "value", None)
    cached_value = getattr(value_cell, "value", None)
    is_formula = getattr(formula_cell, "data_type", None) == "f" or (
        isinstance(formula_value, str) and formula_value.startswith("=")
    )
    raw_lexeme = lexemes.value if lexemes else None
    numeric, numeric_status = _decimal(raw_lexeme, lexemes.cell_type if lexemes else None)
    formula = None
    if is_formula:
        formula = TargetFormulaSnapshot(
            formula=str(formula_value) if formula_value is not None else None,
            cached_value=cached_value,
            formula_data_type=getattr(formula_cell, "data_type", None),
            cached_data_type=getattr(value_cell, "data_type", None),
            cache_state=(
                "CACHE_UNTRUSTED"
                if not caches_trusted
                else "FORMULA_WITH_CACHED_VALUE"
                if cached_value is not None
                else "FORMULA_WITHOUT_CACHED_VALUE"
            ),
            raw_formula_lexeme=lexemes.formula if lexemes else None,
            raw_cached_lexeme=raw_lexeme,
        )
    comment = getattr(formula_cell, "comment", None)
    return TargetCellSnapshot(
        coordinate=coordinate,
        raw_value=formula_value,
        raw_lexeme=raw_lexeme,
        numeric_value=numeric,
        style_id=getattr(formula_cell, "style_id", None),
        number_format=getattr(formula_cell, "number_format", None),
        comment_text=getattr(comment, "text", None) or comment_text,
        formula=formula,
        status="OK" if numeric_status in {"OK", "NOT_NUMERIC_OOXML_CELL"} else numeric_status,
    )


def _worksheet_snapshot(
    sheet, schema: WorksheetSchema, source_path: Path
) -> TargetWorksheetSnapshot:
    structure = read_sheet_structure(source_path, schema.sheet_name)
    return TargetWorksheetSnapshot(
        schema.sheet_name,
        schema.sheet_type,
        structure.dimensions,
        structure.merged_ranges,
        structure.auto_filter_ref,
        read_sheet_comments(source_path, schema.sheet_name),
        structure.freeze_panes,
        sheet.sheet_state != "visible",
    )


def _rows_for_sheet(
    schema: WorksheetSchema,
    bindings: tuple[TargetColumnBinding, ...],
    session: DualWorkbookSession,
    include_empty_rows: bool,
    data_start_row: int | None = None,
    max_rows: int | None = None,
    caches_trusted: bool = True,
) -> tuple[TargetReportRow, ...]:
    start_row = data_start_row or schema.data_start_row
    if start_row is None or not bindings:
        return ()
    formula_sheet = session.formula_workbook[schema.sheet_name]
    value_sheet = session.value_workbook[schema.sheet_name]
    lexemes = read_sheet_lexemes(session.source.local_path, schema.sheet_name)
    comments = dict(read_sheet_comments(session.source.local_path, schema.sheet_name))
    rows: list[TargetReportRow] = []
    for row_number in range(start_row, int(formula_sheet.max_row or 0) + 1):
        if max_rows is not None and len(rows) >= max_rows:
            break
        cells = tuple(
            (
                binding.logical_column,
                _cell_snapshot(
                    f"{binding.column_letter}{row_number}",
                    formula_sheet.cell(row_number, binding.column_index),
                    value_sheet.cell(row_number, binding.column_index),
                    lexemes.get(f"{binding.column_letter}{row_number}"),
                    comments.get(f"{binding.column_letter}{row_number}"),
                    caches_trusted,
                ),
            )
            for binding in bindings
        )
        if not include_empty_rows and all(item.raw_value is None for _, item in cells):
            continue
        values = {key: item.raw_value for key, item in cells}
        numeric = {
            key: TargetNumericCell(
                cell.numeric_value,
                cell.raw_lexeme,
                cell.formula.cache_state if cell.formula else "NOT_FORMULA",
                cell.status,
            )
            for key, cell in cells
        }
        row_kind = (
            "OBJECT_PROCESS" if values.get(LogicalColumn.ROW_NUMBER) is not None else "SUMMARY"
        )
        rows.append(
            TargetReportRow(
                _ROW_VERSION,
                schema.sheet_name,
                schema.sheet_type,
                row_number,
                _text(values.get(LogicalColumn.OBJECT_CODE)),
                _text(values.get(LogicalColumn.OBJECT_NAME)),
                _text(values.get(LogicalColumn.POSITION_CODE)),
                _text(values.get(LogicalColumn.WORK_NAME)),
                cells,
                "OK",
                row_kind=row_kind,
                scope="SELECTED_STAGE" if values.get(LogicalColumn.STAGE) else "UNSCOPED",
                document_index_raw=_text(values.get(LogicalColumn.DOCUMENT_INDEX)),
                document_index_normalized=_text(values.get(LogicalColumn.DOCUMENT_INDEX)),
                stage=_text(values.get(LogicalColumn.STAGE)),
                subobject_code=_text(values.get(LogicalColumn.SUBOBJECT_CODE)),
                subobject_name=_text(values.get(LogicalColumn.SUBOBJECT_NAME)),
                unit=_text(values.get(LogicalColumn.UNIT)),
                document_quantity=numeric.get(LogicalColumn.CUMULATIVE_QUANTITY),
                selected_quantity=numeric.get(LogicalColumn.CURRENT_PERIOD_QUANTITY),
                document_cost=numeric.get(LogicalColumn.CUMULATIVE_COST),
                selected_cost=numeric.get(LogicalColumn.CURRENT_PERIOD_COST),
                writable=row_kind == "OBJECT_PROCESS" and caches_trusted,
            )
        )
    return tuple(rows)


def _text(value: object) -> str | None:
    return str(value).strip() if value is not None and str(value).strip() else None


def _object_blocks(rows: Iterable[TargetReportRow]) -> tuple[TargetObjectBlock, ...]:
    blocks: list[TargetObjectBlock] = []
    active: TargetObjectBlock | None = None
    for row in rows:
        key = (row.sheet_name, row.object_code, row.object_name)
        if active is None or key != (active.sheet_name, active.object_code, active.object_name):
            if active is not None:
                blocks.append(active)
            active = TargetObjectBlock(
                row.sheet_name, row.row_number, row.row_number, row.object_code, row.object_name
            )
        else:
            active = TargetObjectBlock(
                active.sheet_name,
                active.start_row,
                row.row_number,
                active.object_code,
                active.object_name,
            )
    if active is not None:
        blocks.append(active)
    return tuple(blocks)


def read_target_report(
    session: DualWorkbookSession,
    workbook_schema: WorkbookSchema,
    request: TargetReportReadRequest,
) -> TargetReportResult:
    """Build TargetReport-9.0 from validated read-only workbook projections."""

    validate_dual_workbook_session(session)
    stat_before = session.source.local_path.stat()
    fingerprint = _fingerprint(session.source.local_path, session.source.original_file_id)
    selected, diagnostics = _selected_schemas(workbook_schema, request, fingerprint)
    if (
        not selected
        and diagnostics
        and diagnostics[0].code == "AMBIGUOUS_TARGET_SHEET_REQUIRES_OVERRIDE"
    ):
        recovered = tuple(
            item
            for item in workbook_schema.worksheets
            if _semantic_recovery(session, item) is not None
        )
        if len(recovered) == 1:
            selected, diagnostics = recovered, ()
    snapshots = tuple(
        _worksheet_snapshot(
            session.formula_workbook[item.sheet_name], item, session.source.local_path
        )
        for item in selected
    )
    resolved = []
    for item in selected:
        recovered = _semantic_recovery(session, item)
        item_bindings = _bindings(item, request.override)
        if not item_bindings and recovered:
            item_bindings, start_row = recovered
        else:
            start_row = item.data_start_row
        resolved.append((item, item_bindings, start_row))
    bindings = tuple(binding for _, item_bindings, _ in resolved for binding in item_bindings)
    period = request.selected_period_identity or _period(bindings, request.override)
    caches_trusted = formula_caches_trusted(session.source.local_path)
    rows = tuple(
        row
        for item, item_bindings, start_row in resolved
        for row in _rows_for_sheet(
            item,
            item_bindings,
            session,
            request.include_empty_rows,
            start_row,
            request.max_rows,
            caches_trusted,
        )
    )
    blocks = _object_blocks(rows)
    stat_after = session.source.local_path.stat()
    if (stat_before.st_size, stat_before.st_mtime_ns) != (
        stat_after.st_size,
        stat_after.st_mtime_ns,
    ):
        diagnostics += (
            _diagnostic("SOURCE_CHANGED_DURING_READ", "Source bytes changed during read"),
        )
    status = "OK" if not diagnostics else diagnostics[0].code
    schema = TargetReportSchema(
        _VERSION,
        fingerprint,
        period,
        bindings,
        snapshots,
        blocks,
        status,
        diagnostics,
        session.source.original_file_id,
        session.filename,
        fingerprint.digest,
        "1" if period.status == "OK" else "0",
    )
    cell_plans = tuple(
        WritableCellPlan(
            _WRITE_VERSION, row.sheet_name, cell.coordinate, fingerprint.value, cell.raw_lexeme
        )
        for row in rows
        for logical_column, cell in row.cells
        if not diagnostics
        and caches_trusted
        and (
            (row.writable and cell.numeric_value is not None)
            or (row.row_kind == "SUMMARY" and cell.formula is not None)
        )
        and (
            logical_column
            in {LogicalColumn.CURRENT_PERIOD_QUANTITY, LogicalColumn.CURRENT_PERIOD_COST}
            or row.row_kind == "SUMMARY"
        )
    )
    return TargetReportResult(
        schema,
        rows,
        cell_plans,
        StructuralMutationPlan(_STRUCTURAL_VERSION, fingerprint.value),
        PackageSanitizationPlan(
            _SANITIZATION_VERSION, fingerprint.value, package_entries(session.source.local_path)
        ),
        status,
        diagnostics,
    )

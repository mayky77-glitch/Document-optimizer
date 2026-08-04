"""Bounded, read-only extraction of comparable facts from KS-2 workbooks."""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path, PurePosixPath

import openpyxl

from report_processor.schema.text_normalization import clean_display_text, normalize_header_text

from .models import PackageIssue, PackageWorkbookFacts, WorkbookRowFact, WorkbookSheetFacts

_HEADER_ROWS = 100
_HEADER_COLUMNS = 128
_EMPTY_ROW_LIMIT = 20
_ACT_RE = re.compile(r"\b(?:акт|кс\s*[- ]?2)\s*(?:№|n(?:o|омер)?\.?)?\s*([\w./-]+)", re.IGNORECASE)
_PERIOD_RE = re.compile(r"(?:отчетн\w*\s+)?период\w*\s*(?:с|за|:)?\s*([^\n]{3,80})", re.IGNORECASE)
_PERIOD_VALUE_RE = re.compile(
    r"^(?:\d{1,2}[./-]\d{1,2}[./-]\d{2,4}|\d{4}[./-]\d{1,2}[./-]\d{1,2})$"
)
_OBJECT_RE = re.compile(r"(?:код|шифр)\s+объект\w*\s*(?:[:№-]\s*)?([\w./-]+)", re.IGNORECASE)


def _safe_relative_path(value: PurePosixPath | Path | str) -> PurePosixPath:
    path = PurePosixPath(value)
    if path.is_absolute() or ".." in path.parts or path == PurePosixPath("."):
        raise ValueError("workbook path must be a safe relative path")
    return path


def _resolve_workbook(source_root: Path, relative_path: PurePosixPath) -> Path:
    root = Path(source_root)
    if root.is_symlink():
        raise ValueError("symlinked source root is not allowed")
    resolved_root = root.resolve(strict=True)
    candidate = resolved_root.joinpath(*relative_path.parts)
    if candidate.is_symlink():
        raise ValueError("symlinked workbook input is not allowed")
    resolved = candidate.resolve(strict=True)
    try:
        resolved.relative_to(resolved_root)
    except ValueError as error:
        raise ValueError("workbook path escapes the source root") from error
    if not resolved.is_file() or resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("workbook must be an .xlsx or .xlsm regular file")
    return resolved


def _decimal(value: object) -> Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    if isinstance(value, (int, float)):
        try:
            parsed = Decimal(str(value))
        except InvalidOperation:
            return None
        return parsed if parsed.is_finite() else None
    text = clean_display_text(value).replace("\u00a0", " ").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        parsed = Decimal(text)
    except InvalidOperation:
        return None
    return parsed if parsed.is_finite() else None


def _header_field(text: str) -> str | None:
    normalized = normalize_header_text(text)
    tokens = set(normalized.split())
    if not tokens:
        return None
    if "чертеж" in normalized and ("шифр" in tokens or "код" in tokens):
        return "drawing_code"
    if "обоснован" in normalized:
        return "basis"
    if ("код" in tokens or "шифр" in tokens) and "объект" in normalized:
        return "object_code"
    if "единица" in tokens or ("ед" in tokens and "изм" in tokens):
        return "unit"
    if "наименование" in tokens and ("работ" in normalized or "затрат" in normalized):
        return "work_name"
    if "позици" in normalized or ("номер" in tokens and "работ" in normalized):
        return "work_code"
    if "количество" in tokens or "объем" in tokens:
        return "quantity"
    if "стоимость" in tokens and not ({"единицу", "расценка", "цена"} & tokens):
        return "total_cost"
    return None


def _candidate_headers(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]] | None:
    best: tuple[int, int, dict[str, int]] | None = None
    for end_index, _row in enumerate(rows):
        for span in (1, 2, 3):
            start_index = end_index - span + 1
            if start_index < 0:
                continue
            columns: dict[str, int] = {}
            max_width = max(len(value) for value in rows[start_index : end_index + 1])
            for column_index in range(max_width):
                parts = [
                    clean_display_text(candidate[column_index])
                    for candidate in rows[start_index : end_index + 1]
                    if column_index < len(candidate) and clean_display_text(candidate[column_index])
                ]
                field = _header_field(" ".join(dict.fromkeys(parts)))
                if field is not None and field not in columns:
                    columns[field] = column_index
            score = len(columns) + int("work_code" in columns) + int("work_name" in columns)
            if "work_name" not in columns or score < 3:
                continue
            candidate = (score, end_index + 1, columns)
            if (
                best is None
                or candidate[0] > best[0]
                or (candidate[0] == best[0] and candidate[1] < best[1])
            ):
                best = candidate
    return None if best is None else (best[1], best[2])


def _metadata_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_display_text(value)


def _period_date(value: object) -> str | None:
    if normalize_header_text(value) in {"с", "по"}:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    text = clean_display_text(value)
    return text if _PERIOD_VALUE_RE.fullmatch(text) else None


def _structural_period(rows: list[tuple[object, ...]]) -> str | None:
    """Read the common ``Отчетный период / с / по`` cell layout by coordinates."""

    for label_row, row in enumerate(rows):
        for label_column, value in enumerate(row):
            normalized = normalize_header_text(value)
            if "отчет" not in normalized or "период" not in normalized:
                continue
            bounds = range(label_row + 1, min(label_row + 5, len(rows)))
            values_by_marker: dict[str, str] = {}
            for row_index in bounds:
                candidate = rows[row_index]
                upper_column = min(label_column + 6, len(candidate))
                for marker_column in range(max(0, label_column - 1), upper_column):
                    marker = normalize_header_text(candidate[marker_column])
                    if marker not in {"с", "по"}:
                        continue
                    for next_row_index in range(row_index + 1, min(row_index + 4, len(rows))):
                        next_row = rows[next_row_index]
                        if marker_column >= len(next_row):
                            continue
                        period_value = _period_date(next_row[marker_column])
                        if period_value is not None:
                            values_by_marker.setdefault(marker, period_value)
                            break
                    if marker in values_by_marker:
                        continue
                    for value_column in range(marker_column + 1, upper_column):
                        period_value = _period_date(candidate[value_column])
                        if period_value is not None:
                            values_by_marker.setdefault(marker, period_value)
                            break
            if values_by_marker:
                return " ".join(
                    f"{marker} {values_by_marker[marker]}"
                    for marker in ("с", "по")
                    if marker in values_by_marker
                )
    return None


def _metadata(rows: list[tuple[object, ...]]) -> tuple[str | None, str | None, str | None]:
    text = "\n".join(
        _metadata_text(value) for row in rows for value in row if _metadata_text(value)
    )
    act = _ACT_RE.search(text)
    period = _PERIOD_RE.search(text)
    object_code = _OBJECT_RE.search(text)
    return (
        act.group(1).strip() if act else None,
        _structural_period(rows) or (period.group(1).strip(" .;:") if period else None),
        object_code.group(1).strip() if object_code else None,
    )


def _value(row: tuple[object, ...], columns: dict[str, int], field: str) -> object:
    index = columns.get(field)
    return row[index] if index is not None and index < len(row) else None


def _text(row: tuple[object, ...], columns: dict[str, int], field: str) -> str | None:
    value = clean_display_text(_value(row, columns, field))
    return value or None


def _sheet_facts(worksheet: object, workbook_path: PurePosixPath) -> WorkbookSheetFacts:
    max_column = min(max(int(getattr(worksheet, "max_column", 1) or 1), 1), _HEADER_COLUMNS)
    header_rows = list(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(max(int(getattr(worksheet, "max_row", 1) or 1), 1), _HEADER_ROWS),
            min_col=1,
            max_col=max_column,
            values_only=True,
        )
    )
    act_number, period, object_code = _metadata(header_rows)
    candidate = _candidate_headers(header_rows)
    if candidate is None:
        issue = PackageIssue(
            "HEADER_NOT_FOUND",
            "Не найдена структурная строка заголовков в первых 100 строках",
            workbook_path=workbook_path,
            sheet_name=worksheet.title,
        )
        return WorkbookSheetFacts(worksheet.title, act_number, period, object_code, (), (issue,))

    header_end, columns = candidate
    rows: list[WorkbookRowFact] = []
    issues: list[PackageIssue] = []
    empty_streak = 0
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=header_end + 1,
            min_col=1,
            max_col=max_column,
            values_only=True,
        ),
        start=header_end + 1,
    ):
        if not any(value is not None and clean_display_text(value) for value in row):
            empty_streak += 1
            if empty_streak >= _EMPTY_ROW_LIMIT:
                break
            continue
        empty_streak = 0
        work_code = _text(row, columns, "work_code")
        work_name = _text(row, columns, "work_name")
        quantity = _decimal(_value(row, columns, "quantity"))
        total_cost = _decimal(_value(row, columns, "total_cost"))
        if not any((work_code, work_name, quantity is not None, total_cost is not None)):
            continue
        fact = WorkbookRowFact(
            sheet_name=worksheet.title,
            row_number=row_number,
            act_number=act_number,
            period=period,
            object_code=_text(row, columns, "object_code") or object_code,
            work_code=work_code,
            drawing_code=_text(row, columns, "drawing_code"),
            basis=_text(row, columns, "basis"),
            work_name=work_name,
            unit=_text(row, columns, "unit"),
            quantity=quantity,
            total_cost=total_cost,
        )
        rows.append(fact)
        if not work_code:
            issues.append(
                PackageIssue(
                    "MISSING_WORK_CODE",
                    "Строка не содержит стабильный код позиции/работы",
                    workbook_path=workbook_path,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )
        if not work_name or (quantity is None and total_cost is None):
            issues.append(
                PackageIssue(
                    "MISSING_COMPARABLE_FACTS",
                    "Строка не содержит наименование работы и хотя бы количество или стоимость",
                    workbook_path=workbook_path,
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )
    return WorkbookSheetFacts(
        worksheet.title,
        act_number,
        period,
        object_code,
        tuple(rows),
        tuple(issues),
    )


def extract_package_workbook_facts(
    source_root: Path,
    workbook_path: PurePosixPath | Path | str,
) -> PackageWorkbookFacts:
    """Extract KS-2 facts without writing or retaining an open workbook handle."""

    relative_path = _safe_relative_path(workbook_path)
    resolved = _resolve_workbook(source_root, relative_path)
    workbook = openpyxl.load_workbook(
        resolved,
        read_only=True,
        data_only=True,
        keep_links=False,
        keep_vba=resolved.suffix.lower() == ".xlsm",
    )
    try:
        sheets = tuple(_sheet_facts(worksheet, relative_path) for worksheet in workbook.worksheets)
    finally:
        workbook.close()
    return PackageWorkbookFacts(workbook_path=relative_path, sheets=sheets)

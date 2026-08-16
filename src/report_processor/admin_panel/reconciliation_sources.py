"""Fail-soft, per-workbook source selection for reconciliation."""

from __future__ import annotations

import re
import unicodedata
from bisect import bisect_right
from dataclasses import dataclass, replace
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter

from report_processor.extraction.models import CanonicalSourceRow, SourceLocation
from report_processor.metadata.periods import extract_period_from_filename
from report_processor.normalization import NormalizedSourceRow, normalize_training_rows
from report_processor.schema import (
    ComposedHeader,
    LogicalColumn,
    SheetType,
    resolve_logical_columns,
)
from report_processor.schema.column_aliases import DEFAULT_COLUMN_ALIASES
from report_processor.schema.text_normalization import normalize_header_text
from report_processor.training_data import prepare_training_data

from .reconciliation_identity import resolve_source_identity, source_basename_identities

_UNIT_ALIASES = frozenset({"ед изм", "единица измерения", "единица"})
_HIERARCHY_VALUE_RE = re.compile(r"^\d+(?:\.\d+)+\.?$")


class FormulaCacheUnavailableError(ValueError):
    """An otherwise usable source metric cannot be verified from its cache."""


class SourceLayoutAmbiguousError(ValueError):
    """More than one structurally viable source layout exists."""


@dataclass(frozen=True, slots=True)
class ReconciliationSourceDescriptor:
    """Safe upload metadata; the private workbook path stays with the adapter."""

    safe_basename: str
    document_index: str | None = None
    document_period: str | None = None
    document_index_candidates: tuple[str, ...] = ()

    def __post_init__(self) -> None:
        if not self.safe_basename or self.safe_basename != Path(self.safe_basename).name:
            raise ValueError("safe_basename must be a basename")


@dataclass(frozen=True, slots=True)
class _HeaderGraph:
    """Bounded header cells with real merged-cell propagation and parent spans."""

    rows: tuple[tuple[object, ...], ...]
    spans: dict[tuple[int, int], tuple[int, int, int, int]]

    def parent_span(self, row: int, column: int) -> tuple[int, int]:
        _top, left, _bottom, right = self.spans.get((row, column), (row, column, row, column))
        return left, right

    def span_at(self, row: int, column: int) -> tuple[int, int, int, int]:
        return self.spans.get((row, column), (row, column, row, column))


@dataclass(frozen=True, slots=True)
class _MetricRegion:
    """One physical quantity/total-cost region in a bounded header band."""

    quantity_column: int
    cost_column: int
    header_end: int
    metric_span: tuple[int, int, int, int]
    band_start: int


@dataclass(frozen=True, slots=True)
class _StreamedLayout:
    """A later bounded-column layout promoted into normal candidate evaluation."""

    region: _MetricRegion
    work_column: int
    unit_column: int
    cumulative: bool


@dataclass(frozen=True, slots=True)
class _SparseRegionIndex:
    values: dict[tuple[int, int], object]
    formula_cells: frozenset[tuple[int, int]]
    spans: tuple[tuple[int, int, int, int], ...]
    row_values: dict[int, dict[int, object]]
    column_values: dict[int, tuple[tuple[int, object], ...]]
    occupied_rows: frozenset[int]
    occupied_merge_rows: tuple[tuple[int, int], ...]
    spans_by_top: dict[int, tuple[tuple[int, int, int, int], ...]]
    spans_by_bottom: dict[int, tuple[tuple[int, int, int, int], ...]]
    span_by_origin: dict[tuple[int, int], tuple[int, int, int, int]]
    span_starts: tuple[int, ...]
    spans_by_left: tuple[tuple[int, int, int, int], ...]
    covering_span_cache: dict[int, tuple[tuple[int, int, int, int], ...]]
    columns: tuple[int, ...]
    rows: tuple[int, ...]


_REGION_CELL_LIMIT = 500_000
_REGION_MERGE_LIMIT = 1_000
_REGION_CANDIDATE_LIMIT = 256


def descriptor_from_upload_basename(safe_basename: str) -> ReconciliationSourceDescriptor:
    """Infer optional metadata solely from one validated upload basename."""
    period = extract_period_from_filename(safe_basename).value
    candidates = source_basename_identities(safe_basename)
    return ReconciliationSourceDescriptor(
        safe_basename=safe_basename,
        document_index=candidates[0] if len(candidates) == 1 else None,
        document_period=period.normalized if period is not None else None,
        document_index_candidates=candidates,
    )


def document_index_from_basename(safe_basename: str) -> str | None:
    """Return an unambiguous bounded filename identity."""
    candidates = source_basename_identities(safe_basename)
    return candidates[0] if len(candidates) == 1 else None


def resolve_descriptor_identity(
    descriptor: ReconciliationSourceDescriptor, target_identities: set[str] | frozenset[str]
) -> ReconciliationSourceDescriptor:
    """Bind a source basename only to one terminal identity in the selected stage."""
    candidates = descriptor.document_index_candidates or (
        (descriptor.document_index,) if descriptor.document_index else ()
    )
    identity = resolve_source_identity(candidates, target_identities)
    return replace(descriptor, document_index=identity)


@dataclass(frozen=True, slots=True)
class ReconciliationSourceIssue:
    """Controlled, presentation-safe extraction outcome for one workbook."""

    code: str
    safe_basename: str
    comment: str
    repair_hint: str
    can_continue: bool


@dataclass(frozen=True, slots=True)
class ReconciliationSourceSelection:
    """Private selection summary without paths, sheets, formulas or exceptions."""

    safe_basename: str
    source_type: str
    usable_row_count: int


@dataclass(frozen=True, slots=True)
class ReconciliationSourceBatch:
    """Normalized rows and controlled outcomes from independently read workbooks."""

    rows: tuple[NormalizedSourceRow, ...]
    issues: tuple[ReconciliationSourceIssue, ...]
    selections: tuple[ReconciliationSourceSelection, ...]
    terminal_identities: tuple[tuple[str, str], ...] = ()


class AllReconciliationSourcesUnusableError(ValueError):
    """Every upload failed safely; callers may expose only ``issues``."""

    def __init__(self, issues: tuple[ReconciliationSourceIssue, ...]) -> None:
        super().__init__("RECONCILIATION_SOURCES_UNUSABLE")
        self.issues = issues


def extract_reconciliation_sources(
    workbooks: tuple[tuple[Path, str, ReconciliationSourceDescriptor], ...],
    *,
    require_document_index: bool = False,
) -> ReconciliationSourceBatch:
    """Choose one usable cumulative source per workbook without cross-file failure."""
    rows: list[NormalizedSourceRow] = []
    issues: list[ReconciliationSourceIssue] = []
    selections: list[ReconciliationSourceSelection] = []
    identities: list[tuple[str, str]] = []
    for path, source_id, descriptor in workbooks:
        if require_document_index and not _has_usable_document_index(descriptor):
            issues.append(_issue("DOCUMENT_INDEX_MISSING", descriptor))
            continue
        try:
            selected = _extract_one(path, source_id, descriptor)
        except FormulaCacheUnavailableError:
            issues.append(_issue("FORMULA_CACHE_UNAVAILABLE", descriptor))
            continue
        except SourceLayoutAmbiguousError:
            issues.append(_issue("SOURCE_LAYOUT_AMBIGUOUS", descriptor))
            continue
        except Exception:
            issues.append(_issue("WORKBOOK_UNREADABLE", descriptor))
            continue
        if selected is None:
            issues.append(_issue("NO_USABLE_RECONCILIATION_SOURCE", descriptor))
            continue
        source_type, normalized = selected
        rows.extend(normalized)
        if descriptor.document_index is not None:
            identities.append((source_id, descriptor.document_index))
        selections.append(
            ReconciliationSourceSelection(
                safe_basename=descriptor.safe_basename,
                source_type=source_type,
                usable_row_count=len(normalized),
            )
        )
    batch = ReconciliationSourceBatch(
        tuple(rows), tuple(issues), tuple(selections), tuple(sorted(identities))
    )
    if not batch.rows:
        raise AllReconciliationSourcesUnusableError(batch.issues)
    return batch


def _extract_one(
    path: Path, source_id: str, descriptor: ReconciliationSourceDescriptor
) -> tuple[str, tuple[NormalizedSourceRow, ...]] | None:
    # Candidate enumeration needs repeated, deterministic passes over each sheet.
    workbook = load_workbook(path, read_only=False, data_only=True)
    formulas = load_workbook(path, read_only=False, data_only=False)
    try:
        candidates: list[tuple[str, str, tuple[NormalizedSourceRow, ...]]] = []
        for sheet in workbook.worksheets:
            layouts = _indexed_sheet_layouts(sheet, formulas[sheet.title], source_id, descriptor)
            for source_type, source_layouts in (("ks6a", layouts[0]), ("ks2", layouts[1])):
                for _physical_key, canonical in source_layouts:
                    normalized = normalize_training_rows(prepare_training_data(canonical).rows).rows
                    if normalized:
                        candidates.append((source_type, sheet.title, normalized))
        cumulative = [candidate for candidate in candidates if candidate[0] == "ks6a"]
        viable = cumulative or candidates
        if len(viable) > 1:
            raise SourceLayoutAmbiguousError("SOURCE_LAYOUT_AMBIGUOUS")
        if viable:
            source_type, _sheet_name, normalized = viable[0]
            return source_type, normalized
    finally:
        formulas.close()
        workbook.close()
    return None


def _extract_ks6a_rows(
    sheet, formula_sheet, source_id: str, descriptor: ReconciliationSourceDescriptor
):
    """Read the cumulative pair from a structural multi-row КС-6а header."""
    return _unique_layout_rows(_ks6a_layouts(sheet, formula_sheet, source_id, descriptor))


def _extract_ks2_rows(
    sheet, formula_sheet, source_id: str, descriptor: ReconciliationSourceDescriptor
):
    """Read a structural КС-2 detail table only when its direct metrics are explicit."""
    return _unique_layout_rows(_ks2_layouts(sheet, formula_sheet, source_id, descriptor))


def _ks6a_layouts(sheet, formula_sheet, source_id: str, descriptor: ReconciliationSourceDescriptor):
    """Return each physical cumulative layout; normalization is deliberately deferred."""
    return _indexed_sheet_layouts(sheet, formula_sheet, source_id, descriptor)[0]


def _ks2_layouts(sheet, formula_sheet, source_id: str, descriptor: ReconciliationSourceDescriptor):
    """Return each physical direct layout; normalization is deliberately deferred."""
    return _indexed_sheet_layouts(sheet, formula_sheet, source_id, descriptor)[1]


def _indexed_sheet_layouts(
    sheet, formula_sheet, source_id: str, descriptor: ReconciliationSourceDescriptor
):
    """Discover every bounded sparse physical layout once, then probe its own interval once."""
    index = _sparse_region_index(sheet, formula_sheet)
    structural = _indexed_structural_layouts(index)
    output: dict[bool, list[tuple[tuple[object, ...], tuple[CanonicalSourceRow, ...]]]] = {
        True: [],
        False: [],
    }
    probed: list[tuple[_StreamedLayout, tuple[CanonicalSourceRow, ...]]] = []
    for layout in structural:
        region = layout.region
        end_row = _indexed_detail_end(region, structural, index.rows)
        rows = _canonical_index_rows(
            index,
            sheet.title,
            source_id,
            descriptor,
            source_type="ks6a" if layout.cumulative else "ks2",
            start_row=region.header_end + 1,
            end_row=end_row,
            work_column=layout.work_column,
            unit_column=layout.unit_column,
            quantity_column=region.quantity_column,
            cost_column=region.cost_column,
            cumulative=layout.cumulative,
        )
        probed.append((layout, rows))
    for layout, rows in probed:
        if not rows and any(
            other.region.metric_span[0] < layout.region.metric_span[0]
            and other.region.header_end < layout.region.metric_span[0]
            and _spans_overlap(
                other.region.metric_span[1],
                other.region.metric_span[3],
                layout.region.metric_span[1],
                layout.region.metric_span[3],
            )
            for other in structural
        ):
            raise SourceLayoutAmbiguousError("SOURCE_LAYOUT_AMBIGUOUS")
    for layout, rows in probed:
        if rows:
            region = layout.region
            output[layout.cumulative].append(
                (
                    (
                        layout.cumulative,
                        layout.work_column,
                        layout.unit_column,
                        region.metric_span,
                        region.quantity_column,
                        region.cost_column,
                        region.header_end,
                    ),
                    rows,
                )
            )
    return output[True], output[False]


def _sparse_region_index(sheet, formula_sheet) -> _SparseRegionIndex:
    coordinates: set[tuple[int, int]] = set()
    for cell_map in (sheet._cells, formula_sheet._cells):
        for coordinate, cell in cell_map.items():
            if isinstance(cell, MergedCell):
                continue
            coordinates.add(coordinate)
            if len(coordinates) > _REGION_CELL_LIMIT:
                raise SourceLayoutAmbiguousError("SOURCE_LAYOUT_AMBIGUOUS")
    spans = tuple(
        sorted(
            (item.min_row, item.min_col, item.max_row, item.max_col)
            for item in sheet.merged_cells.ranges
        )
    )
    if len(spans) > _REGION_MERGE_LIMIT:
        raise SourceLayoutAmbiguousError("SOURCE_LAYOUT_AMBIGUOUS")
    cells = {
        (row, column): cell.value
        for (row, column), cell in sheet._cells.items()
        if cell.value is not None
    }
    formulas = frozenset(
        (row, column)
        for (row, column), cell in formula_sheet._cells.items()
        if cell.data_type == "f"
    )
    columns = sorted(
        {column for _row, column in cells}
        | {span[1] for span in spans}
        | {span[3] for span in spans}
    )
    row_values: dict[int, dict[int, object]] = {}
    column_values: dict[int, list[tuple[int, object]]] = {}
    for (row, column), value in cells.items():
        row_values.setdefault(row, {})[column] = value
        column_values.setdefault(column, []).append((row, value))
    spans_by_top: dict[int, list[tuple[int, int, int, int]]] = {}
    spans_by_bottom: dict[int, list[tuple[int, int, int, int]]] = {}
    for span in spans:
        spans_by_top.setdefault(span[0], []).append(span)
        spans_by_bottom.setdefault(span[2], []).append(span)
    spans_by_left = tuple(sorted(spans, key=lambda item: (item[1], item[3], item[0], item[2])))
    return _SparseRegionIndex(
        cells,
        formulas,
        spans,
        row_values,
        {column: tuple(sorted(values)) for column, values in column_values.items()},
        frozenset(row_values),
        _coalesced_row_intervals(spans),
        {row: tuple(items) for row, items in spans_by_top.items()},
        {row: tuple(items) for row, items in spans_by_bottom.items()},
        {(top, left): span for span in spans for top, left, _bottom, _right in (span,)},
        tuple(span[1] for span in spans_by_left),
        spans_by_left,
        {},
        tuple(columns),
        tuple(sorted(row_values)),
    )


def _coalesced_row_intervals(
    spans: tuple[tuple[int, int, int, int], ...],
) -> tuple[tuple[int, int], ...]:
    intervals: list[tuple[int, int]] = []
    for top, _left, bottom, _right in spans:
        if intervals and top <= intervals[-1][1] + 1:
            intervals[-1] = (intervals[-1][0], max(intervals[-1][1], bottom))
        else:
            intervals.append((top, bottom))
    return tuple(intervals)


def _indexed_structural_layouts(index: _SparseRegionIndex) -> tuple[_StreamedLayout, ...]:
    layouts: list[_StreamedLayout] = []
    for root in index.spans:
        top, left, _bottom, right = root
        if right <= left or not _role_text(_text(index.values.get((top, left))), "cumulative"):
            continue
        for quantity, cost, header_end in _indexed_cumulative_leaves(index, root, frozenset()):
            roles = _indexed_roles(index, top, header_end, left, right, SheetType.KS6A)
            if roles is not None:
                layouts.append(
                    _StreamedLayout(
                        _MetricRegion(quantity, cost, header_end, root, top), *roles, True
                    )
                )
    for row in index.rows:
        for quantity, cost in _indexed_metric_pairs(index, row, 1, max(index.columns, default=0)):
            band_start = _indexed_band_start(index, row)
            if _indexed_cumulative_ancestor(
                index, row, quantity, band_start
            ) or _indexed_cumulative_ancestor(index, row, cost, band_start):
                continue
            roles = _indexed_roles(index, band_start, row, quantity, cost, SheetType.KS2)
            if roles is not None:
                layouts.append(
                    _StreamedLayout(
                        _MetricRegion(quantity, cost, row, (row, quantity, row, cost), band_start),
                        *roles,
                        False,
                    )
                )
    unique = tuple(dict.fromkeys(layouts))
    if len(unique) > _REGION_CANDIDATE_LIMIT:
        raise SourceLayoutAmbiguousError("SOURCE_LAYOUT_AMBIGUOUS")
    return unique


def _indexed_metric_pairs(
    index: _SparseRegionIndex, row: int, start: int, end: int
) -> list[tuple[int, int]]:
    values = {
        column: _text(value)
        for column, value in index.row_values.get(row, {}).items()
        if start <= column <= end
    }
    quantities = [
        column
        for column, value in values.items()
        if any(stem in value for stem in ("колич", "объем", "объём"))
    ]
    costs = [column for column, value in values.items() if _cost_text(value)]
    return [(quantity, cost) for quantity in quantities for cost in costs if cost == quantity + 1]


def _indexed_cumulative_leaves(
    index: _SparseRegionIndex, current, seen: frozenset[tuple[int, int, int, int]]
):
    if current in seen:
        return []
    _top, left, bottom, right = current
    leaf_row = bottom + 1
    leaves = [
        (quantity, cost, leaf_row)
        for quantity, cost in _indexed_metric_pairs(index, leaf_row, left, right)
    ]
    for nested in index.spans_by_top.get(leaf_row, ()):
        if nested != current and nested[0] == leaf_row and left <= nested[1] <= nested[3] <= right:
            leaves.extend(_indexed_cumulative_leaves(index, nested, seen | {current}))
    return leaves


def _indexed_roles(
    index: _SparseRegionIndex,
    start: int,
    end: int,
    metric_left: int,
    metric_right: int,
    sheet_type: SheetType,
):
    headers = []
    for column in index.columns:
        if metric_left <= column <= metric_right:
            continue
        lineage = _indexed_lineage(index, column, start, end)
        if lineage and not _price_lineage(lineage):
            headers.append(
                ComposedHeader(
                    column,
                    get_column_letter(column),
                    (lineage,),
                    lineage,
                    normalize_header_text(lineage),
                    False,
                    (),
                    (),
                )
            )
    return _resolved_region_roles(tuple(headers), sheet_type)


def _indexed_lineage(index: _SparseRegionIndex, column: int, start: int, end: int) -> str:
    values = []
    seen = set()
    for row, value in index.column_values.get(column, ()):
        if not start <= row <= end:
            continue
        span = index.span_by_origin.get((row, column))
        origin = (span[0], span[1]) if span else (row, column)
        if origin in seen:
            continue
        seen.add(origin)
        text = _text(index.values.get(origin, value))
        if text:
            values.append(text)
    for top, left, bottom, right in _covering_spans(index, column):
        if not (left <= column <= right and top <= end and bottom >= start):
            continue
        origin = (top, left)
        if origin in seen:
            continue
        seen.add(origin)
        text = _text(index.values.get(origin))
        if text:
            values.append(text)
    return " ".join(values)


def _indexed_band_start(index: _SparseRegionIndex, row: int) -> int:
    while row > 1:
        previous = row - 1
        if previous in index.occupied_rows:
            row = previous
            continue
        interval = _merge_interval_containing(index.occupied_merge_rows, previous)
        if interval is None:
            return row
        row = interval[0]
    return row


def _indexed_row_occupied(index: _SparseRegionIndex, row: int) -> bool:
    if row in index.occupied_rows:
        return True
    return _merge_interval_containing(index.occupied_merge_rows, row) is not None


def _merge_interval_containing(
    intervals: tuple[tuple[int, int], ...], row: int
) -> tuple[int, int] | None:
    position = bisect_right(intervals, (row, float("inf"))) - 1
    if position >= 0 and intervals[position][0] <= row <= intervals[position][1]:
        return intervals[position]
    return None


def _covering_spans(
    index: _SparseRegionIndex, column: int
) -> tuple[tuple[int, int, int, int], ...]:
    cached = index.covering_span_cache.get(column)
    if cached is not None:
        return cached
    cutoff = bisect_right(index.span_starts, column)
    result = tuple(span for span in index.spans_by_left[:cutoff] if span[3] >= column)
    index.covering_span_cache[column] = result
    return result


def _indexed_cumulative_ancestor(
    index: _SparseRegionIndex, row: int, column: int, band_start: int
) -> bool:
    current = (row, column, row, column)
    seen = set()
    while current not in seen:
        seen.add(current)
        top, left, _bottom, right = current
        parents = [
            span
            for span in index.spans_by_bottom.get(top - 1, ())
            if span[0] >= band_start and span[1] <= left <= right <= span[3]
        ]
        if len(parents) != 1:
            return bool(parents)
        current = parents[0]
        if _role_text(_text(index.values.get((current[0], current[1]))), "cumulative"):
            return True
    return True


def _indexed_detail_end(
    region: _MetricRegion, layouts: tuple[_StreamedLayout, ...], rows: tuple[int, ...]
) -> int:
    _top, left, _bottom, right = region.metric_span
    starts = [
        item.region.metric_span[0]
        for item in layouts
        if item.region != region
        and item.region.metric_span[0] > region.header_end
        and _spans_overlap(left, right, item.region.metric_span[1], item.region.metric_span[3])
    ]
    return min(starts) - 1 if starts else max(rows, default=region.header_end)


def _canonical_index_rows(
    index: _SparseRegionIndex,
    sheet_name: str,
    source_id: str,
    descriptor: ReconciliationSourceDescriptor,
    *,
    source_type: str,
    start_row: int,
    end_row: int,
    work_column: int,
    unit_column: int,
    quantity_column: int,
    cost_column: int,
    cumulative: bool,
) -> tuple[CanonicalSourceRow, ...]:
    rows = []
    for row_number in index.rows:
        if not start_row <= row_number <= end_row:
            continue
        work_name = _text(index.values.get((row_number, work_column)))
        unit = _text(index.values.get((row_number, unit_column)))
        quantity = _decimal(index.values.get((row_number, quantity_column)))
        cost = _decimal(index.values.get((row_number, cost_column)))
        quantity_formula = (row_number, quantity_column) in index.formula_cells
        cost_formula = (row_number, cost_column) in index.formula_cells
        if not work_name or not unit or _HIERARCHY_VALUE_RE.fullmatch(unit) is not None:
            continue
        if (quantity is None and not quantity_formula) or (cost is None and not cost_formula):
            continue
        if (quantity_formula and quantity is None) or (cost_formula and cost is None):
            raise FormulaCacheUnavailableError("FORMULA_CACHE_UNAVAILABLE")
        location = SourceLocation(
            source_id, descriptor.safe_basename, sheet_name, source_type, row_number
        )
        rows.append(
            CanonicalSourceRow(
                row_id=f"{source_id}:{source_type}:{row_number}",
                source_type=source_type,
                source_location=location,
                document_index=descriptor.document_index,
                document_period=descriptor.document_period,
                object_code_raw=None,
                object_name_raw=None,
                subobject_code_raw=None,
                subobject_name_raw=None,
                position_code_raw=_text(index.values.get((row_number, 2))),
                work_name_raw=work_name,
                unit_raw=unit,
                contract_quantity=None,
                current_period_quantity=None if cumulative else quantity,
                cumulative_quantity=quantity if cumulative else None,
                remaining_quantity=None,
                unit_price=None,
                contract_cost=None,
                current_period_cost=None if cumulative else cost,
                cumulative_cost=cost if cumulative else None,
                total_cost=None,
                basis_code_raw=None,
                drawing_code_raw=_text(index.values.get((row_number, 7))),
                cost_type_code_raw=_text(index.values.get((row_number, 3))),
                source_values=(),
                status="OK",
                warnings=(),
            )
        )
    return tuple(
        replace(
            row,
            current_period_quantity=row.cumulative_quantity,
            current_period_cost=row.cumulative_cost,
        )
        if cumulative
        else row
        for row in rows
    )


def _physical_layouts(
    sheet,
    formula_sheet,
    source_id: str,
    descriptor: ReconciliationSourceDescriptor,
    header: _HeaderGraph,
    *,
    cumulative: bool,
    source_type: str,
):
    """Bind roles only after a physical metric region has been identified."""
    candidates = []
    all_regions = _all_physical_metric_regions(header)
    streamed_layouts = _streamed_layouts(sheet, header)
    for region in _physical_metric_regions(header, cumulative=cumulative):
        roles = _region_roles(
            header,
            region,
            sheet_type=SheetType.KS6A if cumulative else SheetType.KS2,
        )
        if roles is None:
            continue
        work_column, unit_column = roles
        rows = _canonical_rows(
            sheet,
            source_id,
            descriptor,
            source_type=source_type,
            start_row=region.header_end + 1,
            end_row=_detail_end(sheet, header, region, all_regions),
            work_column=work_column,
            unit_column=unit_column,
            quantity_column=region.quantity_column,
            cost_column=region.cost_column,
            cumulative=cumulative,
            formula_sheet=formula_sheet,
        )
        if rows:
            candidates.append(
                (
                    (
                        work_column,
                        unit_column,
                        region.metric_span,
                        region.quantity_column,
                        region.cost_column,
                        region.header_end,
                    ),
                    rows,
                )
            )
    streamed_regions = tuple(item.region for item in streamed_layouts)
    for layout in streamed_layouts:
        if layout.cumulative != cumulative:
            continue
        region = layout.region
        rows = _canonical_rows(
            sheet,
            source_id,
            descriptor,
            source_type=source_type,
            start_row=region.header_end + 1,
            end_row=_streamed_detail_end(sheet, region, (*all_regions, *streamed_regions)),
            work_column=layout.work_column,
            unit_column=layout.unit_column,
            quantity_column=region.quantity_column,
            cost_column=region.cost_column,
            cumulative=cumulative,
            formula_sheet=formula_sheet,
        )
        if rows:
            candidates.append(
                (
                    (
                        layout.work_column,
                        layout.unit_column,
                        region.metric_span,
                        region.quantity_column,
                        region.cost_column,
                        region.header_end,
                    ),
                    rows,
                )
            )
    return candidates


def _header_graph(sheet, *, maximum: int) -> _HeaderGraph:
    """Materialize only bounded header cells, filling each merged range from its top-left."""
    max_row = min(int(sheet.max_row or 0), maximum)
    max_column = int(sheet.max_column or 0)
    values = [
        [sheet.cell(row, column).value for column in range(1, max_column + 1)]
        for row in range(1, max_row + 1)
    ]
    spans: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for merged in sheet.merged_cells.ranges:
        if merged.min_row > max_row or merged.min_col > max_column:
            continue
        top, left = merged.min_row, merged.min_col
        bottom, right = min(merged.max_row, max_row), min(merged.max_col, max_column)
        value = sheet.cell(top, left).value
        for row in range(top, bottom + 1):
            for column in range(left, right + 1):
                values[row - 1][column - 1] = value
                spans[(row, column)] = (top, left, bottom, right)
    return _HeaderGraph(tuple(tuple(row) for row in values), spans)


def _physical_metric_regions(header: _HeaderGraph, *, cumulative: bool) -> list[_MetricRegion]:
    """Find physical metric leaves before considering descriptive header columns."""
    regions = _cumulative_metric_regions(header) if cumulative else _direct_metric_regions(header)
    return list(dict.fromkeys(regions))


def _all_physical_metric_regions(header: _HeaderGraph) -> tuple[_MetricRegion, ...]:
    return tuple(
        dict.fromkeys(
            (
                *_physical_metric_regions(header, cumulative=True),
                *_physical_metric_regions(header, cumulative=False),
            )
        )
    )


def _cumulative_metric_regions(header: _HeaderGraph) -> list[_MetricRegion]:
    regions: list[_MetricRegion] = []
    for parent in sorted(set(header.spans.values())):
        top, left, bottom, right = parent
        if right == left or bottom >= len(header.rows):
            continue
        # A cumulative region must be nominated by the actual merged-cell origin,
        # not by a word inherited into one of its metric descendants.
        if not _role_text(_text_at(header.rows[top - 1], left), "cumulative"):
            continue
        for quantity, cost, leaf_row in _cumulative_metric_leaves(header, parent):
            regions.append(
                _MetricRegion(
                    quantity,
                    cost,
                    leaf_row,
                    parent,
                    top,
                )
            )
    return regions


def _cumulative_metric_leaves(
    header: _HeaderGraph, root: tuple[int, int, int, int]
) -> list[tuple[int, int, int]]:
    """Enumerate every exact nested path; viability later decides physical ambiguity."""
    return _cumulative_metric_leaves_from(header, root, seen=frozenset())


def _cumulative_metric_leaves_from(
    header: _HeaderGraph,
    current: tuple[int, int, int, int],
    *,
    seen: frozenset[tuple[int, int, int, int]],
) -> list[tuple[int, int, int]]:
    if current in seen:
        return []
    _top, left, bottom, right = current
    if bottom >= len(header.rows):
        return []
    leaf_row = bottom + 1
    leaves = [
        (quantity, cost, leaf_row)
        for quantity, cost in _physical_metric_pairs(header, leaf_row, left, right)
    ]
    for nested in _adjacent_nested_spans(header, current):
        leaves.extend(_cumulative_metric_leaves_from(header, nested, seen=seen | {current}))
    return leaves


def _adjacent_nested_spans(
    header: _HeaderGraph, parent: tuple[int, int, int, int]
) -> tuple[tuple[int, int, int, int], ...]:
    """Return only explicit, immediately-adjacent spans geometrically inside ``parent``."""
    _top, left, bottom, right = parent
    return tuple(
        sorted(
            span
            for span in set(header.spans.values())
            if span != parent and span[0] == bottom + 1 and left <= span[1] <= span[3] <= right
        )
    )


def _direct_metric_regions(header: _HeaderGraph) -> list[_MetricRegion]:
    regions: list[_MetricRegion] = []
    for row_number, row in enumerate(header.rows, 1):
        for quantity, cost in _physical_metric_pairs(header, row_number, 1, len(row)):
            band_start = _header_band_start(header, row_number)
            if _has_cumulative_ancestor(
                header, row_number, quantity, band_start=band_start
            ) or _has_cumulative_ancestor(header, row_number, cost, band_start=band_start):
                continue
            quantity_span = header.span_at(row_number, quantity)
            cost_span = header.span_at(row_number, cost)
            regions.append(
                _MetricRegion(
                    quantity,
                    cost,
                    max(quantity_span[2], cost_span[2]),
                    (
                        min(quantity_span[0], cost_span[0]),
                        quantity,
                        max(quantity_span[2], cost_span[2]),
                        cost,
                    ),
                    band_start,
                )
            )
    return regions


def _physical_metric_pairs(
    header: _HeaderGraph, row_number: int, start: int, end: int
) -> list[tuple[int, int]]:
    """Return adjacent leaves whose labels live at their physical origins."""
    row = header.rows[row_number - 1]
    return [
        (quantity, cost)
        for quantity, cost in _metric_pairs(row, start, end)
        if _is_origin(header, row_number, quantity) and _is_origin(header, row_number, cost)
    ]


def _is_origin(header: _HeaderGraph, row: int, column: int) -> bool:
    top, left, _bottom, _right = header.span_at(row, column)
    return (top, left) == (row, column)


def _has_cumulative_ancestor(
    header: _HeaderGraph, row: int, column: int, *, band_start: int
) -> bool:
    """Reject direct leaves with a cumulative span anywhere in their exact ancestry."""
    current = header.span_at(row, column)
    seen: set[tuple[int, int, int, int]] = set()
    while current not in seen:
        seen.add(current)
        parents = _adjacent_containing_spans(header, current, band_start=band_start)
        if len(parents) > 1:
            return True
        if not parents:
            return False
        current = parents[0]
        top, left, _bottom, _right = current
        if _role_text(_text_at(header.rows[top - 1], left), "cumulative"):
            return True
    return True


def _adjacent_containing_spans(
    header: _HeaderGraph,
    child: tuple[int, int, int, int],
    *,
    band_start: int,
) -> tuple[tuple[int, int, int, int], ...]:
    top, left, _bottom, right = child
    return tuple(
        sorted(
            span
            for span in set(header.spans.values())
            if span != child
            and span[0] >= band_start
            and span[2] == top - 1
            and span[1] <= left <= right <= span[3]
        )
    )


def _detail_end(
    sheet,
    header: _HeaderGraph,
    region: _MetricRegion,
    regions: tuple[_MetricRegion, ...],
) -> int:
    """Stop before the next overlapping physical metric region on this worksheet."""
    _top, left, _bottom, right = region.metric_span
    next_starts = [
        other.metric_span[0]
        for other in regions
        if other.metric_span[0] > region.header_end
        and _spans_overlap(left, right, other.metric_span[1], other.metric_span[3])
        and _bounded_region_viable(sheet, header, other)
    ]
    bounded_start = min(next_starts) if next_starts else None
    streamed_start = _next_streamed_region_start(
        sheet,
        header,
        region,
        stop_before=bounded_start,
    )
    starts = [start for start in (bounded_start, streamed_start) if start is not None]
    return min(starts) - 1 if starts else int(sheet.max_row or 0)


def _streamed_layouts(sheet, header: _HeaderGraph) -> tuple[_StreamedLayout, ...]:
    """Promote viable regions after the initial bounded header window to candidates."""
    column_limit = len(header.rows[0])
    first_row = len(header.rows) + 1
    merged_ranges = tuple(sheet.merged_cells.ranges)
    layouts: list[_StreamedLayout] = []
    for root in merged_ranges:
        if (
            root.min_row < first_row
            or root.max_col > column_limit
            or root.max_col <= root.min_col
            or not _role_text(_text(sheet.cell(root.min_row, root.min_col).value), "cumulative")
        ):
            continue
        for quantity, cost, header_end in _streamed_cumulative_leaves(
            sheet, merged_ranges, root, column_limit=column_limit, seen=frozenset()
        ):
            roles = _streamed_region_roles(
                sheet,
                merged_ranges,
                band_start=root.min_row,
                header_end=header_end,
                metric_left=root.min_col,
                metric_right=root.max_col,
                column_limit=column_limit,
                sheet_type=SheetType.KS6A,
            )
            if roles is None or not _streamed_detail_viable(
                sheet,
                start_row=header_end + 1,
                work_column=roles[0],
                unit_column=roles[1],
                quantity_column=quantity,
                cost_column=cost,
                column_limit=column_limit,
            ):
                continue
            layouts.append(
                _StreamedLayout(
                    _MetricRegion(
                        quantity,
                        cost,
                        header_end,
                        (root.min_row, root.min_col, root.max_row, root.max_col),
                        root.min_row,
                    ),
                    roles[0],
                    roles[1],
                    True,
                )
            )
    band_start = first_row
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=first_row, max_col=column_limit, values_only=True), first_row
    ):
        if not any(_text(value) for value in values):
            band_start = row_number + 1
            continue
        for quantity, cost in _metric_pairs(values, 1, column_limit):
            if _streamed_has_cumulative_ancestor(
                sheet, merged_ranges, row_number, quantity, band_start=band_start
            ) or _streamed_has_cumulative_ancestor(
                sheet, merged_ranges, row_number, cost, band_start=band_start
            ):
                continue
            roles = _streamed_region_roles(
                sheet,
                merged_ranges,
                band_start=band_start,
                header_end=row_number,
                metric_left=quantity,
                metric_right=cost,
                column_limit=column_limit,
                sheet_type=SheetType.KS2,
            )
            if roles is None or not _streamed_detail_viable(
                sheet,
                start_row=row_number + 1,
                work_column=roles[0],
                unit_column=roles[1],
                quantity_column=quantity,
                cost_column=cost,
                column_limit=column_limit,
            ):
                continue
            layouts.append(
                _StreamedLayout(
                    _MetricRegion(
                        quantity,
                        cost,
                        row_number,
                        (row_number, quantity, row_number, cost),
                        band_start,
                    ),
                    roles[0],
                    roles[1],
                    False,
                )
            )
    return tuple(dict.fromkeys(layouts))


def _streamed_has_cumulative_ancestor(
    sheet, merged_ranges, row: int, column: int, *, band_start: int
) -> bool:
    """Apply the same exact adjacent-span ancestry exclusion to streamed leaves."""
    current = (row, column, row, column)
    seen: set[tuple[int, int, int, int]] = set()
    while current not in seen:
        seen.add(current)
        top, left, _bottom, right = current
        parents = tuple(
            sorted(
                (
                    item
                    for item in merged_ranges
                    if item.min_row >= band_start
                    and item.max_row == top - 1
                    and item.min_col <= left <= right <= item.max_col
                ),
                key=lambda item: (item.min_row, item.min_col, item.max_row, item.max_col),
            )
        )
        if len(parents) != 1:
            return bool(parents)
        parent = parents[0]
        current = (parent.min_row, parent.min_col, parent.max_row, parent.max_col)
        if _role_text(_text(sheet.cell(parent.min_row, parent.min_col).value), "cumulative"):
            return True
    return True


def _streamed_detail_end(sheet, region: _MetricRegion, regions: tuple[_MetricRegion, ...]) -> int:
    _top, left, _bottom, right = region.metric_span
    starts = [
        other.metric_span[0]
        for other in regions
        if other != region
        and other.metric_span[0] > region.header_end
        and _spans_overlap(left, right, other.metric_span[1], other.metric_span[3])
    ]
    return min(starts) - 1 if starts else int(sheet.max_row or 0)


def _bounded_region_viable(sheet, header: _HeaderGraph, region: _MetricRegion) -> bool:
    """A physical header becomes a boundary only with exact roles and one detail row."""
    cumulative = any(
        span == region.metric_span
        and _role_text(_text_at(header.rows[span[0] - 1], span[1]), "cumulative")
        for span in set(header.spans.values())
    )
    roles = _region_roles(
        header,
        region,
        sheet_type=SheetType.KS6A if cumulative else SheetType.KS2,
    )
    return roles is not None and _streamed_detail_viable(
        sheet,
        start_row=region.header_end + 1,
        work_column=roles[0],
        unit_column=roles[1],
        quantity_column=region.quantity_column,
        cost_column=region.cost_column,
        column_limit=len(header.rows[0]),
    )


def _next_streamed_region_start(
    sheet,
    header: _HeaderGraph,
    region: _MetricRegion,
    *,
    stop_before: int | None,
) -> int | None:
    """Stream established columns and bound only structurally viable later regions."""
    _top, left, _bottom, right = region.metric_span
    column_limit = len(header.rows[0])
    merged_ranges = tuple(sheet.merged_cells.ranges)
    cumulative_start = _next_streamed_cumulative_start(
        sheet,
        merged_ranges,
        region,
        column_limit=column_limit,
    )
    limits = [value - 1 for value in (stop_before, cumulative_start) if value is not None]
    upper = min(limits) if limits else None
    band_start = region.header_end + 1
    for row_number, values in enumerate(
        sheet.iter_rows(
            min_row=region.header_end + 1,
            max_row=upper,
            max_col=column_limit,
            values_only=True,
        ),
        region.header_end + 1,
    ):
        if not any(_text(value) for value in values):
            band_start = row_number + 1
            continue
        for quantity, cost in _metric_pairs(values, 1, column_limit):
            if _spans_overlap(left, right, quantity, cost) and _streamed_direct_region_viable(
                sheet,
                merged_ranges,
                band_start=band_start,
                header_row=row_number,
                quantity_column=quantity,
                cost_column=cost,
                column_limit=column_limit,
            ):
                return row_number
    return cumulative_start


def _next_streamed_cumulative_start(
    sheet, merged_ranges, region: _MetricRegion, *, column_limit: int
):
    """Validate later cumulative roots before using them as an immutable boundary."""
    _top, left, _bottom, right = region.metric_span
    starts = []
    for root in merged_ranges:
        if (
            root.min_row <= region.header_end
            or root.max_col > column_limit
            or root.max_col <= root.min_col
            or not _spans_overlap(left, right, root.min_col, root.max_col)
            or not _role_text(_text(sheet.cell(root.min_row, root.min_col).value), "cumulative")
        ):
            continue
        for quantity, cost, header_end in _streamed_cumulative_leaves(
            sheet, merged_ranges, root, column_limit=column_limit, seen=frozenset()
        ):
            roles = _streamed_region_roles(
                sheet,
                merged_ranges,
                band_start=root.min_row,
                header_end=header_end,
                metric_left=root.min_col,
                metric_right=root.max_col,
                column_limit=column_limit,
                sheet_type=SheetType.KS6A,
            )
            if roles is not None and _streamed_detail_viable(
                sheet,
                start_row=header_end + 1,
                work_column=roles[0],
                unit_column=roles[1],
                quantity_column=quantity,
                cost_column=cost,
                column_limit=column_limit,
            ):
                starts.append(root.min_row)
    return min(starts) if starts else None


def _streamed_cumulative_leaves(
    sheet,
    merged_ranges,
    current,
    *,
    column_limit: int,
    seen: frozenset[object],
) -> list[tuple[int, int, int]]:
    if current in seen:
        return []
    leaf_row = current.max_row + 1
    values = tuple(sheet.cell(leaf_row, column).value for column in range(1, column_limit + 1))
    leaves = [
        (quantity, cost, leaf_row)
        for quantity, cost in _metric_pairs(values, current.min_col, current.max_col)
    ]
    for nested in merged_ranges:
        if (
            nested != current
            and nested.min_row == leaf_row
            and current.min_col <= nested.min_col <= nested.max_col <= current.max_col
        ):
            leaves.extend(
                _streamed_cumulative_leaves(
                    sheet,
                    merged_ranges,
                    nested,
                    column_limit=column_limit,
                    seen=seen | {current},
                )
            )
    return leaves


def _streamed_direct_region_viable(
    sheet,
    merged_ranges,
    *,
    band_start: int,
    header_row: int,
    quantity_column: int,
    cost_column: int,
    column_limit: int,
) -> bool:
    roles = _streamed_region_roles(
        sheet,
        merged_ranges,
        band_start=band_start,
        header_end=header_row,
        metric_left=quantity_column,
        metric_right=cost_column,
        column_limit=column_limit,
        sheet_type=SheetType.KS2,
    )
    return roles is not None and _streamed_detail_viable(
        sheet,
        start_row=header_row + 1,
        work_column=roles[0],
        unit_column=roles[1],
        quantity_column=quantity_column,
        cost_column=cost_column,
        column_limit=column_limit,
    )


def _streamed_region_roles(
    sheet,
    merged_ranges,
    *,
    band_start: int,
    header_end: int,
    metric_left: int,
    metric_right: int,
    column_limit: int,
    sheet_type: SheetType,
) -> tuple[int, int] | None:
    candidates = []
    for column in range(1, column_limit + 1):
        if metric_left <= column <= metric_right:
            continue
        lineage = _streamed_lineage(sheet, merged_ranges, column, band_start, header_end)
        if lineage and not _price_lineage(lineage):
            candidates.append(
                ComposedHeader(
                    column_index=column,
                    column_letter=get_column_letter(column),
                    parts=(lineage,),
                    raw_text=lineage,
                    normalized_text=normalize_header_text(lineage),
                    is_empty=False,
                    source_coordinates=(),
                    merged_sources=(),
                )
            )
    return _resolved_region_roles(tuple(candidates), sheet_type)


def _streamed_lineage(sheet, merged_ranges, column: int, start: int, end: int) -> str:
    values: list[str] = []
    seen: set[tuple[int, int]] = set()
    for row in range(start, end + 1):
        merged = next(
            (
                item
                for item in merged_ranges
                if item.min_row <= row <= item.max_row and item.min_col <= column <= item.max_col
            ),
            None,
        )
        origin = (merged.min_row, merged.min_col) if merged is not None else (row, column)
        if origin in seen:
            continue
        seen.add(origin)
        value = sheet.cell(*origin).value
        text = _text(value)
        if text:
            values.append(text)
    return " ".join(values)


def _streamed_detail_viable(
    sheet,
    *,
    start_row: int,
    work_column: int,
    unit_column: int,
    quantity_column: int,
    cost_column: int,
    column_limit: int,
) -> bool:
    for values in sheet.iter_rows(min_row=start_row, max_col=column_limit, values_only=True):
        work = _text_at(values, work_column)
        unit = _text_at(values, unit_column)
        if (
            work
            and unit
            and _HIERARCHY_VALUE_RE.fullmatch(unit) is None
            and _decimal(_value_at(values, quantity_column)) is not None
            and _decimal(_value_at(values, cost_column)) is not None
        ):
            return True
    return False


def _spans_overlap(left: int, right: int, other_left: int, other_right: int) -> bool:
    return left <= other_right and other_left <= right


def _header_band_start(header: _HeaderGraph, leaf_row: int) -> int:
    """Use the contiguous physical header block, rather than a row-distance guess."""
    for row_number in range(leaf_row - 1, 0, -1):
        if not any(_text(value) for value in header.rows[row_number - 1]):
            return row_number + 1
    return 1


def _region_roles(
    header: _HeaderGraph,
    region: _MetricRegion,
    *,
    sheet_type: SheetType,
) -> tuple[int, int] | None:
    """Resolve unique work/unit roles through the public shared schema ontology."""
    _top, metric_left, _bottom, metric_right = region.metric_span
    candidates: list[ComposedHeader] = []
    for column in range(1, len(header.rows[0]) + 1):
        if metric_left <= column <= metric_right:
            continue
        lineage = _physical_lineage(header, column, region.band_start, region.header_end)
        if not lineage or _price_lineage(lineage):
            continue
        candidates.append(
            ComposedHeader(
                column_index=column,
                column_letter=get_column_letter(column),
                parts=(lineage,),
                raw_text=lineage,
                normalized_text=normalize_header_text(lineage),
                is_empty=False,
                source_coordinates=(),
                merged_sources=(),
            )
        )
    return _resolved_region_roles(tuple(candidates), sheet_type)


def _resolved_region_roles(
    candidates: tuple[ComposedHeader, ...], sheet_type: SheetType
) -> tuple[int, int] | None:
    """Accept only unambiguous public-schema work and unit resolutions."""
    role_rules = tuple(
        rule
        for rule in DEFAULT_COLUMN_ALIASES
        if rule.logical_column in {LogicalColumn.WORK_NAME, LogicalColumn.UNIT}
    )
    resolutions = {
        resolution.logical_column: resolution
        for resolution in resolve_logical_columns(tuple(candidates), sheet_type, role_rules)
    }
    work = resolutions.get(LogicalColumn.WORK_NAME)
    unit = resolutions.get(LogicalColumn.UNIT)
    if (
        work is None
        or unit is None
        or work.status != "OK"
        or unit.status != "OK"
        or work.column_index is None
        or unit.column_index is None
        or work.column_index == unit.column_index
    ):
        return None
    return work.column_index, unit.column_index


def _physical_lineage(header: _HeaderGraph, column: int, start: int, end: int) -> str:
    """Join physical origins touching one column inside one exact header band."""
    values: list[str] = []
    seen: set[tuple[int, int]] = set()
    for row in range(start, end + 1):
        top, left, bottom, right = header.span_at(row, column)
        origin = (top, left)
        if origin in seen or not (start <= top <= end) or not (left <= column <= right):
            continue
        if bottom < start or top > end:
            continue
        seen.add(origin)
        value = _text_at(header.rows[top - 1], left)
        if value:
            values.append(value)
    return " ".join(values)


def _price_lineage(lineage: str) -> bool:
    return any(
        token in lineage
        for token in ("цен", "тариф", "расцен", "unit price", "unit cost", "стоимост единиц")
    )


def _rows(header: _HeaderGraph | tuple[tuple[object, ...], ...]):
    return header.rows if isinstance(header, _HeaderGraph) else header


def _layout_columns(header, *, cumulative: bool) -> list[tuple[int, int, int | None]]:
    rows = _rows(header)
    work = _role_columns(rows, "work")
    units = _unit_columns(rows)
    anchors = _role_columns(rows, "cumulative") if cumulative else [None]
    return [
        (work_column, unit_column, anchor)
        for work_column in work
        for unit_column in units
        for anchor in anchors
        if work_column != unit_column
    ]


def _unique_layout_rows(candidates):
    # Merged parent labels legitimately nominate every covered child column;
    # normalize those candidates to their parent-left physical boundary.
    unique = {key: rows for key, rows in candidates}
    if len(unique) > 1:
        raise SourceLayoutAmbiguousError("SOURCE_LAYOUT_AMBIGUOUS")
    return next(iter(unique.values()), ())


def _column_with(rows: tuple[tuple[object, ...], ...], token: str) -> int | None:
    matches: list[tuple[int, int]] = []
    for row_number, row in enumerate(rows, 1):
        for column, value in enumerate(row, 1):
            if token in _text(value):
                matches.append((row_number, column))
    return min((column for _row, column in matches), default=None)


def _header_path(rows: tuple[tuple[object, ...], ...], row_number: int, column: int) -> str:
    """Join all non-empty ancestor labels in a variable-depth header column."""
    return " ".join(
        _text_at(rows[number - 1], column)
        for number in range(1, row_number + 1)
        if _text_at(rows[number - 1], column)
    )


def _role_text(value: str, role: str) -> bool:
    if role == "work":
        return any(stem in value for stem in ("наименован", "описан", "работ"))
    if role == "cumulative":
        return ("выполн" in value or "освоен" in value) and (
            "весь" in value or "нараст" in value or "итог" in value
        )
    return False


def _role_column(rows: tuple[tuple[object, ...], ...], role: str) -> int | None:
    columns = _role_columns(rows, role)
    return columns[0] if len(columns) == 1 else None


def _role_columns(rows: tuple[tuple[object, ...], ...], role: str) -> list[int]:
    matches = [
        column
        for row_number, row in enumerate(rows, 1)
        for column in range(1, len(row) + 1)
        if _role_text(_header_path(rows, row_number, column), role)
    ]
    return sorted(set(matches))


def _role_header_row(rows: tuple[tuple[object, ...], ...], column: int, role: str) -> int:
    rows = _rows(rows)
    return min(
        (
            number
            for number in range(1, len(rows) + 1)
            if _role_text(_header_path(rows, number, column), role)
        ),
        default=1,
    )


def _unit_column(rows: tuple[tuple[object, ...], ...]) -> int | None:
    """Find one semantic unit column across variable hierarchical headers."""
    matches = _unit_columns(rows)
    return matches[0] if len(matches) == 1 else None


def _unit_columns(rows: tuple[tuple[object, ...], ...]) -> list[int]:
    matches = [
        column
        for row_number, row in enumerate(rows, 1)
        for column in range(1, len(row) + 1)
        if _unit_text(_header_path(rows, row_number, column))
    ]
    return sorted(set(matches))


def _unit_text(value: str) -> bool:
    compact = _header_text(value)
    return compact in _UNIT_ALIASES or ("ед" in compact and "измер" in compact) or "unit" in compact


def _ks2_metric_pair(
    rows: tuple[tuple[object, ...], ...],
) -> tuple[int | None, int | None, int]:
    """Require one explicit quantity / total-cost pair; unit prices are ineligible."""
    pairs = _metric_pairs_for_row(rows)
    if len(pairs) == 1:
        return pairs[0]
    if len(pairs) > 1:
        raise SourceLayoutAmbiguousError("SOURCE_LAYOUT_AMBIGUOUS")
    return None, None, 1


def _metric_pairs_for_row(header) -> list[tuple[int, int, int]]:
    rows = _rows(header)
    return [
        (quantity, cost, row_number)
        for row_number, row in enumerate(rows, 1)
        for quantity, cost in _metric_pairs(row, 1, len(row))
    ]


def _token_header_row(rows: tuple[tuple[object, ...], ...], column: int, token: str) -> int:
    return max(
        (number for number, row in enumerate(rows, 1) if token in _text_at(row, column)),
        default=1,
    )


def _unit_header_row(rows: tuple[tuple[object, ...], ...], column: int) -> int:
    rows = _rows(rows)
    return max(
        (number for number, row in enumerate(rows, 1) if _unit_text(_text_at(row, column))),
        default=1,
    )


def _metric_pair(
    rows: tuple[tuple[object, ...], ...], anchor: int
) -> tuple[int | None, int | None, int]:
    candidates = _metric_pairs_for_anchor(rows, anchor)
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        raise SourceLayoutAmbiguousError("SOURCE_LAYOUT_AMBIGUOUS")
    return None, None, 1


def _metric_pairs_for_anchor(header, anchor: int) -> list[tuple[int, int, int]]:
    return [
        (quantity, cost, leaf_row)
        for quantity, cost, leaf_row, _span in _metric_regions_for_anchor(header, anchor)
    ]


def _metric_regions_for_anchor(
    header, anchor: int
) -> list[tuple[int, int, int, tuple[int, int, int, int]]]:
    """Bind leaves to the exact cumulative header cell that nominated them."""

    rows = _rows(header)
    candidates: dict[
        tuple[int, int, int, tuple[int, int, int, int]],
        tuple[int, int, int, tuple[int, int, int, int]],
    ] = {}
    for row_number, _row in enumerate(rows, 1):
        path = _header_path(rows, row_number, anchor)
        if not _role_text(path, "cumulative"):
            continue
        local = _text_at(rows[row_number - 1], anchor)
        previous = _header_path(rows, row_number - 1, anchor) if row_number > 1 else ""
        if not _role_text(local, "cumulative") and _role_text(previous, "cumulative"):
            continue

        exact_span = None
        if isinstance(header, _HeaderGraph):
            exact_span = header.spans.get((row_number, anchor))
        if exact_span is not None and exact_span[3] > exact_span[1]:
            top, start, bottom, end = exact_span
            pairs = _nearest_metric_pairs(rows, bottom + 1, start, end, anchor)
            regions = [
                (quantity, cost, leaf_row, (top, start, bottom, end))
                for quantity, cost, leaf_row in pairs
            ]
        else:
            pairs = _nearest_metric_pairs(rows, row_number + 1, 1, len(_row), anchor)
            regions = [
                (quantity, cost, leaf_row, (row_number, quantity, row_number, cost))
                for quantity, cost, leaf_row in pairs
                if quantity <= anchor <= cost
            ]
        candidates.update((region, region) for region in regions)
    return list(candidates.values())


def _nearest_metric_pairs(
    rows: tuple[tuple[object, ...], ...],
    first_row: int,
    start: int,
    end: int,
    anchor: int,
) -> list[tuple[int, int, int]]:
    """Return the nearest explicit adjacent leaves before another cumulative region."""

    for leaf_row in range(first_row, len(rows) + 1):
        if _role_text(_text_at(rows[leaf_row - 1], anchor), "cumulative"):
            break
        pairs = _metric_pairs(rows[leaf_row - 1], start, end)
        if pairs:
            return [(quantity, cost, leaf_row) for quantity, cost in pairs]
    return []


def _detail_start(
    sheet,
    formula_sheet,
    header_end: int,
    work_column: int,
    unit_column: int,
    quantity_column: int,
    cost_column: int,
) -> int:
    """Find first semantic detail row; never assume a fixed header depth."""
    values_rows = sheet.iter_rows(min_row=header_end + 1, values_only=True)
    for row_number, values in enumerate(values_rows, header_end + 1):
        work, unit = _text_at(values, work_column), _text_at(values, unit_column)
        formulas_present = (
            formula_sheet.cell(row_number, quantity_column).data_type == "f"
            or formula_sheet.cell(row_number, cost_column).data_type == "f"
        )
        if (
            work
            and unit
            and _HIERARCHY_VALUE_RE.fullmatch(unit) is None
            and (
                formulas_present
                or (
                    _decimal(_value_at(values, quantity_column)) is not None
                    and _decimal(_value_at(values, cost_column)) is not None
                )
            )
        ):
            return row_number
    return header_end + 1


def _column_within(row: tuple[object, ...], start: int, end: int, token: str) -> int | None:
    return next(
        (column for column in range(start, end + 1) if token in _text_at(row, column)), None
    )


def _metric_pairs(row: tuple[object, ...], start: int, end: int) -> list[tuple[int, int]]:
    quantities = [
        column
        for column in range(start, min(end, len(row)) + 1)
        if any(stem in _text_at(row, column) for stem in ("колич", "объем", "объём"))
    ]
    costs = [
        column
        for column in range(start, min(end, len(row)) + 1)
        if _cost_text(_text_at(row, column))
    ]
    return [(quantity, cost) for quantity in quantities for cost in costs if cost == quantity + 1]


def _cost_text(value: str) -> bool:
    return ("стоим" in value or "сумм" in value or "затрат" in value) and not (
        "единиц" in value or "цен" in value
    )


def _canonical_rows(
    sheet,
    source_id: str,
    descriptor: ReconciliationSourceDescriptor,
    *,
    source_type: str,
    start_row: int,
    end_row: int | None = None,
    work_column: int,
    unit_column: int,
    quantity_column: int,
    cost_column: int,
    cumulative: bool,
    formula_sheet=None,
) -> tuple[CanonicalSourceRow, ...]:
    rows: list[CanonicalSourceRow] = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True), start_row
    ):
        work_name = _text_at(values, work_column)
        unit = _text_at(values, unit_column)
        quantity = _decimal(_value_at(values, quantity_column))
        cost = _decimal(_value_at(values, cost_column))
        quantity_formula = (
            formula_sheet is not None
            and formula_sheet.cell(row_number, quantity_column).data_type == "f"
        )
        cost_formula = (
            formula_sheet is not None
            and formula_sheet.cell(row_number, cost_column).data_type == "f"
        )
        if (
            not work_name
            or not unit
            or _HIERARCHY_VALUE_RE.fullmatch(unit) is not None
            or (quantity is None and not quantity_formula)
            or (cost is None and not cost_formula)
        ):
            continue
        if formula_sheet is not None:
            for column, cached in ((quantity_column, quantity), (cost_column, cost)):
                formula_cell = formula_sheet.cell(row_number, column)
                if formula_cell.data_type == "f" and cached is None:
                    raise FormulaCacheUnavailableError("FORMULA_CACHE_UNAVAILABLE")
        location = SourceLocation(
            source_file_id=source_id,
            filename=descriptor.safe_basename,
            sheet_name=sheet.title,
            sheet_type=source_type,
            row_number=row_number,
        )
        rows.append(
            CanonicalSourceRow(
                row_id=f"{source_id}:{source_type}:{row_number}",
                source_type=source_type,
                source_location=location,
                document_index=descriptor.document_index,
                document_period=descriptor.document_period,
                object_code_raw=None,
                object_name_raw=None,
                subobject_code_raw=None,
                subobject_name_raw=None,
                position_code_raw=_text_at(values, 2),
                work_name_raw=work_name,
                unit_raw=unit,
                contract_quantity=None,
                current_period_quantity=None if cumulative else quantity,
                cumulative_quantity=quantity if cumulative else None,
                remaining_quantity=None,
                unit_price=None,
                contract_cost=None,
                current_period_cost=None if cumulative else cost,
                cumulative_cost=cost if cumulative else None,
                total_cost=None,
                basis_code_raw=None,
                drawing_code_raw=_text_at(values, 7),
                cost_type_code_raw=_text_at(values, 3),
                source_values=(),
                status="OK",
                warnings=(),
            )
        )
    return tuple(
        replace(
            row,
            current_period_quantity=row.cumulative_quantity,
            current_period_cost=row.cumulative_cost,
        )
        if cumulative
        else row
        for row in rows
    )


def _value_at(row: tuple[object, ...], column: int) -> object | None:
    return row[column - 1] if column <= len(row) else None


def _text_at(row: tuple[object, ...], column: int) -> str:
    return _text(_value_at(row, column))


def _text(value: object | None) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or "")).replace("\u00a0", " ")
    return " ".join(re.sub(r"[^\w]+", " ", normalized, flags=re.UNICODE).casefold().split())


def _header_text(value: object | None) -> str:
    return _text(value)


def _decimal(value: object | None) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    try:
        parsed = Decimal(str(value).replace("\u00a0", "").replace(" ", "").replace(",", "."))
    except InvalidOperation:
        return None
    return parsed if parsed.is_finite() else None


def _issue(code: str, descriptor: ReconciliationSourceDescriptor) -> ReconciliationSourceIssue:
    if code == "DOCUMENT_INDEX_MISSING":
        return ReconciliationSourceIssue(
            code=code,
            safe_basename=descriptor.safe_basename,
            comment="В имени файла не найден точный трёх- или четырёхзначный индекс документа.",
            repair_hint=(
                "Добавьте один индекс из целевого отчёта, например «1234», "
                "в имя файла и загрузите его снова."
            ),
            can_continue=True,
        )
    if code == "WORKBOOK_UNREADABLE":
        return ReconciliationSourceIssue(
            code=code,
            safe_basename=descriptor.safe_basename,
            comment="Не удалось безопасно прочитать источник для сверки.",
            repair_hint="Загрузите файл Excel повторно или выберите исправную копию.",
            can_continue=True,
        )
    if code == "FORMULA_CACHE_UNAVAILABLE":
        return ReconciliationSourceIssue(
            code=code,
            safe_basename=descriptor.safe_basename,
            comment="В расчётных ячейках источника нет проверяемых значений формул.",
            repair_hint="Пересчитайте книгу в Excel и загрузите сохранённую копию.",
            can_continue=True,
        )
    if code == "SOURCE_LAYOUT_AMBIGUOUS":
        return ReconciliationSourceIssue(
            code=code,
            safe_basename=descriptor.safe_basename,
            comment="В источнике найдено несколько несовместимых табличных структур.",
            repair_hint="Оставьте один однозначный лист с таблицей и загрузите книгу снова.",
            can_continue=True,
        )
    return ReconciliationSourceIssue(
        code=code,
        safe_basename=descriptor.safe_basename,
        comment="В источнике не найдены пригодные накопительные строки КС-6а или КС-2.",
        repair_hint="Загрузите источник с заполненными строками КС-6а или КС-2.",
        can_continue=True,
    )


def _has_usable_document_index(descriptor: ReconciliationSourceDescriptor) -> bool:
    raw = descriptor.document_index or ""
    return re.fullmatch(r"\d{3,4}", raw) is not None

"""Adapter for the documented additional-report reconciliation layout."""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import tempfile
from collections.abc import Iterable
from contextlib import suppress
from dataclasses import dataclass, replace
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl.cell.read_only import ReadOnlyCell
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet._reader import WorkSheetParser

from report_processor.excel import WorkbookOpenRequest, open_dual_workbook
from report_processor.processing.adapters import _materialized
from report_processor.schema import (
    LogicalColumn,
    SheetType,
    analyze_workbook_schema,
    resolve_logical_columns,
)
from report_processor.schema.column_aliases import DEFAULT_COLUMN_ALIASES
from report_processor.target_report import (
    TargetCellSnapshot,
    TargetColumnBinding,
    TargetObjectBlock,
    TargetReportReadRequest,
    TargetReportRow,
)
from report_processor.target_report.ooxml import (
    formula_caches_trusted,
    read_sheet_comments,
    read_sheet_lexemes,
    read_sheet_structure,
)
from report_processor.target_report.reader import _cell_snapshot

from .reconciliation_identity import terminal_identity
from .reconciliation_target_measure import TargetMeasurePair, discover_target_measures

_STAGE_RE = re.compile(r"этап\s*([0-9]+(?:\.[0-9]+)*)", re.IGNORECASE)
_SHA256_RE = re.compile(r"[0-9a-f]{64}\Z")
_BASE_ROLES = (
    LogicalColumn.DOCUMENT_INDEX,
    LogicalColumn.STAGE,
    LogicalColumn.ROW_NUMBER,
    LogicalColumn.WORK_NAME,
    LogicalColumn.UNIT,
)
_CORE_ROLES = _BASE_ROLES[:3]
_MAX_ROLE_SCAN_ROWS = 100_000
_MAX_ROLE_SCAN_CELLS = 500_000
_MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


@dataclass(frozen=True, slots=True)
class _PhysicalWorksheetSnapshot:
    cells: dict[tuple[int, int], object]
    rows: tuple[int, ...]
    inspected: tuple[int, int]


class _SnapshotWorksheet:
    """Request-local worksheet view backed only by parsed physical cells."""

    def __init__(self, worksheet, snapshot: _PhysicalWorksheetSnapshot) -> None:
        self._worksheet = worksheet
        self._snapshot = snapshot
        self.title = worksheet.title
        # Discovery APIs deliberately use this physical index too; it avoids a
        # read-only worksheet's per-cell XML reparse while preserving sparse rows.
        self._cells = snapshot.cells

    @property
    def max_row(self) -> int:
        return int(getattr(self._worksheet, "max_row", 0) or 0)

    @property
    def max_column(self) -> int:
        return int(getattr(self._worksheet, "max_column", 0) or 0)

    def cell(self, row: int, column: int, value=None):
        if value is not None:
            raise TypeError("snapshot worksheet is read-only")
        return _cell_at(self._snapshot, self._worksheet, row, column)

    def iter_rows(
        self,
        min_row: int | None = None,
        max_row: int | None = None,
        min_col: int | None = None,
        max_col: int | None = None,
        values_only: bool = False,
    ):
        start_row = max(1, min_row or 1)
        end_row = min(int(max_row or self.max_row), start_row + _MAX_ROLE_SCAN_ROWS - 1)
        start_column = max(1, min_col or 1)
        end_column = min(int(max_col or self.max_column), 16_384)
        width = max(1, end_column - start_column + 1)
        end_row = min(end_row, start_row + (_MAX_ROLE_SCAN_CELLS // width) - 1)
        for row_number in range(start_row, end_row + 1):
            row = tuple(
                _cell_at(self._snapshot, self._worksheet, row_number, column)
                for column in range(start_column, end_column + 1)
            )
            yield tuple(cell.value for cell in row) if values_only else row

    def __getattr__(self, name):
        return getattr(self._worksheet, name)


class _SnapshotWorkbook:
    """Workbook facade that prevents downstream readers from opening XML per cell."""

    def __init__(self, workbook, snapshots: dict[str, _PhysicalWorksheetSnapshot]) -> None:
        self._workbook = workbook
        self._worksheets = {
            name: _SnapshotWorksheet(workbook[name], snapshots[name])
            for name in workbook.sheetnames
        }

    @property
    def sheetnames(self):
        return self._workbook.sheetnames

    @property
    def worksheets(self):
        return tuple(self._worksheets[name] for name in self.sheetnames)

    @property
    def active(self):
        return self._worksheets[self._workbook.active.title]

    def __getitem__(self, key):
        return self._worksheets[key]

    def __getattr__(self, name):
        return getattr(self._workbook, name)


@dataclass(frozen=True, slots=True)
class ReconciliationTargetIdentity:
    """Digest binding review input to one immutable target interpretation."""

    original_target_digest: str
    selected_stage: str
    period: object | None = None
    plan_digest: str | None = None
    contract_version: str = "ReconciliationTargetIdentity-1.0"

    def __post_init__(self) -> None:
        if self.contract_version != "ReconciliationTargetIdentity-1.0":
            raise ValueError("RECONCILIATION_TARGET_IDENTITY_INVALID")
        if not isinstance(self.original_target_digest, str) or not _SHA256_RE.fullmatch(
            self.original_target_digest
        ):
            raise ValueError("RECONCILIATION_TARGET_IDENTITY_INVALID")
        if not isinstance(self.selected_stage, str) or not self.selected_stage:
            raise ValueError("RECONCILIATION_TARGET_IDENTITY_INVALID")
        period_value = getattr(self.period, "value", self.period)
        if period_value is not None and not isinstance(period_value, str):
            raise ValueError("RECONCILIATION_TARGET_IDENTITY_INVALID")
        if (period_value is None) != (self.plan_digest is None):
            raise ValueError("RECONCILIATION_TARGET_IDENTITY_INVALID")
        if period_value is not None:
            from .reconciliation_period import ReportingPeriod

            try:
                ReportingPeriod.parse(period_value)
            except ValueError as error:
                raise ValueError("RECONCILIATION_TARGET_IDENTITY_INVALID") from error
            object.__setattr__(self, "period", period_value)
        if self.plan_digest is not None and (
            not isinstance(self.plan_digest, str) or not _SHA256_RE.fullmatch(self.plan_digest)
        ):
            raise ValueError("RECONCILIATION_TARGET_IDENTITY_INVALID")

    def canonical_bytes(self) -> bytes:
        return json.dumps(
            {
                "contract_version": self.contract_version,
                "original_target_digest": self.original_target_digest,
                "period": getattr(self.period, "value", self.period),
                "plan_digest": self.plan_digest,
                "selected_stage": self.selected_stage,
            },
            ensure_ascii=True,
            sort_keys=True,
            separators=(",", ":"),
        ).encode()

    @property
    def target_identity_digest(self) -> str:
        return hashlib.sha256(self.canonical_bytes()).hexdigest()

    @property
    def reporting_period(self) -> str | None:
        return self.period


@dataclass(frozen=True, slots=True)
class ReconciliationTargetPreview:
    """Read-only historical target projection and frozen insertion evidence."""

    schema: object
    rows: tuple[object, ...]
    period: object | None
    plan: object | None
    target_identity: ReconciliationTargetIdentity

    @property
    def target_identity_digest(self) -> str:
        return self.target_identity.target_identity_digest


class ReconciliationTargetScopeError(ValueError):
    """The target cannot supply one safe reconciliation stage."""


class ReconciliationTargetInputError(ValueError):
    """The selected target type is unsafe for reconciliation output."""


def publish_unchanged_target(source, output, expected_sha256: str) -> str:
    """Atomically publish one verified byte-identical target copy without clobbering."""
    source_path, output_path = Path(source), Path(output)
    _validate_reconciliation_target_type(source_path)
    if output_path.exists() or output_path.is_symlink():
        raise ValueError("RECONCILIATION_OUTPUT_EXISTS")
    source_sha256 = _sha256(source_path)
    if source_sha256 != expected_sha256:
        raise ValueError("RECONCILIATION_TARGET_CHANGED")
    _reopen_xlsx(source_path)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=".reconciliation-", suffix=".xlsx", dir=output_path.parent
    )
    temporary = Path(temporary_name)
    published = False
    completed = False
    linked_identity: tuple[int, int] | None = None
    try:
        with os.fdopen(descriptor, "wb") as destination, source_path.open("rb") as input_stream:
            shutil.copyfileobj(input_stream, destination, length=1_048_576)
            destination.flush()
            os.fsync(destination.fileno())
        if _sha256(temporary) != source_sha256 or _sha256(source_path) != source_sha256:
            raise ValueError("RECONCILIATION_TARGET_CHANGED")
        _reopen_xlsx(temporary)
        linked_identity = _file_identity(temporary)
        os.link(temporary, output_path)
        published = True
        output_sha256 = _sha256(output_path)
        if output_sha256 != source_sha256:
            raise ValueError("RECONCILIATION_OUTPUT_VERIFY_FAILED")
        _reopen_xlsx(output_path)
        completed = True
        return output_sha256
    except OSError as error:
        if error.errno == 17:
            raise ValueError("RECONCILIATION_OUTPUT_EXISTS") from None
        raise RuntimeError("RECONCILIATION_NO_CHANGE_PUBLISH_FAILED") from error
    finally:
        if published and not completed and _same_inode(output_path, linked_identity):
            output_path.unlink(missing_ok=True)
        temporary.unlink(missing_ok=True)


def read_reconciliation_target(path, digest: str, stage: str | None):
    """Read one structurally proven current-period pair per selected sheet."""
    _validate_reconciliation_target_type(Path(path))
    source = _materialized(path, f"target:{digest}")
    with open_dual_workbook(WorkbookOpenRequest(source)) as session:
        formula_all, value_all = _request_snapshots(session)
        adapted = _snapshot_session(session, formula_all, value_all)
        generic = __import__("report_processor.target_report", fromlist=["read_target_report"])
        workbook_schema = analyze_workbook_schema(adapted)
        roles = _base_roles(workbook_schema)
        formula_snapshots, value_snapshots = _session_snapshots(adapted, roles)
        stages = _enumerate_stages(adapted.formula_workbook, roles, formula_snapshots)
        selected_stage = resolve_reconciliation_stage(stages, stage)
        detail_rows = _first_detail_rows(
            adapted.formula_workbook, selected_stage, roles, formula_snapshots
        )
        if not detail_rows:
            raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_STAGE_EMPTY")
        measure_pairs = discover_target_measures(
            adapted.formula_workbook,
            detail_rows,
            {
                sheet_name: read_sheet_structure(
                    session.source.local_path, sheet_name
                ).merged_ranges
                for sheet_name in session.formula_workbook.sheetnames
            },
        )
        report = generic.read_target_report(
            adapted,
            workbook_schema,
            TargetReportReadRequest(selected_stage=selected_stage, max_rows=0),
        )
        bindings = _bindings(roles, measure_pairs)
        rows = tuple(
            _rows(adapted, selected_stage, measure_pairs, roles, formula_snapshots, value_snapshots)
        )
    if not rows:
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_STAGE_EMPTY")
    schema = replace(report.schema, column_bindings=bindings, object_blocks=_object_blocks(rows))
    return schema, rows


def category_id(label: str) -> str:
    return "category:" + " ".join(label.casefold().split())


def terminal_index(value: object) -> str | None:
    """Compatibility adapter for reconciliation terminal identities."""
    return terminal_identity(value)


def enumerate_reconciliation_stages(session) -> tuple[str, ...]:
    """Return stages only after logical role binding succeeds."""
    formula_all, value_all = _request_snapshots(session)
    adapted = _snapshot_session(session, formula_all, value_all)
    roles = _base_roles(analyze_workbook_schema(adapted))
    formula_snapshots, _value_snapshots = _session_snapshots(adapted, roles)
    return _enumerate_stages(adapted.formula_workbook, roles, formula_snapshots)


def structurally_valid_reconciliation_stages(session, *, maximum: int) -> tuple[str, ...]:
    """Find stages with at least one target row in one workbook pass.

    The criteria mirror ``_rows`` without retaining sheet names, row numbers,
    or other provenance.  ``maximum + 1`` lets callers detect an overlarge
    ambiguous selection without scanning or opening the workbook per stage.
    """

    if not isinstance(maximum, int) or maximum < 1:
        raise ValueError("maximum must be positive")
    formula_all, value_all = _request_snapshots(session)
    adapted = _snapshot_session(session, formula_all, value_all)
    roles = _base_roles(analyze_workbook_schema(adapted))
    formula_snapshots, _value_snapshots = _session_snapshots(adapted, roles)
    return _valid_stages(adapted.formula_workbook, roles, formula_snapshots, maximum)


def _valid_stages(workbook, roles, snapshots, maximum: int) -> tuple[str, ...]:
    stages: set[str] = set()
    for sheet in workbook.worksheets:
        columns = roles.get(sheet.title)
        if columns is None:
            continue
        active_index = active_stage = None
        for row_number in _role_rows(snapshots[sheet.title], columns):
            active_index, active_stage, values = _carried_role_values(
                snapshots[sheet.title], row_number, columns, active_index, active_stage
            )
            if not _semantic_detail(active_stage, active_index, values):
                continue
            stages.add(active_stage)
            if len(stages) > maximum:
                return tuple(sorted(stages))
    return tuple(sorted(stages))


def resolve_reconciliation_stage(stages: tuple[str, ...], requested: str | None) -> str:
    """Resolve only explicit existing stage or exactly one discovered stage."""
    if requested is not None:
        if requested in stages:
            return requested
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_STAGE_MISSING")
    if len(stages) == 1:
        return stages[0]
    if not stages:
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_STAGE_EMPTY")
    raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_STAGE_AMBIGUOUS")


def writer_calculations(calculations: Iterable[object]) -> tuple[object, ...]:
    """Adapt reconciliation values to the target's two-decimal, million-RUB cells."""
    return tuple(
        replace(
            calculation,
            quantity=_quantize(calculation.quantity),
            cost=_million_rub(calculation.cost),
        )
        for calculation in calculations
    )


def _quantize(value: Decimal | None) -> Decimal | None:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if value is not None else None


def _million_rub(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    return (value / Decimal(1_000_000)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _bindings(
    roles: dict[str, dict[LogicalColumn, object]],
    measure_pairs: tuple[TargetMeasurePair, ...] = (),
) -> tuple[TargetColumnBinding, ...]:
    canonical_roles = roles[sorted(roles)[0]]
    bindings = [
        TargetColumnBinding(
            role,
            resolution.column_index,
            resolution.column_letter,
            resolution.header_text,
            "RECONCILIATION_SCHEMA",
        )
        for role, resolution in canonical_roles.items()
    ]
    seen = {(binding.logical_column, binding.column_index) for binding in bindings}
    for pair in measure_pairs:
        for logical_column, column, letter, header in (
            (
                LogicalColumn.CURRENT_PERIOD_QUANTITY,
                pair.quantity_column,
                pair.quantity_letter,
                pair.quantity_header,
            ),
            (
                LogicalColumn.CURRENT_PERIOD_COST,
                pair.cost_column,
                pair.cost_letter,
                pair.cost_header,
            ),
        ):
            if (logical_column, column) not in seen:
                seen.add((logical_column, column))
                bindings.append(
                    TargetColumnBinding(logical_column, column, letter, header, "TARGET_MEASURE")
                )
    return tuple(bindings)


def _preview_bindings(roles, plan) -> tuple[TargetColumnBinding, ...]:
    pairs = tuple(
        TargetMeasurePair(
            anchor.sheet_name,
            anchor.cost_column + 1,
            anchor.cost_column + 2,
            "",
            "",
        )
        for anchor in plan.anchors
    )
    return _bindings(roles, pairs)


def _base_roles(workbook_schema) -> dict[str, dict[LogicalColumn, object]]:
    """Bind reconciliation facts only through existing logical schema evidence."""

    result: dict[str, dict[LogicalColumn, object]] = {}
    for worksheet in workbook_schema.worksheets:
        resolutions = resolve_logical_columns(
            worksheet.headers, SheetType.ADDITIONAL_REPORT, DEFAULT_COLUMN_ALIASES
        )
        by_role = {item.logical_column: item for item in resolutions}
        bound_roles = {
            role
            for role in _BASE_ROLES
            if by_role.get(role) is not None
            and by_role[role].status == "OK"
            and by_role[role].column_index is not None
            and by_role[role].column_letter is not None
        }
        core_count = len(bound_roles.intersection(_CORE_ROLES))
        if not (core_count >= 2 or (core_count >= 1 and len(bound_roles) >= 3)):
            continue
        resolved: dict[LogicalColumn, object] = {}
        for role in _BASE_ROLES:
            resolution = by_role.get(role)
            if (
                resolution is None
                or resolution.status == "COLUMN_NOT_FOUND"
                or (resolution.status != "OK" and "PHYSICAL_COLUMN_CONFLICT" in resolution.warnings)
            ):
                raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_BASE_ROLE_MISSING")
            if (
                resolution.status != "OK"
                or resolution.column_index is None
                or resolution.column_letter is None
            ):
                raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_BASE_ROLE_AMBIGUOUS")
            resolved[role] = resolution
        result[worksheet.sheet_name] = resolved
    if not result:
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_BASE_ROLE_MISSING")
    mappings = {
        tuple((role, item.column_index) for role, item in values.items())
        for values in result.values()
    }
    if len(mappings) != 1:
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_BASE_ROLE_HETEROGENEOUS")
    return result


def _enumerate_stages(workbook, roles, snapshots) -> tuple[str, ...]:
    stages: set[str] = set()
    for sheet in workbook.worksheets:
        if sheet.title not in roles:
            continue
        columns = roles[sheet.title]
        active_index = active_stage = None
        snapshot = snapshots[sheet.title]
        for row_number in _role_rows(snapshot, columns):
            active_index, active_stage, _values = _carried_role_values(
                snapshot, row_number, columns, active_index, active_stage
            )
            if active_stage is None or active_index is None:
                continue
            stages.add(active_stage)
    return tuple(sorted(stages))


def _request_snapshots(
    session,
) -> tuple[dict[str, _PhysicalWorksheetSnapshot], dict[str, _PhysicalWorksheetSnapshot]]:
    cache_key = "reconciliation-physical-cells-v1"
    cached = session.structure_cache.get(cache_key)
    if cached is not None:
        return cached
    result = (
        _view_snapshots(session.formula_workbook),
        _view_snapshots(session.value_workbook),
    )
    session.structure_cache[cache_key] = result
    return result


def _snapshot_session(session, formula_snapshots, value_snapshots):
    return replace(
        session,
        formula_workbook=_SnapshotWorkbook(session.formula_workbook, formula_snapshots),
        value_workbook=_SnapshotWorkbook(session.value_workbook, value_snapshots),
    )


def _session_snapshots(
    session, roles
) -> tuple[dict[str, _PhysicalWorksheetSnapshot], dict[str, _PhysicalWorksheetSnapshot]]:
    formula_all, value_all = _request_snapshots(session)
    return (
        {sheet_name: formula_all[sheet_name] for sheet_name in roles},
        {sheet_name: value_all[sheet_name] for sheet_name in roles},
    )


def _view_snapshots(workbook) -> dict[str, _PhysicalWorksheetSnapshot]:
    return {sheet.title: _physical_snapshot(sheet) for sheet in workbook.worksheets}


def _physical_snapshot(sheet) -> _PhysicalWorksheetSnapshot:
    cached = getattr(sheet, "_reconciliation_physical_snapshot", None)
    if cached is not None:
        return cached
    cells = getattr(sheet, "_cells", None)
    if isinstance(cells, dict):
        rows = {row for row, _ in cells}
        if len(rows) > _MAX_ROLE_SCAN_ROWS or len(cells) > _MAX_ROLE_SCAN_CELLS:
            raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_LIMIT")
        snapshot = _PhysicalWorksheetSnapshot(
            dict(cells),
            tuple(sorted(rows)),
            (len(rows), len(cells)),
        )
    else:
        snapshot = _read_only_physical_snapshot(sheet)
    with suppress(AttributeError, TypeError):
        sheet._reconciliation_physical_snapshot = snapshot
    return snapshot


def _read_only_physical_snapshot(sheet) -> _PhysicalWorksheetSnapshot:
    source = getattr(sheet, "_get_source", None)
    if source is None:
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_UNSUPPORTED")
    parser = WorkSheetParser(
        None,
        sheet._shared_strings,
        data_only=sheet.parent.data_only,
        epoch=sheet.parent.epoch,
        date_formats=sheet.parent._date_formats,
        timedelta_formats=sheet.parent._timedelta_formats,
    )
    cells: dict[tuple[int, int], object] = {}
    rows: list[int] = []
    inspected_rows = inspected_cells = previous_row = 0
    with source() as stream:
        for _event, element in ET.iterparse(stream, events=("end",)):
            if element.tag != f"{_MAIN_NS}row":
                continue
            inspected_rows += 1
            if inspected_rows > _MAX_ROLE_SCAN_ROWS:
                raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_LIMIT")
            row_number = _ooxml_row_number(element.attrib.get("r"), previous_row)
            previous_row = row_number
            previous_column = 0
            for cell in element:
                if cell.tag != f"{_MAIN_NS}c":
                    continue
                inspected_cells += 1
                if inspected_cells > _MAX_ROLE_SCAN_CELLS:
                    raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_LIMIT")
                previous_column = _ooxml_cell_column(
                    cell.attrib.get("r"), row_number, previous_column
                )
            parsed_row, parsed_cells = parser.parse_row(element)
            if parsed_row != row_number:
                raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_INVALID")
            rows.append(row_number)
            for parsed in parsed_cells:
                cells[parsed["row"], parsed["column"]] = ReadOnlyCell(
                    sheet,
                    parsed["row"],
                    parsed["column"],
                    parsed["value"],
                    parsed["data_type"],
                    parsed["style_id"],
                )
            element.clear()
    return _PhysicalWorksheetSnapshot(cells, tuple(rows), (inspected_rows, inspected_cells))


def _role_rows(snapshot: _PhysicalWorksheetSnapshot, columns) -> tuple[int, ...]:
    role_columns = {item.column_index for item in columns.values()}
    return tuple(
        row
        for row in snapshot.rows
        if any((row, column) in snapshot.cells for column in role_columns)
    )


def _ooxml_row_number(reference: str | None, previous_row: int) -> int:
    if not isinstance(reference, str) or not reference.isdecimal():
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_INVALID")
    row_number = int(reference)
    if not 1 <= row_number <= 1_048_576 or row_number <= previous_row:
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_INVALID")
    return row_number


def _ooxml_cell_column(reference: str | None, row_number: int, previous_column: int) -> int:
    if not isinstance(reference, str):
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_INVALID")
    try:
        column_letter, cell_row = coordinate_from_string(reference)
        column = column_index_from_string(column_letter)
    except ValueError as error:
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_INVALID") from error
    if cell_row != row_number or not 1 <= column <= 16_384 or column <= previous_column:
        raise ReconciliationTargetScopeError("RECONCILIATION_TARGET_ROLE_SCAN_INVALID")
    return column


def _carried_role_values(snapshot, row_number, columns, active_index, active_stage):
    values = {
        role: _cell_value(snapshot, row_number, resolution.column_index)
        for role, resolution in columns.items()
    }
    raw_index = values[LogicalColumn.DOCUMENT_INDEX]
    if raw_index is not None:
        active_index = terminal_index(raw_index)
    raw_stage = values[LogicalColumn.STAGE]
    if raw_stage is not None and str(raw_stage).strip():
        stage_name = str(raw_stage).strip()
        stage_match = _STAGE_RE.search(stage_name)
        active_stage = stage_match.group(1) if stage_match else stage_name
    return active_index, active_stage, values


def _semantic_detail(
    active_stage,
    active_index,
    values,
    *,
    selected_stage: str | None = None,
) -> bool:
    return (
        active_index is not None
        and active_stage is not None
        and (selected_stage is None or active_stage == selected_stage)
        and _nonempty(values[LogicalColumn.ROW_NUMBER])
        and _nonempty(values[LogicalColumn.WORK_NAME])
        and _nonempty(values[LogicalColumn.UNIT])
    )


def _cell_value(snapshot: _PhysicalWorksheetSnapshot, row_number: int, column: int) -> object:
    return getattr(snapshot.cells.get((row_number, column)), "value", None)


def _cell_at(snapshot: _PhysicalWorksheetSnapshot, sheet, row_number: int, column: int):
    cell = snapshot.cells.get((row_number, column))
    return cell if cell is not None else ReadOnlyCell(sheet, row_number, column, None)


def _nonempty(value: object) -> bool:
    return value is not None and bool(str(value).strip())


def _first_detail_rows(workbook, selected_stage: str, roles, snapshots) -> dict[str, int]:
    rows: dict[str, int] = {}
    for sheet in workbook.worksheets:
        if sheet.title not in roles:
            continue
        columns = roles[sheet.title]
        active_index = active_stage = None
        snapshot = snapshots[sheet.title]
        for row_number in _role_rows(snapshot, columns):
            active_index, active_stage, values = _carried_role_values(
                snapshot, row_number, columns, active_index, active_stage
            )
            if _semantic_detail(active_stage, active_index, values, selected_stage=selected_stage):
                rows[sheet.title] = row_number
                break
    return rows


def _rows(
    session,
    selected_stage: str,
    measure_pairs: tuple[TargetMeasurePair, ...],
    roles,
    formula_snapshots,
    value_snapshots,
):
    return _rows_for_pairs(
        session,
        selected_stage,
        measure_pairs,
        roles,
        formula_snapshots,
        value_snapshots,
        writable=True,
    )


def _preview_rows(session, selected_stage: str, plan, roles, formula_snapshots, value_snapshots):
    pairs = tuple(
        TargetMeasurePair(
            anchor.sheet_name,
            anchor.cost_column + 1,
            anchor.cost_column + 2,
            "",
            "",
        )
        for anchor in plan.anchors
    )
    return _rows_for_pairs(
        session,
        selected_stage,
        pairs,
        roles,
        formula_snapshots,
        value_snapshots,
        writable=False,
        virtual=True,
    )


def _rows_for_pairs(
    session,
    selected_stage: str,
    measure_pairs: tuple[TargetMeasurePair, ...],
    roles,
    formula_snapshots,
    value_snapshots,
    *,
    writable: bool,
    virtual: bool = False,
):
    formula = session.formula_workbook
    values = session.value_workbook
    trusted = formula_caches_trusted(session.source.local_path)
    pair_by_sheet = {pair.sheet_name: pair for pair in measure_pairs}
    for sheet_name in formula.sheetnames:
        formula_sheet, value_sheet = formula[sheet_name], values[sheet_name]
        pair = pair_by_sheet.get(sheet_name)
        if pair is None or sheet_name not in roles:
            continue
        columns = roles[sheet_name]
        formula_snapshot = formula_snapshots[sheet_name]
        value_snapshot = value_snapshots[sheet_name]
        bindings = _bindings({sheet_name: columns}, (pair,))
        lexemes = read_sheet_lexemes(session.source.local_path, sheet_name)
        comments = dict(read_sheet_comments(session.source.local_path, sheet_name))
        active_index = active_name = active_stage = None
        for row_number in _role_rows(formula_snapshot, columns):
            active_index, active_stage, role_values = _carried_role_values(
                formula_snapshot, row_number, columns, active_index, active_stage
            )
            raw_stage = role_values[LogicalColumn.STAGE]
            if raw_stage is not None and str(raw_stage).strip():
                active_name = str(raw_stage).strip()
            if not _semantic_detail(
                active_stage, active_index, role_values, selected_stage=selected_stage
            ):
                continue
            work_name = role_values[LogicalColumn.WORK_NAME]
            position = role_values[LogicalColumn.ROW_NUMBER]
            cells = tuple(
                (binding.logical_column, _virtual_cell(binding.column_letter, row_number))
                if virtual
                and binding.logical_column
                in {LogicalColumn.CURRENT_PERIOD_QUANTITY, LogicalColumn.CURRENT_PERIOD_COST}
                else (
                    binding.logical_column,
                    _cell_snapshot(
                        f"{binding.column_letter}{row_number}",
                        _cell_at(formula_snapshot, formula_sheet, row_number, binding.column_index),
                        _cell_at(value_snapshot, value_sheet, row_number, binding.column_index),
                        lexemes.get(f"{binding.column_letter}{row_number}"),
                        comments.get(f"{binding.column_letter}{row_number}"),
                        trusted,
                    ),
                )
                for binding in bindings
            )
            by_key = dict(cells)
            yield TargetReportRow(
                "ReconciliationTarget-2.0",
                sheet_name,
                SheetType.ADDITIONAL_REPORT,
                row_number,
                None,
                active_name,
                str(position).strip(),
                str(work_name).strip(),
                cells,
                "OK",
                row_kind="OBJECT_PROCESS",
                scope="SELECTED_STAGE",
                document_index_raw=active_index,
                document_index_normalized=active_index,
                stage=active_stage,
                unit=str(role_values[LogicalColumn.UNIT]).strip() or None,
                selected_quantity=_numeric(by_key.get(LogicalColumn.CURRENT_PERIOD_QUANTITY)),
                selected_cost=_numeric(by_key.get(LogicalColumn.CURRENT_PERIOD_COST)),
                writable=writable and trusted,
            )


def _virtual_cell(column_letter: str, row_number: int) -> TargetCellSnapshot:
    return TargetCellSnapshot(
        f"{column_letter}{row_number}",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        "VIRTUAL_FUTURE_CELL",
    )


def _numeric(cell):
    from report_processor.target_report import TargetNumericCell

    if cell is None:
        return None
    return TargetNumericCell(
        cell.numeric_value,
        cell.raw_lexeme,
        cell.formula.cache_state if cell.formula is not None else "NOT_FORMULA",
        cell.status,
    )


def _object_blocks(rows: Iterable[TargetReportRow]) -> tuple[TargetObjectBlock, ...]:
    blocks: list[TargetObjectBlock] = []
    active: TargetObjectBlock | None = None
    for row in rows:
        key = (row.sheet_name, row.object_code, row.object_name)
        if active is None or key != (active.sheet_name, active.object_code, active.object_name):
            if active is not None:
                blocks.append(active)
            active = TargetObjectBlock(
                row.sheet_name, row.row_number, row.row_number, row.object_code, row.object_name
            )
        else:
            active = TargetObjectBlock(
                active.sheet_name,
                active.start_row,
                row.row_number,
                active.object_code,
                active.object_name,
            )
    if active is not None:
        blocks.append(active)
    return tuple(blocks)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1_048_576), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _reopen_xlsx(path: Path) -> None:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=True)
        workbook.close()
    except Exception as error:
        raise ValueError("RECONCILIATION_OUTPUT_VERIFY_FAILED") from error


def _same_inode(path: Path, identity: tuple[int, int] | None) -> bool:
    if identity is None:
        return False
    try:
        current = path.stat()
    except OSError:
        return False
    return (current.st_dev, current.st_ino) == identity


def _file_identity(path: Path) -> tuple[int, int]:
    current = path.stat()
    return current.st_dev, current.st_ino


def _validate_reconciliation_target_type(path: Path) -> None:
    if path.suffix.casefold() == ".xlsm":
        raise ReconciliationTargetInputError(
            "Целевой отчёт .xlsm пока не поддерживается: загрузите целевой файл .xlsx."
        )

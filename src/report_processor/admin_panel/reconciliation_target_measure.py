"""Fail-closed discovery of a reconciliation target's current-period pair."""

from __future__ import annotations

import hashlib
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, range_boundaries

from report_processor.target_report.ooxml import worksheet_parts

_HEADER_ROWS = 80
_MAX_HEADER_WINDOW_CELLS = 500_000
_MAX_SUFFIX_INSPECTED_CELLS = 100_000
_MAX_SUFFIX_CELLS = 50_000
_MAX_SUFFIX_ROW = 1_048_576
_MAX_SUFFIX_COLUMN = 16_384
_MAX_VALIDATED_MERGE_RANGES = 4_096
_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_Q = lambda name: f"{{{_MAIN_NS}}}{name}"  # noqa: E731
_RUSSIAN_MONTH_TOKENS = {
    "январь": 1,
    "января": 1,
    "февраль": 2,
    "февраля": 2,
    "март": 3,
    "марта": 3,
    "апрель": 4,
    "апреля": 4,
    "май": 5,
    "мая": 5,
    "июнь": 6,
    "июня": 6,
    "июль": 7,
    "июля": 7,
    "август": 8,
    "августа": 8,
    "сентябрь": 9,
    "сентября": 9,
    "октябрь": 10,
    "октября": 10,
    "ноябрь": 11,
    "ноября": 11,
    "декабрь": 12,
    "декабря": 12,
}
_YEAR_MONTH = re.compile(r"(?<!\d)((?:19|20)\d{2})\s*[-./]\s*(0?[1-9]|1[0-2])(?!\d)")
_MONTH_YEAR = re.compile(r"(?<!\d)(0?[1-9]|1[0-2])\s*[./-]\s*((?:19|20)\d{2})(?!\d)")
_DATE = re.compile(r"(?<!\d)\d{1,2}\s*[./-]\s*(0?[1-9]|1[0-2])\s*[./-]\s*((?:19|20)\d{2})(?!\d)")
_RUB_CURRENCY = re.compile(r"\bруб\w*\b", flags=re.UNICODE)
_SCALED_RUB = re.compile(
    r"\b(?:тыс(?:яч\w*)?|млн|миллион\w*|млрд|миллиард\w*)\b",
    flags=re.UNICODE,
)
_UNIT_DENOMINATOR = re.compile(
    r"/\s*(?:"
    r"м(?:\d+)?|мм|см|дм|км|"
    r"п\.?м\.?|пог\.?м\.?|"
    r"шт\.?|штук\w*|ед\.?|единиц\w*|"
    r"ч|час\w*|день\w*|сут\w*|"
    r"т|кг|г|л|"
    r"unit\w*|piece\w*|pcs?"
    r")\b",
    flags=re.UNICODE,
)


class ReconciliationTargetMeasureError(ValueError):
    """The selected target does not expose one safe current-period measure."""


def raw_worksheet_merge_ranges(source_path: Path, sheet_name: str) -> tuple[str, ...]:
    """Read one worksheet's unnormalised merge topology from its OOXML part."""

    try:
        part = worksheet_parts(Path(source_path)).get(sheet_name)
        if part is None:
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        with zipfile.ZipFile(source_path) as archive:
            root = ET.fromstring(archive.read(part))
    except ReconciliationTargetMeasureError:
        raise
    except Exception as error:
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID") from error
    containers = root.findall(_Q("mergeCells"))
    if not containers:
        return ()
    if len(containers) != 1:
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    container = containers[0]
    count = container.attrib.get("count")
    if count is None or not count.isdecimal():
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    normalized_count = count.lstrip("0") or "0"
    maximum_count = str(_MAX_VALIDATED_MERGE_RANGES)
    if len(normalized_count) > len(maximum_count) or len(container) > _MAX_VALIDATED_MERGE_RANGES:
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    declared_count = int(normalized_count)
    if declared_count > _MAX_VALIDATED_MERGE_RANGES:
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    if any(child.tag != _Q("mergeCell") or set(child.attrib) != {"ref"} for child in container):
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    references = tuple(child.attrib["ref"] for child in container)
    if declared_count != len(references):
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    return validated_merge_ranges(references)


def validated_merge_ranges(references: tuple[str, ...]) -> tuple[str, ...]:
    """Fail closed on non-canonical, duplicate, or overlapping merge evidence."""

    ranges: list[tuple[str, tuple[int, int, int, int]]] = []
    seen: set[str] = set()
    for reference in references:
        if len(ranges) >= _MAX_VALIDATED_MERGE_RANGES:
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        if not isinstance(reference, str):
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        try:
            left, top, right, bottom = range_boundaries(reference)
        except ValueError as error:
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID") from error
        canonical = f"{get_column_letter(left)}{top}:{get_column_letter(right)}{bottom}"
        if (
            reference != canonical
            or (left == right and top == bottom)
            or not 1 <= left <= right <= _MAX_SUFFIX_COLUMN
            or not 1 <= top <= bottom <= _MAX_SUFFIX_ROW
            or reference in seen
        ):
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        seen.add(reference)
        ranges.append((reference, (left, top, right, bottom)))
    _reject_overlapping_merge_ranges(bounds for _reference, bounds in ranges)
    return tuple(reference for reference, _bounds in ranges)


def _reject_overlapping_merge_ranges(bounds) -> None:
    events: dict[int, list[tuple[int, int, int]]] = {}
    for left, top, right, bottom in bounds:
        events.setdefault(top, []).append((1, left, right))
        events.setdefault(bottom + 1, []).append((-1, left, right))
    index = _MergeRangeIndex()
    for row in sorted(events):
        for delta, left, right in events[row]:
            if delta < 0:
                index.add(left, right, delta)
        for delta, left, right in events[row]:
            if delta > 0:
                if index.maximum(left, right):
                    raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
                index.add(left, right, delta)


class _MergeRangeIndex:
    """Fixed-column range index, keeping raw merge validation subquadratic."""

    def __init__(self) -> None:
        self._maximum = [0] * (_MAX_SUFFIX_COLUMN * 4)
        self._lazy = [0] * (_MAX_SUFFIX_COLUMN * 4)

    def add(self, left: int, right: int, delta: int) -> None:
        self._update(1, 1, _MAX_SUFFIX_COLUMN, left, right, delta)

    def maximum(self, left: int, right: int) -> int:
        return self._query(1, 1, _MAX_SUFFIX_COLUMN, left, right)

    def _update(self, node: int, start: int, end: int, left: int, right: int, delta: int) -> None:
        if left <= start and end <= right:
            self._maximum[node] += delta
            self._lazy[node] += delta
            return
        middle = (start + end) // 2
        if left <= middle:
            self._update(node * 2, start, middle, left, right, delta)
        if right > middle:
            self._update(node * 2 + 1, middle + 1, end, left, right, delta)
        self._maximum[node] = self._lazy[node] + max(
            self._maximum[node * 2], self._maximum[node * 2 + 1]
        )

    def _query(self, node: int, start: int, end: int, left: int, right: int) -> int:
        if left <= start and end <= right:
            return self._maximum[node]
        middle = (start + end) // 2
        result = 0
        if left <= middle:
            result = self._query(node * 2, start, middle, left, right)
        if right > middle:
            result = max(result, self._query(node * 2 + 1, middle + 1, end, left, right))
        return self._lazy[node] + result


@dataclass(frozen=True, slots=True)
class BoundedHeaderWindow:
    """Physical columns permitted for one exact header-row window."""

    start_row: int
    end_row: int
    columns: tuple[int, ...]


@dataclass(frozen=True, slots=True)
class TargetMeasurePair:
    """One sheet-local quantity/total-cost pair with its header provenance."""

    sheet_name: str
    quantity_column: int
    cost_column: int
    quantity_header: str
    cost_header: str

    @property
    def quantity_letter(self) -> str:
        return get_column_letter(self.quantity_column)

    @property
    def cost_letter(self) -> str:
        return get_column_letter(self.cost_column)


@dataclass(frozen=True, slots=True)
class HistoricalTargetMeasureEvidence:
    """Immutable structural proof for one documentary insertion anchor."""

    sheet_name: str
    quantity_column: int
    cost_column: int
    parent_span: tuple[int, int, int, int]
    historical_parent_label: str
    quantity_leaf_row: int
    cost_leaf_row: int
    quantity_leaf_label: str
    cost_leaf_label: str
    suffix_nonempty_count: int
    suffix_first_coordinate: str
    suffix_last_coordinate: str
    suffix_rightmost_coordinate: str
    suffix_coordinate_sha256: str

    @property
    def quantity_letter(self) -> str:
        return get_column_letter(self.quantity_column)

    @property
    def cost_letter(self) -> str:
        return get_column_letter(self.cost_column)


@dataclass(frozen=True, slots=True)
class _Label:
    text: str
    key: tuple[int, int, int, int]


def discover_target_measures(
    workbook,
    first_detail_rows: dict[str, int],
    merged_ranges_by_sheet: dict[str, tuple[str, ...]] | None = None,
    header_windows: dict[str, BoundedHeaderWindow] | None = None,
) -> tuple[TargetMeasurePair, ...]:
    """Discover exactly one current-period pair for every selected-stage sheet.

    The detector intentionally considers only the header immediately above a
    sheet's first semantic detail row.  It never uses a physical column as a
    fallback: an absent or competing structure is a controlled technical
    condition.
    """

    pairs: list[TargetMeasurePair] = []
    for sheet_name, first_detail_row in sorted(first_detail_rows.items()):
        window = _bound_header_window(
            workbook[sheet_name],
            first_detail_row,
            (merged_ranges_by_sheet or {}).get(sheet_name, ()),
            (header_windows or {}).get(sheet_name),
        )
        candidates = _sheet_candidates(
            workbook[sheet_name],
            first_detail_row,
            (merged_ranges_by_sheet or {}).get(sheet_name, ()),
            window,
        )
        if not candidates:
            raise ReconciliationTargetMeasureError("TARGET_CURRENT_PERIOD_PAIR_MISSING")
        if len(candidates) != 1:
            raise ReconciliationTargetMeasureError("TARGET_CURRENT_PERIOD_PAIR_AMBIGUOUS")
        pairs.append(candidates[0])
    if not pairs:
        raise ReconciliationTargetMeasureError("TARGET_CURRENT_PERIOD_PAIR_MISSING")
    return tuple(pairs)


def discover_historical_target_measures(
    workbook,
    first_detail_rows: dict[str, int],
    merged_ranges_by_sheet: dict[str, tuple[str, ...]] | None = None,
    header_windows: dict[str, BoundedHeaderWindow] | None = None,
) -> tuple[HistoricalTargetMeasureEvidence, ...]:
    """Return one documentary/historical adjacent measure pair per selected sheet.

    This is deliberately separate from the current-period reader: a pair is an
    insertion anchor only when its header path positively says it is historical.
    """

    pairs: list[HistoricalTargetMeasureEvidence] = []
    for sheet_name, first_detail_row in sorted(first_detail_rows.items()):
        window = _bound_header_window(
            workbook[sheet_name],
            first_detail_row,
            (merged_ranges_by_sheet or {}).get(sheet_name, ()),
            (header_windows or {}).get(sheet_name),
        )
        candidates = _sheet_historical_candidates(
            workbook[sheet_name],
            first_detail_row,
            (merged_ranges_by_sheet or {}).get(sheet_name, ()),
            window,
        )
        if len(candidates) != 1:
            raise ReconciliationTargetMeasureError("TARGET_HISTORICAL_PAIR_MISSING")
        pairs.append(candidates[0])
    if not pairs:
        raise ReconciliationTargetMeasureError("TARGET_HISTORICAL_PAIR_MISSING")
    return tuple(pairs)


def bounded_header_windows(
    workbook,
    first_detail_rows: dict[str, int],
    merged_ranges_by_sheet: dict[str, tuple[str, ...]] | None = None,
) -> dict[str, BoundedHeaderWindow]:
    """Derive all detector inputs from physical cells in exact header rows."""

    return {
        sheet_name: bounded_header_window(
            workbook[sheet_name],
            first_detail_row,
            (merged_ranges_by_sheet or {}).get(sheet_name, ()),
        )
        for sheet_name, first_detail_row in sorted(first_detail_rows.items())
    }


def bounded_header_window(
    sheet, first_detail_row: int, merged_ranges: tuple[str, ...] = ()
) -> BoundedHeaderWindow:
    """Return only physical or merged columns above the first detail row."""

    if not isinstance(first_detail_row, int) or not 2 <= first_detail_row <= _MAX_SUFFIX_ROW:
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    end = first_detail_row - 1
    start = max(1, end - _HEADER_ROWS + 1)
    cells = getattr(sheet, "_cells", None)
    if not isinstance(cells, dict):
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    columns: set[int] = set()
    for inspected, (coordinate, _cell) in enumerate(cells.items(), start=1):
        if inspected > _MAX_HEADER_WINDOW_CELLS:
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        if not isinstance(coordinate, tuple) or len(coordinate) != 2:
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        row, column = coordinate
        if not isinstance(row, int) or not isinstance(column, int):
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        if not 1 <= row <= _MAX_SUFFIX_ROW or not 1 <= column <= _MAX_SUFFIX_COLUMN:
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        if start <= row <= end:
            columns.add(column)
    for left, _top, right, _bottom in _header_merged_ranges(sheet, merged_ranges, start, end):
        columns.update(range(left, right + 1))
        if len(columns) > _MAX_HEADER_WINDOW_CELLS:
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    return BoundedHeaderWindow(start, end, tuple(sorted(columns)))


def _bound_header_window(
    sheet,
    first_detail_row: int,
    merged_ranges: tuple[str, ...],
    supplied: BoundedHeaderWindow | None,
) -> BoundedHeaderWindow:
    derived = bounded_header_window(sheet, first_detail_row, merged_ranges)
    if supplied is not None and supplied != derived:
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    _validate_header_work_budget(derived)
    return derived


def _validate_header_work_budget(window: BoundedHeaderWindow) -> None:
    depth = window.end_row - window.start_row + 1
    columns = frozenset(window.columns)
    candidate_count = sum(column + 1 in columns for column in columns)
    work = depth * (len(window.columns) + 2 * candidate_count)
    if depth < 1 or work > _MAX_HEADER_WINDOW_CELLS:
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")


def _header_merged_ranges(sheet, merged_ranges, start, end):
    ranges = tuple(merged_ranges) or getattr(getattr(sheet, "merged_cells", None), "ranges", ())
    result = []
    for merged in ranges:
        try:
            if isinstance(merged, tuple):
                left, top, right, bottom = merged
            else:
                left, top, right, bottom = range_boundaries(str(merged))
        except ValueError as error:
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID") from error
        if not (1 <= left <= right <= _MAX_SUFFIX_COLUMN and 1 <= top <= bottom <= _MAX_SUFFIX_ROW):
            raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
        if top <= end and bottom >= start:
            result.append((left, top, right, bottom))
    return tuple(result)


def _sheet_candidates(
    sheet,
    first_detail_row: int,
    merged_ranges: tuple[str, ...],
    window: BoundedHeaderWindow,
) -> tuple[TargetMeasurePair, ...]:
    start, end = _validated_window(window, first_detail_row)
    _values, paths = _header_paths(sheet, window, merged_ranges)
    columns = frozenset(window.columns)
    candidates: dict[tuple[object, ...], HistoricalTargetMeasureEvidence] = {}
    for row in range(start, end + 1):
        for quantity_column in window.columns:
            cost_column = quantity_column + 1
            if cost_column not in columns:
                continue
            quantity = paths[row, quantity_column]
            cost = paths[row, cost_column]
            if (
                not quantity
                or not cost
                or not _leaf_at_row(quantity, row)
                or not _leaf_at_row(cost, row)
            ):
                continue
            quantity_text = " ".join(label.text for label in quantity)
            cost_text = " ".join(label.text for label in cost)
            if not _quantity_leaf(quantity[-1].text) or not _total_cost_leaf(cost[-1].text):
                continue
            if (
                _unit_price(quantity_text)
                or _unit_price(cost_text)
                or _historical(quantity_text + " " + cost_text)
            ):
                continue
            common = {label.key for label in quantity} & {label.key for label in cost}
            common_labels = tuple(
                label for label in quantity if label.key in common and label.key[1] != label.key[3]
            )
            common_current = any(_current_scope(label.text) for label in common_labels)
            quantity_periods = _period_mentions(quantity_text)
            cost_periods = _period_mentions(cost_text)
            if _period_conflict(quantity_periods, cost_periods):
                continue
            same_period = _same_unambiguous_period(quantity_periods, cost_periods)
            if not common_current and not same_period:
                continue
            key = (
                sheet.title,
                quantity_column,
                cost_column,
                quantity[-1].key,
                cost[-1].key,
            )
            candidates[key] = TargetMeasurePair(
                sheet.title,
                quantity_column,
                cost_column,
                quantity_text,
                cost_text,
            )
    return tuple(candidates.values())


def _sheet_historical_candidates(
    sheet,
    first_detail_row: int,
    merged_ranges: tuple[str, ...],
    window: BoundedHeaderWindow,
) -> tuple[HistoricalTargetMeasureEvidence, ...]:
    start, end = _validated_window(window, first_detail_row)
    _values, paths = _header_paths(sheet, window, merged_ranges)
    columns = frozenset(window.columns)
    candidates: dict[tuple[object, ...], HistoricalTargetMeasureEvidence] = {}
    for row in range(start, end + 1):
        for quantity_column in window.columns:
            cost_column = quantity_column + 1
            if cost_column not in columns:
                continue
            quantity = paths[row, quantity_column]
            cost = paths[row, cost_column]
            if (
                not quantity
                or not cost
                or not _leaf_at_row(quantity, row)
                or not _leaf_at_row(cost, row)
            ):
                continue
            quantity_text = " ".join(label.text for label in quantity)
            cost_text = " ".join(label.text for label in cost)
            if (
                not _quantity_leaf(quantity[-1].text)
                or not _total_cost_leaf(cost[-1].text)
                or _unit_price(quantity_text)
                or _unit_price(cost_text)
            ):
                continue
            common = [label for label in quantity if label.key in {item.key for item in cost}]
            historical_parents = [
                label
                for label in common
                if (
                    label.key[1] != label.key[3]
                    and label.key[1] <= quantity_column
                    and label.key[3] >= cost_column
                    and _historical(label.text)
                )
            ]
            if len(historical_parents) != 1:
                continue
            parent = historical_parents[0]
            suffix = _suffix_evidence(sheet, cost_column)
            if suffix is None:
                continue
            key = (sheet.title, quantity_column, cost_column, parent.key)
            candidates[key] = HistoricalTargetMeasureEvidence(
                sheet.title,
                quantity_column,
                cost_column,
                parent.key,
                parent.text,
                quantity[-1].key[0],
                cost[-1].key[0],
                str(sheet.cell(quantity[-1].key[0], quantity_column).value or ""),
                str(sheet.cell(cost[-1].key[0], cost_column).value or ""),
                *suffix,
            )
    return tuple(candidates.values())


def _suffix_evidence(sheet, cost_column: int) -> tuple[int, str, str, str, str] | None:
    """Return bounded, value-free proof that content follows an insertion point.

    ``Worksheet._cells`` is the materialized-cell index populated by the normal
    planning reader.  Iterating it avoids allocating a ``max_row * max_column``
    rectangle for sparse sheets.  A read-only worksheet deliberately has no
    such index and is not a valid planning input; the caller already opens the
    workbook in normal mode to preserve merged-header provenance.
    """

    cells = getattr(sheet, "_cells", None)
    if not isinstance(cells, dict):
        raise ReconciliationTargetMeasureError("TARGET_HISTORICAL_PAIR_MISSING")
    if len(cells) > _MAX_SUFFIX_INSPECTED_CELLS:
        raise ReconciliationTargetMeasureError("TARGET_HISTORICAL_PAIR_MISSING")
    coordinates: list[tuple[int, int]] = []
    for (row, column), cell in cells.items():
        if column <= cost_column or cell.value is None:
            continue
        if not 1 <= row <= _MAX_SUFFIX_ROW or not 1 <= column <= _MAX_SUFFIX_COLUMN:
            raise ReconciliationTargetMeasureError("TARGET_HISTORICAL_PAIR_MISSING")
        coordinates.append((row, column))
        if len(coordinates) > _MAX_SUFFIX_CELLS:
            raise ReconciliationTargetMeasureError("TARGET_HISTORICAL_PAIR_MISSING")
    if not coordinates:
        return None
    coordinates.sort()
    references = tuple(f"{get_column_letter(column)}{row}" for row, column in coordinates)
    digest = hashlib.sha256("\n".join(references).encode()).hexdigest()
    return (
        len(references),
        references[0],
        references[-1],
        max(references, key=lambda reference: _rightmost_coordinate_key(reference)),
        digest,
    )


def _rightmost_coordinate_key(reference: str) -> tuple[int, int]:
    column, row = coordinate_from_string(reference)
    return column_index_from_string(column), row


def _validated_window(window: BoundedHeaderWindow, first_detail_row: int) -> tuple[int, int]:
    end = first_detail_row - 1
    start = max(1, end - _HEADER_ROWS + 1)
    if (
        window.start_row != start
        or window.end_row != end
        or window.columns != tuple(sorted(window.columns))
        or len(set(window.columns)) != len(window.columns)
    ):
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    if any(
        not isinstance(column, int) or not 1 <= column <= _MAX_SUFFIX_COLUMN
        for column in window.columns
    ):
        raise ReconciliationTargetMeasureError("TARGET_HEADER_WINDOW_INVALID")
    return start, end


def _values_by_column(window: BoundedHeaderWindow) -> frozenset[int]:
    return frozenset(window.columns)


def _header_paths(sheet, window: BoundedHeaderWindow, merged_ranges: tuple[str, ...]):
    values, spans = _header_cells(sheet, window, merged_ranges)
    paths: dict[tuple[int, int], tuple[_Label, ...]] = {}
    for column in window.columns:
        labels: list[_Label] = []
        seen: set[tuple[int, int, int, int]] = set()
        for row in range(window.start_row, window.end_row + 1):
            value = values.get((row, column))
            text = _text(value)
            key = spans.get((row, column), (row, column, row, column))
            if text and key not in seen:
                seen.add(key)
                labels.append(_Label(text, key))
            paths[row, column] = tuple(labels)
    return values, paths


def _header_cells(sheet, window: BoundedHeaderWindow, merged_ranges: tuple[str, ...]):
    start, end = window.start_row, window.end_row
    values = {
        (row, column): sheet.cell(row, column).value
        for row in range(start, end + 1)
        for column in window.columns
    }
    spans: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    allowed = _values_by_column(window)
    for left, top, right, bottom in _header_merged_ranges(sheet, merged_ranges, start, end):
        key = (top, left, bottom, right)
        value = sheet.cell(top, left).value
        for row in range(max(start, top), min(end, bottom) + 1):
            for column in range(left, right + 1):
                if column not in allowed:
                    continue
                values[row, column] = value
                spans[row, column] = key
    return values, spans


def _labels(values, spans, start: int, row: int, column: int) -> tuple[_Label, ...]:
    labels: list[_Label] = []
    seen: set[tuple[int, int, int, int]] = set()
    for header_row in range(start, row + 1):
        value = values.get((header_row, column))
        text = _text(value)
        if not text:
            continue
        key = spans.get((header_row, column), (header_row, column, header_row, column))
        if key not in seen:
            seen.add(key)
            labels.append(_Label(text, key))
    return tuple(labels)


def _leaf_at_row(labels: tuple[_Label, ...], row: int) -> bool:
    return bool(labels) and labels[-1].key[0] == row


def _quantity_leaf(value: str) -> bool:
    return any(stem in value for stem in ("колич", "объем", "объём"))


def _total_cost_leaf(value: str) -> bool:
    return (
        any(stem in value for stem in ("стоим", "сумм", "затрат"))
        or bool(_SCALED_RUB.search(value) and _RUB_CURRENCY.search(value))
    ) and not _unit_price(value)


def _unit_price(value: str) -> bool:
    return any(stem in value for stem in ("цен", "тариф", "расцен", "единиц")) or bool(
        _RUB_CURRENCY.search(value) and _UNIT_DENOMINATOR.search(value)
    )


def _historical(value: str) -> bool:
    tokens = set(value.split())
    return (
        any(stem in value for stem in ("истор", "документ", "накоп", "нараст", "предыдущ", "прошл"))
        or {"весь", "период"}.issubset(tokens)
        or ({"с", "начала"}.issubset(tokens))
        or any(token.startswith("итог") for token in tokens)
    )


def _current_scope(value: str) -> bool:
    return "период" in value and any(stem in value for stem in ("текущ", "отчетн", "отчётн"))


def _period_mentions(value: str) -> frozenset[tuple[int, int | None]]:
    result = {(int(month), int(year)) for year, month in _YEAR_MONTH.findall(value)}
    result.update((int(month), int(year)) for month, year in _MONTH_YEAR.findall(value))
    result.update((int(month), int(year)) for month, year in _DATE.findall(value))
    tokens = re.findall(r"\w+", value, flags=re.UNICODE)
    for index, token in enumerate(tokens):
        month = _RUSSIAN_MONTH_TOKENS.get(token)
        if month is None:
            continue
        next_token = tokens[index + 1] if index + 1 < len(tokens) else ""
        year = int(next_token) if _year(next_token) else None
        result.add((month, year))
    return frozenset(result)


def calendar_identities(value: str) -> frozenset[tuple[int, int | None]]:
    """Authoritative broad calendar evidence used by target and insertion planning."""

    return _period_mentions(_text(value))


def _same_unambiguous_period(
    quantity_periods: frozenset[tuple[int, int | None]],
    cost_periods: frozenset[tuple[int, int | None]],
) -> bool:
    return len(quantity_periods) == len(cost_periods) == 1 and quantity_periods == cost_periods


def _period_conflict(
    quantity_periods: frozenset[tuple[int, int | None]],
    cost_periods: frozenset[tuple[int, int | None]],
) -> bool:
    """Reject multiple or contradictory calendar evidence before parent fallback."""

    if len(quantity_periods) > 1 or len(cost_periods) > 1:
        return True
    return bool(quantity_periods and cost_periods and quantity_periods != cost_periods)


def _year(value: str) -> bool:
    return bool(re.fullmatch(r"(?:19|20)\d{2}", value))


def _text(value: object | None) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or "")).replace("\u00a0", " ")
    return " ".join(re.sub(r"[^\w./-]+", " ", normalized, flags=re.UNICODE).casefold().split())

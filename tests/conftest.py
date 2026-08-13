from __future__ import annotations

import zipfile
from contextlib import contextmanager
from dataclasses import replace
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter

from report_processor.domain.models import FileManifest, FileManifestEntry
from report_processor.extraction.models import CanonicalSourceRow, SourceLocation
from report_processor.inventory.file_classifier import classify_file_by_name
from report_processor.inventory.file_manifest import build_manifest_summary
from report_processor.storage import DuckDBStore
from report_processor.training_data import (
    DataQualityStatus,
    FormulaErrorCode,
    TrainingDataRow,
)


@pytest.fixture(scope="session")
def duckdb_over_default_limit_seed(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """Create the expensive 1,001-row boundary database once; consumers copy it."""
    database_path = tmp_path_factory.mktemp("duckdb-boundary") / "rows.duckdb"
    template = CanonicalSourceRow(
        row_id="row-0000",
        source_type="ks2",
        source_location=SourceLocation("file", "source.xlsx", "КС-2", "ks2", 4),
        document_index="1006",
        document_period="2026-07",
        object_code_raw="1006",
        object_name_raw="Объект",
        subobject_code_raw=None,
        subobject_name_raw=None,
        position_code_raw="0004",
        work_name_raw="Монтаж",
        unit_raw="м",
        contract_quantity=Decimal("123.45"),
        current_period_quantity=Decimal("12.3"),
        cumulative_quantity=None,
        remaining_quantity=None,
        unit_price=None,
        contract_cost=None,
        current_period_cost=None,
        cumulative_cost=None,
        total_cost=None,
        basis_code_raw=None,
        drawing_code_raw=None,
        cost_type_code_raw=None,
        source_values=(),
        status="OK",
        warnings=(),
    )
    rows = (replace(template, row_id=f"row-{index:04d}") for index in range(1_001))
    with DuckDBStore(database_path) as store:
        store.write_rows(rows)
    return database_path


@pytest.fixture
def make_training_row():
    def factory(
        *,
        source_file_id: str = "source-a",
        source_row_id: str = "source-a:17",
        object_code: str = "0007",
        work_name: str = "Монтаж трубопровда",
        unit: str = "пог. м",
        period_cost: Decimal = Decimal("1234.50"),
        warnings: tuple[str, ...] = ("SOURCE_WARNING",),
    ) -> TrainingDataRow:
        return TrainingDataRow(
            document_type="ks2",
            document_period="2026-06",
            source_file_id=source_file_id,
            source_filename="КС-2 № 01.xlsx",
            source_sheet="Лист １",
            source_row=17,
            source_row_id=source_row_id,
            object_code=object_code,
            subobject_code="0003",
            position_code="000042",
            cost_type_code="СМР",
            drawing_code="Ч-007",
            basis_code="ГЭСН 01-01",
            work_name_raw=work_name,
            work_name_normalized=work_name.casefold(),
            unit_raw=unit,
            unit_normalized=unit.casefold(),
            contract_quantity=Decimal("001.250"),
            period_quantity=Decimal("2.50"),
            cumulative_quantity=None,
            remaining_quantity=None,
            unit_price=Decimal("10.00"),
            contract_cost=None,
            period_cost=period_cost,
            cumulative_cost=None,
            total_cost=Decimal("1234.50"),
            is_detail=True,
            is_total=False,
            is_outdated=False,
            formula_error=FormulaErrorCode.NONE,
            data_quality_status=DataQualityStatus.WARNING,
            line_id="block-7-source-dependent-id",
            warnings=warnings,
        )

    return factory


@pytest.fixture
def workbook_session_factory(tmp_path: Path):
    @contextmanager
    def factory(
        rows_by_sheet: dict[str, list[list[object]]],
        *,
        filename: str = "source.xlsx",
        formulas: dict[tuple[str, str], str] | None = None,
    ):
        from report_processor.excel import DualWorkbookSession, ExcelFormatResult
        from report_processor.materialization.models import MaterializedSource

        path = tmp_path / filename
        workbook = Workbook()
        for index, (sheet_name, rows) in enumerate(rows_by_sheet.items()):
            sheet = workbook.active if index == 0 else workbook.create_sheet()
            sheet.title = sheet_name
            for row in rows:
                sheet.append(row)
        for (sheet_name, coordinate), formula in (formulas or {}).items():
            workbook[sheet_name][coordinate] = formula
        workbook.save(path)
        workbook.close()
        formula_book = load_workbook(path, read_only=True, data_only=False)
        values_book = load_workbook(path, read_only=True, data_only=True)
        source = MaterializedSource(
            local_path=path,
            original_file_id="file-001",
            original_relative_path=filename,
            source_kind="file",
            archive_path=None,
            was_extracted=False,
            temporary=False,
            size_bytes=path.stat().st_size,
            extension=".xlsx",
            cleanup_required=False,
            warnings=(),
        )
        session = DualWorkbookSession(
            formula_book,
            values_book,
            source,
            ExcelFormatResult(".xlsx", "xlsx", True, False),
            False,
        )
        try:
            yield session, path
        finally:
            session.close()

    return factory


@pytest.fixture
def schema_factory():
    def factory(
        sheet_name,
        sheet_type,
        logical_columns,
        *,
        data_start_row=2,
        headers=None,
        physical_columns=None,
    ):
        from report_processor.schema import ColumnResolution, WorksheetSchema

        headers = headers or [item.value for item in logical_columns]
        physical_columns = physical_columns or list(range(1, len(logical_columns) + 1))
        columns = tuple(
            ColumnResolution(
                logical,
                physical,
                get_column_letter(physical),
                header,
                1.0,
                "fixture",
                (),
                "OK",
            )
            for logical, physical, header in zip(
                logical_columns,
                physical_columns,
                headers,
                strict=True,
            )
        )
        return WorksheetSchema(
            sheet_name=sheet_name,
            sheet_type=sheet_type,
            classification=None,
            header_start_row=1,
            header_end_row=1,
            data_start_row=data_start_row,
            first_table_column=None,
            last_table_column=None,
            headers=(),
            columns=columns,
            confidence=1.0,
            status="OK",
        )

    return factory


@pytest.fixture
def make_entry():
    def factory(
        filename: str,
        *,
        file_id: str | None = None,
        relative_path: str | None = None,
        document_type: str | None = None,
        extension: str | None = None,
        is_archive_entry: bool = False,
        source_type: str | None = None,
        modified_at: datetime | None = None,
    ) -> FileManifestEntry:
        classification = classify_file_by_name(filename)
        return FileManifestEntry(
            file_id=file_id or filename,
            source_type=source_type or ("archive_entry" if is_archive_entry else "file"),
            source_root="/redacted-fixture/source",
            relative_path=relative_path or filename,
            filename=filename,
            extension=extension or ("." + filename.rsplit(".", 1)[-1].lower()),
            size_bytes=None,
            compressed_size_bytes=None,
            modified_at=modified_at or datetime(2026, 7, 1, tzinfo=UTC),
            crc32=None,
            is_archive_entry=is_archive_entry,
            archive_path="/redacted-fixture/archive.zip" if is_archive_entry else None,
            document_type=document_type or classification.document_type,
            document_markers=classification.document_markers,
            is_temporary=classification.is_temporary,
            is_probable_copy=classification.is_probable_copy,
            is_probably_outdated=classification.is_probably_outdated,
            status="OK",
        )

    return factory


@pytest.fixture
def make_manifest():
    def factory(entries: list[FileManifestEntry]) -> FileManifest:
        return FileManifest(
            source_path="/redacted-fixture/source",
            source_kind="directory",
            created_at=datetime(2026, 7, 1, tzinfo=UTC),
            entries=entries,
            summary=build_manifest_summary(entries, "directory"),
        )

    return factory


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "sample.xlsx"
    workbook = Workbook()
    data = workbook.active
    data.title = "Данные"
    data["A1"] = 10
    data["A2"] = 20
    data["A3"] = "=SUM(A1:A2)"
    data["B1"] = "т"
    data["D1"] = "#NAME?"
    data["D1"].data_type = "s"
    data["E1"] = "#DIV/0!"
    data["E1"].data_type = "e"
    hidden = workbook.create_sheet("Скрытый")
    hidden.sheet_state = "hidden"
    very_hidden = workbook.create_sheet("Очень скрытый")
    very_hidden.sheet_state = "veryHidden"
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def schema_workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "1006 (682)_КС-6а.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "КС-6а"
    worksheet["A5"] = "№ п/п"
    worksheet["B5"] = "Наименование этапа выполнения работ"
    worksheet["C5"] = "Ед. изм."
    worksheet.merge_cells("D5:E5")
    worksheet["D5"] = "Выполнено за отчетный период"
    worksheet["D6"] = "Количество"
    worksheet["E6"] = "Стоимость"
    worksheet["A7"] = 1
    worksheet["B7"] = "Монтаж трубопровода"
    worksheet["C7"] = "м"
    worksheet["D7"] = 12.5
    worksheet["E7"] = 1000
    workbook.save(path)
    workbook.close()
    return path


def regular_entry(path: Path, **overrides: object) -> FileManifestEntry:
    entry = make_entry.__wrapped__()(
        path.name, relative_path=path.name, extension=path.suffix.lower()
    )
    entry.source_root = str(path)
    entry.size_bytes = path.stat().st_size if path.exists() else None
    for key, value in overrides.items():
        setattr(entry, key, value)
    return entry


def zip_entry(
    archive: Path, inner_path: str, size: int, crc: int, **overrides: object
) -> FileManifestEntry:
    entry = make_entry.__wrapped__()(
        Path(inner_path).name,
        relative_path=inner_path,
        extension=Path(inner_path).suffix.lower(),
        is_archive_entry=True,
    )
    entry.source_root = str(archive)
    entry.archive_path = str(archive)
    entry.size_bytes = size
    entry.crc32 = crc
    for key, value in overrides.items():
        setattr(entry, key, value)
    return entry


def candidate(entry: FileManifestEntry):
    from report_processor.selection.models import SourceCandidate

    return SourceCandidate(
        file_id=entry.file_id,
        entry=entry,
        score=0,
        rank=None,
        accepted=True,
        rejection_reasons=(),
        score_components=(),
        warnings=(),
    )


def schema_candidate(path: Path):
    return candidate(regular_entry(path))


def create_zip_with_workbook(
    archive: Path, workbook: Path, inner: str = "nested/sample.xlsx"
) -> FileManifestEntry:
    with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED) as output:
        output.write(workbook, inner)
        output.writestr("other/ignore.txt", "do not extract")
    with zipfile.ZipFile(archive) as source:
        info = source.getinfo(inner)
    return zip_entry(archive, inner, info.file_size, info.CRC)

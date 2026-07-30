from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from report_processor.cli import main
from report_processor.inventory.file_manifest import build_file_manifest, save_manifest_json


def _write_manifest(workbook_path: Path) -> tuple[Path, str]:
    manifest = build_file_manifest(workbook_path)
    path = workbook_path.with_name("manifest.json")
    save_manifest_json(manifest, path)
    return path, manifest.entries[0].file_id


def test_cli_extract_rows_jsonl(tmp_path: Path, capsys):
    workbook_path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "КС-2"
    sheet.append(["Наименование", "Количество", "Стоимость"])
    sheet.append(["Работа", "2,5", "100,50"])
    workbook.save(workbook_path)
    workbook.close()

    manifest_path, file_id = _write_manifest(workbook_path)
    schema = {
        "source_file_id": file_id,
        "filename": workbook_path.name,
        "sheets": [
            {
                "sheet_name": "КС-2",
                "sheet_type": "ks2",
                "header_start_row": 1,
                "header_end_row": 1,
                "data_start_row": 2,
                "columns": [
                    {
                        "logical_column": "work_name",
                        "physical_column": 1,
                        "column_letter": "A",
                        "header_text": "Наименование",
                    },
                    {
                        "logical_column": "current_period_quantity",
                        "physical_column": 2,
                        "column_letter": "B",
                        "header_text": "Количество",
                    },
                    {
                        "logical_column": "current_period_cost",
                        "physical_column": 3,
                        "column_letter": "C",
                        "header_text": "Стоимость",
                    },
                ],
            }
        ],
    }
    schema_path = tmp_path / "schema.json"
    schema_path.write_text(json.dumps(schema, ensure_ascii=False), encoding="utf-8")
    output = tmp_path / "extracted_rows.jsonl"

    exit_code = main(
        [
            "extract-rows",
            "--manifest",
            str(manifest_path),
            "--file-id",
            file_id,
            "--schema",
            str(schema_path),
            "--output",
            str(output),
        ]
    )
    captured = capsys.readouterr()
    assert exit_code == 0
    assert output.exists()
    assert output.with_suffix(".meta.json").exists()
    row = json.loads(output.read_text(encoding="utf-8").splitlines()[0])
    assert row["work_name_raw"] == "Работа"
    assert row["current_period_quantity"] == "2.5"
    assert row["document_index"] is None
    assert row["document_period"] is None
    assert "Извлечено: 1" in captured.out


def test_cli_requires_file_id_with_manifest(tmp_path: Path, capsys):
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text('{"entries": []}', encoding="utf-8")
    schema_path = tmp_path / "schema.json"
    schema_path.write_text('{"sheets": []}', encoding="utf-8")
    exit_code = main(
        [
            "extract-rows",
            "--manifest",
            str(manifest_path),
            "--schema",
            str(schema_path),
            "--output",
            str(tmp_path / "rows.jsonl"),
        ]
    )
    assert exit_code == 9
    assert capsys.readouterr().out == ""


def test_cli_rejects_schema_for_another_file(tmp_path: Path, capsys):
    workbook_path = tmp_path / "source.xlsx"
    workbook = Workbook()
    workbook.save(workbook_path)
    workbook.close()
    manifest_path, file_id = _write_manifest(workbook_path)
    schema_path = tmp_path / "schema.json"
    schema_path.write_text(
        json.dumps({"source_file_id": "other", "sheets": []}),
        encoding="utf-8",
    )
    output = tmp_path / "rows.jsonl"
    exit_code = main(
        [
            "extract-rows",
            "--manifest",
            str(manifest_path),
            "--file-id",
            file_id,
            "--schema",
            str(schema_path),
            "--output",
            str(output),
        ]
    )
    assert exit_code == 9
    assert capsys.readouterr().out == ""
    assert not output.exists()


@pytest.mark.parametrize(
    "schema",
    [
        {"sheets": [], "warnings": None},
        {
            "sheets": [
                {
                    "sheet_name": "КС-2",
                    "sheet_type": "ks2",
                    "columns": [
                        {
                            "logical_column": "unit",
                            "column_index": None,
                            "status": "COLUMN_NOT_FOUND",
                            "confidence": 10**400,
                        }
                    ],
                }
            ]
        },
    ],
)
def test_cli_controls_malformed_schema_error_without_output(tmp_path: Path, capsys, schema: dict):
    workbook_path = tmp_path / "source.xlsx"
    Workbook().save(workbook_path)
    manifest_path, file_id = _write_manifest(workbook_path)
    schema_path = tmp_path / "schema.json"
    schema_path.write_text(json.dumps(schema), encoding="utf-8")
    output = tmp_path / "rows.jsonl"

    exit_code = main(
        [
            "extract-rows",
            "--manifest",
            str(manifest_path),
            "--file-id",
            file_id,
            "--schema",
            str(schema_path),
            "--output",
            str(output),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 9
    assert captured.out == ""
    assert "Ошибка схемы" in captured.err
    assert "Traceback" not in captured.err
    assert not output.exists()

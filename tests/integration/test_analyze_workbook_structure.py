from pathlib import Path

from openpyxl import Workbook

from conftest import schema_candidate
from report_processor.schema import (
    ColumnOverride,
    LogicalColumn,
    SheetType,
    WorksheetSchemaOverride,
    analyze_workbook_schema,
)
from report_processor.workflow import prepared_workbook_session


def test_analyze_multiline_merged_structure(schema_workbook_path: Path) -> None:
    with prepared_workbook_session(schema_candidate(schema_workbook_path)) as session:
        schema = analyze_workbook_schema(session)
    worksheet = schema.worksheets[0]
    resolved = {
        item.logical_column: item.column_letter for item in worksheet.columns if item.status == "OK"
    }
    assert worksheet.sheet_type == SheetType.KS6A
    assert (worksheet.header_start_row, worksheet.header_end_row) == (5, 6)
    assert worksheet.data_start_row == 7
    assert resolved[LogicalColumn.WORK_NAME] == "B"
    assert resolved[LogicalColumn.CURRENT_PERIOD_QUANTITY] == "D"
    assert resolved[LogicalColumn.CURRENT_PERIOD_COST] == "E"
    assert worksheet.status == "OK"


def test_unknown_sheet_has_explicit_status(tmp_path: Path) -> None:
    path = tmp_path / "unknown.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист1"
    worksheet["A1"] = "неизвестное содержимое"
    workbook.save(path)
    workbook.close()
    with prepared_workbook_session(schema_candidate(path)) as session:
        schema = analyze_workbook_schema(session)
    assert schema.worksheets[0].sheet_type == SheetType.UNKNOWN
    assert schema.worksheets[0].status == "UNKNOWN_SHEET_TYPE"


def test_duplicate_sheet_type_has_no_primary(tmp_path: Path) -> None:
    path = tmp_path / "duplicates.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "КС-6а"
    second = workbook.create_sheet("КС-6а корректировка")
    for worksheet in (first, second):
        worksheet["A1"] = "Наименование этапа выполнения работ"
        worksheet["B1"] = "Ед. изм."
        worksheet["C1"] = "Количество по проекту"
        worksheet["A2"] = "Работа"
        worksheet["B2"] = "м"
        worksheet["C2"] = 1
    workbook.save(path)
    workbook.close()
    with prepared_workbook_session(schema_candidate(path)) as session:
        schema = analyze_workbook_schema(session)
    assert schema.primary_sheets["ks6a"] is None
    assert "MULTIPLE_SHEETS_OF_SAME_TYPE:ks6a" in schema.warnings


def test_manual_override_is_explicit(tmp_path: Path) -> None:
    path = tmp_path / "manual.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист1"
    worksheet["A3"] = "Работа"
    worksheet["B3"] = "Единица"
    worksheet["C3"] = "Значение"
    worksheet["A4"] = "Монтаж"
    worksheet["B4"] = "м"
    worksheet["C4"] = 2
    workbook.save(path)
    workbook.close()
    override = WorksheetSchemaOverride(
        sheet_name="Лист1",
        sheet_type=SheetType.ADDITIONAL_REPORT,
        header_start_row=3,
        header_end_row=3,
        column_overrides=(ColumnOverride(LogicalColumn.WORK_NAME, "A"),),
    )
    with prepared_workbook_session(schema_candidate(path)) as session:
        schema = analyze_workbook_schema(session, overrides=(override,))
    worksheet_schema = schema.worksheets[0]
    manual = next(
        item for item in worksheet_schema.columns if item.logical_column == LogicalColumn.WORK_NAME
    )
    assert manual.is_manual is True
    assert manual.column_letter == "A"
    assert "MANUAL_OVERRIDE_APPLIED" in worksheet_schema.warnings


def test_twenty_sheets_use_bounded_scan(tmp_path: Path) -> None:
    path = tmp_path / "twenty.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index in range(20):
        worksheet = workbook.create_sheet(f"КС-2 {index + 1}")
        worksheet["A5"] = "Наименование работ"
        worksheet["B5"] = "Ед. изм."
        worksheet["C5"] = "Цена за единицу"
        worksheet["A6"] = "Работа"
        worksheet["B6"] = "шт"
        worksheet["C6"] = 10
        worksheet["A5000"] = "форматированная дальняя строка"
    workbook.save(path)
    workbook.close()
    with prepared_workbook_session(schema_candidate(path)) as session:
        schema = analyze_workbook_schema(session)
    assert len(schema.worksheets) == 20
    assert all(item.header_start_row == 5 for item in schema.worksheets)

"""Real XLSX safety evidence for ExcelWriterEngine-15.0."""

from __future__ import annotations

import os
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from report_processor.excel_writer import (
    ExcelWriterAtomicError,
    ExcelWriterSafetyError,
    WriteStatus,
    write_target_report,
)

from report_processor.business_rules import load_default_rule_set
from report_processor.calculation import calculate_matches
from report_processor.matching import match_rows
from report_processor.quality_control import WriteDecision, evaluate_quality_control


def _block12_helpers():
    try:
        from test_block12_real_data import _fingerprint, _normalized_source_rows, _target_rows
    except ImportError as exc:  # pragma: no cover - pytest exposes sibling modules
        raise RuntimeError("Block 12 real-data helpers are unavailable") from exc
    return _fingerprint, _normalized_source_rows, _target_rows


def _real_inputs() -> tuple[Path, Path]:
    source_value = os.environ.get("DOCUMENT_OPTIMIZER_REAL_SOURCE_XLSX")
    target_value = os.environ.get("DOCUMENT_OPTIMIZER_REAL_TARGET_XLSX")
    if not source_value or not target_value:
        pytest.skip("real XLSX paths are not set")
    return Path(source_value), Path(target_value)


def _real_pipeline(target_path: Path):
    _, normalized_source_rows, target_rows = _block12_helpers()
    source_path = _real_inputs()[0]
    target_result = target_rows(target_path)
    validation = load_default_rule_set()
    assert validation.valid and validation.rule_set is not None
    context = {
        "target_source_id": target_result.schema.source_file_id,
        "target_fingerprint": target_result.schema.source_fingerprint.value,
    }
    matches = match_rows(
        normalized_source_rows(source_path), target_result.rows, validation.rule_set, **context
    )
    return (
        target_result,
        matches,
        calculate_matches(matches, validation.rule_set),
        validation.rule_set,
    )


def _package_entries(path: Path) -> tuple[str, ...]:
    with ZipFile(path) as package:
        return tuple(sorted(package.namelist()))


def _matched_subset(target_path: Path):
    target_result, matches, calculations, rules = _real_pipeline(target_path)
    subset_matches = tuple(match for match in matches if match.selected_candidate is not None)
    assert len(subset_matches) == 1
    matched = subset_matches[0]
    subset_calculations = tuple(
        calculation
        for calculation in calculations
        if calculation.match_result_id == matched.result_id
        and calculation.target_row_id == matched.target_row_id
    )
    assert len(subset_calculations) == 1
    assert evaluate_quality_control(subset_matches, subset_calculations, rules).decision is (
        WriteDecision.ALLOW_WRITE
    )
    return target_result, subset_calculations


def test_real_full_manual_review_skips_without_creating_output(tmp_path: Path) -> None:
    source_path, target_path = _real_inputs()
    fingerprint, _, _ = _block12_helpers()
    before = (fingerprint(source_path), fingerprint(target_path))
    target_result, matches, calculations, rules = _real_pipeline(target_path)
    decision = evaluate_quality_control(matches, calculations, rules).decision
    assert decision is WriteDecision.REQUIRE_MANUAL_REVIEW

    output_path = tmp_path / "manual-review.xlsx"
    result = write_target_report(
        target_path, output_path, decision, calculations, target_result.schema
    )

    assert result.status is WriteStatus.SKIPPED_DECISION
    assert result.output_path is None
    assert not output_path.exists()
    assert (fingerprint(source_path), fingerprint(target_path)) == before


def test_real_matched_subset_writes_only_d30_to_a_temporary_output(tmp_path: Path) -> None:
    source_path, target_path = _real_inputs()
    fingerprint, _, _ = _block12_helpers()
    before = (fingerprint(source_path), fingerprint(target_path))
    target_result, subset_calculations = _matched_subset(target_path)

    output_path = tmp_path / "matched-subset.xlsx"
    result = write_target_report(
        target_path,
        output_path,
        WriteDecision.ALLOW_WRITE,
        subset_calculations,
        target_result.schema,
    )

    assert result.status is WriteStatus.WRITTEN
    assert output_path.exists()
    assert tuple(
        (cell.sheet_name, cell.coordinate, cell.decimal_text) for cell in result.written_cells
    ) == (("Лист", "D30", "0"),)
    assert _package_entries(output_path) == _package_entries(target_path)
    with load_workbook(output_path, data_only=False, read_only=False) as workbook:
        sheet = workbook["Лист"]
        assert sheet["D30"].value == 0
        assert sum(cell.data_type == "f" for row in sheet.iter_rows() for cell in row) == 14
        assert len(sheet.merged_cells.ranges) == 128
    assert (fingerprint(source_path), fingerprint(target_path)) == before


@pytest.mark.parametrize(
    "seam",
    (
        "_write_temp_package",
        "_verify_temp_package",
        "_publish_no_clobber",
        "_reopen_published_output",
    ),
)
def test_real_failure_injection_never_publishes_or_mutates_inputs(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, seam: str
) -> None:
    import report_processor.excel_writer.engine as engine

    source_path, target_path = _real_inputs()
    fingerprint, _, _ = _block12_helpers()
    before = (fingerprint(source_path), fingerprint(target_path))
    target_result, calculations = _matched_subset(target_path)
    output_path = tmp_path / f"{seam}.xlsx"

    def fail(*_args: object, **_kwargs: object) -> None:
        raise ExcelWriterAtomicError("ATOMIC_PUBLISH_FAILED", seam)

    monkeypatch.setattr(engine, seam, fail)
    with pytest.raises(ExcelWriterAtomicError, match="ATOMIC_PUBLISH_FAILED"):
        write_target_report(
            target_path,
            output_path,
            WriteDecision.ALLOW_WRITE,
            calculations,
            target_result.schema,
        )

    assert not output_path.exists()
    assert (fingerprint(source_path), fingerprint(target_path)) == before


def test_real_existing_output_is_never_clobbered(tmp_path: Path) -> None:
    _, target_path = _real_inputs()
    target_result, calculations = _matched_subset(target_path)
    output_path = tmp_path / "existing.xlsx"
    output_path.write_bytes(b"must-not-change")

    with pytest.raises(ExcelWriterSafetyError, match="OUTPUT_EXISTS"):
        write_target_report(
            target_path,
            output_path,
            WriteDecision.ALLOW_WRITE,
            calculations,
            target_result.schema,
        )

    assert output_path.read_bytes() == b"must-not-change"

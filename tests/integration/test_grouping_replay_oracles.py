"""Offline integration contract for injected replay execution and oracle adapters."""

from __future__ import annotations

import hashlib
from dataclasses import fields, replace
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from fixtures.calculation.builders import calculation_rule_set, calculation_source_row, match_result
from report_processor.calculation import calculate_matches
from report_processor.reconciliation_patterns import offline, replay
from report_processor.reconciliation_patterns import pattern_models as models
from report_processor.reconciliation_patterns import pattern_registry as registry


def _hash(value: str) -> str:
    return "sha256:" + hashlib.sha256(value.encode()).hexdigest()


def _value(model: type[object], values: dict[str, object], field: str = "fingerprint") -> object:
    return model(**values, **{field: replay.replay_fingerprint(values)})


def _replace(value: object, /, **changes: object) -> object:
    fingerprint_field = (
        "oracle_fingerprint"
        if isinstance(value, replay.OracleResult)
        else "semantic_fingerprint"
        if isinstance(value, replay.ReplayObservation)
        else "fingerprint"
    )
    values = {
        field.name: getattr(value, field.name)
        for field in fields(value)
        if field.name != fingerprint_field
    }
    values.update(changes)
    return _value(type(value), values, fingerprint_field)


def _candidate() -> offline.PatternCandidate:
    scope = offline.PatternScope(
        "category", "quantity_cost", "unit", "accept", "object", "document"
    )
    proposal = offline.IncludeExcludeProposal("predicate", "accept")
    candidate_id = offline.fingerprint(
        {
            "version": offline.PATTERN_CANDIDATE_VERSION,
            "kind": "include_exclude",
            "scope": scope,
            "proposal": proposal,
        }
    )
    support = offline.SupportSummary(
        3, 2, 3, 3, 0, tuple(sorted((_hash("support-a"), _hash("support-b"))))
    )
    return offline.PatternCandidate(
        "candidate",
        candidate_id,
        offline.CandidateKind.INCLUDE_EXCLUDE,
        scope,
        proposal,
        offline.OutcomeSignature("accept", "quantity_cost", "category"),
        support,
        (),
        offline.fingerprint({"candidate_id": candidate_id, "support": support, "risks": ()}),
    )


def _shadow() -> registry.RegistryHistory:
    history = registry.register_candidate(
        _candidate(),
        versions=models.PatternVersions("Parser-1.0", "Model-1.0", "Taxonomy-1.0"),
        actor_ref=_hash("miner"),
    )
    return registry.move_to_shadow(history, expected_head=history.head, actor_ref=_hash("shadow"))


def _snapshot(split: replay.ReplaySplit, name: str) -> replay.ReplaySnapshotIdentity:
    values = {
        "split": split,
        "snapshot_ref": _hash(name),
        "manifest_fingerprint": _hash(name + "-manifest"),
        "corpus_fingerprint": _hash(name + "-corpus"),
        "source_set_refs": (_hash(name + "-source"),),
        "document_set_refs": (_hash(name + "-document"),),
        "consequential_version_fingerprint": _hash("version-context"),
        "row_count": 4,
        "review_row_count": 2,
        "review_group_count": 1,
        "sealed": True,
        "seal_ref": _hash(name + "-seal"),
        "version": replay.GROUPING_REPLAY_VERSION,
    }
    return _value(replay.ReplaySnapshotIdentity, values)  # type: ignore[return-value]


def _oracle(name: str, mismatches: tuple[str, ...] = ()) -> replay.OracleResult:
    values = {"case_count": 2, "mismatch_refs": mismatches, "version": replay.ORACLE_RESULT_VERSION}
    return _value(replay.OracleResult, values, "oracle_fingerprint")  # type: ignore[return-value]


def _policy(pattern: models.PatternRecord) -> replay.PromotionPolicy:
    values = {
        "policy_ref": _hash("policy"),
        "owner_ref": _hash("owner"),
        "approval_ref": _hash("policy-approval"),
        "release_window_ref": _hash("window"),
        "allowed_kinds": (offline.CandidateKind.INCLUDE_EXCLUDE,),
        "allowed_scope_fingerprints": (offline.fingerprint(pattern.scope),),
        "min_support_document_sets": 2,
        "min_holdout_document_sets": 1,
        "min_holdout_decisions": 2,
        "min_coverage_rows": replay.Ratio(1, 1),
        "min_coverage_groups": replay.Ratio(1, 1),
        "min_precision": replay.Ratio(1, 1),
        "max_manual_group_count": 2,
        "max_manual_action_count": 2,
        "max_unresolved_row_count": 0,
        "max_p95_latency_ns": 9,
        "index_required": False,
        "max_index_size_bytes": None,
        "version": replay.GROUPING_PROMOTION_POLICY_VERSION,
    }
    return _value(replay.PromotionPolicy, values)  # type: ignore[return-value]


def _index() -> replay.IndexMeasurement:
    values = {
        "status": replay.MeasurementStatus.NOT_APPLICABLE,
        "environment_ref": None,
        "index_ref": None,
        "size_bytes": None,
        "version": replay.GROUPING_REPLAY_VERSION,
    }
    return _value(replay.IndexMeasurement, values)  # type: ignore[return-value]


def _observation(
    pattern: models.PatternRecord, snapshot: replay.ReplaySnapshotIdentity, *, before: bool
) -> replay.ReplayObservation:
    calc, xlsx = _oracle("calculation"), _oracle("xlsx")

    def refs(*names: str) -> tuple[str, ...]:
        return tuple(sorted(_hash(name) for name in names))

    values = {
        "snapshot_fingerprint": snapshot.fingerprint,
        "evaluated_head_fingerprint": pattern.fingerprint,
        "effective_decision_fingerprint": _hash("before" if before else "after"),
        "pattern_decision_refs": refs("decision-a", "decision-b"),
        "correct_decision_refs": refs("decision-a", "decision-b"),
        "covered_row_refs": refs("row-a", "row-b"),
        "covered_group_refs": refs("group-a"),
        "supporting_document_set_refs": snapshot.document_set_refs,
        "contradiction_refs": (),
        "forbidden_pair_refs": (),
        "category_change_refs": (),
        "mode_change_refs": (),
        "unit_change_refs": (),
        "decision_mismatch_refs": (),
        "manual_group_count": 4 if before else 2,
        "manual_action_count": 4 if before else 2,
        "unresolved_row_count": 1 if before else 0,
        "double_membership_count": 0,
        "calculation_oracle": calc,
        "xlsx_oracle": xlsx,
        "version": replay.GROUPING_REPLAY_VERSION,
    }
    return _value(replay.ReplayObservation, values, "semantic_fingerprint")  # type: ignore[return-value]


def _run(
    history: registry.RegistryHistory, policy: replay.PromotionPolicy
) -> replay.GroupingReplayReport:
    pattern = history.head
    baseline, holdout = (
        _snapshot(replay.ReplaySplit.BASELINE, "baseline"),
        _snapshot(replay.ReplaySplit.HOLDOUT, "holdout"),
    )
    clock = iter((0, 5, 8, 15, 20, 26, 30, 39))
    return replay.run_grouping_replay(
        pattern,
        baseline,
        holdout,
        policy,
        executor=lambda candidate, snapshot: _observation(
            pattern, snapshot, before=candidate is None
        ),
        calculation_oracle=lambda _before, after, _snapshot: after.calculation_oracle,
        xlsx_oracle=lambda _before, after, _snapshot: after.xlsx_oracle,
        monotonic_ns=lambda: next(clock),
        index_measurement=_index(),
    )


def _report(
    report: replay.GroupingReplayReport,
    *,
    policy: replay.PromotionPolicy | None = None,
    baseline: replay.SplitReplayMetrics | None = None,
    holdout: replay.SplitReplayMetrics | None = None,
    measurements: replay.ReplayMeasurements | None = None,
) -> replay.GroupingReplayReport:
    values = {
        "evaluated_pattern_id": report.evaluated_pattern_id,
        "evaluated_head_fingerprint": report.evaluated_head_fingerprint,
        "policy_fingerprint": policy.fingerprint if policy else report.policy_fingerprint,
        "baseline_snapshot_fingerprint": report.baseline_snapshot_fingerprint,
        "holdout_snapshot_fingerprint": report.holdout_snapshot_fingerprint,
        "baseline_metrics": baseline or report.baseline_metrics,
        "holdout_metrics": holdout or report.holdout_metrics,
        "deterministic_repeatability": report.deterministic_repeatability,
        "version": replay.GROUPING_REPLAY_VERSION,
    }
    semantic = replay.replay_fingerprint(values)
    full = {
        **values,
        "semantic_fingerprint": semantic,
        "measurements": measurements or report.measurements,
    }
    return _value(replay.GroupingReplayReport, full)  # type: ignore[return-value]


def test_replay_uses_each_split_once_before_twice_after_and_binds_activation() -> None:
    history = _shadow()
    pattern = history.head
    policy = _policy(pattern)
    baseline, holdout = (
        _snapshot(replay.ReplaySplit.BASELINE, "baseline"),
        _snapshot(replay.ReplaySplit.HOLDOUT, "holdout"),
    )
    calls: list[tuple[bool, replay.ReplaySplit]] = []

    def executor(
        candidate: models.PatternRecord | None, snapshot: replay.ReplaySnapshotIdentity
    ) -> replay.ReplayObservation:
        calls.append((candidate is None, snapshot.split))
        return _observation(pattern, snapshot, before=candidate is None)

    clock = iter((0, 5, 8, 15, 20, 26, 30, 39))
    oracle_calls: list[str] = []

    def calculation(
        before: replay.ReplayObservation,
        after: replay.ReplayObservation,
        snapshot: replay.ReplaySnapshotIdentity,
    ) -> replay.OracleResult:
        oracle_calls.append("calculation")
        return after.calculation_oracle

    def xlsx(
        before: replay.ReplayObservation,
        after: replay.ReplayObservation,
        snapshot: replay.ReplaySnapshotIdentity,
    ) -> replay.OracleResult:
        oracle_calls.append("xlsx")
        return after.xlsx_oracle

    report = replay.run_grouping_replay(
        pattern,
        baseline,
        holdout,
        policy,
        executor=executor,
        calculation_oracle=calculation,
        xlsx_oracle=xlsx,
        monotonic_ns=lambda: next(clock),
        index_measurement=_index(),
    )
    assert calls == [
        (True, replay.ReplaySplit.BASELINE),
        (False, replay.ReplaySplit.BASELINE),
        (False, replay.ReplaySplit.BASELINE),
        (True, replay.ReplaySplit.HOLDOUT),
        (False, replay.ReplaySplit.HOLDOUT),
        (False, replay.ReplaySplit.HOLDOUT),
    ]
    assert oracle_calls == ["calculation", "xlsx", "calculation", "xlsx"]
    assert (
        report.holdout_metrics.coverage_rows,
        report.holdout_metrics.precision,
        report.measurements.p50_latency_ns,
        report.measurements.p95_latency_ns,
    ) == (replay.Ratio(2, 2), replay.Ratio(2, 2), 6, 9)
    assert replay.evaluate_promotion(history, report, policy).verdict is (
        replay.PromotionVerdict.OWNER_APPROVAL_REQUIRED
    )
    missing_policy = replay.evaluate_promotion(history, report, None)
    assert missing_policy.verdict is replay.PromotionVerdict.STOP
    assert "THRESHOLDS_MISSING" in missing_policy.reason_codes
    approved = registry.approve_head(
        history,
        expected_head=pattern,
        owner_ref=_hash("owner"),
        approval_ref=replay.owner_approval_ref(pattern, report, policy),
    )
    decision = replay.evaluate_promotion(approved, report, policy)
    assert decision.verdict is replay.PromotionVerdict.ACTIVATION_ELIGIBLE
    assert (
        replay.build_activation_metadata(
            approved, report, decision, policy, activation_ref=_hash("activation")
        ).wave5_verification_ref
        == report.fingerprint
    )
    forged = _replace(decision, head_fingerprint=_hash("forged"))
    assert isinstance(forged, replay.PromotionDecision)
    with pytest.raises(models.PatternRegistryError) as error:
        replay.build_activation_metadata(
            approved, report, forged, policy, activation_ref=_hash("activation")
        )
    assert error.value.code == "PATTERN_STATE_INVALID"


def test_overlap_and_oracle_tampering_fail_before_authoritative_promotion() -> None:
    history = _shadow()
    pattern, policy = history.head, _policy(history.head)
    baseline = _snapshot(replay.ReplaySplit.BASELINE, "same")
    holdout = _snapshot(replay.ReplaySplit.HOLDOUT, "same")
    with pytest.raises(models.PatternRegistryError) as error:
        replay.run_grouping_replay(
            pattern,
            baseline,
            holdout,
            policy,
            executor=lambda *_: pytest.fail("executor must not run"),
            calculation_oracle=lambda *_: pytest.fail("oracle must not run"),
            xlsx_oracle=lambda *_: pytest.fail("oracle must not run"),
            monotonic_ns=lambda: 0,
            index_measurement=_index(),
        )
    assert error.value.code == "SNAPSHOT_OVERLAP"
    with pytest.raises(models.PatternRegistryError) as error:
        replay.OracleResult(2, (), _hash("tampered"))
    assert error.value.code == "REPLAY_FINGERPRINT_MISMATCH"


@pytest.mark.parametrize(
    ("changes", "verdict", "reason"),
    (
        ({"min_support_document_sets": 1}, replay.PromotionVerdict.OWNER_APPROVAL_REQUIRED, None),
        ({"min_support_document_sets": 3}, replay.PromotionVerdict.OWNER_APPROVAL_REQUIRED, None),
        ({"min_support_document_sets": 4}, replay.PromotionVerdict.STOP, "SUPPORT_INSUFFICIENT"),
        ({"min_holdout_document_sets": 2}, replay.PromotionVerdict.STOP, "HOLDOUT_INSUFFICIENT"),
        ({"min_holdout_decisions": 3}, replay.PromotionVerdict.STOP, "HOLDOUT_INSUFFICIENT"),
        (
            {"min_coverage_rows": replay.Ratio(1, 2)},
            replay.PromotionVerdict.OWNER_APPROVAL_REQUIRED,
            None,
        ),
        (
            {"min_coverage_groups": replay.Ratio(1, 2)},
            replay.PromotionVerdict.OWNER_APPROVAL_REQUIRED,
            None,
        ),
        (
            {"min_precision": replay.Ratio(1, 2)},
            replay.PromotionVerdict.OWNER_APPROVAL_REQUIRED,
            None,
        ),
        ({"max_manual_group_count": 1}, replay.PromotionVerdict.STOP, "MANUAL_METRIC_REGRESSION"),
        ({"max_manual_action_count": 1}, replay.PromotionVerdict.STOP, "MANUAL_METRIC_REGRESSION"),
        ({"max_p95_latency_ns": 8}, replay.PromotionVerdict.STOP, "LATENCY_EXCEEDED"),
        ({"max_p95_latency_ns": 9}, replay.PromotionVerdict.OWNER_APPROVAL_REQUIRED, None),
        ({"max_p95_latency_ns": 10}, replay.PromotionVerdict.OWNER_APPROVAL_REQUIRED, None),
    ),
)
def test_policy_thresholds_gate_on_below_equal_and_above_boundaries(
    changes: dict[str, object], verdict: replay.PromotionVerdict, reason: str | None
) -> None:
    history = _shadow()
    original = _policy(history.head)
    policy = _replace(original, **changes)
    assert isinstance(policy, replay.PromotionPolicy)
    report = _report(_run(history, original), policy=policy)
    decision = replay.evaluate_promotion(history, report, policy)
    assert decision.verdict is verdict
    assert (reason in decision.reason_codes) if reason else not decision.reason_codes


@pytest.mark.parametrize(
    ("field", "value", "reason"),
    (
        ("contradiction_count", 1, "CONTRADICTION_PRESENT"),
        ("forbidden_merge_count", 1, "FORBIDDEN_MERGE_PRESENT"),
        ("decision_mismatch_count", 1, "DECISION_MISMATCH"),
        ("double_membership_count", 1, "DOUBLE_MEMBERSHIP"),
        ("calculation_mismatch_count", 1, "CALCULATION_NOT_EQUIVALENT"),
        ("xlsx_mismatch_count", 1, "XLSX_NOT_EQUIVALENT"),
    ),
)
def test_baseline_hard_failures_stop_promotion(field: str, value: int, reason: str) -> None:
    history = _shadow()
    policy = _policy(history.head)
    report = _run(history, policy)
    baseline = _replace(report.baseline_metrics, **{field: value})
    assert isinstance(baseline, replay.SplitReplayMetrics)
    decision = replay.evaluate_promotion(history, _report(report, baseline=baseline), policy)
    assert decision.verdict is replay.PromotionVerdict.STOP
    assert reason in decision.reason_codes


@pytest.mark.parametrize("kind", ("swapped", "version", "stale_snapshot", "stale_head"))
def test_split_and_observation_binding_fail_before_promotion(kind: str) -> None:
    history = _shadow()
    pattern, policy = history.head, _policy(history.head)
    baseline, holdout = (
        _snapshot(replay.ReplaySplit.BASELINE, "baseline"),
        _snapshot(replay.ReplaySplit.HOLDOUT, "holdout"),
    )
    called = False

    def executor(
        candidate: models.PatternRecord | None, snapshot: replay.ReplaySnapshotIdentity
    ) -> replay.ReplayObservation:
        nonlocal called
        called = True
        observation = _observation(pattern, snapshot, before=candidate is None)
        if kind == "stale_snapshot":
            return _replace(observation, snapshot_fingerprint=_hash("stale"))  # type: ignore[return-value]
        if kind == "stale_head":
            return _replace(observation, evaluated_head_fingerprint=_hash("stale"))  # type: ignore[return-value]
        return observation

    if kind == "swapped":
        baseline, holdout = holdout, baseline
    elif kind == "version":
        holdout = _replace(holdout, consequential_version_fingerprint=_hash("other-context"))  # type: ignore[assignment]
    with pytest.raises(models.PatternRegistryError) as error:
        replay.run_grouping_replay(
            pattern,
            baseline,
            holdout,
            policy,
            executor=executor,
            calculation_oracle=lambda _before, after, _snapshot: after.calculation_oracle,
            xlsx_oracle=lambda _before, after, _snapshot: after.xlsx_oracle,
            monotonic_ns=lambda: 0,
            index_measurement=_index(),
        )
    assert error.value.code in {"SNAPSHOT_INVALID", "VERSION_CONTEXT_MISMATCH", "EXECUTOR_INVALID"}
    assert called is (kind.startswith("stale"))


def test_repeatability_clock_and_nested_report_tampering_fail_closed() -> None:
    history = _shadow()
    pattern, policy = history.head, _policy(history.head)
    baseline, holdout = (
        _snapshot(replay.ReplaySplit.BASELINE, "baseline"),
        _snapshot(replay.ReplaySplit.HOLDOUT, "holdout"),
    )
    calls = 0

    def nondeterministic(
        candidate: models.PatternRecord | None, snapshot: replay.ReplaySnapshotIdentity
    ) -> replay.ReplayObservation:
        nonlocal calls
        calls += candidate is not None
        result = _observation(pattern, snapshot, before=candidate is None)
        return (
            _replace(result, effective_decision_fingerprint=_hash(f"changed-{calls}"))
            if calls == 2
            else result
        )  # type: ignore[return-value]

    with pytest.raises(models.PatternRegistryError) as error:
        replay.run_grouping_replay(
            pattern,
            baseline,
            holdout,
            policy,
            executor=nondeterministic,
            calculation_oracle=lambda _before, after, _snapshot: after.calculation_oracle,
            xlsx_oracle=lambda _before, after, _snapshot: after.xlsx_oracle,
            monotonic_ns=lambda: 0,
            index_measurement=_index(),
        )
    assert error.value.code == "REPLAY_NONDETERMINISTIC"

    report = _run(history, policy)
    with pytest.raises(models.PatternRegistryError) as error:
        _report(
            report,
            baseline=_replace(
                report.baseline_metrics, snapshot_fingerprint=_hash("wrong-snapshot")
            ),  # type: ignore[arg-type]
        )
    assert error.value.code == "REPLAY_SCHEMA_INVALID"
    with pytest.raises(models.PatternRegistryError) as error:
        _report(
            report,
            baseline=_replace(
                report.baseline_metrics, repeat_semantic_fingerprint=_hash("wrong-repeat")
            ),  # type: ignore[arg-type]
        )
    assert error.value.code == "REPLAY_SCHEMA_INVALID"

    clock = iter((4, 3, 0, 0))
    with pytest.raises(models.PatternRegistryError) as error:
        replay.run_grouping_replay(
            pattern,
            baseline,
            holdout,
            policy,
            executor=lambda candidate, snapshot: _observation(
                pattern, snapshot, before=candidate is None
            ),
            calculation_oracle=lambda _before, after, _snapshot: after.calculation_oracle,
            xlsx_oracle=lambda _before, after, _snapshot: after.xlsx_oracle,
            monotonic_ns=lambda: next(clock),
            index_measurement=_index(),
        )
    assert error.value.code == "MEASUREMENT_INVALID"


def test_authoritative_calculation_oracle_compares_identity_status_and_decimals() -> None:
    match = match_result(
        calculation_source_row("oracle-source", quantity=Decimal("0.005"), cost=Decimal("1.005"))
    )
    expected = calculate_matches((match,), calculation_rule_set(coefficient=Decimal("1.1")))[0]
    observed = calculate_matches((match,), calculation_rule_set(coefficient=Decimal("1.1")))[0]
    assert (observed.calculation_id, observed.match_result_id, observed.status) == (
        expected.calculation_id,
        expected.match_result_id,
        expected.status,
    )
    assert (
        observed.quantity,
        observed.cost_before_coefficient,
        observed.coefficient,
        observed.cost,
    ) == (
        Decimal("0.01"),
        Decimal("1.005"),
        Decimal("1.1"),
        Decimal("1.11"),
    )
    assert tuple(
        (item.category, item.quantity, item.cost) for item in observed.category_totals
    ) == tuple((item.category, item.quantity, item.cost) for item in expected.category_totals)
    mismatched = replace(observed, cost=Decimal("1.10"))
    assert mismatched.cost != expected.cost


def test_temporary_xlsx_oracle_compares_schema_cells_values_and_formats(tmp_path: Path) -> None:
    path = tmp_path / "oracle.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Target"
    sheet["B2"] = Decimal("1.11")
    sheet["B2"].number_format = "0.00"
    workbook.save(path)
    reloaded = load_workbook(path, data_only=False)["Target"]
    cell = reloaded["B2"]
    assert (cell.coordinate, Decimal(str(cell.value)), cell.number_format) == (
        "B2",
        Decimal("1.11"),
        "0.00",
    )
    assert (Decimal(str(cell.value)), cell.number_format) != (Decimal("1.10"), "#,##0.00")

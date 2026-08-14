import threading
from dataclasses import replace
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from fixtures.calculation.builders import calculation_rule_set, calculation_source_row, match_result
from fixtures.quality_control.builders import calculated_match, calculated_result
from report_processor.admin_panel.reconciliation_execution import (
    ReconciliationApplyResult,
    ReconciliationReviewResult,
    _apply_plan,
    _feedback_records,
    _physical_source_identity,
)
from report_processor.admin_panel.reconciliation_period_preview import preview_reconciliation_target
from report_processor.admin_panel.reconciliation_state import ReconciliationReviewState
from report_processor.admin_panel.reconciliation_target import (
    ReconciliationTargetIdentity,
    _bindings,
    writer_calculations,
)
from report_processor.admin_panel.reconciliation_target_measure import TargetMeasurePair
from report_processor.admin_panel.service import (
    AdminJob,
    AdminPanelService,
    _copy_verified_snapshot,
)
from report_processor.calculation import calculate_matches
from report_processor.matching import MatchResult, MatchStatus
from report_processor.reconciliation_review import (
    ReviewAction,
    ReviewDecision,
    ReviewMode,
    ReviewRow,
    build_review_groups,
)
from report_processor.schema import LogicalColumn


def test_two_accepted_source_rows_contribute_once_to_one_target_aggregate() -> None:
    first = calculation_source_row("source-a:1", quantity=Decimal("2"), cost=Decimal("10"))
    second = calculation_source_row("source-b:1", quantity=Decimal("3"), cost=Decimal("15"))
    first_match = match_result(first, candidate_id="candidate-a", target_row_id="target-1")
    second_match = match_result(second, candidate_id="candidate-b", target_row_id="target-1")
    combined = MatchResult(
        result_id="target-1",
        target_row_id="target-1",
        target_row=first_match.target_row,
        status=MatchStatus.MATCHED,
        selected_candidate=None,
        selected_candidates=(first_match.selected_candidate, second_match.selected_candidate),
        candidates=(first_match.selected_candidate, second_match.selected_candidate),
        warnings=(),
        explanation=("authoritative",),
    )

    (result,) = calculate_matches((combined,), calculation_rule_set())

    assert result.quantity == Decimal("5.00")
    assert result.cost == Decimal("25.00")


def test_physical_source_identity_uses_normalized_source_row_provenance() -> None:
    source = calculation_source_row("source:upload-a:17")

    assert _physical_source_identity(source) == (
        source.source_file_id,
        source.source_sheet,
        source.source_row_number,
    )


def test_reconciliation_target_binds_discovered_cells_and_scales_writer_values() -> None:
    source = calculation_source_row(
        "source:1", quantity=Decimal("1.005"), cost=Decimal("1000000.005")
    )
    calculated = calculate_matches(
        (match_result(source),), calculation_rule_set(coefficient=Decimal("2.7"))
    )

    (written,) = writer_calculations(calculated)

    pair = TargetMeasurePair("Отчёт", 12, 13, "август 2026 количество", "август 2026 стоимость")
    roles = {
        "Отчёт": {
            role: SimpleNamespace(
                column_index=index,
                column_letter=letter,
                header_text=role.value,
            )
            for role, index, letter in (
                (LogicalColumn.OBJECT_CODE, 1, "A"),
                (LogicalColumn.DOCUMENT_INDEX, 2, "B"),
                (LogicalColumn.STAGE, 3, "C"),
                (LogicalColumn.ROW_NUMBER, 4, "D"),
                (LogicalColumn.WORK_NAME, 5, "E"),
                (LogicalColumn.UNIT, 6, "F"),
            )
        }
    }
    assert [
        (binding.logical_column.value, binding.column_letter)
        for binding in _bindings(roles, (pair,))
    ] == [
        ("object_code", "A"),
        ("document_index", "B"),
        ("stage", "C"),
        ("row_number", "D"),
        ("work_name", "E"),
        ("unit", "F"),
        ("current_period_quantity", "L"),
        ("current_period_cost", "M"),
    ]
    assert written.quantity == Decimal("1.01")
    assert written.cost == Decimal("2.70")


def _historical_target_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["B1"], sheet["C1"], sheet["D1"], sheet["E1"], sheet["F1"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
        "Единица измерения",
    )
    sheet.merge_cells("L1:M1")
    sheet["L1"] = "Документальная отчетность за весь период"
    sheet["L2"], sheet["M2"], sheet["N1"] = "Количество", "Стоимость", "Следующий раздел"
    sheet["B3"], sheet["C3"], sheet["D3"], sheet["E3"], sheet["F3"] = (
        "1234",
        "Этап 13.1",
        "1",
        "Монтаж",
        "м",
    )
    sheet["L3"].font = sheet["M3"].font = Font(bold=True)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def test_historical_actionable_apply_writes_inserted_cells_and_survives_restart(
    tmp_path, monkeypatch
) -> None:
    source = BytesIO()
    Workbook().save(source)
    source_bytes, target_bytes = source.getvalue(), _historical_target_bytes()

    def execute(job: AdminJob) -> ReconciliationReviewResult:
        preview = preview_reconciliation_target(
            job.target, job.target_digest, job.stage, job.reporting_period
        )
        job.target_identity_digest = preview.target_identity_digest
        row = ReviewRow("source:1", "Монтаж", "м", Decimal("1"), Decimal("1"))
        (group,) = build_review_groups((row,))
        return ReconciliationReviewResult(
            ReconciliationReviewState(
                rows={row.row_id: row},
                groups={group.group_id: group},
                categories={"target-1": "Цель"},
                source_digests=job.source_digests,
                target_digest=job.target_digest,
                target_identity_digest=preview.target_identity_digest,
            ),
            None,
        )

    service = AdminPanelService(tmp_path, execute=execute)
    job = service.create_job(
        source_name="source.xlsx",
        source_content=source_bytes,
        target_name="target.xlsx",
        target_content=target_bytes,
        stage="13.1",
        reporting_period="2026-08",
    )
    assert job.review_state is not None
    _accept_group(job, job.review_state.group_snapshot()[0].group_id)

    def selected(_job, _state, _decisions, _rules, targets):
        calculation = replace(
            calculated_result(calculated_match()),
            target_row=replace(targets[0], writable=True),
            quantity=Decimal("7.50"),
            cost=Decimal("8.25"),
        )
        return object(), (), writer_calculations((calculation,))

    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution._calculate_selected", selected
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution._catalog_digest",
        lambda *_args: "generated-catalog",
    )

    applied = service.apply_reconciliation(job.job_id)

    assert applied.status == "ready" and applied.output is not None
    assert applied.source.read_bytes() == source_bytes
    workbook = load_workbook(applied.output, data_only=True)
    try:
        assert workbook["Отчёт"]["N3"].value == 7.5
        assert workbook["Отчёт"]["O3"].value is not None
    finally:
        workbook.close()
    feedback = service.feedback_store.records(job.target_digest)
    restored = AdminPanelService(tmp_path).get_job(job.job_id)
    assert restored.status == "ready" and restored.result_available is True
    assert AdminPanelService(tmp_path).feedback_store.records(job.target_digest) == feedback


def _review_job(
    tmp_path,
    *,
    work_name: str = "Монтаж трубы",
    unit: str = "м",
    category: str = "target-1",
) -> tuple[AdminPanelService, AdminJob, str]:
    service = AdminPanelService(tmp_path)
    directory = tmp_path / "job"
    directory.mkdir()
    source, target = directory / "source.xlsx", directory / "target.xlsx"
    source.write_bytes(b"source")
    target.write_bytes(b"target")
    row = ReviewRow("source:1", work_name, unit, Decimal("1"), Decimal("2"))
    (group,) = build_review_groups((row,))
    state = ReconciliationReviewState(
        rows={row.row_id: row},
        groups={group.group_id: group},
        categories={category: "Цель"},
        source_digests=(sha256(source.read_bytes()).hexdigest(),),
        target_digest=sha256(target.read_bytes()).hexdigest(),
        target_identity_digest=ReconciliationTargetIdentity(
            sha256(target.read_bytes()).hexdigest(), "13.1"
        ).target_identity_digest,
    )
    job = AdminJob(
        job_id="job",
        directory=directory,
        source=source,
        target=target,
        stage="13.1",
        mode="write",
        source_digest=state.source_digests[0],
        target_digest=state.target_digest,
        target_identity_digest=state.target_identity_digest,
        sources=(source,),
        source_digests=state.source_digests,
        status="review_required",
        review_state=state,
    )
    service.jobs[job.job_id] = job
    return service, job, group.group_id


def _accept_group(job: AdminJob, group_id: str, *, category: str = "target-1") -> None:
    assert job.review_state is not None
    version = job.review_state.group_snapshot()[0].version
    job.review_state.put_group(
        group_id,
        ReviewDecision(
            action=ReviewAction.ACCEPT,
            mode=ReviewMode.QUANTITY_COST,
            target_category=category,
            group_id=group_id,
            version=version,
        ),
    )


def test_apply_blocks_unresolved_and_tampered_input_before_write(tmp_path, monkeypatch) -> None:
    service, job, group_id = _review_job(tmp_path)
    called = False

    def should_not_apply(*_args):
        nonlocal called
        called = True
        raise AssertionError("apply must not execute")

    monkeypatch.setattr("report_processor.admin_panel.service.apply_review", should_not_apply)
    with pytest.raises(ValueError, match="incomplete"):
        service.apply_reconciliation(job.job_id)
    assert called is False

    _accept_group(job, group_id)
    job.target.write_bytes(b"tampered")
    with pytest.raises(RuntimeError, match="target upload changed"):
        service.apply_reconciliation(job.job_id)
    assert called is False and job.status == "failed"


def test_feedback_failure_removes_written_output_and_never_marks_job_ready(
    tmp_path, monkeypatch
) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    output.write_bytes(b"written")

    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review", lambda *_args: (output, ())
    )
    monkeypatch.setattr(
        service.feedback_store,
        "commit_apply",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("db")),
    )

    with pytest.raises(OSError, match="db"):
        service.apply_reconciliation(job.job_id)

    assert output.exists() is True
    assert job.output == output and job.status == "applying"
    assert job.result_available is False


@pytest.mark.parametrize("fault", ["before_commit", "after_commit"])
def test_restart_exact_replays_durable_apply_once(tmp_path, monkeypatch, fault) -> None:
    distinctive_work = "DISTINCTIVE PRIVATE WORK 9f44"
    distinctive_unit = "DISTINCTIVE PRIVATE UNIT 2b81"
    distinctive_category = "DISTINCTIVE PRIVATE CATEGORY 7c12"
    service, job, group_id = _review_job(
        tmp_path,
        work_name=distinctive_work,
        unit=distinctive_unit,
        category=distinctive_category,
    )
    _accept_group(job, group_id, category=distinctive_category)
    output = job.directory / "result.xlsx"
    decisions = tuple(job.review_state.core_decisions())
    feedback = _feedback_records(job.review_state, decisions)
    evidence = {
        "catalog_digest": "a" * 64,
        "target_identity_digest": job.target_identity_digest,
        "calculation_digest": "c" * 64,
        "rules_hash": "d" * 64,
        "actionable": True,
        "feedback": feedback,
    }
    plan: dict[str, str] = {}

    def write_result(*_args):
        output.write_bytes(b"result")
        from report_processor.business_rules import load_default_rule_set

        apply_key, plan_hash = _apply_plan(
            job,
            job.review_state,
            load_default_rule_set().rule_set.content_hash,
            _feedback_records(job.review_state, decisions),
            decisions,
        )
        plan.update(apply_key=apply_key, plan_hash=plan_hash)
        return ReconciliationApplyResult(
            output,
            feedback,
            apply_key,
            plan_hash,
            evidence["catalog_digest"],
            evidence["target_identity_digest"],
            evidence["calculation_digest"],
            evidence["rules_hash"],
            True,
        )

    monkeypatch.setattr("report_processor.admin_panel.service.apply_review", write_result)
    monkeypatch.setattr(
        "report_processor.admin_panel.service.rebuild_apply_evidence",
        lambda *_args: {**plan, **evidence},
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.service.prepare_review",
        lambda *_args: __import__(
            "report_processor.admin_panel.reconciliation_execution",
            fromlist=["ReconciliationReviewResult"],
        ).ReconciliationReviewResult(state=job.review_state, source_batch=None),
    )
    if fault == "before_commit":
        monkeypatch.setattr(
            service.feedback_store,
            "commit_apply",
            lambda **_kwargs: (_ for _ in ()).throw(OSError("before commit")),
        )
    else:
        original_save = service._job_store.save
        calls = 0

        def fail_ready_manifest(*args, **kwargs):
            nonlocal calls
            calls += 1
            if calls == 3:
                raise OSError("after commit")
            return original_save(*args, **kwargs)

        monkeypatch.setattr(service._job_store, "save", fail_ready_manifest)

    with pytest.raises(OSError):
        service.apply_reconciliation(job.job_id)
    assert job.status == "applying"
    manifest_bytes = (job.directory / "job-manifest.json").read_bytes()
    for private_value in (distinctive_work, distinctive_unit, distinctive_category):
        assert private_value.encode() not in manifest_bytes
    manifest = service._job_store.load(job.job_id)
    assert manifest is not None
    replay_decision = manifest["apply"]["evidence"]["decisions"][0]
    assert set(replay_decision) == {
        "action",
        "group_id",
        "mode",
        "row_id",
        "target_category_token",
        "version",
    }

    restored = AdminPanelService(tmp_path).get_job(job.job_id)

    assert restored.status == "ready"
    records = AdminPanelService(tmp_path).feedback_store.records(job.target_digest)
    assert records == feedback
    assert AdminPanelService(tmp_path).feedback_store.records(job.target_digest) == records


def _crash_non_actionable_apply(tmp_path, monkeypatch):
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    decisions = tuple(job.review_state.core_decisions())
    feedback = _feedback_records(job.review_state, decisions)
    evidence = {
        "catalog_digest": "a" * 64,
        "target_identity_digest": job.target_identity_digest,
        "calculation_digest": "c" * 64,
        "rules_hash": "d" * 64,
        "actionable": False,
        "feedback": feedback,
    }
    plan: dict[str, str] = {}

    def publish_unchanged(*_args):
        output.write_bytes(job.target.read_bytes())
        from report_processor.business_rules import load_default_rule_set

        apply_key, plan_hash = _apply_plan(
            job,
            job.review_state,
            load_default_rule_set().rule_set.content_hash,
            feedback,
            decisions,
        )
        plan.update(apply_key=apply_key, plan_hash=plan_hash)
        return ReconciliationApplyResult(
            output,
            feedback,
            apply_key,
            plan_hash,
            evidence["catalog_digest"],
            evidence["target_identity_digest"],
            evidence["calculation_digest"],
            evidence["rules_hash"],
            False,
        )

    monkeypatch.setattr("report_processor.admin_panel.service.apply_review", publish_unchanged)
    monkeypatch.setattr(
        "report_processor.admin_panel.service.rebuild_apply_evidence",
        lambda *_args: {**plan, **evidence},
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.service.prepare_review",
        lambda *_args: __import__(
            "report_processor.admin_panel.reconciliation_execution",
            fromlist=["ReconciliationReviewResult"],
        ).ReconciliationReviewResult(state=job.review_state, source_batch=None),
    )
    monkeypatch.setattr(
        service.feedback_store,
        "commit_apply",
        lambda **_kwargs: (_ for _ in ()).throw(OSError("before commit")),
    )
    with pytest.raises(OSError, match="before commit"):
        service.apply_reconciliation(job.job_id)
    return service, job, feedback, plan


def test_non_actionable_recovery_keeps_original_bytes_and_replays_once(
    tmp_path, monkeypatch
) -> None:
    service, job, feedback, _plan = _crash_non_actionable_apply(tmp_path, monkeypatch)
    original = job.target.read_bytes()

    restored = AdminPanelService(tmp_path).get_job(job.job_id)

    assert restored.output is not None and restored.output.read_bytes() == original
    assert service.feedback_store.records(job.target_digest) == feedback
    assert AdminPanelService(tmp_path).feedback_store.records(job.target_digest) == feedback


def test_non_actionable_recovery_rejects_self_consistent_changed_output(
    tmp_path, monkeypatch
) -> None:
    service, job, _feedback, plan = _crash_non_actionable_apply(tmp_path, monkeypatch)
    assert job.output is not None
    job.output.write_bytes(b"self-consistent but not the immutable target")
    changed_digest = sha256(job.output.read_bytes()).hexdigest()
    manifest = service._job_store.load(job.job_id)
    assert manifest is not None
    manifest["output_digest"] = changed_digest
    manifest["apply"]["output_digest"] = changed_digest
    manifest["apply"]["payload_hash"] = sha256(
        f"{plan['plan_hash']}:output-sha256:{changed_digest}".encode()
    ).hexdigest()
    service._job_store.save(job.job_id, manifest)

    recovered = AdminPanelService(tmp_path)

    with pytest.raises(KeyError):
        recovered.get_job(job.job_id)
    assert recovered.feedback_store.records(job.target_digest) == ()


def test_apply_recovery_rejects_tampered_category_replay_token(tmp_path, monkeypatch) -> None:
    service, job, _feedback, _plan = _crash_non_actionable_apply(tmp_path, monkeypatch)
    manifest = service._job_store.load(job.job_id)
    assert manifest is not None
    manifest["apply"]["evidence"]["decisions"][0]["target_category_token"] = "f" * 64
    service._job_store.save(job.job_id, manifest)

    recovered = AdminPanelService(tmp_path)

    with pytest.raises(KeyError):
        recovered.get_job(job.job_id)
    assert recovered.feedback_store.records(job.target_digest) == ()


def test_hostile_apply_hash_cannot_commit_feedback_or_publish_ready(tmp_path, monkeypatch) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    decisions = tuple(job.review_state.core_decisions())
    feedback = _feedback_records(job.review_state, decisions)

    def write_result(*_args):
        output.write_bytes(b"result")
        from report_processor.business_rules import load_default_rule_set

        apply_key, plan_hash = _apply_plan(
            job,
            job.review_state,
            load_default_rule_set().rule_set.content_hash,
            feedback,
            decisions,
        )
        return ReconciliationApplyResult(output, feedback, apply_key, plan_hash)

    monkeypatch.setattr("report_processor.admin_panel.service.apply_review", write_result)
    monkeypatch.setattr(
        service.feedback_store,
        "commit_apply",
        lambda **_kwargs: (_ for _ in ()).throw(OSError("before commit")),
    )
    with pytest.raises(OSError):
        service.apply_reconciliation(job.job_id)
    manifest = service._job_store.load(job.job_id)
    assert manifest is not None
    manifest["apply"]["apply_key"] = "f" * 64
    service._job_store.save(job.job_id, manifest)
    monkeypatch.setattr(
        "report_processor.admin_panel.service.prepare_review",
        lambda *_args: __import__(
            "report_processor.admin_panel.reconciliation_execution",
            fromlist=["ReconciliationReviewResult"],
        ).ReconciliationReviewResult(state=job.review_state, source_batch=None),
    )

    recovered = AdminPanelService(tmp_path)

    with pytest.raises(KeyError):
        recovered.get_job(job.job_id)
    assert recovered.feedback_store.records(job.target_digest) == ()


def test_repeated_apply_keeps_verified_ready_result_unchanged(tmp_path, monkeypatch) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    calls = 0

    def write_once(*_args):
        nonlocal calls
        calls += 1
        output.write_bytes(b"verified-result")
        return output, ()

    monkeypatch.setattr("report_processor.admin_panel.service.apply_review", write_once)

    first = service.apply_reconciliation(job.job_id)
    before = output.read_bytes()
    second = service.apply_reconciliation(job.job_id)

    assert first is second is job
    assert calls == 1
    assert output.read_bytes() == before
    assert job.status == "ready" and job.result_available is True
    manifest = service._job_store.load(job.job_id)
    assert manifest is not None and "apply" not in manifest
    restored = AdminPanelService(tmp_path).get_job(job.job_id)
    assert restored.status == "ready" and restored.result_available is True


def test_apply_validates_and_chmods_output_before_feedback_commit(tmp_path, monkeypatch) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    output.write_bytes(b"result")
    events: list[str] = []
    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review", lambda *_args: (output, ())
    )
    original_fchmod = __import__("os").fchmod

    def fchmod(descriptor, mode):
        events.append("chmod")
        return original_fchmod(descriptor, mode)

    monkeypatch.setattr("report_processor.admin_panel.service.os.fchmod", fchmod)
    monkeypatch.setattr(
        service.feedback_store,
        "commit_apply",
        lambda **_kwargs: events.append("commit") or True,
    )

    service.apply_reconciliation(job.job_id)

    assert events[-1] == "commit" and "chmod" in events
    assert job.status == "ready" and job.output == output


def test_apply_chmod_failure_never_commits_feedback_and_removes_owned_output(
    tmp_path, monkeypatch
) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    output.write_bytes(b"result")
    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review", lambda *_args: (output, ())
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.service.os.fchmod",
        lambda _descriptor, _mode: (_ for _ in ()).throw(OSError("chmod")),
    )
    committed = False

    def should_not_commit(**_kwargs):
        nonlocal committed
        committed = True
        return True

    monkeypatch.setattr(service.feedback_store, "commit_apply", should_not_commit)

    with pytest.raises(OSError, match="chmod"):
        service.apply_reconciliation(job.job_id)

    assert committed is False and output.exists() is True and job.status == "failed"


def test_apply_failure_keeps_concurrent_output_replacement(tmp_path, monkeypatch) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    output.write_bytes(b"attempt")
    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review", lambda *_args: (output, ())
    )

    def replace_then_fail(**_kwargs):
        replacement = job.directory / "replacement.xlsx"
        replacement.write_bytes(b"replacement")
        replacement.replace(output)
        raise OSError("database")

    monkeypatch.setattr(service.feedback_store, "commit_apply", replace_then_fail)

    with pytest.raises(OSError, match="database"):
        service.apply_reconciliation(job.job_id)

    assert output.read_bytes() == b"replacement"


def test_apply_rejects_wrong_output_path_before_feedback_commit(tmp_path, monkeypatch) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    wrong = job.directory / "other.xlsx"
    wrong.write_bytes(b"result")
    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review", lambda *_args: (wrong, ())
    )
    committed = False

    def should_not_commit(**_kwargs):
        nonlocal committed
        committed = True
        return True

    monkeypatch.setattr(service.feedback_store, "commit_apply", should_not_commit)
    with pytest.raises(RuntimeError, match="OUTPUT_INVALID"):
        service.apply_reconciliation(job.job_id)
    assert committed is False and wrong.exists()


def test_apply_precommit_rejects_replaced_source_and_preserves_owned_result(
    tmp_path, monkeypatch
) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    output.write_bytes(b"result")
    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review", lambda *_args: (output, ())
    )

    def replace_source_then_validate(**kwargs):
        replacement = job.directory / "replacement-source.xlsx"
        replacement.write_bytes(b"different source")
        replacement.replace(job.source)
        kwargs["precommit_validator"]()
        raise AssertionError("commit must not continue")

    monkeypatch.setattr(service.feedback_store, "commit_apply", replace_source_then_validate)
    with pytest.raises(RuntimeError, match="source upload changed"):
        service.apply_reconciliation(job.job_id)
    assert output.exists() is True


def test_apply_precommit_rejects_replaced_output_and_preserves_replacement(
    tmp_path, monkeypatch
) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    output.write_bytes(b"attempt")
    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review", lambda *_args: (output, ())
    )

    def replace_output_then_validate(**kwargs):
        replacement = job.directory / "replacement.xlsx"
        replacement.write_bytes(b"replacement")
        replacement.replace(output)
        kwargs["precommit_validator"]()

    monkeypatch.setattr(service.feedback_store, "commit_apply", replace_output_then_validate)
    with pytest.raises(RuntimeError, match="OUTPUT_INVALID"):
        service.apply_reconciliation(job.job_id)
    assert output.read_bytes() == b"replacement"


@pytest.mark.parametrize("mutation", ["content", "mode"])
def test_apply_precommit_rejects_in_place_output_mutation(tmp_path, monkeypatch, mutation) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    output.write_bytes(b"attempt")
    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review", lambda *_args: (output, ())
    )

    def mutate_then_validate(**kwargs):
        if mutation == "content":
            output.write_bytes(b"changed")
        else:
            output.chmod(0o644)
        kwargs["precommit_validator"]()

    monkeypatch.setattr(service.feedback_store, "commit_apply", mutate_then_validate)
    with pytest.raises(RuntimeError, match="OUTPUT_INVALID"):
        service.apply_reconciliation(job.job_id)
    assert output.exists()


def test_keyboard_interrupt_during_apply_marks_job_failed_and_keeps_output_safe(
    tmp_path, monkeypatch
) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    output.write_bytes(b"existing")
    monkeypatch.setattr(
        "report_processor.admin_panel.service.apply_review",
        lambda *_args: (_ for _ in ()).throw(KeyboardInterrupt()),
    )
    with pytest.raises(KeyboardInterrupt):
        service.apply_reconciliation(job.job_id)
    assert job.status == "failed" and output.read_bytes() == b"existing"


def test_failed_snapshot_copy_keeps_concurrent_replacement(tmp_path, monkeypatch) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    destination = tmp_path / "private-snapshot.xlsx"

    def replace_then_fail(_descriptor, _content):
        replacement = tmp_path / "replacement.xlsx"
        replacement.write_bytes(b"replacement")
        replacement.replace(destination)
        raise OSError("write")

    monkeypatch.setattr("report_processor.admin_panel.service.os.write", replace_then_fail)
    with pytest.raises(OSError, match="write"):
        _copy_verified_snapshot(source, destination, sha256(source.read_bytes()).hexdigest())
    assert destination.read_bytes() == b"replacement"


def test_decision_mutation_cannot_interleave_authoritative_apply(tmp_path, monkeypatch) -> None:
    service, job, group_id = _review_job(tmp_path)
    _accept_group(job, group_id)
    output = job.directory / "result.xlsx"
    entered, release, mutation_done = threading.Event(), threading.Event(), threading.Event()
    errors: list[Exception] = []

    def blocked_apply(*_args):
        output.write_bytes(b"result")
        entered.set()
        assert release.wait(2)
        return output, ()

    monkeypatch.setattr("report_processor.admin_panel.service.apply_review", blocked_apply)
    apply_thread = threading.Thread(target=lambda: service.apply_reconciliation(job.job_id))
    apply_thread.start()
    assert entered.wait(2)

    def mutate():
        try:
            service.put_reconciliation_group(
                job.job_id, group_id, ReviewDecision(ReviewAction.REJECT)
            )
        except Exception as error:
            errors.append(error)
        finally:
            mutation_done.set()

    mutation_thread = threading.Thread(target=mutate)
    mutation_thread.start()
    release.set()
    apply_thread.join(2)
    mutation_thread.join(2)
    assert job.status == "ready" and len(errors) == 1 and isinstance(errors[0], ValueError)

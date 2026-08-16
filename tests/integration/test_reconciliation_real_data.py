"""Synthetic and opt-in private-workbook acceptance for reconciliation."""

from __future__ import annotations

import os
from hashlib import sha256
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from report_processor.admin_panel.reconciliation_execution import _sources, prepare_review
from report_processor.admin_panel.reconciliation_sources import (
    AllReconciliationSourcesUnusableError,
    ReconciliationSourceDescriptor,
    extract_reconciliation_sources,
)


def _ks2(path: Path, *, work_name: str = "Монтаж трубы") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("№", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"))
    sheet.append(("1", work_name, "м", "1.25", "1250.50"))
    workbook.save(path)
    workbook.close()


def _source_input(path: Path, source_id: str, basename: str):
    return path, source_id, ReconciliationSourceDescriptor(basename)


def test_partial_source_failure_keeps_usable_synthetic_source_and_safe_guidance(
    tmp_path: Path,
) -> None:
    usable = tmp_path / "source-1234.xlsx"
    unusable = tmp_path / "bad-source.xlsx"
    _ks2(usable)
    unusable.write_bytes(b"not-an-xlsx")

    batch = extract_reconciliation_sources(
        (
            _source_input(unusable, "source:bad", "bad-source.xlsx"),
            _source_input(usable, "source:good", "source-1234.xlsx"),
        )
    )

    assert len(batch.rows) == 1
    assert batch.selections[0].safe_basename == "source-1234.xlsx"
    assert batch.issues[0].code == "WORKBOOK_UNREADABLE"
    assert batch.issues[0].safe_basename == "bad-source.xlsx"
    assert batch.issues[0].can_continue is True
    assert str(tmp_path) not in repr(batch.issues)


def test_all_bad_sources_return_only_controlled_safe_basename_guidance(tmp_path: Path) -> None:
    bad = tmp_path / "input.xlsx"
    bad.write_bytes(b"not-an-xlsx")

    with pytest.raises(AllReconciliationSourcesUnusableError) as raised:
        extract_reconciliation_sources((_source_input(bad, "source:bad", "input.xlsx"),))

    (issue,) = raised.value.issues
    assert issue.code == "WORKBOOK_UNREADABLE"
    assert issue.safe_basename == "input.xlsx"
    assert issue.comment and issue.repair_hint and issue.can_continue is True
    assert str(tmp_path) not in repr(issue)


def test_missing_document_index_is_excluded_with_safe_repair_guidance(tmp_path: Path) -> None:
    usable = tmp_path / "source-1234.xlsx"
    missing_index = tmp_path / "source.xlsx"
    _ks2(usable)
    _ks2(missing_index, work_name="Монтаж без индекса")

    batch = extract_reconciliation_sources(
        (
            (
                usable,
                "source:good",
                ReconciliationSourceDescriptor("source-1234.xlsx", document_index="1234"),
            ),
            _source_input(missing_index, "source:missing", "source.xlsx"),
        ),
        require_document_index=True,
    )

    assert len(batch.rows) == 1
    assert batch.selections[0].safe_basename == "source-1234.xlsx"
    assert batch.issues[0].code == "DOCUMENT_INDEX_MISSING"
    assert batch.issues[0].safe_basename == "source.xlsx"
    assert "индекс" in batch.issues[0].comment.casefold()
    assert "индекс" in batch.issues[0].repair_hint.casefold()
    assert str(tmp_path) not in repr(batch.issues)


def test_bare_four_digit_upload_name_survives_production_source_path(tmp_path: Path) -> None:
    usable = tmp_path / "stored-source.xlsx"
    _ks2(usable)
    job = SimpleNamespace(
        sources=(usable,),
        source=usable,
        source_names=("source-1234.xlsx",),
        source_digests=(sha256(usable.read_bytes()).hexdigest(),),
    )

    batch = _sources(job)

    assert len(batch.rows) == 1
    assert batch.selections[0].safe_basename == "source-1234.xlsx"
    assert batch.issues == ()


def test_hierarchical_cumulative_header_keeps_first_detail_row(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("", "Наименование", "Единица", "Выполнено за весь период", "", ""))
    sheet.append(("", "работ и этапов", "измерения", "за весь период", "", ""))
    sheet.append(("№", "", "", "Количество", "Общая стоимость", ""))
    sheet.append(("1", "Первая работа", "м", "1.25", "1250.50", ""))
    sheet.merge_cells("D1:E2")
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert len(batch.rows) == 1
    assert batch.selections[0].source_type == "ks6a"


def _cumulative_sheet(sheet, name: str = "Работа") -> None:
    sheet.append(("", "Описание", "Единица", "Освоено нарастающим итогом", "", ""))
    sheet.append(("", "строительных работ", "измерения", "нарастающим итогом", "", ""))
    sheet.append(("№", "", "", "Объём", "Сумма затрат", ""))
    sheet.append(("1", name, "м", "1.25", "1250.50", ""))
    sheet.merge_cells(
        start_row=sheet.max_row - 3, start_column=4, end_row=sheet.max_row - 2, end_column=5
    )


def test_two_viable_cumulative_sheets_fail_controlled_ambiguity(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    _cumulative_sheet(workbook.active)
    _cumulative_sheet(workbook.create_sheet("Копия"), "Другая работа")
    workbook.save(path)
    workbook.close()

    with pytest.raises(AllReconciliationSourcesUnusableError) as raised:
        extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert raised.value.issues[0].code == "SOURCE_LAYOUT_AMBIGUOUS"


def test_unique_cumulative_candidate_outranks_direct_candidate(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    _cumulative_sheet(workbook.active)
    direct = workbook.create_sheet("КС-2")
    _ks2_sheet = (
        ("№", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"),
        ("1", "Прямая работа", "м", "2", "20"),
    )
    for row in _ks2_sheet:
        direct.append(row)
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert len(batch.rows) == 1
    assert batch.selections[0].source_type == "ks6a"


def test_cumulative_detail_interval_stops_before_later_overlapping_direct_table(
    tmp_path: Path,
) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Выполнено нарастающим итогом", ""),
        ("", "", "", "Количество", "Общая стоимость"),
        ("1", "Первая работа", "м", "1", "10"),
        (),
        ("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"),
        ("2", "Вторая работа", "м", "2", "20"),
    ):
        sheet.append(row)
    sheet.merge_cells("D1:E1")
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert batch.selections[0].source_type == "ks6a"
    assert len(batch.rows) == 1
    assert batch.rows[0].work_name == "первая работа"


def test_cumulative_detail_interval_streams_past_initial_header_window(
    tmp_path: Path,
) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Выполнено нарастающим итогом", ""),
        ("", "", "", "Количество", "Общая стоимость"),
        ("1", "Первая работа", "м", "1", "10"),
    ):
        sheet.append(row)
    sheet.merge_cells("D1:E1")
    sheet.cell(row=59, column=1, value="")
    assert sheet.max_row == 59
    sheet.append(("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"))
    sheet.append(("2", "Вторая работа", "м", "2", "20"))
    assert sheet.max_row == 61
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert batch.selections[0].source_type == "ks6a"
    assert len(batch.rows) == 1
    assert batch.rows[0].work_name == "первая работа"


def test_later_cumulative_region_past_header_window_fails_controlled_ambiguity(
    tmp_path: Path,
) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Выполнено нарастающим итогом", ""),
        ("", "", "", "Количество", "Общая стоимость"),
        ("1", "Первая работа", "м", "1", "10"),
    ):
        sheet.append(row)
    sheet.merge_cells("D1:E1")
    sheet.cell(row=58, column=1, value="")
    sheet.append(("", "Наименование работ", "Ед. изм.", "Выполнено нарастающим итогом", ""))
    sheet.append(("", "", "", "Количество", "Общая стоимость"))
    sheet.append(("2", "Вторая работа", "м", "2", "20"))
    assert sheet.max_row == 61
    sheet.merge_cells("D59:E59")
    workbook.save(path)
    workbook.close()

    with pytest.raises(AllReconciliationSourcesUnusableError) as raised:
        extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert raised.value.issues[0].code == "SOURCE_LAYOUT_AMBIGUOUS"


def test_later_direct_region_past_header_window_fails_controlled_ambiguity(
    tmp_path: Path,
) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"),
        ("1", "Первая работа", "м", "1", "10"),
    ):
        sheet.append(row)
    sheet.cell(row=89, column=1, value="")
    sheet.append(("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"))
    sheet.append(("2", "Вторая работа", "м", "2", "20"))
    assert sheet.max_row == 91
    workbook.save(path)
    workbook.close()

    with pytest.raises(AllReconciliationSourcesUnusableError) as raised:
        extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert raised.value.issues[0].code == "SOURCE_LAYOUT_AMBIGUOUS"


def test_cumulative_root_straddling_initial_window_is_discovered(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=49, column=2, value="Наименование работ")
    sheet.cell(row=49, column=3, value="Ед. изм.")
    sheet.cell(row=49, column=4, value="Выполнено нарастающим итогом")
    sheet.merge_cells("D49:E50")
    sheet.cell(row=51, column=4, value="Количество")
    sheet.cell(row=51, column=5, value="Общая стоимость")
    sheet.cell(row=52, column=2, value="Работа")
    sheet.cell(row=52, column=3, value="м")
    sheet.cell(row=52, column=4, value=1)
    sheet.cell(row=52, column=5, value=10)
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert batch.selections[0].source_type == "ks6a"
    assert len(batch.rows) == 1


def test_direct_header_band_straddling_initial_window_is_discovered(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=79, column=2, value="Наименование")
    sheet.cell(row=79, column=3, value="Единица")
    sheet.cell(row=80, column=2, value="работ")
    sheet.cell(row=80, column=3, value="измерения")
    sheet.cell(row=81, column=4, value="Количество")
    sheet.cell(row=81, column=5, value="Общая стоимость")
    sheet.cell(row=82, column=2, value="Работа")
    sheet.cell(row=82, column=3, value="м")
    sheet.cell(row=82, column=4, value=1)
    sheet.cell(row=82, column=5, value=10)
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert batch.selections[0].source_type == "ks2"
    assert len(batch.rows) == 1


def test_late_formula_without_cache_remains_controlled_issue(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=59, column=2, value="Наименование работ")
    sheet.cell(row=59, column=3, value="Ед. изм.")
    sheet.cell(row=59, column=4, value="Количество")
    sheet.cell(row=59, column=5, value="Общая стоимость")
    sheet.cell(row=60, column=2, value="Работа")
    sheet.cell(row=60, column=3, value="м")
    sheet.cell(row=60, column=4, value="=1+1")
    sheet.cell(row=60, column=5, value=10)
    workbook.save(path)
    workbook.close()

    with pytest.raises(AllReconciliationSourcesUnusableError) as raised:
        extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert raised.value.issues[0].code == "FORMULA_CACHE_UNAVAILABLE"


def test_inflated_worksheet_dimension_uses_sparse_region_index(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"),
        ("1", "Работа", "м", 1, 10),
    ):
        sheet.append(row)
    sheet["XFD1048576"] = "dimension sentinel"
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert len(batch.rows) == 1


def test_metric_shaped_noise_without_roles_does_not_truncate_cumulative_detail(
    tmp_path: Path,
) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Выполнено нарастающим итогом", ""),
        ("", "", "", "Количество", "Общая стоимость"),
        ("1", "Первая работа", "м", "1", "10"),
        (),
        ("", "Произвольный текст", "", "Количество", "Общая стоимость"),
        ("2", "Вторая работа", "м", "2", "20"),
    ):
        sheet.append(row)
    sheet.merge_cells("D1:E1")
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert batch.selections[0].source_type == "ks6a"
    assert {row.work_name for row in batch.rows} == {"первая работа", "вторая работа"}


def test_two_cumulative_regions_in_one_sheet_fail_controlled_ambiguity(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    _cumulative_sheet(sheet)
    sheet.append(())
    _cumulative_sheet(sheet, "Вторая работа")
    workbook.save(path)
    workbook.close()

    with pytest.raises(AllReconciliationSourcesUnusableError) as raised:
        extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert raised.value.issues[0].code == "SOURCE_LAYOUT_AMBIGUOUS"


def test_two_direct_regions_in_one_sheet_fail_controlled_ambiguity(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"),
        ("1", "Первая", "м", "1", "10"),
        (),
        ("", "Описание работ", "Единица измерения", "Объём", "Сумма затрат"),
        ("2", "Вторая", "м", "2", "20"),
    ):
        sheet.append(row)
    workbook.save(path)
    workbook.close()

    with pytest.raises(AllReconciliationSourcesUnusableError) as raised:
        extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert raised.value.issues[0].code == "SOURCE_LAYOUT_AMBIGUOUS"


def test_empty_physical_candidate_does_not_compete_with_one_normalized_layout(
    tmp_path: Path,
) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    rejected = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"),
        ("1", "Итого", "м", "1", "10"),
    ):
        rejected.append(row)
    viable = workbook.create_sheet("Данные")
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"),
        ("1", "Учитывать", "м", "2", "20"),
    ):
        viable.append(row)
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert len(batch.rows) == 1
    assert batch.rows[0].work_name == "учитывать"


def test_footer_formula_does_not_invalidate_eligible_source_rows(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("№", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"))
    sheet.append(("1", "Работа", "м", "1", "10"))
    sheet.append(("", "Итого", "", "=SUM(D2:D2)", "=SUM(E2:E2)"))
    workbook.save(path)
    workbook.close()

    batch = extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert len(batch.rows) == 1


def test_formula_metric_without_cache_is_controlled_source_issue(tmp_path: Path) -> None:
    path = tmp_path / "source.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("№", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"))
    sheet.append(("1", "Работа", "м", "=1+1", "10"))
    workbook.save(path)
    workbook.close()

    with pytest.raises(AllReconciliationSourcesUnusableError) as raised:
        extract_reconciliation_sources((_source_input(path, "source:one", "source.xlsx"),))

    assert raised.value.issues[0].code == "FORMULA_CACHE_UNAVAILABLE"


def test_opt_in_real_workbooks_leave_all_input_bytes_unchanged() -> None:
    source_values = tuple(
        value
        for value in os.environ.get("RECONCILIATION_REAL_SOURCE_PATHS", "").split(os.pathsep)
        if value
    )
    target_value = os.environ.get("RECONCILIATION_REAL_TARGET_PATH")
    stage = os.environ.get("RECONCILIATION_REAL_STAGE")
    if not source_values or not target_value or not stage:
        pytest.skip(
            "set RECONCILIATION_REAL_SOURCE_PATHS, RECONCILIATION_REAL_TARGET_PATH "
            "and RECONCILIATION_REAL_STAGE"
        )

    sources = tuple(Path(value) for value in source_values)
    target = Path(target_value)
    assert all(path.is_file() for path in (*sources, target))
    before = {path: sha256(path.read_bytes()).hexdigest() for path in (*sources, target)}
    job = SimpleNamespace(
        sources=sources,
        source=sources[0],
        source_names=tuple(path.name for path in sources),
        source_digests=tuple(before[path] for path in sources),
        target=target,
        target_digest=before[target],
        stage=stage,
        rules_path=None,
    )

    result = prepare_review(job, ())

    assert result.state is not None or result.source_issues or result.target_error
    assert {path: sha256(path.read_bytes()).hexdigest() for path in (*sources, target)} == before

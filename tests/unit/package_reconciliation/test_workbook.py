from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from report_processor.package_reconciliation import extract_package_workbook_facts


def _workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "КС-2 август"
    worksheet["A1"] = "Акт № KS-2/17"
    worksheet["A2"] = "Отчетный период за август 2026"
    worksheet["A3"] = "Шифр объекта OBJ-01"
    headers = [
        "Позиция",
        "Наименование работ",
        "Ед. изм.",
        "Количество",
        "Стоимость всего",
        "Шифр чертежа",
        "Обоснование",
    ]
    for column, value in enumerate(headers, start=1):
        worksheet.cell(7, column, value)
    worksheet.append([])
    worksheet.cell(8, 1, "1.02")
    worksheet.cell(8, 2, "Монтаж опор")
    worksheet.cell(8, 3, "шт")
    worksheet.cell(8, 4, 2.5)
    worksheet.cell(8, 5, 12500)
    worksheet.cell(8, 6, "DWG-1")
    worksheet.cell(8, 7, "ФЕР01")
    worksheet.cell(9, 2, "Строка без кода")
    worksheet.cell(9, 4, 1)
    workbook.save(path)
    workbook.close()


def test_extracts_structural_header_and_comparable_ks2_facts(tmp_path: Path) -> None:
    workbook_path = tmp_path / "act.xlsx"
    _workbook(workbook_path)

    facts = extract_package_workbook_facts(tmp_path, "act.xlsx")

    assert facts.workbook_path.as_posix() == "act.xlsx"
    sheet = facts.sheets[0]
    assert sheet.act_number == "KS-2/17"
    assert sheet.period == "август 2026"
    assert sheet.object_code == "OBJ-01"
    assert len(sheet.rows) == 2
    first = sheet.rows[0]
    assert first.work_code == "1.02"
    assert first.work_name == "Монтаж опор"
    assert first.unit == "шт"
    assert first.quantity == Decimal("2.5")
    assert first.total_cost == Decimal("12500")
    assert first.drawing_code == "DWG-1"
    assert first.basis == "ФЕР01"
    assert [issue.code for issue in sheet.issues] == ["MISSING_WORK_CODE"]


def test_returns_controlled_issue_when_header_is_not_structural(tmp_path: Path) -> None:
    path = tmp_path / "notes.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "Только примечание"
    workbook.save(path)
    workbook.close()

    facts = extract_package_workbook_facts(tmp_path, Path("notes.xlsx"))

    assert facts.sheets[0].rows == ()
    assert facts.sheets[0].issues[0].code == "HEADER_NOT_FOUND"


def test_extracts_multirow_report_period_from_structural_labels(tmp_path: Path) -> None:
    path = tmp_path / "period.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["E2"] = "Отчетный период"
    worksheet["E3"] = "с"
    worksheet["F3"] = date(2026, 8, 1)
    worksheet["E4"] = "по"
    worksheet["F4"] = date(2026, 8, 31)
    for column, value in enumerate(
        ["Позиция", "Наименование работ", "Ед. изм.", "Количество", "Стоимость всего"], start=1
    ):
        worksheet.cell(7, column, value)
    worksheet.append(["2.01", "Устройство основания", "м2", 12, 6000])
    workbook.save(path)
    workbook.close()

    facts = extract_package_workbook_facts(tmp_path, "period.xlsx")

    assert facts.sheets[0].period == "с 01.08.2026 по 31.08.2026"
    assert facts.sheets[0].rows[0].work_code == "2.01"


def test_extracts_period_dates_below_adjacent_from_to_markers(tmp_path: Path) -> None:
    path = tmp_path / "period-below.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["E2"] = "Отчетный период"
    worksheet["E3"] = "с"
    worksheet["F3"] = "по"
    worksheet["E4"] = date(2026, 8, 1)
    worksheet["F4"] = date(2026, 8, 31)
    for column, value in enumerate(
        ["Позиция", "Наименование работ", "Ед. изм.", "Количество", "Стоимость всего"], start=1
    ):
        worksheet.cell(7, column, value)
    worksheet.append(["2.02", "Устройство основания", "м2", 12, 6000])
    workbook.save(path)
    workbook.close()

    facts = extract_package_workbook_facts(tmp_path, "period-below.xlsx")

    assert facts.sheets[0].period == "с 01.08.2026 по 31.08.2026"


def test_rejects_path_escape_and_does_not_accept_symlinked_workbook(tmp_path: Path) -> None:
    _workbook(tmp_path / "act.xlsx")
    link = tmp_path / "link.xlsx"
    link.symlink_to(tmp_path / "act.xlsx")

    with pytest.raises(ValueError, match="safe relative"):
        extract_package_workbook_facts(tmp_path, "../act.xlsx")
    with pytest.raises(ValueError, match="symlinked workbook"):
        extract_package_workbook_facts(tmp_path, "link.xlsx")

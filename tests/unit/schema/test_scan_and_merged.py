from pathlib import Path

from openpyxl import Workbook, load_workbook

from conftest import schema_candidate
from report_processor.schema.config import SheetScanConfig
from report_processor.schema.header_composer import compose_logical_headers
from report_processor.schema.scan_window import get_cached_merged_ranges, scan_worksheet_window
from report_processor.workflow import prepared_workbook_session


def test_readonly_scan_collects_merged_headers(schema_workbook_path: Path) -> None:
    with prepared_workbook_session(schema_candidate(schema_workbook_path)) as session:
        scan = scan_worksheet_window(session, "КС-6а", SheetScanConfig())
        ranges = get_cached_merged_ranges(session, "КС-6а")
        headers = compose_logical_headers(scan, ranges, start_row=5, end_row=6)
    by_letter = {header.column_letter: header.raw_text for header in headers}
    assert "D5:E5" in scan.merged_ranges
    assert by_letter["D"] == "Выполнено за отчетный период Количество"
    assert by_letter["E"] == "Выполнено за отчетный период Стоимость"


def test_scan_is_bounded_by_config(tmp_path: Path) -> None:
    path = tmp_path / "large.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист1"
    worksheet["A1"] = "начало"
    worksheet["A100"] = "не должно читаться"
    worksheet["Z1"] = "далеко"
    workbook.save(path)
    workbook.close()
    config = SheetScanConfig(max_scan_rows=10, max_scan_columns=5)
    with prepared_workbook_session(schema_candidate(path)) as session:
        scan = scan_worksheet_window(session, "Лист1", config)
    assert scan.max_scanned_row == 10
    assert scan.max_scanned_column == 5
    assert all(cell.row <= 10 and cell.column <= 5 for cell in scan.cells)


def test_merged_values_are_not_copied_to_workbook(schema_workbook_path: Path) -> None:
    workbook = load_workbook(schema_workbook_path, read_only=False)
    worksheet = workbook["КС-6а"]
    assert worksheet["E5"].value is None
    workbook.close()


def test_scan_cell_limit_is_reported(tmp_path: Path) -> None:
    path = tmp_path / "cell-limit.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "КС-2"
    for row in range(1, 8):
        for column in range(1, 8):
            worksheet.cell(row, column, f"v-{row}-{column}")
    workbook.save(path)
    workbook.close()
    config = SheetScanConfig(max_nonempty_cells=10, stop_after_empty_rows=15)
    with prepared_workbook_session(schema_candidate(path)) as session:
        scan = scan_worksheet_window(session, "КС-2", config)
    assert scan.nonempty_cell_count == 10
    assert scan.stopped_early is True
    assert "SCAN_CELL_LIMIT_REACHED" in scan.warnings


def test_scan_passes_explicit_bounds_to_openpyxl(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    path = tmp_path / "bounded-spy.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Лист1"
    worksheet["A1"] = "Наименование работ"
    worksheet["B1"] = "Ед. изм."
    worksheet["C1"] = "Количество"
    worksheet["A500"] = "далеко"
    workbook.save(path)
    workbook.close()

    original = ReadOnlyWorksheet.iter_rows
    calls: list[dict[str, int]] = []

    def spy(self, *args, **kwargs):
        calls.append(dict(kwargs))
        return original(self, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", spy)
    config = SheetScanConfig(max_scan_rows=12, max_scan_columns=7)
    with prepared_workbook_session(schema_candidate(path)) as session:
        scan_worksheet_window(session, "Лист1", config)
    assert calls == [{"min_row": 1, "max_row": 12, "min_col": 1, "max_col": 3}]


def test_vertical_merged_parent_is_composed(tmp_path: Path) -> None:
    path = tmp_path / "vertical-merge.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "КС-6а"
    worksheet.merge_cells("A5:A6")
    worksheet["A5"] = "Наименование работ"
    worksheet["B5"] = "Выполнено за июль 2026"
    worksheet["B6"] = "Количество"
    worksheet["A7"] = "Монтаж"
    worksheet["B7"] = 1
    workbook.save(path)
    workbook.close()
    with prepared_workbook_session(schema_candidate(path)) as session:
        scan = scan_worksheet_window(session, "КС-6а", SheetScanConfig())
        ranges = get_cached_merged_ranges(session, "КС-6а")
        headers = compose_logical_headers(scan, ranges, start_row=5, end_row=6)
    by_letter = {item.column_letter: item.raw_text for item in headers}
    assert by_letter["A"] == "Наименование работ"
    assert by_letter["B"] == "Выполнено за июль 2026 Количество"

from pathlib import Path

from openpyxl import Workbook

from conftest import schema_candidate
from report_processor.schema.config import create_default_schema_config
from report_processor.schema.header_candidates import find_header_candidates
from report_processor.schema.scan_window import scan_worksheet_window
from report_processor.schema.sheet_classifier import classify_worksheet
from report_processor.workflow import prepared_workbook_session


def test_one_row_header_and_data_start(tmp_path: Path) -> None:
    path = tmp_path / "ks2.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "КС-2"
    for column, value in enumerate(
        ["Позиция", "Наименование работ", "Ед. изм.", "Количество", "Стоимость"],
        start=1,
    ):
        worksheet.cell(5, column, value)
    worksheet.append([])
    worksheet.cell(6, 1, "1.1")
    worksheet.cell(6, 2, "Работа")
    worksheet.cell(6, 3, "м")
    worksheet.cell(6, 4, 2)
    worksheet.cell(6, 5, 50)
    workbook.save(path)
    workbook.close()
    config = create_default_schema_config()
    with prepared_workbook_session(schema_candidate(path)) as session:
        scan = scan_worksheet_window(session, "КС-2", config.scan)
        classification = classify_worksheet(session, "КС-2", config, scan=scan)
        candidates = find_header_candidates(scan, classification, config.headers)
    assert candidates[0].start_row == 5
    assert candidates[0].end_row == 5

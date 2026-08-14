"""Publication cleanup ownership regression tests."""

from __future__ import annotations

import hashlib
import os
from dataclasses import replace
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from fixtures.quality_control.builders import calculated_match, calculated_result
from report_processor.excel_writer import (
    ExcelWriterIntegrityError,
    engine,
    ooxml,
    write_target_report,
)
from report_processor.quality_control import WriteDecision
from report_processor.schema import LogicalColumn, SheetType
from report_processor.target_report.models import (
    TargetCellSnapshot,
    TargetColumnBinding,
    TargetPeriodIdentity,
    TargetReportSchema,
    TargetSourceFingerprint,
    TargetWorksheetSnapshot,
)


def _schema(path: Path) -> TargetReportSchema:
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    source_file_id = f"synthetic:{digest}"
    return TargetReportSchema(
        version="TargetReport-9.0",
        source_fingerprint=TargetSourceFingerprint(
            "sha256", digest, path.stat().st_size, source_file_id
        ),
        period_identity=TargetPeriodIdentity(status="OK"),
        column_bindings=(
            TargetColumnBinding(LogicalColumn.CURRENT_PERIOD_QUANTITY, 4, "D", "quantity", "test"),
        ),
        worksheets=(
            TargetWorksheetSnapshot("Лист", SheetType.KS6A, "D30", (), None, (), None, False),
        ),
        object_blocks=(),
        status="OK",
        source_file_id=source_file_id,
        filename=path.name,
        source_sha256=digest,
    )


def _calculation() -> object:
    cell = TargetCellSnapshot("D30", 5, "5", Decimal("5"), 1, "0.00", None, None, "OK")
    match = calculated_match()
    target = replace(
        match.target_row,
        sheet_name="Лист",
        row_number=30,
        cells=((LogicalColumn.CURRENT_PERIOD_QUANTITY, cell),),
    )
    calculated = calculated_result(replace(match, target_row=target))
    return replace(calculated, target_row=target, quantity=Decimal("0"), cost=None)


def _workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист"
    sheet["D30"] = 5
    sheet["D30"].font = Font(bold=True)
    workbook.save(path)
    workbook.close()


def test_reopen_failure_never_removes_output_replaced_during_publication(
    tmp_path, monkeypatch
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "result.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    _workbook(source)
    replacement.write_bytes(b"concurrent replacement")

    publish = engine._publish_no_clobber

    def publish_then_replace(temporary: Path, destination: Path) -> None:
        publish(temporary, destination)
        os.replace(replacement, destination)

    def fail_reopen(_path: Path) -> None:
        raise ExcelWriterIntegrityError("REOPEN_FAILED", "injected")

    monkeypatch.setattr(engine, "_publish_no_clobber", publish_then_replace)
    monkeypatch.setattr(
        "report_processor.excel_writer.engine._reopen_published_output", fail_reopen
    )

    with pytest.raises(ExcelWriterIntegrityError, match="REOPEN_FAILED"):
        write_target_report(
            source, output, WriteDecision.ALLOW_WRITE, (_calculation(),), _schema(source)
        )

    assert output.read_bytes() == b"concurrent replacement"


def test_swap_use_restore_of_source_path_cannot_change_snapshot_output(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.xlsx"
    alternate = tmp_path / "alternate.xlsx"
    parked = tmp_path / "parked.xlsx"
    output = tmp_path / "result.xlsx"
    _workbook(source)
    _workbook(alternate)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист"
    sheet["D30"] = 99
    sheet["D30"].font = Font(bold=True)
    workbook.save(alternate)
    workbook.close()
    original_write = engine._write_temp_package

    def swap_during_write(snapshot: int, *args, **kwargs) -> None:
        assert isinstance(snapshot, int)
        os.replace(source, parked)
        os.replace(alternate, source)
        try:
            original_write(snapshot, *args, **kwargs)
        finally:
            os.replace(source, alternate)
            os.replace(parked, source)

    monkeypatch.setattr(engine, "_write_temp_package", swap_during_write)
    result = write_target_report(
        source, output, WriteDecision.ALLOW_WRITE, (_calculation(),), _schema(source)
    )

    assert result.status.value == "written"
    written = load_workbook(output, data_only=True, read_only=True)
    assert written["Лист"]["D30"].value == 0
    written.close()
    assert not tuple(tmp_path.glob(".excel-writer-source-*.xlsx"))


def test_build_write_plan_scans_one_changed_part_once_for_one_thousand_cells(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.xlsx"
    _workbook(source)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист"
    for row in range(1, 1_001):
        cell = sheet.cell(row, 4, row)
        cell.font = Font(bold=True)
    workbook.save(source)
    workbook.close()
    schema = _schema(source)
    template = _calculation()
    calculations = []
    for row in range(1, 1_001):
        snapshot = TargetCellSnapshot(
            f"D{row}", row, str(row), Decimal(row), 1, "0.00", None, None, "OK"
        )
        target = replace(
            template.target_row,
            row_number=row,
            cells=((LogicalColumn.CURRENT_PERIOD_QUANTITY, snapshot),),
        )
        calculations.append(
            replace(
                template,
                calculation_id=f"generated-{row}",
                target_row_id=f"generated-row-{row}",
                target_row=target,
                trace=replace(template.trace, target_row_id=f"generated-row-{row}"),
            )
        )
    original = ooxml._scan_worksheet
    calls = 0

    def counted(payload: bytes, error_code: str = "TARGET_CELL_MISSING"):
        nonlocal calls
        calls += 1
        return original(payload, error_code)

    monkeypatch.setattr(ooxml, "_scan_worksheet", counted)
    parts = engine.worksheet_part_map(source)
    plan = engine._build_write_plan(source, tuple(calculations), schema, parts)

    assert len(plan) == 1_000
    assert calls == 1

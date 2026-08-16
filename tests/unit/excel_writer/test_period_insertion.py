"""Synthetic direct-OOXML reporting-period insertion regressions."""

import zipfile
from copy import copy
from dataclasses import replace
from pathlib import Path
from xml.etree import ElementTree as ET

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from report_processor.admin_panel import reconciliation_target_measure
from report_processor.admin_panel.reconciliation_period import ReconciliationPeriodError
from report_processor.excel_writer import period_insertion
from report_processor.excel_writer.ooxml import inspect_cell, replace_cell_value
from report_processor.excel_writer.period_insertion import (
    _translate_formula,
    _verify_drawing_delta,
    build_period_insertion_plan,
    prepare_period_insertion,
    verify_period_insertion,
)


def _historical_book(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["B3"], sheet["C3"], sheet["D3"], sheet["E3"] = "1", "Этап 1", 1, "Монтаж"
    sheet.merge_cells("L1:M1")
    sheet["L1"] = "Документальная отчетность за весь период"
    sheet["L2"], sheet["M2"] = "Количество", "Стоимость"
    sheet["N3"] = "хвост"
    sheet["N4"] = "=L4+M4"
    workbook.save(path)


def _rewrite_zip_metadata(
    path: Path,
    *,
    flag_bits: int,
    external_attr: int,
    local_flag_bits: int | None = None,
    compression: int | None = None,
) -> None:
    """Generate a ZIP with deliberate matching local/central metadata."""

    with zipfile.ZipFile(path) as source:
        entries = [(copy(info), source.read(info.filename)) for info in source.infolist()]
        comment = source.comment
    with zipfile.ZipFile(path, "w") as rewritten:
        rewritten.comment = comment
        for info, payload in entries:
            if compression is not None:
                info.compress_type = compression
            rewritten.writestr(info, payload)
            info.flag_bits = flag_bits
            info.external_attr = external_attr
            position = rewritten.fp.tell()
            rewritten.fp.seek(info.header_offset + 6)
            written_local_flags = flag_bits if local_flag_bits is None else local_flag_bits
            rewritten.fp.write(written_local_flags.to_bytes(2, "little"))
            rewritten.fp.seek(position)


def _zip_metadata(path: Path) -> tuple[tuple[int, int, int], ...]:
    with zipfile.ZipFile(path) as archive, path.open("rb") as stream:
        result = []
        for info in archive.infolist():
            stream.seek(info.header_offset + 6)
            local_flags = int.from_bytes(stream.read(2), "little")
            result.append((local_flags, info.flag_bits, info.external_attr))
    return tuple(result)


def test_period_insertion_preserves_zero_external_attr_and_local_zip_flags(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _rewrite_zip_metadata(source, flag_bits=0x0006, external_attr=0)

    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    expected = tuple((0x0006, 0x0006, 0) for _ in _zip_metadata(source))
    assert _zip_metadata(source) == expected
    assert _zip_metadata(output) == expected


def test_period_insertion_rejects_unsupported_zip_flags_without_output(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _rewrite_zip_metadata(source, flag_bits=0x0010, external_attr=0)

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})

    assert not output.exists()


def test_period_insertion_rejects_mismatched_local_zip_flags_without_output(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _rewrite_zip_metadata(source, flag_bits=0x0006, local_flag_bits=0x0010, external_attr=0)

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})

    assert not output.exists()


def test_verifier_rejects_tampered_output_local_zip_flags(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _rewrite_zip_metadata(source, flag_bits=0x0006, external_attr=0)
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)
    _rewrite_zip_metadata(output, flag_bits=0x0006, local_flag_bits=0x0010, external_attr=0)

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_DELTA_INVALID"):
        verify_period_insertion(source, output, plan)


def test_period_insertion_rejects_lzma_without_eos_flag_without_output(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _rewrite_zip_metadata(
        source,
        flag_bits=0,
        external_attr=0,
        compression=zipfile.ZIP_LZMA,
    )

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})

    assert not output.exists()


@pytest.mark.parametrize("flag_bits", (0x0002, 0x0802))
def test_period_insertion_preserves_lzma_eos_zip_flags_and_zero_external_attr(
    tmp_path: Path, flag_bits: int
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _rewrite_zip_metadata(
        source,
        flag_bits=flag_bits,
        external_attr=0,
        compression=zipfile.ZIP_LZMA,
    )

    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    expected = tuple((flag_bits, flag_bits, 0) for _ in _zip_metadata(source))
    assert _zip_metadata(source) == expected
    assert _zip_metadata(output) == expected


def test_period_insertion_rejects_stored_deflate_option_flags_without_clobbering_output(
    tmp_path: Path,
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    output.write_bytes(b"user-sentinel")
    _rewrite_zip_metadata(
        source,
        flag_bits=0x0002,
        external_attr=0,
        compression=zipfile.ZIP_STORED,
    )

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})

    assert output.read_bytes() == b"user-sentinel"


def test_inserts_unmerged_period_columns_with_parent_row_labels(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)

    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepared = prepare_period_insertion(source, output, plan)

    assert prepared.output_sha256
    workbook = load_workbook(output, data_only=False)
    sheet = workbook["Отчёт"]
    assert sheet["N1"].value == "Август 2026 Количество"
    assert sheet["O1"].value == "Август 2026 Стоимость"
    assert sheet["N2"].value is None and sheet["O2"].value is None
    assert sheet["P3"].value == "хвост"
    assert sheet["P4"].value == "=L4+M4"


def test_prepared_historical_target_prefixes_remain_writable_by_the_byte_preserving_writer(
    tmp_path: Path,
) -> None:
    source, prepared = tmp_path / "historical.xlsx", tmp_path / "prepared.xlsx"
    _historical_book(source)

    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, prepared, plan)
    with zipfile.ZipFile(prepared) as archive:
        part = archive.read("xl/worksheets/sheet1.xml")

    _, quantity, _, _, _ = inspect_cell(part, "N3")
    updated = replace_cell_value(part, "N3", "7.50")

    assert quantity is None
    assert inspect_cell(updated, "N3")[1] == "7.50"
    assert inspect_cell(updated, "O3")[0] in updated


def test_plan_digest_is_deterministic_and_rejects_a_forged_digest(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _historical_book(source)

    first = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    second = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})

    assert first.plan_digest == second.plan_digest
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PLAN_INVALID"):
        replace(first, plan_digest="0" * 64)


def test_current_equivalent_calendar_identity_is_idempotent_and_conflict_fails(
    tmp_path: Path,
) -> None:
    source = tmp_path / "current.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["B3"], sheet["C3"], sheet["D3"], sheet["E3"] = "1", "Этап 1", 1, "Монтаж"
    sheet["L1"], sheet["M1"] = "08.2026", "2026-08"
    sheet["L2"], sheet["M2"] = "Количество", "Стоимость"
    workbook.save(source)

    assert build_period_insertion_plan(source, "2026-08", {"Отчёт": 3}).idempotent
    with pytest.raises(ReconciliationPeriodError, match="REPORTING_PERIOD_CONFLICT"):
        build_period_insertion_plan(source, "2026-09", {"Отчёт": 3})


def test_sparse_far_suffix_is_compact_and_part_of_the_immutable_plan(tmp_path: Path) -> None:
    source = tmp_path / "sparse.xlsx"
    _historical_book(source)
    workbook = load_workbook(source)
    workbook["Отчёт"]["Z1000000"] = "far suffix"
    workbook.save(source)

    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    (anchor,) = plan.anchors

    assert anchor.suffix_nonempty_count == 3
    assert anchor.suffix_first_coordinate == "N3"
    assert anchor.suffix_last_coordinate == "Z1000000"
    assert anchor.suffix_rightmost_coordinate == "Z1000000"
    assert len(anchor.suffix_coordinate_sha256) == 64


def _duplicate_l1_m1_merge(payload: bytes) -> bytes:
    root = ET.fromstring(payload)
    merges = root.find(_Q("mergeCells"))
    assert merges is not None
    merges.attrib["count"] = str(len(merges) + 1)
    ET.SubElement(merges, _Q("mergeCell"), {"ref": "L1:M1"})
    return ET.tostring(root)


def test_raw_duplicate_merge_rejects_build_prepare_and_verify_without_output(
    tmp_path: Path,
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    with zipfile.ZipFile(source) as archive:
        source_sheet = archive.read("xl/worksheets/sheet1.xml")
    _add_zip_members(source, {"xl/worksheets/sheet1.xml": _duplicate_l1_m1_merge(source_sheet)})

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    with pytest.raises(ReconciliationPeriodError):
        prepare_period_insertion(source, output, plan)
    assert not output.exists()

    clean = tmp_path / "clean.xlsx"
    _historical_book(clean)
    clean_plan = build_period_insertion_plan(clean, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(clean, output, clean_plan)
    with zipfile.ZipFile(output) as archive:
        output_sheet = archive.read("xl/worksheets/sheet1.xml")
    _add_zip_members(output, {"xl/worksheets/sheet1.xml": _duplicate_l1_m1_merge(output_sheet)})
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        verify_period_insertion(clean, output, clean_plan)


@pytest.mark.parametrize("count", (1_000, 2_000, 4_000))
def test_raw_merge_sweep_has_bounded_subquadratic_work(count: int) -> None:
    references = "".join(
        f'<mergeCell ref="{get_column_letter(index * 4 + 1)}1:'
        f'{get_column_letter(index * 4 + 2)}1"/>'
        for index in range(count)
    )
    payload = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<mergeCells count="{count}">{references}</mergeCells>'
        "</worksheet>"
    ).encode()
    operations: list[int] = []

    assert len(period_insertion._raw_sheet_merges(payload, operations)) == count
    assert len(operations) < count * 100


def test_raw_merge_count_limit_rejects_before_overlap_work() -> None:
    count = period_insertion._MAX_RAW_MERGES + 1
    references = "".join(
        f'<mergeCell ref="{get_column_letter(index * 2 + 1)}1:'
        f'{get_column_letter(index * 2 + 2)}1"/>'
        for index in range(count)
    )
    payload = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<mergeCells count="{count}">{references}</mergeCells>'
        "</worksheet>"
    ).encode()
    operations: list[int] = []

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        period_insertion._raw_sheet_merges(payload, operations)

    assert operations == []


def _raw_merges_payload(declared_count: str, physical_count: int) -> bytes:
    references = "".join(
        f'<mergeCell ref="A{index}:B{index}"/>' for index in range(1, physical_count + 1)
    )
    return (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<mergeCells count="{declared_count}">{references}</mergeCells>'
        "</worksheet>"
    ).encode()


@pytest.mark.parametrize(
    ("declared_count", "physical_count"),
    (("0", 0), ("1", 1), ("64", 64), ("٤٠٩٦", 4_096)),
)
def test_raw_merge_count_uses_numeric_limit(declared_count: str, physical_count: int) -> None:
    assert (
        len(period_insertion._raw_sheet_merges(_raw_merges_payload(declared_count, physical_count)))
        == physical_count
    )


def test_raw_merge_count_rejects_numeric_value_above_limit() -> None:
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        period_insertion._raw_sheet_merges(_raw_merges_payload("4097", 0))


@pytest.mark.parametrize(("declared_count", "physical_count"), (("0", 1), ("1", 0)))
def test_raw_merge_count_rejects_declared_physical_mismatch(
    declared_count: str, physical_count: int
) -> None:
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PACKAGE_INVALID"):
        period_insertion._raw_sheet_merges(_raw_merges_payload(declared_count, physical_count))


def test_planner_rejects_wide_header_merge_before_output(tmp_path: Path) -> None:
    source, output = tmp_path / "wide.xlsx", tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet.merge_cells("A1:XFD1")
    sheet.merge_cells("L50:M50")
    sheet["L50"] = "Документальная отчетность за весь период"
    sheet["L51"], sheet["M51"], sheet["N52"] = "Количество", "Стоимость", "suffix"
    workbook.save(source)
    workbook.close()

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_ANCHOR_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 52})

    assert not output.exists()


def test_suffix_rightmost_is_column_major_while_bounds_are_row_major(tmp_path: Path) -> None:
    source = tmp_path / "crossed-axis.xlsx"
    _historical_book(source)
    workbook = load_workbook(source)
    sheet = workbook["Отчёт"]
    sheet["Z4"] = "rightmost"
    sheet["N100"] = "last"
    workbook.save(source)

    (anchor,) = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3}).anchors

    assert anchor.suffix_first_coordinate == "N3"
    assert anchor.suffix_last_coordinate == "N100"
    assert anchor.suffix_rightmost_coordinate == "Z4"


def test_exported_verifier_rejects_forged_redigested_plan_evidence(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)
    (anchor,) = plan.anchors

    forged_plans = (
        replace(
            plan,
            anchors=(replace(anchor, historical_parent_label="поддельная историческая подпись"),),
            plan_digest="",
        ),
        replace(
            plan,
            anchors=(replace(anchor, suffix_coordinate_sha256="0" * 64),),
            plan_digest="",
        ),
        replace(plan, selected_detail_rows=(("Отчёт", 4),), plan_digest=""),
    )

    verify_period_insertion(source, output, plan)
    for forged in forged_plans:
        assert forged.plan_digest != plan.plan_digest
        with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_PLAN_INVALID"):
            verify_period_insertion(source, output, forged)


def test_empty_suffix_fails_closed(tmp_path: Path) -> None:
    source = tmp_path / "empty-suffix.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["B3"], sheet["C3"], sheet["D3"], sheet["E3"] = "1", "Этап 1", 1, "Монтаж"
    sheet.merge_cells("L1:M1")
    sheet["L1"] = "Документальная отчетность за весь период"
    sheet["L2"], sheet["M2"] = "Количество", "Стоимость"
    workbook.save(source)

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_ANCHOR_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})


def test_suffix_over_limit_fails_closed(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source = tmp_path / "over-limit.xlsx"
    _historical_book(source)
    monkeypatch.setattr(reconciliation_target_measure, "_MAX_SUFFIX_CELLS", 1)

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_ANCHOR_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})


def test_materialized_cell_inspection_limit_includes_empty_left_cells(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "inspection-limit.xlsx"
    _historical_book(source)
    assert build_period_insertion_plan(source, "2026-08", {"Отчёт": 3}).anchors
    workbook = load_workbook(source)
    sheet = workbook["Отчёт"]
    for column in range(1, 11):
        sheet.cell(20, column).fill = PatternFill("solid", fgColor="FFFFFF")
    workbook.save(source)
    monkeypatch.setattr(reconciliation_target_measure, "_MAX_SUFFIX_INSPECTED_CELLS", 15)

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_ANCHOR_INVALID"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})


def test_formula_translation_preserves_quoted_coordinate_and_translates_local_range() -> None:
    assert _translate_formula('SUM(L4:N4)+"N4"', 13) == 'SUM(L4:P4)+"N4"'
    assert _translate_formula('="N""4"+N4', 13) == '="N""4"+P4'


@pytest.mark.parametrize("formula", ("SUM(N:N)", "SUM(4:4)", '"N:N"+N4'))
def test_formula_translation_rejects_whole_column_or_row_operands(formula: str) -> None:
    if formula.startswith('"'):
        assert _translate_formula(formula, 13) == '"N:N"+P4'
    else:
        with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_UNSUPPORTED_FEATURE"):
            _translate_formula(formula, 13)


@pytest.mark.parametrize(
    "formula",
    ("'Другой лист'!N4", "Book.xlsx!N4", "MyNamedRange+N4", 'INDIRECT("N4")'),
)
def test_formula_translation_rejects_nonlocal_or_named_operands(formula: str) -> None:
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_UNSUPPORTED_FEATURE"):
        _translate_formula(formula, 13)


def test_existing_output_sentinel_is_never_replaced(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    output.write_bytes(b"user-sentinel")
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_OUTPUT_EXISTS"):
        prepare_period_insertion(source, output, plan)

    assert output.read_bytes() == b"user-sentinel"


@pytest.mark.parametrize(
    "needle,replacement",
    (
        ('r="N1"', 'r="Q1"'),  # inserted-header coordinate / extra gap
        ('r="N2"', 'r="Q2"'),  # missing inserted blank
        ('ref="B1:P4"', 'ref="B1:Q4"'),  # dimension
        ('ref="L1:M1"', 'ref="L1:N1"'),  # merge
    ),
)
def test_independent_verifier_rejects_tampered_worksheet_delta(
    tmp_path: Path, needle: str, replacement: str
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    import zipfile

    with zipfile.ZipFile(output) as original:
        payloads = {info.filename: original.read(info.filename) for info in original.infolist()}
        infos = original.infolist()
    part = "xl/worksheets/sheet1.xml"
    assert needle.encode() in payloads[part]
    payloads[part] = payloads[part].replace(needle.encode(), replacement.encode(), 1)
    with zipfile.ZipFile(output, "w") as replacement_zip:
        for info in infos:
            replacement_zip.writestr(info, payloads[info.filename])

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_DELTA_INVALID"):
        verify_period_insertion(source, output, plan)


@pytest.mark.parametrize("corruption", ("follower_text", "bare_duplicate"))
def test_verifier_shared_formula_checks_are_independent_of_forward_preflight(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, corruption: str
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _add_shared_formula_group(source, "0", "A4:B4", "A1", ("A4", "B4"))
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    monkeypatch.setattr(period_insertion, "_shared_formula_topology", lambda *_args: ())
    verify_period_insertion(source, output, plan)

    def corrupt_shared_follower(root: ET.Element) -> None:
        if corruption == "bare_duplicate":
            row = next(row for row in root.iter(_Q("row")) if row.attrib.get("r") == "4")
            ET.SubElement(row, _Q("c"), {"r": "B4"})
            return
        formula = next(
            cell.find(_Q("f")) for cell in root.iter(_Q("c")) if cell.attrib.get("r") == "B4"
        )
        assert formula is not None
        formula.text = "A1"

    _replace_worksheet(source, corrupt_shared_follower)
    _replace_worksheet(output, corrupt_shared_follower)
    candidate_plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_DELTA_INVALID"):
        verify_period_insertion(source, output, candidate_plan)


def _rewrite_zip_member(path: Path, member: str, mutate) -> None:
    with zipfile.ZipFile(path) as archive:
        entries = [(info, archive.read(info.filename)) for info in archive.infolist()]
        comment = archive.comment
    with zipfile.ZipFile(path, "w") as archive:
        archive.comment = comment
        for info, payload in entries:
            if info.filename == member:
                mutate(info, payload)
            archive.writestr(info, payload)


def _add_zip_members(path: Path, replacements: dict[str, bytes]) -> None:
    with zipfile.ZipFile(path) as archive:
        entries = [(info, archive.read(info.filename)) for info in archive.infolist()]
        comment = archive.comment
    seen = {info.filename for info, _payload in entries}
    with zipfile.ZipFile(path, "w") as archive:
        archive.comment = comment
        for info, payload in entries:
            archive.writestr(info, replacements.pop(info.filename, payload))
        for name, payload in replacements.items():
            assert name not in seen
            archive.writestr(name, payload)


_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_Q = lambda name: f"{{{_MAIN}}}{name}"  # noqa: E731


def _replace_worksheet(path: Path, mutate) -> None:
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    mutate(root)
    _add_zip_members(path, {"xl/worksheets/sheet1.xml": ET.tostring(root)})


def _add_shared_formula_group(
    path: Path,
    si: str,
    reference: str,
    text: str,
    members: tuple[str, ...],
) -> None:
    """Inject a minimal OOXML shared group into the synthetic workbook only."""

    def mutate(root: ET.Element) -> None:
        sheet_data = root.find(_Q("sheetData"))
        assert sheet_data is not None
        rows = {int(row.attrib["r"]): row for row in sheet_data.findall(_Q("row"))}
        for coordinate in members:
            row_number = int("".join(character for character in coordinate if character.isdigit()))
            row = rows.get(row_number)
            if row is None:
                row = ET.SubElement(sheet_data, _Q("row"), {"r": str(row_number)})
                rows[row_number] = row
            for cell in tuple(row.findall(_Q("c"))):
                if cell.attrib.get("r") == coordinate:
                    row.remove(cell)
            cell = ET.SubElement(row, _Q("c"), {"r": coordinate})
            attributes = {"t": "shared", "si": si}
            formula = ET.SubElement(cell, _Q("f"), attributes)
            if coordinate == members[0]:
                formula.attrib["ref"] = reference
                formula.text = text
            row[:] = sorted(row, key=lambda node: node.attrib.get("r", ""))

    _replace_worksheet(path, mutate)


def _shared_formula_state(
    path: Path,
) -> tuple[tuple[str, tuple[tuple[str, str], ...], str | None], ...]:
    with zipfile.ZipFile(path) as archive:
        root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    return tuple(
        sorted(
            (
                cell.attrib["r"],
                tuple(sorted(formula.attrib.items())),
                formula.text,
            )
            for cell in root.iter(_Q("c"))
            for formula in cell.findall(_Q("f"))
            if formula.attrib.get("t") == "shared"
        )
    )


def _source_with_affected_workbook_part(path: Path) -> None:
    _historical_book(path)
    with zipfile.ZipFile(path) as archive:
        workbook = archive.read("xl/workbook.xml")
    workbook = workbook.replace(
        b"</workbook>",
        b'<definedNames><definedName name="_xlnm.Print_Area">'
        b"'&#1054;&#1090;&#1095;&#1105;&#1090;'!$N:$N</definedName></definedNames></workbook>",
    )
    _add_zip_members(path, {"xl/workbook.xml": workbook})


def test_verifier_rejects_tampered_workbook_payload_attribute(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _source_with_affected_workbook_part(source)
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    def mutate(_info: zipfile.ZipInfo, payload: bytes) -> None:
        nonlocal tampered
        tampered = payload.replace(b"<s:workbookPr", b'<s:workbookPr date1904="1"', 1)

    tampered = b""
    _rewrite_zip_member(output, "xl/workbook.xml", mutate)
    # `_rewrite_zip_member` writes the local payload variable, so replace once
    # more with the intentionally altered workbook payload.
    _add_zip_members(output, {"xl/workbook.xml": tampered})
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_DELTA_INVALID"):
        verify_period_insertion(source, output, plan)


@pytest.mark.parametrize(
    "mutation",
    (
        lambda payload: payload.replace(b'<s:c r="P4" i="1"', b'<s:c r="P4" i="1" extra="1"'),
        lambda payload: payload.replace(b"</s:calcChain>", b'<s:c r="P4" i="1"/></s:calcChain>'),
    ),
)
def test_verifier_rejects_calc_chain_extra_attribute_or_node(tmp_path: Path, mutation) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _add_zip_members(
        source,
        {
            "xl/calcChain.xml": (
                b'<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                b'<c r="N4" i="1"/></calcChain>'
            )
        },
    )
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)
    with zipfile.ZipFile(output) as archive:
        payload = archive.read("xl/calcChain.xml")
    _add_zip_members(output, {"xl/calcChain.xml": mutation(payload)})
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_DELTA_INVALID"):
        verify_period_insertion(source, output, plan)


def test_drawing_verifier_rejects_payload_attribute_tamper(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _historical_book(source)
    with zipfile.ZipFile(source) as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml")
    sheet = sheet.replace(
        b"<worksheet ",
        b'<worksheet xmlns:r="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships" ',
        1,
    ).replace(b"</worksheet>", b'<drawing r:id="rIdDrawing"/></worksheet>')
    drawing = (
        b'<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing">'
        b"<xdr:twoCellAnchor><xdr:from><xdr:col>13</xdr:col></xdr:from>"
        b"<xdr:to><xdr:col>14</xdr:col></xdr:to></xdr:twoCellAnchor></xdr:wsDr>"
    )
    rels = (
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        b'<Relationship Id="rIdDrawing" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'
        b"</Relationships>"
    )
    _add_zip_members(
        source,
        {
            "xl/worksheets/sheet1.xml": sheet,
            "xl/worksheets/_rels/sheet1.xml.rels": rels,
            "xl/drawings/drawing1.xml": drawing,
        },
    )
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    (anchor,) = plan.anchors
    tampered = drawing.replace(b"<xdr:to>", b'<xdr:to payload="tampered">')
    with (
        zipfile.ZipFile(source) as archive,
        pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_DELTA_INVALID"),
    ):
        _verify_drawing_delta(
            drawing,
            tampered,
            archive,
            "xl/drawings/drawing1.xml",
            {anchor.sheet_name: anchor},
        )


@pytest.mark.parametrize("attribute,value", (("internal_attr", 1), ("create_version", 63)))
def test_verifier_rejects_tampered_zipinfo_metadata(
    tmp_path: Path, attribute: str, value: int
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    def mutate(info: zipfile.ZipInfo, _payload: bytes) -> None:
        setattr(info, attribute, value)

    _rewrite_zip_member(output, "xl/worksheets/sheet1.xml", mutate)
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_DELTA_INVALID"):
        verify_period_insertion(source, output, plan)


def test_wholly_left_comment_and_external_hyperlink_are_preserved(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    workbook = load_workbook(source)
    sheet = workbook["Отчёт"]
    sheet["A1"].comment = Comment("left", "tester")
    sheet["B1"].hyperlink = "https://example.test/left"
    workbook.save(source)
    with zipfile.ZipFile(source) as archive:
        related = {
            info.filename: archive.read(info.filename)
            for info in archive.infolist()
            if (
                "comments" in info.filename
                or "vmlDrawing" in info.filename
                or info.filename.endswith("sheet1.xml.rels")
            )
        }
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)
    with zipfile.ZipFile(output) as archive:
        assert {name: archive.read(name) for name in related} == related


def _add_threaded_comments(path: Path, reference: str = "A1", *, malformed: bool = False) -> None:
    threaded_part = "xl/threadedComments/threadedComment1.xml"
    person_part = "xl/persons/person.xml"
    rels_part = "xl/worksheets/_rels/sheet1.xml.rels"
    threaded = (
        b"<ThreadedComments"
        b' xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments">'
        b'<threadedComment ref="'
        + reference.encode()
        + b'" personId="person-1" id="comment-1" dT="2026-08-01T10:00:00Z">'
        b'<text xml:space="preserve"> left</text></threadedComment>'
        b"</ThreadedComments>"
    )
    if malformed:
        threaded = threaded[:-1]
    relationship = (
        b'<Relationship Id="rIdThreadedComment" Type="http://schemas.microsoft.com/'
        b'office/2017/10/relationships/threadedComment" '
        b'Target="../threadedComments/threadedComment1.xml"/>'
    )
    person_relationship = (
        b'<Relationship Id="rIdPersons" Type="http://schemas.microsoft.com/office/2017/10/'
        b'relationships/person" Target="persons/person.xml"/>'
    )
    with zipfile.ZipFile(path) as archive:
        rels = (
            archive.read(rels_part)
            if rels_part in archive.namelist()
            else b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b"</Relationships>"
        )
        workbook_rels = archive.read("xl/_rels/workbook.xml.rels")
    rels = rels.replace(b"</Relationships>", relationship + b"</Relationships>")
    workbook_rels = workbook_rels.replace(
        b"</Relationships>", person_relationship + b"</Relationships>"
    )
    _add_zip_members(
        path,
        {
            rels_part: rels,
            "xl/_rels/workbook.xml.rels": workbook_rels,
            threaded_part: threaded,
            person_part: (
                b'<personList xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments">'
                b'<person id="person-1" displayName="Tester" userId="tester" providerId="None"/>'
                b"</personList>"
            ),
        },
    )


def test_wholly_left_threaded_comments_and_person_part_are_preserved(tmp_path: Path) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    workbook = load_workbook(source)
    workbook["Отчёт"]["B1"].comment = Comment("classic left", "tester")
    workbook.save(source)
    _add_threaded_comments(source)
    with zipfile.ZipFile(source) as archive:
        preserved = {
            name: archive.read(name)
            for name in archive.namelist()
            if "comments" in name
            or "vmlDrawing" in name
            or name
            in {
                "xl/threadedComments/threadedComment1.xml",
                "xl/persons/person.xml",
                "xl/worksheets/_rels/sheet1.xml.rels",
                "xl/_rels/workbook.xml.rels",
            }
        }

    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    with zipfile.ZipFile(output) as archive:
        assert {name: archive.read(name) for name in preserved} == preserved
    verify_period_insertion(source, output, plan)


@pytest.mark.parametrize(
    "case",
    (
        "right",
        "malformed",
        "missing",
        "duplicate",
        "escape",
        "duplicate_rel_id",
        "attribute",
        "person",
        "text_attribute",
        "person_attribute",
        "timestamp",
    ),
)
def test_unsafe_threaded_comments_fail_closed_without_output(tmp_path: Path, case: str) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    output.write_bytes(b"user-sentinel")
    _add_threaded_comments(source, "N1" if case == "right" else "A1", malformed=case == "malformed")
    if case == "missing":
        _add_zip_members(
            source,
            {
                "xl/worksheets/_rels/sheet1.xml.rels": b"<Relationships "
                b'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                b'<Relationship Id="rIdThreadedComment" Type="http://schemas.microsoft.com/'
                b'office/2017/10/relationships/threadedComment" '
                b'Target="../threadedComments/missing.xml"/>'
                b"</Relationships>"
            },
        )
    elif case == "duplicate":
        with zipfile.ZipFile(source) as archive:
            rels = archive.read("xl/worksheets/_rels/sheet1.xml.rels")
        _add_zip_members(
            source,
            {
                "xl/worksheets/_rels/sheet1.xml.rels": rels.replace(
                    b"</Relationships>",
                    b'<Relationship Id="rIdThreadedComment2" Type="http://schemas.microsoft.com/'
                    b'office/2017/10/relationships/threadedComment" '
                    b'Target="../threadedComments/threadedComment1.xml"/>'
                    b"</Relationships>",
                )
            },
        )
    elif case == "escape":
        with zipfile.ZipFile(source) as archive:
            rels = archive.read("xl/worksheets/_rels/sheet1.xml.rels")
        _add_zip_members(
            source,
            {
                "xl/worksheets/_rels/sheet1.xml.rels": rels.replace(
                    b"../threadedComments/threadedComment1.xml",
                    b"../../threadedComments/threadedComment1.xml",
                )
            },
        )
    elif case == "duplicate_rel_id":
        with zipfile.ZipFile(source) as archive:
            rels = archive.read("xl/worksheets/_rels/sheet1.xml.rels")
        _add_zip_members(
            source,
            {
                "xl/worksheets/_rels/sheet1.xml.rels": rels.replace(
                    b"</Relationships>",
                    b'<Relationship Id="rIdThreadedComment" Type="http://schemas.openxmlformats.org/'
                    b'officeDocument/2006/relationships/hyperlink" Target="https://example.test" '
                    b'TargetMode="External"/>'
                    b"</Relationships>",
                )
            },
        )
    elif case == "attribute":
        with zipfile.ZipFile(source) as archive:
            threaded = archive.read("xl/threadedComments/threadedComment1.xml")
        _add_zip_members(
            source,
            {
                "xl/threadedComments/threadedComment1.xml": threaded.replace(
                    b' dT="2026-08-01T10:00:00Z"',
                    b' dT="2026-08-01T10:00:00Z" unexpected="1"',
                )
            },
        )
    elif case == "text_attribute":
        with zipfile.ZipFile(source) as archive:
            threaded = archive.read("xl/threadedComments/threadedComment1.xml")
        _add_zip_members(
            source,
            {
                "xl/threadedComments/threadedComment1.xml": threaded.replace(
                    b'xml:space="preserve"', b'xml:space="default"'
                )
            },
        )
    elif case == "person_attribute":
        with zipfile.ZipFile(source) as archive:
            people = archive.read("xl/persons/person.xml")
        _add_zip_members(
            source,
            {"xl/persons/person.xml": people.replace(b"/>", b' unexpected="1"/>')},
        )
    elif case == "timestamp":
        with zipfile.ZipFile(source) as archive:
            threaded = archive.read("xl/threadedComments/threadedComment1.xml")
        _add_zip_members(
            source,
            {
                "xl/threadedComments/threadedComment1.xml": threaded.replace(
                    b"2026-08-01T10:00:00Z", b"not-a-timestamp"
                )
            },
        )
    elif case == "person":
        with zipfile.ZipFile(source) as archive:
            people = archive.read("xl/persons/person.xml")
        _add_zip_members(
            source, {"xl/persons/person.xml": people.replace(b"person-1", b"person-2")}
        )

    with pytest.raises(ReconciliationPeriodError):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    assert output.read_bytes() == b"user-sentinel"


@pytest.mark.parametrize("kind", ("comment", "hyperlink"))
def test_right_comment_or_external_hyperlink_is_rejected(tmp_path: Path, kind: str) -> None:
    source = tmp_path / "source.xlsx"
    _historical_book(source)
    workbook = load_workbook(source)
    sheet = workbook["Отчёт"]
    if kind == "comment":
        sheet["N1"].comment = Comment("right", "tester")
    else:
        sheet["N1"].hyperlink = "https://example.test/right"
    workbook.save(source)
    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_UNSUPPORTED_FEATURE"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})


@pytest.mark.parametrize(
    "groups",
    (
        (("0", "A4", "A1", ("A4",)),),
        (("000", "A4:B4", "A1", ("A4", "B4")),),
        (("0", "A4:B5", "SUM(A1:B1)", ("A4", "B4", "A5", "B5")),),
        (
            ("0", "A4:B4", "A1", ("A4", "B4")),
            ("1", "C4:C5", "B1", ("C4", "C5")),
        ),
    ),
)
def test_wholly_left_complete_shared_formula_groups_are_preserved(
    tmp_path: Path, groups: tuple[tuple[str, str, str, tuple[str, ...]], ...]
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    for si, reference, text, members in groups:
        _add_shared_formula_group(source, si, reference, text, members)
    before = _shared_formula_state(source)

    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    assert _shared_formula_state(output) == before
    verify_period_insertion(source, output, plan)


def test_shared_formula_calc_chain_keeps_left_members_and_shifts_right_entries(
    tmp_path: Path,
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _add_shared_formula_group(source, "0", "A4:B4", "A1", ("A4", "B4"))
    _add_zip_members(
        source,
        {
            "xl/calcChain.xml": (
                b'<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                b'<c r="A4" i="1"/><c r="N4"/></calcChain>'
            )
        },
    )

    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    with zipfile.ZipFile(output) as archive:
        calc_chain = archive.read("xl/calcChain.xml")
    assert b'r="A4"' in calc_chain
    assert b'r="P4"' in calc_chain


@pytest.mark.parametrize(
    "case",
    (
        "missing_member",
        "extra_member",
        "overlap",
        "bare_duplicate_member",
        "duplicate_si",
        "mismatched_si",
        "negative_si",
        "uint32_overflow",
        "oversized_si",
        "blank_anchor",
        "follower_text",
        "follower_ref",
        "array",
        "data_table",
        "reversed_ref",
        "non_top_left_anchor",
        "affected_operand",
        "affected_ref",
        "dynamic",
        "huge_ref",
    ),
)
def test_invalid_shared_formula_groups_fail_closed_without_clobbering_output(
    tmp_path: Path, case: str
) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    output.write_bytes(b"user-sentinel")
    if case == "missing_member":
        _add_shared_formula_group(source, "0", "A4:B4", "A1", ("A4",))
    elif case == "extra_member":
        _add_shared_formula_group(source, "0", "A4", "A1", ("A4", "B4"))
    elif case == "overlap":
        _add_shared_formula_group(source, "0", "A4:B4", "A1", ("A4", "B4"))

        def add_overlapping_group(root: ET.Element) -> None:
            row = next(row for row in root.iter(_Q("row")) if row.attrib.get("r") == "4")
            duplicate = ET.SubElement(row, _Q("c"), {"r": "B4"})
            anchor_formula = ET.SubElement(
                duplicate, _Q("f"), {"t": "shared", "si": "1", "ref": "B4:C4"}
            )
            anchor_formula.text = "A1"
            follower = ET.SubElement(row, _Q("c"), {"r": "C4"})
            ET.SubElement(follower, _Q("f"), {"t": "shared", "si": "1"})

        _replace_worksheet(source, add_overlapping_group)
    elif case == "bare_duplicate_member":
        _add_shared_formula_group(source, "0", "A4:B4", "A1", ("A4", "B4"))

        def add_bare_duplicate(root: ET.Element) -> None:
            row = next(row for row in root.iter(_Q("row")) if row.attrib.get("r") == "4")
            ET.SubElement(row, _Q("c"), {"r": "B4"})

        _replace_worksheet(source, add_bare_duplicate)
    elif case == "duplicate_si":
        _add_shared_formula_group(source, "0", "A4", "A1", ("A4",))
        _add_shared_formula_group(source, "0", "B4", "A1", ("B4",))
    elif case == "negative_si":
        _add_shared_formula_group(source, "-1", "A4", "A1", ("A4",))
    elif case == "uint32_overflow":
        _add_shared_formula_group(source, "4294967296", "A4", "A1", ("A4",))
    elif case == "oversized_si":
        _add_shared_formula_group(source, "9" * 512, "A4", "A1", ("A4",))
    elif case == "blank_anchor":
        _add_shared_formula_group(source, "0", "A4", "   ", ("A4",))
    elif case == "affected_ref":
        _add_shared_formula_group(source, "0", "M4:N4", "M1", ("M4", "N4"))
    elif case == "huge_ref":
        _add_shared_formula_group(source, "0", "A4:XFD1048576", "A1", ("A4",))
    else:
        _add_shared_formula_group(
            source,
            "0",
            "A4:B4",
            "N4" if case == "affected_operand" else "@A1" if case == "dynamic" else "A1",
            ("B4", "A4") if case == "non_top_left_anchor" else ("A4", "B4"),
        )

        def mutate(root: ET.Element) -> None:
            formulas = {
                cell.attrib["r"]: cell.find(_Q("f"))
                for cell in root.iter(_Q("c"))
                if cell.find(_Q("f")) is not None
            }
            if case == "follower_text":
                assert formulas["B4"] is not None
                formulas["B4"].text = "A1"
            elif case == "follower_ref":
                assert formulas["B4"] is not None
                formulas["B4"].attrib["ref"] = "A4:B4"
            elif case == "mismatched_si":
                assert formulas["B4"] is not None
                formulas["B4"].attrib["si"] = "00"
            elif case == "array":
                assert formulas["A4"] is not None
                formulas["A4"].attrib["t"] = "array"
            elif case == "data_table":
                assert formulas["A4"] is not None
                formulas["A4"].attrib["t"] = "dataTable"
            elif case == "reversed_ref":
                assert formulas["A4"] is not None
                formulas["A4"].attrib["ref"] = "B4:A4"

        _replace_worksheet(source, mutate)

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_UNSUPPORTED_FEATURE"):
        build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    assert output.read_bytes() == b"user-sentinel"


@pytest.mark.parametrize("tamper", ("text", "attribute", "member"))
def test_verifier_rejects_shared_formula_topology_tampering(tmp_path: Path, tamper: str) -> None:
    source, output = tmp_path / "source.xlsx", tmp_path / "output.xlsx"
    _historical_book(source)
    _add_shared_formula_group(source, "0", "A4:B4", "A1", ("A4", "B4"))
    plan = build_period_insertion_plan(source, "2026-08", {"Отчёт": 3})
    prepare_period_insertion(source, output, plan)

    def mutate(root: ET.Element) -> None:
        cells = {cell.attrib["r"]: cell for cell in root.iter(_Q("c"))}
        if tamper == "text":
            formula = cells["A4"].find(_Q("f"))
            assert formula is not None
            formula.text = "B1"
        elif tamper == "attribute":
            formula = cells["B4"].find(_Q("f"))
            assert formula is not None
            formula.attrib["ca"] = "1"
        else:
            cells["B4"].attrib["r"] = "C4"

    _replace_worksheet(output, mutate)

    with pytest.raises(ReconciliationPeriodError, match="PERIOD_INSERTION_DELTA_INVALID"):
        verify_period_insertion(source, output, plan)

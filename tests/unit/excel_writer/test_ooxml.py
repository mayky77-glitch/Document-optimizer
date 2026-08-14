"""Low-level no-clobber and targeted XML preservation tests."""

from __future__ import annotations

import struct
import threading
import tracemalloc
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from report_processor.excel_writer import ExcelWriterIntegrityError, ExcelWriterSafetyError, ooxml
from report_processor.excel_writer.ooxml import (
    formula_count,
    inspect_cell,
    inspect_index_cell,
    materialize_formula_cells,
    numeric_formula_values,
    publish_no_clobber,
    replace_cell_value,
    verify_temp_package,
    worksheet_index,
    worksheet_part_map,
    write_temp_package,
)

_SHEET_XML = (
    b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    b'<sheetData><row r="30"><c r="D30" s="5"><v>12.50</v></c>'
    b'<c r="E30" s="7"><f>SUM(D30)</f><v>12.50</v></c></row></sheetData></worksheet>'
)


def test_targeted_cell_replacement_preserves_style_formula_and_other_cell_bytes() -> None:
    updated = replace_cell_value(_SHEET_XML, "D30", "-0.250")

    cell, lexeme, is_formula, has_style, cell_type = inspect_cell(updated, "D30")
    assert b's="5"' in cell
    assert lexeme == "-0.250"
    assert not is_formula
    assert has_style
    assert cell_type in {None, "n"}
    assert _SHEET_XML[_SHEET_XML.index(b'<c r="E30"') :] == updated[updated.index(b'<c r="E30"') :]


def test_self_closing_target_cell_does_not_capture_the_next_cell_value() -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row r="30"><c r="D30" s="5"/>'
        b'<c r="E30" s="7"><v>40</v></c></row></sheetData></worksheet>'
    )

    cell, lexeme, is_formula, has_style, cell_type = inspect_cell(xml, "D30")
    updated = replace_cell_value(xml, "D30", "0.00")

    assert cell == b'<c r="D30" s="5"/>'
    assert lexeme is None and not is_formula and has_style and cell_type is None
    assert b'<c r="D30" s="5"><v>0.00</v></c>' in updated
    assert updated[updated.index(b'<c r="E30"') :] == xml[xml.index(b'<c r="E30"') :]


def test_formula_and_missing_cells_remain_rejectable() -> None:
    _, _, is_formula, _, _ = inspect_cell(_SHEET_XML, "E30")
    assert is_formula
    with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
        inspect_cell(_SHEET_XML, "D31")


def test_publish_no_clobber_keeps_existing_output_and_removes_temp(tmp_path: Path) -> None:
    temp_path = tmp_path / ".candidate.xlsx.tmp"
    output_path = tmp_path / "published.xlsx"
    temp_path.write_bytes(b"new")
    output_path.write_bytes(b"existing")

    with pytest.raises(ExcelWriterSafetyError, match="OUTPUT_EXISTS"):
        publish_no_clobber(temp_path, output_path)

    assert output_path.read_bytes() == b"existing"
    assert not temp_path.exists()


def test_formula_cells_become_numeric_literals_including_shared_formula_cells() -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row r="30"><c r="D30" s="5"><v>0</v></c>'
        b'<c r="E30" s="7"><f t="shared" si="0">D30*2</f><v>0</v></c>'
        b'<c r="F30" s="7"><f t="shared" si="0"/><v>0</v></c>'
        b"</row></sheetData></worksheet>"
    )

    values = numeric_formula_values(xml, ("E30", "F30"))
    updated = materialize_formula_cells(xml, values)

    assert formula_count(xml) == 2
    assert formula_count(updated) == 0
    assert values == {"E30": "0", "F30": "0"}
    for coordinate in values:
        _, value, is_formula, has_style, cell_type = inspect_cell(updated, coordinate)
        assert value == "0"
        assert not is_formula
        assert has_style
        assert cell_type in {None, "n"}


@pytest.mark.parametrize("prefix", (b"", b"s:", b"ns0:", b"arbitrary:"))
def test_namespace_qualified_cells_preserve_their_exact_qnames(prefix: bytes) -> None:
    declaration = (
        b'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        if not prefix
        else b"xmlns:"
        + prefix[:-1]
        + b'="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
    )
    xml = (
        b"<"
        + prefix
        + b"worksheet "
        + declaration
        + b"><"
        + prefix
        + b"sheetData><"
        + prefix
        + b'row r="1"><'
        + prefix
        + b'c r="D1"/><'
        + prefix
        + b'c r="E1"><'
        + prefix
        + b"f>D1*2</"
        + prefix
        + b"f><"
        + prefix
        + b"v/></"
        + prefix
        + b"c></"
        + prefix
        + b"row></"
        + prefix
        + b"sheetData></"
        + prefix
        + b"worksheet>"
    )

    updated = replace_cell_value(xml, "D1", "1.25")
    materialized = materialize_formula_cells(updated, {"E1": "2.50"})

    assert (
        b"<" + prefix + b'c r="D1"><' + prefix + b"v>1.25</" + prefix + b"v></" + prefix + b"c>"
        in updated
    )
    assert (
        b"<" + prefix + b'c r="E1"><' + prefix + b"v>2.50</" + prefix + b"v></" + prefix + b"c>"
        in materialized
    )
    assert formula_count(materialized) == 0
    assert inspect_cell(materialized, "E1")[1] == "2.50"


def test_foreign_lookalikes_are_ignored_and_namespace_calls_are_request_local() -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'xmlns:foreign="urn:foreign"><sheetData><row r="1">'
        b'<foreign:c r="D1"><foreign:f>1</foreign:f><foreign:v>99</foreign:v></foreign:c>'
        b'<c r="E1"><v>2</v></c></row></sheetData></worksheet>'
    )
    before = dict(ET._namespace_map)
    errors: list[Exception] = []

    def replace() -> None:
        try:
            assert inspect_cell(xml, "E1")[1] == "2"
            assert replace_cell_value(xml, "E1", "3").count(b"foreign:c") == 2
        except Exception as error:  # pragma: no cover - asserted below
            errors.append(error)

    threads = [threading.Thread(target=replace) for _ in range(8)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()

    assert not errors
    assert formula_count(xml) == 0
    with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
        inspect_cell(xml, "D1")
    assert ET._namespace_map == before


def test_duplicate_spreadsheet_value_nodes_fail_closed() -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row r="1"><c r="D1"><v>1</v><v>2</v></c>'
        b"</row></sheetData></worksheet>"
    )

    with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
        inspect_cell(xml, "D1")


def test_duplicate_coordinates_and_dtds_fail_closed() -> None:
    duplicate = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row><c r="D1"/><c r="D1"/></row></sheetData></worksheet>'
    )
    dtd = (
        b'<!DOCTYPE worksheet [<!ENTITY value "1">]><worksheet '
        b'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<c r="D1"><v>&value;</v></c></worksheet>'
    )

    with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
        inspect_cell(duplicate, "D1")
    with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
        inspect_cell(dtd, "D1")


@pytest.mark.parametrize("coordinate", ("XFE1", "A1048577"))
def test_out_of_range_excel_a1_coordinates_fail_closed(coordinate: str) -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row><c r="' + coordinate.encode() + b'"><v>1</v></c>'
        b"</row></sheetData></worksheet>"
    )

    with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
        inspect_cell(xml, coordinate)


def test_wrong_root_and_nested_spreadsheet_cells_fail_before_edit() -> None:
    wrong_root = b'<root xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>'
    nested = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<c r="A1"><c r="A2"><v>2</v></c></c></worksheet>'
    )

    for xml in (wrong_root, nested):
        with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
            replace_cell_value(xml, "A1", "3")


@pytest.mark.parametrize(
    "cell_parent",
    (b'<c r="A1"><v>2</v></c>', b'<sheetData><c r="A1"><v>2</v></c></sheetData>'),
)
def test_spreadsheet_cells_outside_rows_fail_before_edit(cell_parent: bytes) -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        + cell_parent
        + b"</worksheet>"
    )

    with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
        replace_cell_value(xml, "A1", "3")


def test_self_closing_value_expansion_preserves_opening_attributes() -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row><c r="D1"><v custom="keep"/></c></row></sheetData></worksheet>'
    )

    updated = replace_cell_value(xml, "D1", "4")

    assert b'<v custom="keep">4</v>' in updated


def test_formula_materialization_preserves_self_closing_value_attributes() -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row><c r="D1"><f>1</f><v custom="keep"/></c>'
        b"</row></sheetData></worksheet>"
    )

    updated = materialize_formula_cells(xml, {"D1": "4"})

    assert b'<v custom="keep">4</v>' in updated


def test_request_local_index_scans_once_for_many_inspections(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    cells = b"".join(f'<c r="A{row}" s="1"><v>{row}</v></c>'.encode() for row in range(1, 1_001))
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b"<sheetData><row>" + cells + b"</row></sheetData></worksheet>"
    )
    original = ooxml._scan_worksheet
    calls = 0

    def counted(payload: bytes, error_code: str = "TARGET_CELL_MISSING"):
        nonlocal calls
        calls += 1
        return original(payload, error_code)

    monkeypatch.setattr(ooxml, "_scan_worksheet", counted)
    index = worksheet_index(xml)
    for row in range(1, 1_001):
        assert inspect_index_cell(index, f"A{row}")[1] == str(row)
    assert calls == 1


def test_compact_index_has_bounded_memory_for_twenty_thousand_cells() -> None:
    cells = b"".join(f'<c r="A{row}" s="1"><v>{row}</v></c>'.encode() for row in range(1, 20_001))
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b"<sheetData><row>" + cells + b"</row></sheetData></worksheet>"
    )

    tracemalloc.start()
    try:
        tracemalloc.reset_peak()
        index = worksheet_index(xml)
        _current, peak = tracemalloc.get_traced_memory()
    finally:
        tracemalloc.stop()

    assert len(index.cells) == 20_000
    # The compact index retains less than 10x the source XML, excluding the
    # caller-owned XML buffer itself; this catches accidental per-cell graphs.
    assert peak - len(xml) < len(xml) * 9


def test_verify_temp_package_scans_changed_part_once_for_one_thousand_cells(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source_path = tmp_path / "source.xlsx"
    temp_path = tmp_path / "temporary.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист"
    for row in range(1, 1_001):
        sheet.cell(row, 1, row)
    workbook.save(source_path)
    workbook.close()
    part = worksheet_part_map(source_path)["Лист"]
    changes = {part: tuple((f"A{row}", str(row + 1_000)) for row in range(1, 1_001))}
    write_temp_package(source_path, temp_path, changes)
    original = ooxml._scan_worksheet
    calls = 0

    def counted(payload: bytes, error_code: str = "TARGET_CELL_MISSING"):
        nonlocal calls
        calls += 1
        return original(payload, error_code)

    monkeypatch.setattr(ooxml, "_scan_worksheet", counted)
    verify_temp_package(source_path, temp_path, changes)

    assert calls == 1


def test_worksheet_and_archive_admission_limits_fail_before_processing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(ooxml, "_MAX_WORKSHEET_XML_BYTES", 10)
    with pytest.raises(ExcelWriterIntegrityError, match="TARGET_CELL_MISSING"):
        worksheet_index(b"<worksheet/>")

    archive_path = tmp_path / "compressed.xlsx"
    with ZipFile(archive_path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"x" * 100)
    monkeypatch.setattr(ooxml, "_MAX_COMPRESSION_RATIO", 1)
    with pytest.raises(ExcelWriterSafetyError, match="INVALID_XLSX_PACKAGE"):
        ooxml.reject_unsupported_package(archive_path)


def test_raw_central_directory_count_must_match_eocd_before_zipfile(tmp_path: Path) -> None:
    archive_path = tmp_path / "forged-count.xlsx"
    with ZipFile(archive_path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("first", b"1")
        archive.writestr("second", b"2")
    payload = bytearray(archive_path.read_bytes())
    eocd = payload.rfind(b"PK\x05\x06")
    struct.pack_into("<H", payload, eocd + 8, 1)
    struct.pack_into("<H", payload, eocd + 10, 1)
    archive_path.write_bytes(payload)

    with pytest.raises(ExcelWriterSafetyError, match="INVALID_XLSX_PACKAGE"):
        ooxml.reject_unsupported_package(archive_path)


def test_raw_central_directory_must_end_exactly_at_eocd_before_zipfile(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A short declared directory must not hide another real CD record."""

    archive_path = tmp_path / "decoy-directory.xlsx"
    with ZipFile(archive_path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("first", b"1")
        archive.writestr("second", b"2")
    payload = bytearray(archive_path.read_bytes())
    eocd = payload.rfind(b"PK\x05\x06")
    central_directory = struct.unpack_from("<L", payload, eocd + 16)[0]
    first_record_size = 46 + sum(struct.unpack_from("<3H", payload, central_directory + 28))
    struct.pack_into("<H", payload, eocd + 8, 1)
    struct.pack_into("<H", payload, eocd + 10, 1)
    struct.pack_into("<L", payload, eocd + 12, first_record_size)
    archive_path.write_bytes(payload)

    def must_not_open(*_args: object, **_kwargs: object) -> object:
        raise AssertionError("ZipFile must not see an archive rejected by raw preflight")

    monkeypatch.setattr(ooxml.zipfile, "ZipFile", must_not_open)
    with pytest.raises(ExcelWriterSafetyError, match="INVALID_XLSX_PACKAGE"):
        ooxml.reject_unsupported_package(archive_path)


def test_semantic_worksheet_size_limit_is_checked_before_custom_part_read(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Worksheet relationship targets need not live under xl/worksheets/."""

    archive_path = tmp_path / "custom-sheet.xlsx"
    with ZipFile(archive_path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("xl/custom/sheet.xml", b"x" * 100)
    monkeypatch.setattr(ooxml, "_MAX_WORKSHEET_XML_BYTES", 10)
    with (
        ZipFile(archive_path) as archive,
        pytest.raises(ExcelWriterSafetyError, match="INVALID_XLSX_PACKAGE"),
    ):
        ooxml.read_archive_part(
            archive,
            "xl/custom/sheet.xml",
            ExcelWriterSafetyError,
            "INVALID_XLSX_PACKAGE",
            worksheet=True,
        )


@pytest.mark.parametrize(
    ("operation", "expected_code"),
    (
        (lambda payload: replace_cell_value(payload, "D1", "2"), "TARGET_CELL_MISSING"),
        (
            lambda payload: numeric_formula_values(payload, ("D1",)),
            "FORMULA_RESULT_NOT_NUMERIC",
        ),
        (
            lambda payload: materialize_formula_cells(payload, {"D1": "2"}),
            "FORMULA_MATERIALIZATION_FAILED",
        ),
    ),
)
def test_duplicate_coordinates_keep_each_callers_controlled_error_code(
    operation, expected_code: str
) -> None:
    duplicate = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row><c r="D1"><f>1</f><v>1</v></c>'
        b'<c r="D1"><f>1</f><v>1</v></c></row></sheetData></worksheet>'
    )

    with pytest.raises(ExcelWriterIntegrityError, match=expected_code):
        operation(duplicate)


def test_dtd_rejection_preserves_the_callers_error_code() -> None:
    dtd = (
        b'<!DOCTYPE worksheet [<!ENTITY value "1">]><worksheet '
        b'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row><c r="D1"><v>&value;</v></c></row></sheetData></worksheet>'
    )

    with pytest.raises(ExcelWriterIntegrityError, match="PRESERVATION_CHECK_FAILED"):
        worksheet_index(dtd, "PRESERVATION_CHECK_FAILED")


@pytest.mark.parametrize(
    "cell_xml",
    (
        b'<c r="E30"><f>D30</f></c>',
        b'<c r="E30" t="str"><f>D30</f><v>text</v></c>',
        b'<c r="E30"><f>D30</f><v>#VALUE!</v></c>',
        b'<c r="E30"><f>D30</f><v>NaN</v></c>',
        b'<c r="E30"><f>D30</f><v>Infinity</v></c>',
    ),
)
def test_invalid_formula_results_are_rejected_before_publication(cell_xml: bytes) -> None:
    xml = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<sheetData><row r="30">' + cell_xml + b"</row></sheetData></worksheet>"
    )

    with pytest.raises(ExcelWriterIntegrityError, match="FORMULA_RESULT_NOT_NUMERIC"):
        numeric_formula_values(xml, ("E30",))


def test_calc_chain_part_and_references_are_removed_from_a_temp_package(tmp_path: Path) -> None:
    source_path = tmp_path / "source.xlsx"
    temp_path = tmp_path / "temporary.xlsx"
    workbook = Workbook()
    workbook.active.title = "Лист"
    workbook.active["D30"] = 5
    workbook.save(source_path)
    workbook.close()
    seeded_path = tmp_path / "seeded.xlsx"
    with ZipFile(source_path) as source, ZipFile(seeded_path, "w", ZIP_DEFLATED) as output:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "[Content_Types].xml":
                payload = payload.replace(
                    b"</Types>",
                    b'<Override PartName="/xl/calcChain.xml" '
                    b'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/>'
                    b"</Types>",
                )
            elif info.filename == "xl/_rels/workbook.xml.rels":
                payload = payload.replace(
                    b"</Relationships>",
                    b'<Relationship Id="rId99" Target="calcChain.xml" '
                    b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain"/>'
                    b"</Relationships>",
                )
            output.writestr(info, payload)
        output.writestr("xl/calcChain.xml", b"<calcChain/>")
    seeded_path.replace(source_path)

    worksheet = worksheet_part_map(source_path)["Лист"]
    changes = {worksheet: (("D30", "0"),)}
    write_temp_package(source_path, temp_path, changes, remove_calc_chain=True)
    verify_temp_package(source_path, temp_path, changes, remove_calc_chain=True)

    with ZipFile(temp_path) as package:
        assert "xl/calcChain.xml" not in package.namelist()
        assert b"calcChain" not in package.read("[Content_Types].xml")
        assert b"calcChain" not in package.read("xl/_rels/workbook.xml.rels")

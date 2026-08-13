"""Contract checks for immutable failed-row OOXML annotations."""

from __future__ import annotations

import hashlib
import re
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from report_processor.excel_writer import (
    ExcelWriterInputError,
    ExcelWriterIntegrityError,
    ExcelWriterSafetyError,
    annotate_failed_rows,
)
from report_processor.excel_writer.ooxml import worksheet_part_map
from report_processor.excel_writer.row_annotations import _red_style_variants

_CELL = re.compile(rb"<c\b[^>]*?(?:/>|>.*?</c>)", re.DOTALL)
_REFERENCE = re.compile(rb"\br\s*=\s*([\"'])([^\"']+)\1")
_STYLE = re.compile(rb"\bs\s*=\s*([\"'])([^\"']+)\1")


def _workbook(path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "КС-2"
    first["A2"] = "neighbor"
    first["A3"] = 12.5
    first["A3"].fill = PatternFill("solid", fgColor="FF00FF00")
    first["B3"] = "=A3*2"
    second = workbook.create_sheet("КС-6а")
    second["A3"] = "other sheet"
    workbook.save(path)
    workbook.close()


def _parts(path: Path) -> dict[str, bytes]:
    with ZipFile(path) as package:
        return {name: package.read(name) for name in package.namelist()}


def _cell(xml: bytes, coordinate: str) -> bytes:
    for match in _CELL.finditer(xml):
        value = match.group(0)
        reference = _REFERENCE.search(value)
        if reference is not None and reference.group(2) == coordinate.encode("ascii"):
            return value
    raise AssertionError(f"cell {coordinate} is missing")


def _red_fill_for_cell(styles: bytes, cell: bytes) -> str | None:
    root = ElementTree.fromstring(styles)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    qualified = lambda name: f"{{{namespace}}}{name}" if namespace else name  # noqa: E731
    style = _STYLE.search(cell)
    style_id = int(style.group(2)) if style is not None else 0
    xf = list(root.find(qualified("cellXfs")))[style_id]  # type: ignore[arg-type]
    fill = list(root.find(qualified("fills")))[int(xf.attrib["fillId"])]  # type: ignore[arg-type]
    pattern = fill.find(qualified("patternFill"))
    foreground = None if pattern is None else pattern.find(qualified("fgColor"))
    return None if foreground is None else foreground.attrib.get("rgb")


def test_annotations_change_only_requested_existing_cells_and_preserve_package(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "annotated.xlsx"
    _workbook(source)
    before_digest = hashlib.sha256(source.read_bytes()).hexdigest()
    before = _parts(source)
    parts = worksheet_part_map(source)

    assert annotate_failed_rows(source, output, {"КС-2": {3}}) == output

    after = _parts(output)
    assert hashlib.sha256(source.read_bytes()).hexdigest() == before_digest
    assert tuple(before) == tuple(after)
    assert all(
        before[name] == after[name]
        for name in before
        if name not in {"xl/styles.xml", parts["КС-2"]}
    )
    source_sheet = before[parts["КС-2"]]
    output_sheet = after[parts["КС-2"]]
    assert _cell(source_sheet, "A2") == _cell(output_sheet, "A2")
    for coordinate in ("A3", "B3"):
        before_cell = _cell(source_sheet, coordinate)
        after_cell = _cell(output_sheet, coordinate)
        assert re.sub(rb"\s+s\s*=\s*([\"'])[^\"']+\1", b"", before_cell) == re.sub(
            rb"\s+s\s*=\s*([\"'])[^\"']+\1", b"", after_cell
        )
        assert _red_fill_for_cell(after["xl/styles.xml"], after_cell) == "FFFF0000"
    assert b"<f>A3*2</f>" in _cell(output_sheet, "B3")
    assert b"<v>12.5</v>" in _cell(output_sheet, "A3")
    assert after[parts["КС-6а"]] == before[parts["КС-6а"]]


def test_annotations_preserve_vba_payload_bytes(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsm"
    seeded = tmp_path / "seeded.xlsx"
    output = tmp_path / "annotated.xlsm"
    _workbook(seeded)
    seeded.replace(source)
    with ZipFile(source, "a") as package:
        package.writestr("xl/vbaProject.bin", b"not-a-real-macro-but-byte-preservation-is-required")

    annotate_failed_rows(source, output, {"КС-2": {3}})

    with ZipFile(source) as original, ZipFile(output) as annotated:
        assert annotated.read("xl/vbaProject.bin") == original.read("xl/vbaProject.bin")


def test_signed_and_unsafe_packages_are_rejected_without_output(tmp_path: Path) -> None:
    signed = tmp_path / "signed.xlsx"
    unsafe = tmp_path / "unsafe.xlsx"
    _workbook(signed)
    _workbook(unsafe)
    with ZipFile(signed, "a") as package:
        package.writestr("_xmlsignatures/sig1.xml", b"<Signature/>")
    with ZipFile(unsafe, "a") as package:
        package.writestr("../outside.xml", b"not safe")

    with pytest.raises(ExcelWriterSafetyError, match="SIGNED_PACKAGE_UNSUPPORTED"):
        annotate_failed_rows(signed, tmp_path / "signed-output.xlsx", {"КС-2": {3}})
    with pytest.raises(ExcelWriterSafetyError, match="INVALID_XLSX_PACKAGE"):
        annotate_failed_rows(unsafe, tmp_path / "unsafe-output.xlsx", {"КС-2": {3}})


@pytest.mark.parametrize(
    "failed_rows",
    (
        {},
        {"КС-2": set()},
        {"КС-2": {0}},
        {"КС-2": {True}},
        {"КС-2": {1_048_577}},
        {"КС-2": "3"},
    ),
)
def test_invalid_failed_rows_are_rejected(tmp_path: Path, failed_rows: object) -> None:
    source = tmp_path / "source.xlsx"
    _workbook(source)

    with pytest.raises(ExcelWriterInputError, match="INVALID_FAILED_ROWS"):
        annotate_failed_rows(source, tmp_path / "output.xlsx", failed_rows)  # type: ignore[arg-type]


def test_unknown_sheet_and_existing_output_are_rejected_without_clobber(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _workbook(source)
    output.write_bytes(b"existing")

    with pytest.raises(ExcelWriterSafetyError, match="OUTPUT_EXISTS"):
        annotate_failed_rows(source, output, {"КС-2": {3}})
    assert output.read_bytes() == b"existing"
    with pytest.raises(ExcelWriterInputError, match="UNKNOWN_WORKSHEET"):
        annotate_failed_rows(source, tmp_path / "unknown.xlsx", {"not-a-sheet": {3}})


def test_missing_or_cellless_failed_row_is_rejected_without_output(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _workbook(source)

    with pytest.raises(ExcelWriterIntegrityError, match="FAILED_ROW_NOT_FOUND"):
        annotate_failed_rows(source, output, {"КС-2": {999}})

    assert not output.exists()


def test_style_patch_preserves_office_ignorable_namespace_declarations() -> None:
    styles = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        b'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac" '
        b'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
        b'mc:Ignorable="x14ac xr"><fills count="1"><fill>'
        b'<patternFill patternType="none"/></fill></fills><cellXfs count="1">'
        b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>'
        b"</cellXfs></styleSheet>"
    )

    updated, variants = _red_style_variants(styles, {0})

    assert variants == {0: 1}
    for declaration in (
        b'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"',
        b'xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"',
        b'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"',
        b'mc:Ignorable="x14ac xr"',
    ):
        assert declaration in updated
    ElementTree.fromstring(updated)


def test_style_patch_handles_adjacent_self_closing_and_paired_xfs() -> None:
    styles = (
        b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
        b'<cellXfs count="2"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>'
        b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0">'
        b'<alignment horizontal="center"/></xf></cellXfs></styleSheet>'
    )

    updated, variants = _red_style_variants(styles, {0, 1})

    root = ElementTree.fromstring(updated)
    namespace = root.tag.partition("}")[0].removeprefix("{")
    cell_xfs = root.find(f"{{{namespace}}}cellXfs")
    assert variants == {0: 2, 1: 3}
    assert len(cell_xfs) == 4  # type: ignore[arg-type]
    assert b'<alignment horizontal="center"/>' in updated
    assert b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>' in updated

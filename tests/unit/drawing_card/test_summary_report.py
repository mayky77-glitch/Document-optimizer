"""Regression contract for the published drawing-card summary workbook."""

from __future__ import annotations

from dataclasses import replace
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from report_processor.drawing_card.models import (
    CATEGORY_DISPLAY_NAMES,
    CATEGORY_ORDER,
    DrawingCardResultRow,
)
from report_processor.drawing_card.output.contract import (
    CARD_HEADERS,
    COST_FORMAT,
    FRACTIONAL_QUANTITY_FORMAT,
    MAIN_CARD_SHEET_NAME,
    SUMMARY_BLOCK_COLUMN_SPAN,
    SUMMARY_BLOCK_ROW_SPAN,
    SUMMARY_HEADERS,
    SUMMARY_SHEET_NAME,
)
from report_processor.drawing_card.output.discrepancies import (
    CONTRACT_COST_ERROR_REASON,
    DISCREPANCY_SHEET_NAME,
    report_issue_text,
)
from report_processor.drawing_card.output.layout import plan_layout
from report_processor.drawing_card.output.summary import summary_block_position, summary_row_count
from report_processor.drawing_card.output.validator import validate_card
from report_processor.drawing_card.output.writer import (
    load_existing_values,
    merge_update_rows,
    write_card,
)
from report_processor.drawing_card.sources.normalization import build_drawing_code
from report_processor.drawing_card.statuses import Status

FIXTURES = Path(__file__).parents[2] / "fixtures" / "drawing_card"


def _rows(indices: tuple[str, ...] = ("1001", "1002", "1003"), *, mixed_unit: bool = False):
    rows = []
    for index_number, object_index in enumerate(indices, start=1):
        drawing = build_drawing_code(f"А-{index_number:03d}")
        for category_number, category in enumerate(CATEGORY_ORDER, start=1):
            unit = "м"
            if mixed_unit and category_number == 1:
                unit = "шт" if object_index == indices[0] else "м"
            rows.append(
                DrawingCardResultRow(
                    object_index=object_index,
                    drawing_code=drawing,
                    category=category,
                    display_name=CATEGORY_DISPLAY_NAMES[category],
                    result_unit=unit,
                    remaining_quantity=Decimal(f"{index_number}.{category_number}"),
                    # Internal costs stay in rubles.  With cost_scale=100 the
                    # published values below are whole millions, making unit
                    # conversion and the two-decimal display contract observable.
                    remaining_total_cost=Decimal(str(index_number * category_number * 100_000_000)),
                    quantity_source_rows=(f"row-{object_index}-{category_number}",),
                    cost_source_rows=(f"row-{object_index}-{category_number}",),
                    quantity_rule_id="test-rule",
                    cost_rule_id="test-rule",
                    quantity_confidence=1.0,
                    cost_confidence=1.0,
                    requires_manual_review=False,
                    status=Status.OK,
                    warnings=(),
                )
            )
    return rows


def _write(tmp_path: Path, *, mixed_unit: bool = False) -> tuple[Path, list]:
    rows = _rows(mixed_unit=mixed_unit)
    layouts = plan_layout(rows)
    output = tmp_path / "card.xlsx"
    write_card(
        base_path=FIXTURES / "default_template.xlsx",
        output_path=output,
        rows=rows,
        layouts=layouts,
        run_id="summary-test",
        cost_scale=100,
    )
    return output, layouts


def test_summary_uses_two_column_index_cards_with_literal_million_ruble_values(
    tmp_path: Path,
) -> None:
    output, layouts = _write(tmp_path)
    validation = validate_card(output, layouts)
    workbook = load_workbook(output, data_only=False)
    try:
        summary = workbook[SUMMARY_SHEET_NAME]
        assert validation["status"] == Status.OK
        assert workbook.sheetnames[0] == MAIN_CARD_SHEET_NAME
        assert summary.max_row == summary_row_count(layouts)

        expected_cards = (
            ("1001", (1, 1), "A1:H1"),
            ("1002", (1, 1 + SUMMARY_BLOCK_COLUMN_SPAN), "K1:R1"),
            ("1003", (1 + SUMMARY_BLOCK_ROW_SPAN, 1), "A12:H12"),
            (
                "Все индексы",
                (1 + SUMMARY_BLOCK_ROW_SPAN, 1 + SUMMARY_BLOCK_COLUMN_SPAN),
                "K12:R12",
            ),
        )
        merged = {str(cell_range) for cell_range in summary.merged_cells.ranges}
        for title, (start_row, start_column), merge in expected_cards:
            assert summary.cell(start_row, start_column).value == (
                title if title == "Все индексы" else f"Индекс объекта: {title}"
            )
            assert merge in merged
            assert (
                tuple(
                    summary.cell(start_row + 1, start_column + offset).value
                    for offset in range(len(SUMMARY_HEADERS))
                )
                == SUMMARY_HEADERS
            )
            assert summary.cell(start_row, start_column).font.name == "Arial"
            assert summary.cell(start_row, start_column).font.bold is True
            assert summary.cell(start_row, start_column).fill.fgColor.rgb.endswith("548235")
            assert summary.cell(start_row, start_column).font.color.rgb.endswith("FFFFFF")
            assert summary.cell(start_row + 1, start_column).fill.fgColor.rgb.endswith("E2F0D9")
            assert tuple(
                summary.cell(start_row + 2 + offset, start_column).value
                for offset in range(len(CATEGORY_ORDER))
            ) == tuple(CATEGORY_DISPLAY_NAMES[category] for category in CATEGORY_ORDER)
            assert all(
                summary.cell(start_row + 2, start_column + offset).has_style
                for offset in range(len(SUMMARY_HEADERS))
            )

        first_row, first_column = summary_block_position(0)
        second_row, second_column = summary_block_position(1)
        third_row, third_column = summary_block_position(2)
        total_row, total_column = summary_block_position(len(layouts))
        assert all(cell.data_type != "f" for row in summary.iter_rows() for cell in row)
        assert all(
            isinstance(summary.cell(first_row + 2, first_column + offset).value, (int, float))
            for offset in range(2, len(SUMMARY_HEADERS))
        )
        assert summary.cell(first_row + 2, first_column + 3).value == 1
        assert summary.cell(second_row + 2, second_column + 3).value == 2
        assert summary.cell(third_row + 2, third_column + 3).value == 3
        assert summary.cell(total_row + 2, total_column + 3).value == 6
        for category_offset in range(len(CATEGORY_ORDER)):
            positions = [summary_block_position(number) for number in range(len(layouts))]
            for metric_offset in range(2, len(SUMMARY_HEADERS)):
                total = summary.cell(
                    total_row + 2 + category_offset, total_column + metric_offset
                ).value
                index_values = [
                    summary.cell(
                        row + 2 + category_offset,
                        column + metric_offset,
                    ).value
                    for row, column in positions
                ]
                assert Decimal(str(total)) == sum(Decimal(str(value)) for value in index_values)
        assert (
            summary.cell(first_row + 2, first_column + 2).number_format
            == FRACTIONAL_QUANTITY_FORMAT
        )
        assert summary.cell(first_row + 2, first_column + 3).number_format == COST_FORMAT
        assert (
            summary.cell(total_row + 2, total_column + 2).number_format
            == FRACTIONAL_QUANTITY_FORMAT
        )
        assert summary.cell(total_row + 2, total_column + 3).number_format == COST_FORMAT
        assert summary.sheet_view.showGridLines is False
        assert workbook[MAIN_CARD_SHEET_NAME].sheet_view.showGridLines is False
        assert workbook[MAIN_CARD_SHEET_NAME]["B2"].fill.fgColor.rgb.endswith("548235")
        assert workbook[MAIN_CARD_SHEET_NAME]["B2"].font.color.rgb.endswith("FFFFFF")
        assert workbook[MAIN_CARD_SHEET_NAME]["B4"].font.name == "Arial"
        assert workbook[MAIN_CARD_SHEET_NAME].freeze_panes == "B4"
        assert COST_FORMAT == "#,##0.00"
        assert "млн руб." in SUMMARY_HEADERS[-1]
        assert "млн руб." in workbook[MAIN_CARD_SHEET_NAME]["F3"].value
    finally:
        workbook.close()


def test_numeric_display_uses_two_decimals_without_rounding_stored_values(tmp_path: Path) -> None:
    rows = _rows()
    rows[0] = replace(
        rows[0],
        remaining_quantity=Decimal("24"),
        remaining_total_cost=Decimal("780901100"),
    )
    rows[1] = replace(rows[1], remaining_quantity=Decimal("2704.7755"))
    layouts = plan_layout(rows)
    output = tmp_path / "two-decimal-card.xlsx"
    write_card(
        base_path=FIXTURES / "default_template.xlsx",
        output_path=output,
        rows=rows,
        layouts=layouts,
        run_id="two-decimal-test",
        cost_scale=100,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        card = workbook[MAIN_CARD_SHEET_NAME]
        summary = workbook[SUMMARY_SHEET_NAME]
        assert Decimal(str(card["E4"].value)) == Decimal("24")
        assert card["E4"].number_format == "0.00"
        assert Decimal(str(card["E5"].value)) == Decimal("2704.7755")
        assert card["E5"].number_format == "0.00"
        assert Decimal(str(card["F4"].value)) == Decimal("7.809011")
        assert card["F4"].number_format == "#,##0.00"
        assert Decimal(str(summary["C3"].value)) == Decimal("24")
        assert summary["C3"].number_format == "0.00"
        assert Decimal(str(summary["D3"].value)) == Decimal("7.809011")
        assert summary["D3"].number_format == "#,##0.00"
    finally:
        workbook.close()


def test_main_card_publishes_contract_and_performed_values_without_scaling_quantity(
    tmp_path: Path,
) -> None:
    rows = _rows()
    rows[0] = replace(
        rows[0],
        contract_quantity=Decimal("11.25"),
        contract_total_cost=Decimal("2500000"),
        performed_quantity=Decimal("9.5"),
        performed_total_cost=Decimal("1750000"),
    )
    layouts = plan_layout(rows)
    output = tmp_path / "contract-values.xlsx"
    write_card(
        base_path=FIXTURES / "default_template.xlsx",
        output_path=output,
        rows=rows,
        layouts=layouts,
        run_id="contract-values",
        cost_scale=1,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        card = workbook[MAIN_CARD_SHEET_NAME]
        headers = {
            card.cell(layouts[0].column_header_row, layouts[0].start_column + offset).value: offset
            for offset in range(len(CARD_HEADERS))
        }
        assert tuple(headers) == CARD_HEADERS
        first_data_row = layouts[0].drawing_code_blocks[0].start_row
        published = {
            header: card.cell(first_data_row, layouts[0].start_column + offset).value
            for header, offset in headers.items()
        }
        assert published[CARD_HEADERS[5]] == 11.25
        assert published[CARD_HEADERS[6]] == 2.5
        assert published[CARD_HEADERS[7]] == 9.5
        assert published[CARD_HEADERS[8]] == 1.75
        summary = workbook[SUMMARY_SHEET_NAME]
        summary_row, summary_column = summary_block_position(0)
        assert (
            tuple(
                summary.cell(summary_row + 1, summary_column + offset).value
                for offset in range(len(SUMMARY_HEADERS))
            )
            == SUMMARY_HEADERS
        )
        assert summary.cell(summary_row + 2, summary_column + 4).value == 11.25
        assert summary.cell(summary_row + 2, summary_column + 5).value == 2.5
        assert summary.cell(summary_row + 2, summary_column + 6).value == 9.5
        assert summary.cell(summary_row + 2, summary_column + 7).value == 1.75
        assert DISCREPANCY_SHEET_NAME not in workbook.sheetnames
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("difference_rub", "has_violation"),
    ((Decimal("1000"), False), (Decimal("1000.01"), True)),
)
def test_contract_cost_violation_uses_strict_ruble_tolerance_and_links_to_only_red_cell(
    tmp_path: Path, difference_rub: Decimal, has_violation: bool
) -> None:
    rows = _rows(indices=("1001",))
    rows[0] = replace(
        rows[0],
        contract_quantity=Decimal("10.1234567890123"),
        contract_total_cost=Decimal("2000000"),
        performed_quantity=Decimal("12.9876543210987"),
        performed_total_cost=Decimal("2000000") + difference_rub,
    )
    layouts = plan_layout(rows)
    output = tmp_path / f"tolerance-{difference_rub}.xlsx"
    write_card(
        base_path=FIXTURES / "default_template.xlsx",
        output_path=output,
        rows=rows,
        layouts=layouts,
        run_id="contract-tolerance",
        cost_scale=1,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        card = workbook[MAIN_CARD_SHEET_NAME]
        headers = {
            card.cell(layouts[0].column_header_row, layouts[0].start_column + offset).value: offset
            for offset in range(len(CARD_HEADERS))
        }
        row_number = layouts[0].drawing_code_blocks[0].start_row
        contract_cost_column = layouts[0].start_column + headers[CARD_HEADERS[6]]
        contract_cost_cell = card.cell(row_number, contract_cost_column)
        if not has_violation:
            assert DISCREPANCY_SHEET_NAME not in workbook.sheetnames
            assert not contract_cost_cell.fill.fgColor.rgb.endswith("FFC7CE")
            return

        assert DISCREPANCY_SHEET_NAME in workbook.sheetnames
        assert contract_cost_cell.fill.fgColor.rgb.endswith("FFC7CE")
        discrepancy = workbook[DISCREPANCY_SHEET_NAME]
        assert discrepancy.max_row == 2
        assert tuple(cell.value for cell in discrepancy[1]) == (
            "Индекс объекта",
            "Шифр чертежа",
            "Этап / категория",
            "Ед. изм.",
            CARD_HEADERS[5],
            CARD_HEADERS[6],
            CARD_HEADERS[7],
            CARD_HEADERS[8],
            "Причина ошибки",
            "Разница стоимости, млн руб.",
            "Ссылка на договорную стоимость",
        )
        expected_target = f"#'{MAIN_CARD_SHEET_NAME}'!{contract_cost_cell.coordinate}"
        assert discrepancy.cell(2, 9).value == CONTRACT_COST_ERROR_REASON
        assert discrepancy.cell(2, 11).hyperlink.target == expected_target
        red_coordinates = {
            cell.coordinate
            for cell in card[row_number]
            if cell.fill.fill_type == "solid" and cell.fill.fgColor.rgb.endswith("FFC7CE")
        }
        assert red_coordinates == {contract_cost_cell.coordinate}
        assert Decimal(str(discrepancy.cell(2, 5).value)) == Decimal("10.1234567890123")
        assert discrepancy.cell(2, 6).value == 2
        assert Decimal(str(discrepancy.cell(2, 7).value)) == Decimal("12.9876543210987")
        assert Decimal(str(discrepancy.cell(2, 8).value)) == Decimal("2.00100001")
        assert Decimal(str(discrepancy.cell(2, 10).value)) == Decimal("0.00100001")
    finally:
        workbook.close()


def test_report_comments_explain_warnings_without_internal_row_hashes(tmp_path: Path) -> None:
    rows = _rows(indices=("1001",))
    rows[0] = replace(
        rows[0],
        status=Status.WARNING,
        warnings=("POSSIBLE_DUPLICATE:private-row-hash-1:private-row-hash-2",),
    )
    layouts = plan_layout(rows)
    output = tmp_path / "readable-warning.xlsx"
    write_card(
        base_path=FIXTURES / "default_template.xlsx",
        output_path=output,
        rows=rows,
        layouts=layouts,
        run_id="readable-warning",
        cost_scale=1,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        card = workbook[MAIN_CARD_SHEET_NAME]
        row_number = layouts[0].drawing_code_blocks[0].start_row
        comment = card.cell(row_number, layouts[0].start_column + 1).comment
        assert comment is not None
        assert comment.text == "В исходных данных найдены возможные дубли этой работы."
        assert "private-row-hash" not in comment.text
        assert "POSSIBLE_DUPLICATE" not in comment.text
    finally:
        workbook.close()


def test_dimensional_formula_repair_has_a_plain_russian_reason() -> None:
    assert report_issue_text(
        ("REMAINING_QUANTITY_REPAIRED_FROM_DIMENSIONAL_FORMULA",),
        Status.WARNING,
    ) == (
        "В исходной книге формула остаточного объёма ссылалась на стоимость. "
        "Объём пересчитан по договорному количеству и вычитаемым объёмам."
    )


def test_update_clears_resolved_contract_cost_highlight(tmp_path: Path) -> None:
    rows = _rows(indices=("1001",))
    layouts = plan_layout(rows)
    first_output = tmp_path / "first-violation.xlsx"
    second_output = tmp_path / "resolved-violation.xlsx"
    violating = replace(
        rows[0],
        contract_total_cost=Decimal("1000000"),
        performed_total_cost=Decimal("1001000.01"),
    )
    write_card(
        base_path=FIXTURES / "default_template.xlsx",
        output_path=first_output,
        rows=[violating, *rows[1:]],
        layouts=layouts,
        run_id="first-violation",
        cost_scale=1,
    )
    resolved = replace(
        rows[0],
        contract_total_cost=Decimal("1000000"),
        performed_total_cost=Decimal("1001000"),
    )
    write_card(
        base_path=first_output,
        output_path=second_output,
        rows=[resolved, *rows[1:]],
        layouts=layouts,
        run_id="resolved-violation",
        cost_scale=1,
    )

    workbook = load_workbook(second_output, data_only=False)
    try:
        card = workbook[MAIN_CARD_SHEET_NAME]
        contract_cost_cell = card.cell(
            layouts[0].drawing_code_blocks[0].start_row,
            layouts[0].start_column + CARD_HEADERS.index("По договору — стоимость, млн руб."),
        )
        assert DISCREPANCY_SHEET_NAME not in workbook.sheetnames
        assert not contract_cost_cell.fill.fgColor.rgb.endswith("FFC7CE")
    finally:
        workbook.close()


def test_update_reads_million_ruble_card_back_as_internal_rubles_without_double_scaling(
    tmp_path: Path,
) -> None:
    output, _layouts = _write(tmp_path)
    existing = load_existing_values(output, cost_scale=100)
    source_key = ("1001", "А-001", CATEGORY_DISPLAY_NAMES[CATEGORY_ORDER[0]].casefold())

    assert existing[source_key][2] == Decimal("100000000")
    assert ("1002", "А-002", CATEGORY_DISPLAY_NAMES[CATEGORY_ORDER[0]].casefold()) in existing
    assert ("1003", "А-003", CATEGORY_DISPLAY_NAMES[CATEGORY_ORDER[0]].casefold()) in existing

    preserved, warnings = merge_update_rows([], existing, policy="keep_existing")
    updated = tmp_path / "updated-card.xlsx"
    write_card(
        base_path=output,
        output_path=updated,
        rows=preserved,
        layouts=plan_layout(preserved),
        run_id="summary-update",
        cost_scale=100,
    )

    assert warnings == []
    assert load_existing_values(updated, cost_scale=100)[source_key][2] == Decimal("100000000")


def test_three_occupied_slots_trim_only_the_fourth_template_slot(tmp_path: Path) -> None:
    output, layouts = _write(tmp_path)
    workbook = load_workbook(output)
    try:
        sheet = workbook[layouts[-1].sheet_name]
        assert [layout.start_column for layout in layouts] == [2, 12, 22]
        assert [
            sheet.cell(layout.drawing_code_blocks[0].start_row, layout.start_column).value
            for layout in layouts
        ] == ["А-001", "А-002", "А-003"]
        slot_span = layouts[1].start_column - layouts[0].start_column
        fourth_slot_start = layouts[-1].start_column + slot_span
        assert sheet.cell(4, fourth_slot_start).value is None
        assert not sheet.cell(4, fourth_slot_start).has_style
    finally:
        workbook.close()


def test_mixed_units_leave_all_indices_quantity_blank_and_fail_validation(tmp_path: Path) -> None:
    rows = _rows(mixed_unit=True)
    layouts = plan_layout(rows)
    output = tmp_path / "mixed-units.xlsx"

    with pytest.raises(ValueError, match="SUMMARY_MIXED_UNIT"):
        write_card(
            base_path=FIXTURES / "default_template.xlsx",
            output_path=output,
            rows=rows,
            layouts=layouts,
            run_id="summary-mixed-unit",
            cost_scale=100,
        )

    assert not output.exists()


def test_missing_unit_for_one_index_rejects_all_indices_summary(tmp_path: Path) -> None:
    rows = _rows()
    rows[0] = replace(rows[0], result_unit=None)
    layouts = plan_layout(rows)
    output = tmp_path / "missing-unit.xlsx"

    with pytest.raises(ValueError, match="SUMMARY_MISSING_UNIT"):
        write_card(
            base_path=FIXTURES / "default_template.xlsx",
            output_path=output,
            rows=rows,
            layouts=layouts,
            run_id="summary-missing-unit",
            cost_scale=100,
        )

    assert not output.exists()


def test_multiple_units_for_one_category_reject_summary(tmp_path: Path) -> None:
    rows = _rows()
    rows[8] = replace(rows[8], result_unit="шт")
    layouts = plan_layout(rows)
    output = tmp_path / "incompatible-units.xlsx"

    with pytest.raises(ValueError, match="SUMMARY_MIXED_UNIT"):
        write_card(
            base_path=FIXTURES / "default_template.xlsx",
            output_path=output,
            rows=rows,
            layouts=layouts,
            run_id="summary-incompatible-units",
            cost_scale=100,
        )

    assert not output.exists()

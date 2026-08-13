from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from report_processor.admin_panel.app import _result_media_type, _safe_download_name
from report_processor.admin_panel.presentation import job_payload
from report_processor.admin_panel.reconciliation_execution import _Catalog, _review_row_id
from report_processor.admin_panel.reconciliation_numeric_verification import (
    NumericVerificationFailure,
    _authorizations,
    _matches_target,
    _reject_duplicate_source_identities,
    _reject_duplicate_target_keys,
    _validate_units,
    verify_numeric,
)
from report_processor.admin_panel.reconciliation_sources import ReconciliationSourceIssue
from report_processor.admin_panel.reconciliation_verification import (
    VerificationTechnicalFailure,
    _write_artifact,
    verify_reconciliation,
)
from report_processor.admin_panel.service import AdminPanelService
from report_processor.excel_writer import ExcelWriterSafetyError
from report_processor.reconciliation_review import (
    AppliedOverride,
    ReviewAction,
    ReviewDecision,
    ReviewMode,
    ReviewRow,
    build_review_groups,
)
from report_processor.target_report import TargetNumericCell


def _job(tmp_path: Path) -> SimpleNamespace:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    return SimpleNamespace(
        source=source,
        sources=(source,),
        source_names=("source.xlsx",),
        source_digests=("source-digest",),
        target_digest="target-digest",
        directory=tmp_path,
    )


def _xlsx(path: Path, value: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист1"
    sheet["A2"] = value
    workbook.save(path)
    workbook.close()


def _review(job, *, safe: bool, action: ReviewAction | None = None, issues=()):
    row = ReviewRow("source-row", "Монтаж трубы", "м", None, None, "target")
    (group,) = build_review_groups((row,))
    review_row_id = _review_row_id(job, "source-row")
    state_row = ReviewRow(
        review_row_id,
        row.display_name,
        row.unit,
        None,
        None,
        row.proposed_category,
    )
    group = next(iter(build_review_groups((state_row,))))
    decision = (
        SimpleNamespace(action=action, row_id=review_row_id, group_id=None) if action else None
    )
    state = SimpleNamespace(
        rows={review_row_id: state_row},
        groups={group.group_id: group},
        grouping=SimpleNamespace(
            packages=(SimpleNamespace(safe=safe, member_group_ids=(group.group_id,)),)
        ),
        effective_decisions=lambda: () if decision is None else (decision,),
    )
    source = SimpleNamespace(
        source_row_id="source-row",
        source_sheet="Лист1",
        source_row_number=7,
        source_filename="source.xlsx",
    )
    return SimpleNamespace(state=state, source_batch=SimpleNamespace(rows=(source,), issues=issues))


def test_safe_package_passes_without_an_artifact(tmp_path: Path, monkeypatch) -> None:
    job = _job(tmp_path)
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_verification.prepare_review",
        lambda *_args: _review(job, safe=True),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_verification.verify_numeric",
        lambda *_args: (frozenset(), 1),
    )

    result = verify_reconciliation(job, ())

    assert result.verification_status == "passed"
    assert result.message == "Все документы проверены. Ошибок не найдено."
    assert result.checked_row_count == 1 and result.failed_row_count == 0
    assert result.output is None


@pytest.mark.parametrize(
    ("safe", "action", "expected_failed"),
    ((False, None, 1), (False, ReviewAction.ACCEPT, 0), (True, ReviewAction.REJECT, 1)),
)
def test_safe_package_and_latest_authoritative_feedback_have_expected_precedence(
    tmp_path: Path, monkeypatch, safe: bool, action: ReviewAction | None, expected_failed: int
) -> None:
    job = _job(tmp_path)
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_verification.prepare_review",
        lambda *_args: _review(job, safe=safe, action=action),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_verification.verify_numeric",
        lambda *_args: (
            frozenset() if expected_failed == 0 else frozenset({_review_row_id(job, "source-row")}),
            1,
        ),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_verification._write_artifact",
        lambda *_args: (tmp_path / "verification.xlsx", "Проверено_source.xlsx"),
    )

    result = verify_reconciliation(job, ())

    assert result.failed_row_count == expected_failed
    assert result.verification_status == ("failed" if expected_failed else "passed")


def test_numeric_comparison_uses_writer_scale_and_cost_only_omits_quantity() -> None:
    target = SimpleNamespace(
        selected_quantity=SimpleNamespace(value=Decimal("999.999")),
        selected_cost=SimpleNamespace(value=Decimal("2.695")),
    )
    calculation = SimpleNamespace(target_row=target, quantity=Decimal("1.01"), cost=Decimal("2.70"))

    assert _matches_target(calculation, ReviewMode.COST_ONLY) is True
    assert _matches_target(calculation, ReviewMode.QUANTITY_COST) is False


def test_numeric_comparison_requires_finite_target_value() -> None:
    calculation = SimpleNamespace(
        target_row=SimpleNamespace(
            selected_quantity=SimpleNamespace(value=Decimal("1.00")),
            selected_cost=SimpleNamespace(value=Decimal("NaN")),
        ),
        quantity=Decimal("1.00"),
        cost=Decimal("1.00"),
    )

    with pytest.raises(NumericVerificationFailure, match="TARGET_VALUE_UNAVAILABLE"):
        _matches_target(calculation, ReviewMode.COST_ONLY)


@pytest.mark.parametrize("mode", (ReviewMode.COST_ONLY, ReviewMode.QUANTITY_COST))
def test_missing_calculated_value_is_an_ordinary_numeric_mismatch(mode: ReviewMode) -> None:
    calculation = SimpleNamespace(
        target_row=SimpleNamespace(
            selected_quantity=TargetNumericCell(Decimal("1.00"), "1", "NOT_FORMULA", "OK"),
            selected_cost=TargetNumericCell(Decimal("2.70"), "2.7", "NOT_FORMULA", "OK"),
        ),
        quantity=Decimal("1.00"),
        cost=None,
    )

    assert _matches_target(calculation, mode) is False


def test_numeric_comparison_reads_target_numeric_cell_value() -> None:
    calculation = SimpleNamespace(
        target_row=SimpleNamespace(
            selected_quantity=TargetNumericCell(Decimal("1.00"), "1", "NOT_FORMULA", "OK"),
            selected_cost=TargetNumericCell(Decimal("2.70"), "2.7", "NOT_FORMULA", "OK"),
        ),
        quantity=Decimal("1.00"),
        cost=Decimal("2.70"),
    )

    assert _matches_target(calculation, ReviewMode.QUANTITY_COST) is True


def test_cost_only_does_not_require_source_or_target_unit() -> None:
    target = SimpleNamespace(unit=None)
    overrides = {"row": AppliedOverride("row", "category:target", False, True, ReviewAction.ACCEPT)}
    source_rows = {"row": SimpleNamespace(source_filename="source.xlsx", unit=None)}
    catalog = _Catalog({"category:target": "Target"}, {("1234", "category:target"): target})

    _validate_units(overrides, source_rows, catalog, SimpleNamespace())


def test_explicit_rejection_wins_over_safe_package_authorization() -> None:
    row = ReviewRow("row", "Монтаж", "м", Decimal("1"), Decimal("1"))
    (group,) = build_review_groups((row,))
    rejected = ReviewDecision(ReviewAction.REJECT, group_id=group.group_id)
    state = SimpleNamespace(
        rows={row.row_id: row},
        groups={group.group_id: group},
        grouping=SimpleNamespace(
            packages=(
                SimpleNamespace(
                    safe=True,
                    package_key=("category:target", "quantity_cost"),
                    member_group_ids=(group.group_id,),
                ),
            )
        ),
        effective_decisions=lambda: (rejected,),
    )

    assert _authorizations(state)[row.row_id] is rejected


@pytest.mark.parametrize("reversed_rows", (False, True))
def test_duplicate_physical_source_identity_is_technical_failure(reversed_rows: bool) -> None:
    first = SimpleNamespace(source_filename="first.xlsx", source_sheet="Sheet", source_row_number=7)
    second = SimpleNamespace(
        source_filename="second.xlsx", source_sheet="Sheet", source_row_number=7
    )

    with pytest.raises(NumericVerificationFailure, match="SOURCE_IDENTITY_AMBIGUOUS"):
        _reject_duplicate_source_identities(
            (second, first) if reversed_rows else (first, second),
            {"first.xlsx": "same-content", "second.xlsx": "same-content"},
        )


def test_duplicate_target_key_is_technical_failure() -> None:
    target = SimpleNamespace(document_index_normalized="1234", work_name="Монтаж")

    with pytest.raises(NumericVerificationFailure, match="TARGET_BINDING_AMBIGUOUS"):
        _reject_duplicate_target_keys((target, target))


@pytest.mark.parametrize(
    ("quantity", "cost", "mismatch"),
    (
        (Decimal("5.00"), Decimal("2.70"), False),
        (Decimal("4.99"), Decimal("2.70"), True),
        (Decimal("5.00"), Decimal("2.69"), True),
    ),
)
def test_numeric_oracle_aggregates_two_contributions_and_marks_each_mismatch(
    tmp_path: Path, monkeypatch, quantity: Decimal, cost: Decimal, mismatch: bool
) -> None:
    job = SimpleNamespace(
        source_names=("first.xlsx", "second.xlsx"),
        source_digests=("digest-first", "digest-second"),
        target=tmp_path / "target.xlsx",
        target_digest="target-digest",
        stage="13.1",
    )
    first_source = SimpleNamespace(
        source_row_id="first-source",
        source_filename="first.xlsx",
        source_sheet="Sheet",
        source_row_number=7,
    )
    second_source = SimpleNamespace(
        source_row_id="second-source",
        source_filename="second.xlsx",
        source_sheet="Sheet",
        source_row_number=8,
    )
    first, second = (
        _review_row_id(job, source.source_row_id) for source in (first_source, second_source)
    )
    rows = tuple(
        ReviewRow(row_id, "Монтаж", "м", Decimal("1"), Decimal("1")) for row_id in (first, second)
    )
    (group,) = build_review_groups(rows)
    decision = ReviewDecision(
        ReviewAction.ACCEPT,
        ReviewMode.QUANTITY_COST,
        "category:target",
        group_id=group.group_id,
    )
    state = SimpleNamespace(
        rows={row.row_id: row for row in rows},
        groups={group.group_id: group},
        grouping=SimpleNamespace(packages=()),
        effective_decisions=lambda: (decision,),
    )
    target = SimpleNamespace(document_index_normalized="1234", work_name="Target")
    calculation = SimpleNamespace(
        target_row=SimpleNamespace(
            selected_quantity=TargetNumericCell(Decimal("5.00"), "5", "NOT_FORMULA", "OK"),
            selected_cost=TargetNumericCell(Decimal("2.70"), "2.7", "NOT_FORMULA", "OK"),
        ),
        quantity=quantity,
        cost=cost,
        trace=SimpleNamespace(
            contributions=(
                SimpleNamespace(candidate_id="candidate-first", source_row_id="first-source"),
                SimpleNamespace(candidate_id="candidate-second", source_row_id="second-source"),
            )
        ),
        status=SimpleNamespace(value="calculated"),
    )
    review = SimpleNamespace(
        state=state, source_batch=SimpleNamespace(rows=(first_source, second_source))
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_numeric_verification.read_reconciliation_target",
        lambda *_args: (None, (target,)),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_numeric_verification._catalog",
        lambda *_args: SimpleNamespace(targets={}),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_numeric_verification._validate_units",
        lambda *_args: None,
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_numeric_verification._selected_matches",
        lambda *_args: (
            SimpleNamespace(
                effective_selected_candidates=(
                    SimpleNamespace(candidate_id="candidate-first", source_row_id=first),
                    SimpleNamespace(candidate_id="candidate-second", source_row_id=second),
                )
            ),
        ),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_numeric_verification._candidate_inclusions",
        lambda *_args: {},
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_numeric_verification.calculate_matches",
        lambda *_args: (calculation,),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_numeric_verification.writer_calculations",
        lambda values: tuple(values),
    )

    failed, checked = verify_numeric(job, review)

    assert failed == (frozenset({first, second}) if mismatch else frozenset())
    assert checked == 2


def test_partial_source_input_is_a_technical_failure(tmp_path: Path, monkeypatch) -> None:
    job = _job(tmp_path)
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_verification.prepare_review",
        lambda *_args: _review(job, safe=True, issues=(object(),)),
    )

    with pytest.raises(VerificationTechnicalFailure, match="INPUT_UNUSABLE") as raised:
        verify_reconciliation(job, ())

    assert len(raised.value.issues) == 1


def test_technical_verification_payload_keeps_bounded_repair_guidance(tmp_path: Path) -> None:
    issue = ReconciliationSourceIssue(
        "WORKBOOK_UNREADABLE",
        "source-1234.xlsx",
        "Не удалось прочитать исходную книгу.",
        "Проверьте файл и повторите загрузку.",
        False,
    )
    job = SimpleNamespace(
        job_id="job-1",
        operation="verify",
        status="failed",
        verification_status=None,
        verification_message=None,
        checked_row_count=0,
        failed_row_count=0,
        source_issues=(issue,),
    )

    payload = job_payload(job)

    assert payload["status"] == "failed" and payload["verification_status"] is None
    assert payload["source_issues"] == [
        {
            "basename": "source-1234.xlsx",
            "comment": "Не удалось прочитать исходную книгу.",
            "repair_hint": "Проверьте файл и повторите загрузку.",
            "can_continue": False,
        }
    ]
    assert str(tmp_path) not in repr(payload)


def test_service_never_marks_an_unreadable_verification_input_as_passed(tmp_path: Path) -> None:
    service = AdminPanelService(tmp_path / "jobs")

    job = service.create_job(
        source_name="source-1234.xlsx",
        source_content=b"PK\x03\x04not-a-workbook",
        target_name="target.xlsx",
        target_content=b"PK\x03\x04not-a-workbook",
        stage="13.1",
        operation="verify",
    )

    payload = job_payload(job)
    assert job.status == "failed" and payload["verification_status"] is None
    assert payload["source_issues"]
    assert "paths" not in repr(payload) and str(tmp_path) not in repr(payload)


def test_single_source_artifact_is_a_red_annotated_copy_with_safe_name(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    job_dir = tmp_path / "job"
    inputs.mkdir()
    job_dir.mkdir()
    source = inputs / "source.xlsx"
    _xlsx(source, "failed")
    original = source.read_bytes()
    job = SimpleNamespace(
        source=source,
        sources=(source,),
        source_names=("Исходник.xlsx",),
        directory=job_dir,
    )

    output, result_name = _write_artifact(job, {source: {"Лист1": {2}}})

    assert result_name == "Проверено_Исходник.xlsx"
    assert output == job_dir / "verification.xlsx"
    assert source.read_bytes() == original
    annotated = load_workbook(output)
    assert annotated["Лист1"]["A2"].fill.fgColor.rgb == "FFFF0000"
    annotated.close()


def test_multiple_source_artifact_contains_each_original_name_once(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    job_dir = tmp_path / "job"
    inputs.mkdir()
    job_dir.mkdir()
    failed_source = inputs / "failed.xlsx"
    clean_source = inputs / "clean.xlsx"
    _xlsx(failed_source, "failed")
    _xlsx(clean_source, "clean")
    clean_bytes = clean_source.read_bytes()
    job = SimpleNamespace(
        source=failed_source,
        sources=(failed_source, clean_source),
        source_names=("Первый.xlsx", "Второй.xlsx"),
        directory=job_dir,
    )

    output, result_name = _write_artifact(job, {failed_source: {"Лист1": {2}}})

    assert result_name == "Проверка_документов.zip"
    with ZipFile(output) as archive:
        assert archive.namelist() == ["Первый.xlsx", "Второй.xlsx"]
        assert archive.read("Второй.xlsx") == clean_bytes
        annotated = load_workbook(BytesIO(archive.read("Первый.xlsx")))
        assert annotated["Лист1"]["A2"].fill.fgColor.rgb == "FFFF0000"
        annotated.close()


def test_multiple_source_artifact_does_not_clobber_existing_output(tmp_path: Path) -> None:
    inputs = tmp_path / "inputs"
    job_dir = tmp_path / "job"
    inputs.mkdir()
    job_dir.mkdir()
    failed_source = inputs / "failed.xlsx"
    clean_source = inputs / "clean.xlsx"
    _xlsx(failed_source, "failed")
    _xlsx(clean_source, "clean")
    output = job_dir / "verification-documents.zip"
    output.write_bytes(b"existing-output")
    job = SimpleNamespace(
        source=failed_source,
        sources=(failed_source, clean_source),
        source_names=("Первый.xlsx", "Второй.xlsx"),
        directory=job_dir,
    )

    with pytest.raises(ExcelWriterSafetyError, match="OUTPUT_EXISTS"):
        _write_artifact(job, {failed_source: {"Лист1": {2}}})

    assert output.read_bytes() == b"existing-output"


def test_multiple_source_artifact_does_not_touch_legacy_temporary_name(
    tmp_path: Path, monkeypatch
) -> None:
    inputs = tmp_path / "inputs"
    job_dir = tmp_path / "job"
    inputs.mkdir()
    job_dir.mkdir()
    failed_source = inputs / "failed.xlsx"
    clean_source = inputs / "clean.xlsx"
    _xlsx(failed_source, "failed")
    _xlsx(clean_source, "clean")
    legacy = job_dir / "verification-source-01.xlsx"
    legacy.write_bytes(b"existing-temporary")
    job = SimpleNamespace(
        source=failed_source,
        sources=(failed_source, clean_source),
        source_names=("Первый.xlsx", "Второй.xlsx"),
        directory=job_dir,
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_verification._annotate",
        lambda *_args: (_ for _ in ()).throw(RuntimeError("annotation failed")),
    )

    with pytest.raises(RuntimeError, match="annotation failed"):
        _write_artifact(job, {failed_source: {"Лист1": {2}}})

    assert legacy.read_bytes() == b"existing-temporary"


@pytest.mark.parametrize(
    ("filename", "expected"),
    (
        ("Проверено.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ("Проверено.xlsm", "application/vnd.ms-excel.sheet.macroEnabled.12"),
        ("Проверка.zip", "application/zip"),
    ),
)
def test_verification_result_mime_types(filename: str, expected: str) -> None:
    assert _result_media_type(filename) == expected


def test_long_macro_workbook_download_name_preserves_extension_and_mime() -> None:
    safe_name = _safe_download_name("Проверено_" + "а" * 150 + ".xlsm")

    assert len(safe_name) == 120
    assert safe_name.endswith(".xlsm")
    assert _result_media_type(safe_name) == "application/vnd.ms-excel.sheet.macroEnabled.12"

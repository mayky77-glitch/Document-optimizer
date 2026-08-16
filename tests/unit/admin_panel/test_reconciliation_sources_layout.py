"""Structural source-header regression contracts."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from report_processor.admin_panel.reconciliation_sources import (
    ReconciliationSourceDescriptor,
    SourceLayoutAmbiguousError,
    _extract_ks2_rows,
    _extract_ks6a_rows,
    _sparse_region_index,
)


def test_merged_cumulative_parent_binds_adjacent_quantity_and_total_cost() -> None:
    workbook = Workbook()
    sheet = workbook.active
    formulas = workbook.copy_worksheet(sheet)
    sheet.merge_cells("B1:B3")
    sheet.merge_cells("C1:C3")
    sheet.merge_cells("D1:E2")
    for candidate in (sheet, formulas):
        candidate["B1"] = "Наименование работ"
        candidate["C1"] = "Единица измерения"
        candidate["D1"] = "Выполнено нарастающим итогом"
        candidate["D3"] = "Количество"
        candidate["E3"] = "Общая стоимость"
        candidate.append(("", "Монтаж", "м", 2, 10))

    rows = _extract_ks6a_rows(
        sheet, formulas, "source:one", ReconciliationSourceDescriptor("source-1234.xlsx")
    )

    workbook.close()
    assert len(rows) == 1
    assert rows[0].cumulative_quantity == 2
    assert rows[0].cumulative_cost == 10


def test_unrelated_later_merge_does_not_invalidate_cumulative_parent() -> None:
    workbook = Workbook()
    sheet = workbook.active
    formulas = workbook.copy_worksheet(sheet)
    for candidate in (sheet, formulas):
        candidate["B1"] = "Перечень строительных работ"
        candidate["C1"] = "Единица измерения"
        candidate["D1"] = "Выполненные работы нарастающим итогом"
        candidate["D3"] = "Количество"
        candidate["E3"] = "Общая стоимость"
        candidate.append(("", "Монтаж", "м", 2, 10))
    sheet.merge_cells("D1:E2")
    sheet.merge_cells("D10:E10")

    rows = _extract_ks6a_rows(
        sheet, formulas, "source:one", ReconciliationSourceDescriptor("source-1234.xlsx")
    )

    workbook.close()
    assert len(rows) == 1
    assert rows[0].work_name_raw == "монтаж"


def test_broad_work_header_is_selected_outside_cumulative_metric_span() -> None:
    workbook = Workbook()
    sheet = workbook.active
    formulas = workbook.copy_worksheet(sheet)
    for candidate in (sheet, formulas):
        candidate["B1"] = "Перечень строительных работ"
        candidate["C1"] = "Единица измерения"
        candidate["D1"] = "Выполненные работы нарастающим итогом"
        candidate["D3"] = "Объём"
        candidate["E3"] = "Сумма затрат"
        candidate.append(("", "Устройство основания", "м²", 3, 30))
    sheet.merge_cells("D1:E2")

    rows = _extract_ks6a_rows(
        sheet, formulas, "source:one", ReconciliationSourceDescriptor("source-1234.xlsx")
    )

    workbook.close()
    assert len(rows) == 1
    assert rows[0].work_name_raw == "устройство основания"
    assert rows[0].cumulative_quantity == 3
    assert rows[0].cumulative_cost == 30


def test_unit_price_leaf_does_not_form_a_metric_pair() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("", "Наименование работ", "Единица", "Количество", "Цена за единицу"))
    sheet.append(("", "Монтаж", "м", 2, 10))

    assert (
        _extract_ks2_rows(sheet, sheet, "source:one", ReconciliationSourceDescriptor("x.xlsx"))
        == ()
    )
    workbook.close()


def test_cumulative_parent_work_stem_does_not_contaminate_metric_children() -> None:
    workbook = Workbook()
    sheet = workbook.active
    formulas = workbook.copy_worksheet(sheet)
    for candidate in (sheet, formulas):
        candidate.append(("", "Наименование", "Ед. изм.", "Выполнено работ нарастающим итогом", ""))
        candidate.append(("", "работ", "", "Количество", "Общая стоимость"))
        candidate.append(("1", "Монтаж", "м", 2, 10))
    sheet.merge_cells("D1:E1")

    rows = _extract_ks6a_rows(
        sheet, formulas, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )

    workbook.close()
    assert len(rows) == 1
    assert rows[0].work_name_raw == "монтаж"


def test_split_local_work_and_unit_roles_bind_inside_direct_region_band() -> None:
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование", "Единица", "", ""),
        ("", "работ", "измерения", "Количество", "Общая стоимость"),
        ("1", "Монтаж", "м", 2, 10),
    ):
        sheet.append(row)

    rows = _extract_ks2_rows(
        sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )

    workbook.close()
    assert len(rows) == 1
    assert rows[0].work_name_raw == "монтаж"


def test_price_lineage_cannot_supply_role_like_unit_descendant() -> None:
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Цена", "", ""),
        ("", "", "Ед. изм.", "Количество", "Общая стоимость"),
        ("1", "Монтаж", "м", 2, 10),
    ):
        sheet.append(row)

    assert (
        _extract_ks2_rows(sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx"))
        == ()
    )
    workbook.close()


def test_shared_schema_selects_canonical_work_role_among_broad_work_headers() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        (
            "",
            "Код работ",
            "Наименование работ и затрат",
            "Ед. изм.",
            "Количество",
            "Общая стоимость",
        )
    )
    sheet.append(("1", "A-01", "Монтаж", "м", 2, 10))

    rows = _extract_ks2_rows(
        sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )

    workbook.close()
    assert len(rows) == 1
    assert rows[0].work_name_raw == "монтаж"


def test_equal_shared_work_role_candidates_fail_closed() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        (
            "",
            "Наименование работ",
            "Наименование работ",
            "Ед. изм.",
            "Количество",
            "Общая стоимость",
        )
    )
    sheet.append(("1", "Монтаж 1", "Монтаж 2", "м", 2, 10))

    assert (
        _extract_ks2_rows(sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx"))
        == ()
    )
    workbook.close()


def test_direct_metric_leaves_below_cumulative_parent_are_rejected() -> None:
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Выполнено нарастающим итогом", ""),
        ("", "", "", "Количество", "Общая стоимость"),
        ("1", "Монтаж", "м", 2, 10),
    ):
        sheet.append(row)
    sheet.merge_cells("D1:E1")

    assert (
        _extract_ks2_rows(sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx"))
        == ()
    )
    workbook.close()


def test_nested_cumulative_span_chain_is_not_admitted_as_direct() -> None:
    workbook = Workbook()
    sheet = workbook.active
    formulas = workbook.copy_worksheet(sheet)
    for candidate in (sheet, formulas):
        candidate.append(("", "Наименование работ", "Ед. изм.", "Выполнено нарастающим итогом", ""))
        candidate.append(("", "", "", "Показатели", ""))
        candidate.append(("", "", "", "Количество", "Общая стоимость"))
        candidate.append(("1", "Монтаж", "м", 2, 10))
    sheet.merge_cells("D1:E1")
    sheet.merge_cells("D2:E2")

    cumulative = _extract_ks6a_rows(
        sheet, formulas, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )
    direct = _extract_ks2_rows(
        sheet, formulas, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )

    workbook.close()
    assert len(cumulative) == 1
    assert direct == ()


def test_cumulative_direct_and_nested_metric_branches_fail_closed_as_ambiguous() -> None:
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ("", "Наименование работ", "Ед. изм.", "Выполнено нарастающим итогом", "", "", ""),
        ("", "", "", "Количество", "Общая стоимость", "Подгруппа", ""),
        ("", "", "", "", "", "Количество", "Общая стоимость"),
        ("1", "Монтаж", "м", 1, 10, 2, 20),
    ):
        sheet.append(row)
    sheet.merge_cells("D1:G1")
    sheet.merge_cells("F2:G2")

    with pytest.raises(SourceLayoutAmbiguousError):
        _extract_ks6a_rows(
            sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx")
        )
    workbook.close()


def test_column_permutation_does_not_change_physical_role_binding() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("", "Ед. изм.", "Описание работ", "Количество", "Общая стоимость"))
    sheet.append(("1", "м", "Монтаж", 2, 10))

    rows = _extract_ks2_rows(
        sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )

    workbook.close()
    assert len(rows) == 1
    assert rows[0].work_name_raw == "монтаж"


def test_direct_roles_merged_vertically_cover_the_full_header_band() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("", "Наименование работ", "Ед. изм.", "", ""))
    sheet.append(("", "", "", "", ""))
    sheet.append(("", "", "", "Количество", "Общая стоимость"))
    sheet.append(("", "", "", "", ""))
    sheet.append(("1", "Монтаж", "м", 2, 10))
    sheet.merge_cells("B1:B3")
    sheet.merge_cells("C1:C3")

    rows = _extract_ks2_rows(
        sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )

    workbook.close()
    assert len(rows) == 1


def test_formula_only_materialized_coordinates_obey_sparse_cap(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import report_processor.admin_panel.reconciliation_sources as sources

    data = Workbook()
    formulas = Workbook()
    formulas.active["A1"] = "=1+1"
    formulas.active["B1"] = "=2+2"
    monkeypatch.setattr(sources, "_REGION_CELL_LIMIT", 1)

    with pytest.raises(SourceLayoutAmbiguousError):
        _sparse_region_index(data.active, formulas.active)

    data.close()
    formulas.close()


def test_styled_empty_cells_do_not_consume_the_sparse_cell_cap(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import report_processor.admin_panel.reconciliation_sources as sources

    workbook = Workbook()
    sheet = workbook.active
    for column in range(1, 4):
        sheet.cell(row=1, column=column).number_format = "0.00"
    monkeypatch.setattr(sources, "_REGION_CELL_LIMIT", 1)

    index = _sparse_region_index(sheet, sheet)

    workbook.close()
    assert index.values == {}
    assert index.columns == ()


def test_nonempty_data_coordinates_obey_sparse_cap(monkeypatch: pytest.MonkeyPatch) -> None:
    import report_processor.admin_panel.reconciliation_sources as sources

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "one"
    sheet["B1"] = "two"
    monkeypatch.setattr(sources, "_REGION_CELL_LIMIT", 1)

    with pytest.raises(SourceLayoutAmbiguousError):
        _sparse_region_index(sheet, sheet)

    workbook.close()


def test_sparse_band_start_jumps_over_tall_merged_coverage(monkeypatch: pytest.MonkeyPatch) -> None:
    import report_processor.admin_panel.reconciliation_sources as sources

    spans = ((1, 2, 1_000_000, 2), (1, 3, 1_000_000, 3))
    index = sources._SparseRegionIndex(
        values={},
        formula_cells=frozenset(),
        spans=spans,
        row_values={},
        column_values={},
        column_rows={},
        occupied_rows=frozenset(),
        occupied_merge_rows=((1, 1_000_000),),
        spans_by_top={1: spans},
        spans_by_bottom={1_000_000: spans},
        span_by_origin={(1, 2): spans[0], (1, 3): spans[1]},
        span_starts=(2, 3),
        spans_by_left=spans,
        covering_span_cache={},
        visit_count=[0],
        columns=(),
        rows=(),
        max_column=3,
        last_sparse_row=1_000_000,
    )
    calls = 0
    original = sources._merge_interval_containing

    def counted(intervals, row):
        nonlocal calls
        calls += 1
        return original(intervals, row)

    monkeypatch.setattr(sources, "_merge_interval_containing", counted)

    assert sources._indexed_band_start(index, 1_000_001) == 1
    assert calls <= 2


def test_rejected_metric_shapes_stop_before_unbounded_role_probes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import report_processor.admin_panel.reconciliation_sources as sources

    workbook = Workbook()
    sheet = workbook.active
    for row in range(1, 6):
        sheet.cell(row=row, column=4, value="Количество")
        sheet.cell(row=row, column=5, value="Общая стоимость")
    monkeypatch.setattr(sources, "_REGION_PROBE_LIMIT", 2)
    index = _sparse_region_index(sheet, sheet)

    with pytest.raises(SourceLayoutAmbiguousError):
        sources._indexed_structural_layouts(index)

    workbook.close()


def test_sparse_region_visit_budget_fails_before_role_cartesian_work(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import report_processor.admin_panel.reconciliation_sources as sources

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("", "Наименование работ", "Ед. изм.", "Количество", "Общая стоимость"))
    sheet.append(("1", "Монтаж", "м", 2, 10))
    monkeypatch.setattr(sources, "_REGION_VISIT_LIMIT", 1)

    with pytest.raises(SourceLayoutAmbiguousError):
        _extract_ks2_rows(sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx"))

    workbook.close()


def test_horizontal_work_merge_nominates_each_material_detail_column() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("", "Наименование работ", "", "Количество", "Общая стоимость", "Ед. изм."))
    sheet.append(("1", "Первая", "Вторая", 1, 10, "м"))
    sheet.merge_cells("B1:C1")

    assert (
        _extract_ks2_rows(sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx"))
        == ()
    )
    workbook.close()


def test_vertically_merged_direct_metric_leaves_end_after_their_spans() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("", "", "", "Количество", "Общая стоимость"))
    sheet.append(("", "Наименование работ", "Ед. изм.", "", ""))
    sheet.append(("1", "Монтаж", "м", 2, 10))
    sheet.merge_cells("D1:D2")
    sheet.merge_cells("E1:E2")

    rows = _extract_ks2_rows(
        sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )

    workbook.close()
    assert len(rows) == 1


def test_vertically_merged_cumulative_metric_leaves_end_after_their_spans() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["B1"] = "Наименование работ"
    sheet["C1"] = "Ед. изм."
    sheet["D1"] = "Выполнено нарастающим итогом"
    sheet["D2"] = "Количество"
    sheet["E2"] = "Общая стоимость"
    sheet.append(("", "", "", "", ""))
    sheet.append(("1", "Монтаж", "м", 2, 10))
    sheet.merge_cells("B1:B3")
    sheet.merge_cells("C1:C3")
    sheet.merge_cells("D1:E1")
    sheet.merge_cells("D2:D3")
    sheet.merge_cells("E2:E3")

    rows = _extract_ks6a_rows(
        sheet, sheet, "source:one", ReconciliationSourceDescriptor("source.xlsx")
    )

    workbook.close()
    assert len(rows) == 1


def test_metric_pair_scan_charges_only_the_adjacent_material_columns() -> None:
    import report_processor.admin_panel.reconciliation_sources as sources

    workbook = Workbook()
    sheet = workbook.active
    for column in range(1, 81, 2):
        sheet.cell(row=1, column=column, value="Количество")
        sheet.cell(row=1, column=column + 1, value="Общая стоимость")
    index = _sparse_region_index(sheet, sheet)

    assert len(sources._indexed_metric_pairs(index, 1, 1, index.max_column)) == 40
    assert index.visit_count[0] == 80
    workbook.close()


def test_sparse_band_start_jumps_over_consecutive_material_rows(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import report_processor.admin_panel.reconciliation_sources as sources

    workbook = Workbook()
    sheet = workbook.active
    for row in range(1, 129):
        sheet.cell(row=row, column=1, value=row)
    index = _sparse_region_index(sheet, sheet)
    calls = 0
    original = sources._merge_interval_containing

    def counted(intervals, row):
        nonlocal calls
        calls += 1
        return original(intervals, row)

    monkeypatch.setattr(sources, "_merge_interval_containing", counted)

    assert sources._indexed_band_start(index, 129) == 1
    assert calls <= 2
    assert index.visit_count[0] <= 2
    workbook.close()

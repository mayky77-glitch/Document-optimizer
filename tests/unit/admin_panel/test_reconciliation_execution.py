from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from report_processor.admin_panel.reconciliation_execution import (
    _Catalog,
    _catalog,
    _feedback_records,
    _normalized_source_digests,
    _review_row_id,
    _selected_matches,
    apply_review,
    calculation_semantic_digest,
    prepare_review,
)
from report_processor.admin_panel.reconciliation_semantic_assist import RUBERT_TINY2_MODEL_REVISION
from report_processor.admin_panel.reconciliation_target import (
    ReconciliationTargetIdentity,
    publish_unchanged_target,
)
from report_processor.reconciliation_review import (
    FeedbackRecord,
    ReviewAction,
    ReviewDecision,
    ReviewMode,
    ReviewRow,
    build_review_groups,
)


def test_prepare_review_hides_leading_zero_row_without_hiding_later_nonzero_row(
    tmp_path, monkeypatch
) -> None:
    zero = ReviewRow("zero", "Нулевая работа", "м", Decimal("0"), Decimal("0"), "target-1")
    nonzero = ReviewRow("nonzero", "Ненулевая работа", "м", Decimal("1"), Decimal("2"), "target-1")
    job = SimpleNamespace(
        source_digests=("source",),
        target_digest="b" * 64,
        target=tmp_path / "target.xlsx",
        stage="13.1",
        directory=tmp_path,
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution._sources",
        lambda _job: SimpleNamespace(rows=(), issues=()),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution.read_reconciliation_target",
        lambda *_args: (object(), ()),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution._catalog",
        lambda _targets: _Catalog({"target-1": "Цель"}, {}),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution._review_rows",
        lambda *_args: (zero, nonzero),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution._available_categories",
        lambda *_args: {"zero": frozenset({"target-1"}), "nonzero": frozenset({"target-1"})},
    )

    result = prepare_review(job, (FeedbackRecord("нулевая работа", "м", ReviewAction.REJECT),))

    assert result.state is not None
    assert tuple(result.state.rows) == ("nonzero",)
    assert result.state.grouping.partition.hidden_rows == (zero,)
    assert result.state.grouping.version_context.model_revision == RUBERT_TINY2_MODEL_REVISION
    assert (
        result.state.target_identity_digest
        == ReconciliationTargetIdentity(job.target_digest, job.stage).target_identity_digest
    )
    assert result.state.group_decisions == {} and result.state.row_decisions == {}


def test_package_version_source_digests_are_normalized_sorted_and_unique() -> None:
    assert _normalized_source_digests((" B ", "a", "b", "A")) == ("a", "b")
    with pytest.raises(ValueError, match="required"):
        _normalized_source_digests(("source", None))


def test_calculation_semantic_digest_binds_calculation_identity_and_order() -> None:
    first = SimpleNamespace(
        calculation_id="calculation-a",
        target_row_id="target-a",
        status=SimpleNamespace(value="calculated"),
        quantity=Decimal("0.00"),
        cost=Decimal("0.00"),
    )
    second = SimpleNamespace(
        calculation_id="calculation-b",
        target_row_id="target-b",
        status=SimpleNamespace(value="calculated"),
        quantity=None,
        cost=Decimal("1.00"),
    )

    assert calculation_semantic_digest((first, second)) == calculation_semantic_digest(
        (second, first)
    )
    assert calculation_semantic_digest((first,)) != calculation_semantic_digest(
        (SimpleNamespace(**{**first.__dict__, "calculation_id": "calculation-drift"}),)
    )


def test_catalog_rejects_duplicate_document_index_and_category() -> None:
    first = SimpleNamespace(work_name="Монтаж", document_index_normalized="1001")
    second = SimpleNamespace(work_name="Монтаж", document_index_normalized="1001")

    with pytest.raises(ValueError, match="DUPLICATE_TARGET_CATEGORY"):
        _catalog((first, second))


def test_prepare_review_projects_duplicate_target_category_as_controlled_target_error(
    tmp_path, monkeypatch
) -> None:
    job = SimpleNamespace(
        source_digests=("source",),
        target_digest="b" * 64,
        target=tmp_path / "target.xlsx",
        stage="13.1",
        directory=tmp_path,
    )
    duplicate = SimpleNamespace(work_name="Монтаж", document_index_normalized="1001")
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution._sources",
        lambda _job: SimpleNamespace(rows=(), issues=()),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution.read_reconciliation_target",
        lambda *_args: (object(), (duplicate, duplicate)),
    )

    result = prepare_review(job, ())

    assert result.state is None and result.target_error is True


def test_selected_matches_rejects_same_physical_source_row_across_upload_ordinals() -> None:
    target = SimpleNamespace(
        document_index_normalized=("1001", ""),
        work_name="Монтаж",
        sheet_name="Отчёт",
        row_number=10,
    )
    catalog = _Catalog({"category": "Монтаж"}, {("1001", "category"): target})
    first = SimpleNamespace(
        source_file_id="source:" + "a" * 64,
        source_sheet="КС-6а",
        source_row_number=20,
        source_filename="source-1001.xlsx",
    )
    second = SimpleNamespace(
        source_filename="copy-1001.xlsx",
        source_file_id="source:" + "a" * 64,
        source_sheet="КС-6а",
        source_row_number=20,
    )
    overrides = {
        "first": SimpleNamespace(action=ReviewAction.ACCEPT, target_category="category"),
        "second": SimpleNamespace(action=ReviewAction.ACCEPT, target_category="category"),
    }
    job = SimpleNamespace(target_digest="b" * 64)

    with pytest.raises(ValueError, match="DUPLICATE_SOURCE_IDENTITY"):
        _selected_matches(
            None,
            overrides,
            catalog,
            job,
            {"first": first, "second": second},
            ((first.source_file_id, "1001"), (second.source_file_id, "1001")),
        )


def test_feedback_records_restore_row_feedback_over_group_feedback() -> None:
    first = ReviewRow("source-a:1", "Монтаж трубы", "м", Decimal("1"), Decimal("2"))
    second = ReviewRow("source-b:1", "Монтаж трубы", "м", Decimal("1"), Decimal("2"))
    (group,) = build_review_groups((first, second))
    state = SimpleNamespace(
        groups={group.group_id: group},
        rows={first.row_id: first, second.row_id: second},
        group_decisions={
            group.group_id: ReviewDecision(
                ReviewAction.ACCEPT,
                ReviewMode.QUANTITY_COST,
                "target-1",
                group_id=group.group_id,
            )
        },
        row_decisions={
            second.row_id: ReviewDecision(
                ReviewAction.REJECT,
                row_id=second.row_id,
            )
        },
    )

    feedback = _feedback_records(state)

    assert feedback == (
        FeedbackRecord(
            group.normalized_name or "",
            group.normalized_unit,
            ReviewAction.ACCEPT,
            "target-1",
            ReviewMode.QUANTITY_COST,
            1,
        ),
        FeedbackRecord("монтаж трубы", "м", ReviewAction.REJECT, sequence=2),
    )


def test_review_row_id_is_stable_and_hides_source_provenance() -> None:
    source_digest = "a" * 64
    source_row_id = f"source:0:{source_digest}:ks6a:455"
    job = SimpleNamespace(target_digest="b" * 64, source_digests=(source_digest,))

    row_id = _review_row_id(job, source_row_id)

    assert row_id == _review_row_id(job, source_row_id)
    assert row_id.startswith("review-row-")
    assert source_digest not in row_id
    assert "ks6a" not in row_id
    assert ":455" not in row_id


def test_unchanged_publisher_removes_link_after_final_verification_failure(
    tmp_path, monkeypatch
) -> None:
    source = tmp_path / "target.xlsx"
    output = tmp_path / "result.xlsx"
    workbook = Workbook()
    workbook.save(source)
    workbook.close()
    digest = sha256(source.read_bytes()).hexdigest()
    calls = 0

    def fail_final_reopen(_path):
        nonlocal calls
        calls += 1
        if calls == 3:
            raise ValueError("final reopen failed")

    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_target._reopen_xlsx",
        fail_final_reopen,
    )

    with pytest.raises(ValueError, match="final reopen failed"):
        publish_unchanged_target(source, output, digest)

    assert output.exists() is False
    assert sha256(source.read_bytes()).hexdigest() == digest


def test_unchanged_publisher_never_links_output_when_identity_probe_fails(
    tmp_path, monkeypatch
) -> None:
    source = tmp_path / "target.xlsx"
    output = tmp_path / "result.xlsx"
    workbook = Workbook()
    workbook.save(source)
    workbook.close()
    digest = sha256(source.read_bytes()).hexdigest()
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_target._file_identity",
        lambda _path: (_ for _ in ()).throw(OSError("stat failed")),
    )

    with pytest.raises(RuntimeError, match="PUBLISH_FAILED"):
        publish_unchanged_target(source, output, digest)

    assert output.exists() is False
    assert sha256(source.read_bytes()).hexdigest() == digest


def test_apply_review_all_rejected_writes_unchanged_target_and_feedback(
    tmp_path, monkeypatch
) -> None:
    row = ReviewRow("source:1", "Монтаж трубы", "м", Decimal("1"), Decimal("2"))
    (group,) = build_review_groups((row,))
    rejected = SimpleNamespace(
        rows={row.row_id: row},
        groups={group.group_id: group},
        group_decisions={
            group.group_id: ReviewDecision(
                ReviewAction.REJECT,
                group_id=group.group_id,
            )
        },
        row_decisions={},
        core_decisions=lambda: (
            ReviewDecision(ReviewAction.REJECT, group_id=group.group_id, version=group.version),
        ),
    )
    job = SimpleNamespace(
        target=tmp_path / "target.xlsx",
        target_digest="b" * 64,
        stage="13.1",
        directory=tmp_path,
        rules_path=None,
    )
    workbook = Workbook()
    workbook.active["A1"] = "unchanged"
    input_stream = BytesIO()
    workbook.save(input_stream)
    workbook.close()
    input_bytes = input_stream.getvalue()
    job.target.write_bytes(input_bytes)
    job.target_digest = sha256(input_bytes).hexdigest()
    output = tmp_path / "result.xlsx"
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution.read_reconciliation_target",
        lambda *_args: (object(), ()),
    )
    monkeypatch.setattr(
        "report_processor.admin_panel.reconciliation_execution._sources",
        lambda _job: SimpleNamespace(rows=()),
    )

    def write_unchanged(source, destination, *_args):
        destination.write_bytes(source.read_bytes())
        return SimpleNamespace(output_sha256=sha256(destination.read_bytes()).hexdigest())

    monkeypatch.setattr("report_processor.excel_writer.write_target_report", write_unchanged)

    written, feedback = apply_review(job, rejected)

    assert written == output
    assert written.read_bytes() == input_bytes
    assert sha256(written.read_bytes()).hexdigest() == sha256(input_bytes).hexdigest()
    reopened = load_workbook(written, read_only=True, data_only=True)
    try:
        assert reopened.active["A1"].value == "unchanged"
    finally:
        reopened.close()
    assert feedback[0].action is ReviewAction.REJECT

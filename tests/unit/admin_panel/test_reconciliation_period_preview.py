"""Read-only structural historical-period target preview regressions."""

from __future__ import annotations

import json
import zipfile
from contextlib import contextmanager
from hashlib import sha256
from xml.etree import ElementTree as ET

import pytest
from openpyxl import Workbook, load_workbook

from report_processor.admin_panel import reconciliation_period_preview
from report_processor.admin_panel import reconciliation_target as reconciliation_target_module
from report_processor.admin_panel.reconciliation_period import ReportingPeriod
from report_processor.admin_panel.reconciliation_period_preview import (  # type: ignore[import-not-found]
    preview_reconciliation_target,
)
from report_processor.admin_panel.reconciliation_target import (
    ReconciliationTargetIdentity,
    ReconciliationTargetScopeError,
    read_reconciliation_target,
)
from report_processor.admin_panel.reconciliation_target_measure import (
    TARGET_MEASURE_SEMANTICS_VERSION,
)
from report_processor.excel_writer import engine as writer_engine
from report_processor.schema import LogicalColumn, SheetType
from report_processor.work_semantics import (
    REPORTING_SCOPE_VERSION,
    TERM_CANONICALIZATION_VERSION,
    UNIT_ONTOLOGY_VERSION,
)


def _digest(value: str) -> str:
    return sha256(value.encode()).hexdigest()


def _target(path, *, second_sheet: bool = False, current: bool = False) -> None:
    workbook = Workbook()
    sheets = [workbook.active]
    if second_sheet:
        sheets.append(workbook.create_sheet())
    for index, sheet in enumerate(sheets, start=1):
        sheet.title = f"Отчёт {index}"
        sheet["C1"], sheet["D1"], sheet["E1"], sheet["F1"], sheet["G1"] = (
            "Индекс документа",
            "Номер этапа",
            "Номер п/п",
            "Наименование работ",
            "Единица измерения",
        )
        sheet["C3"], sheet["D3"], sheet["E3"], sheet["F3"], sheet["G3"] = (
            f"12{index}4",
            "Этап 13.1",
            "1",
            f"Монтаж {index}",
            "м",
        )
        if current:
            sheet.merge_cells("L1:M1")
            sheet["L1"] = "Отчетный период"
        else:
            sheet.merge_cells("L1:M1")
            sheet["L1"] = "Документальная отчетность за весь период"
            sheet["N1"] = "Следующий раздел"
        sheet["L2"], sheet["M2"] = "Количество", "Стоимость"
    workbook.save(path)
    workbook.close()


def _malform_raw_merge_inventory(path, case: str) -> None:
    if case == "split_competing_pair":
        workbook = load_workbook(path)
        sheet = workbook["Отчёт 1"]
        sheet["N1"], sheet["N2"], sheet["O2"] = (
            "Отчетный период",
            "Количество",
            "Стоимость",
        )
        workbook.save(path)
        workbook.close()
    with zipfile.ZipFile(path) as archive:
        members = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    root = ET.fromstring(members["xl/worksheets/sheet1.xml"])
    merges = root.find(f"{namespace}mergeCells")
    assert merges is not None
    if case == "count_mismatch":
        merges.attrib["count"] = str(len(merges) + 1)
    elif case == "multiple_containers":
        ET.SubElement(root, f"{namespace}mergeCells", {"count": "0"})
    elif case == "split_competing_pair":
        additional = ET.SubElement(root, f"{namespace}mergeCells", {"count": "1"})
        ET.SubElement(additional, f"{namespace}mergeCell", {"ref": "N1:O1"})
    else:
        raise AssertionError(case)
    members["xl/worksheets/sheet1.xml"] = ET.tostring(root)
    with zipfile.ZipFile(path, "w") as archive:
        for name, payload in members.items():
            archive.writestr(name, payload)


def test_historical_preview_projects_future_cells_and_leaves_source_unchanged(tmp_path) -> None:
    target = tmp_path / "historical.xlsx"
    _target(target)
    original = target.read_bytes()

    preview = preview_reconciliation_target(
        target, sha256(original).hexdigest(), "13.1", ReportingPeriod.parse("2026-08")
    )

    assert preview.period.value == "2026-08"
    assert preview.plan is not None
    assert preview.plan.source_sha256 == sha256(original).hexdigest()
    assert target.read_bytes() == original
    (row,) = preview.rows
    assert row.work_name == "Монтаж 1"
    assert row.document_index_normalized == "1214"
    assert row.unit == "м"
    assert row.writable is False
    assert row.cell_for(LogicalColumn.CURRENT_PERIOD_QUANTITY).coordinate == "N3"
    assert row.cell_for(LogicalColumn.CURRENT_PERIOD_COST).coordinate == "O3"
    assert row.cell_for(LogicalColumn.CURRENT_PERIOD_QUANTITY).raw_value is None
    assert row.cell_for(LogicalColumn.CURRENT_PERIOD_COST).status == "VIRTUAL_FUTURE_CELL"


def test_historical_preview_maps_each_sheet_to_its_own_anchor(tmp_path) -> None:
    target = tmp_path / "historical.xlsx"
    _target(target, second_sheet=True)

    preview = preview_reconciliation_target(
        target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
    )

    assert [anchor.sheet_name for anchor in preview.plan.anchors] == ["Отчёт 1", "Отчёт 2"]
    assert [
        (row.sheet_name, row.cell_for(LogicalColumn.CURRENT_PERIOD_COST).coordinate)
        for row in preview.rows
    ] == [
        ("Отчёт 1", "O3"),
        ("Отчёт 2", "O3"),
    ]
    assert [
        binding.logical_column
        for binding in preview.schema.column_bindings
        if binding.logical_column
        in {
            LogicalColumn.DOCUMENT_INDEX,
            LogicalColumn.STAGE,
            LogicalColumn.ROW_NUMBER,
            LogicalColumn.WORK_NAME,
            LogicalColumn.UNIT,
        }
    ].count(LogicalColumn.DOCUMENT_INDEX) == 1


def test_historical_plan_uses_physical_boundary_and_is_stage_independent(tmp_path) -> None:
    target = tmp_path / "late-stage-historical.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["C1"], sheet["D1"], sheet["E1"], sheet["F1"], sheet["G1"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
        "Единица измерения",
    )
    sheet.merge_cells("L1:M1")
    sheet["L1"], sheet["L2"], sheet["M2"], sheet["N1"] = (
        "Документальная отчетность за весь период",
        "Количество",
        "Стоимость",
        "Следующий раздел",
    )
    for row, index, stage, name in (
        (3, "1234", "Этап 13.1", "Ранняя строка"),
        (84, "5678", "Этап 13.2", "Поздняя строка"),
    ):
        sheet.cell(row, 3).value = index
        sheet.cell(row, 4).value = stage
        sheet.cell(row, 5).value = str(row - 2)
        sheet.cell(row, 6).value = name
        sheet.cell(row, 7).value = "м"
    unrelated = workbook.create_sheet("Без этапа")
    unrelated["C1"], unrelated["D1"], unrelated["E1"], unrelated["F1"], unrelated["G1"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
        "Единица измерения",
    )
    unrelated.merge_cells("L1:M1")
    unrelated["L1"], unrelated["L2"], unrelated["M2"], unrelated["N1"] = (
        "Документальная отчетность за весь период",
        "Количество",
        "Стоимость",
        "Следующий раздел",
    )
    unrelated["C3"], unrelated["E3"], unrelated["F3"], unrelated["G3"] = (
        "9999",
        "1",
        "Не участник",
        "м",
    )
    workbook.save(target)
    workbook.close()
    digest = sha256(target.read_bytes()).hexdigest()

    early = preview_reconciliation_target(target, digest, "13.1", "2026-08")
    late = preview_reconciliation_target(target, digest, "13.2", "2026-08")

    assert early.plan is not None
    assert late.plan is not None
    assert early.plan.plan_digest == late.plan.plan_digest
    assert [(anchor.sheet_name, anchor.first_detail_row) for anchor in late.plan.anchors] == [
        ("Отчёт", 3)
    ]
    assert [(row.sheet_name, row.row_number, row.work_name) for row in late.rows] == [
        ("Отчёт", 84, "Поздняя строка")
    ]
    assert late.rows[0].cell_for(LogicalColumn.CURRENT_PERIOD_QUANTITY).coordinate == "N84"


def test_partial_unrelated_sheet_is_ignored_but_heterogeneous_participant_fails(tmp_path) -> None:
    target = tmp_path / "historical.xlsx"
    _target(target, second_sheet=True)
    workbook = load_workbook(target)
    unrelated = workbook.create_sheet("Примечания")
    unrelated["A1"] = "Наименование работ"
    workbook.save(target)
    workbook.close()

    preview_reconciliation_target(
        target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
    )

    workbook = load_workbook(target)
    workbook["Отчёт 2"].insert_cols(1)
    workbook.save(target)
    workbook.close()
    with pytest.raises(ReconciliationTargetScopeError, match="BASE_ROLE_HETEROGENEOUS"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )


@pytest.mark.parametrize("header", ("C1", "D1", "E1", "F1", "G1"))
def test_participating_sheet_missing_any_base_role_fails_closed(tmp_path, header) -> None:
    target = tmp_path / "historical.xlsx"
    _target(target, second_sheet=True)
    workbook = load_workbook(target)
    workbook["Отчёт 2"][header] = None
    workbook.save(target)
    workbook.close()

    with pytest.raises(ReconciliationTargetScopeError, match="BASE_ROLE_MISSING"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )


def test_blank_unit_summary_is_not_semantic_detail(tmp_path) -> None:
    target = tmp_path / "current.xlsx"
    _target(target, current=True)
    workbook = load_workbook(target)
    workbook["Отчёт 1"]["G3"] = ""
    workbook.save(target)
    workbook.close()

    with pytest.raises(ReconciliationTargetScopeError, match="STAGE_EMPTY"):
        read_reconciliation_target(target, sha256(target.read_bytes()).hexdigest(), "13.1")


def test_existing_current_pair_keeps_strict_physical_target_identity(tmp_path) -> None:
    target = tmp_path / "current.xlsx"
    _target(target, current=True)
    digest = sha256(target.read_bytes()).hexdigest()

    preview = preview_reconciliation_target(target, digest, "13.1", None)

    assert preview.period is None
    assert preview.plan is None
    assert preview.rows[0].cell_for(LogicalColumn.CURRENT_PERIOD_QUANTITY).coordinate == "L3"
    assert preview.target_identity == ReconciliationTargetIdentity(digest, "13.1")


def test_reconciliation_schema_envelopes_are_writer_coherent(tmp_path) -> None:
    current = tmp_path / "current.xlsx"
    _target(current, current=True)
    schema, rows = read_reconciliation_target(
        current, sha256(current.read_bytes()).hexdigest(), "13.1"
    )

    assert schema.status == "OK"
    assert schema.diagnostics == ()
    assert schema.pair_cardinality == "1"
    assert {row.sheet_name for row in rows} == {item.sheet_name for item in schema.worksheets}
    assert {item.sheet_type for item in schema.worksheets} == {SheetType.ADDITIONAL_REPORT}
    writer_engine._validate_schema_identity(
        current, writer_engine._source_identity(current), schema
    )

    historical = tmp_path / "historical.xlsx"
    _target(historical)
    preview = preview_reconciliation_target(
        historical, sha256(historical.read_bytes()).hexdigest(), "13.1", "2026-08"
    )

    assert preview.schema.status == "OK"
    assert preview.schema.diagnostics == ()
    assert preview.schema.pair_cardinality == str(len(preview.plan.anchors))
    assert preview.schema.period_identity.current_period == "2026-08"
    assert {row.sheet_name for row in preview.rows} == {
        item.sheet_name for item in preview.schema.worksheets
    }


def test_target_identity_is_canonical_and_period_plan_bound() -> None:
    original, plan, changed = _digest("original"), _digest("plan"), _digest("changed")
    first = ReconciliationTargetIdentity(original, "13.1", "2026-08", plan)

    assert (
        first.target_identity_digest
        == ReconciliationTargetIdentity(original, "13.1", "2026-08", plan).target_identity_digest
    )
    assert (
        first.target_identity_digest
        != ReconciliationTargetIdentity(original, "13.1", "2026-09", plan).target_identity_digest
    )
    assert (
        first.target_identity_digest
        != ReconciliationTargetIdentity(changed, "13.1", "2026-08", plan).target_identity_digest
    )
    assert (
        first.target_identity_digest
        != ReconciliationTargetIdentity(original, "13.2", "2026-08", plan).target_identity_digest
    )
    assert (
        first.target_identity_digest
        != ReconciliationTargetIdentity(
            original, "13.1", "2026-08", _digest("changed-plan")
        ).target_identity_digest
    )


def test_target_identity_binds_semantic_contract_versions() -> None:
    original, plan = _digest("original"), _digest("plan")
    identity = ReconciliationTargetIdentity(original, "13.1", "2026-08", plan)

    assert json.loads(identity.canonical_bytes()) == {
        "contract_version": "ReconciliationTargetIdentity-2.0",
        "original_target_digest": original,
        "period": "2026-08",
        "plan_digest": plan,
        "reporting_scope_version": REPORTING_SCOPE_VERSION,
        "selected_stage": "13.1",
        "target_measure_semantics_version": TARGET_MEASURE_SEMANTICS_VERSION,
        "term_canonicalization_version": TERM_CANONICALIZATION_VERSION,
        "unit_ontology_version": UNIT_ONTOLOGY_VERSION,
    }
    assert identity.canonical_bytes() == (
        b'{"contract_version":"ReconciliationTargetIdentity-2.0",'
        b'"original_target_digest":"0682c5f2076f099c34cfdd15a9e063849ed437a49677e6fcc5b4198c76575be5",'
        b'"period":"2026-08",'
        b'"plan_digest":"64879f7d6b960a01909762d911a32d4582c20010c5641ee90278b644a9e3b525",'
        b'"reporting_scope_version":"ReportingScope-2.1",'
        b'"selected_stage":"13.1",'
        b'"target_measure_semantics_version":"ReconciliationTargetMeasure-3.0",'
        b'"term_canonicalization_version":"TermCanonicalization-2.0",'
        b'"unit_ontology_version":"UnitOntology-1.1"}'
    )
    assert (
        identity.target_identity_digest
        == "f14390610bffd4d34229782a5f23b0428526b4e6db7a4895ea1fda85c5693b33"
    )


def test_target_identity_rejects_v1_contract() -> None:
    with pytest.raises(ValueError, match="TARGET_IDENTITY_INVALID"):
        ReconciliationTargetIdentity(
            _digest("original"),
            "13.1",
            contract_version="ReconciliationTargetIdentity-1.0",
        )


@pytest.mark.parametrize(
    ("original", "period", "plan"),
    (
        ("A" * 64, None, None),
        (_digest("original"), None, _digest("plan")),
        (_digest("original"), "2026-08", None),
    ),
)
def test_target_identity_rejects_noncanonical_digest_or_unpaired_period(
    original, period, plan
) -> None:
    with pytest.raises(ValueError, match="TARGET_IDENTITY_INVALID"):
        ReconciliationTargetIdentity(original, "13.1", period, plan)


def test_target_identity_canonicalizes_mutable_reporting_period() -> None:
    class MutablePeriod:
        value = "2026-08"

    source = MutablePeriod()
    identity = ReconciliationTargetIdentity(_digest("original"), "13.1", source, _digest("plan"))
    expected = identity.target_identity_digest
    source.value = "2026-09"

    assert identity.reporting_period == "2026-08"
    assert identity.target_identity_digest == expected


def test_missing_or_tied_base_roles_fail_closed(tmp_path) -> None:
    target = tmp_path / "missing.xlsx"
    _target(target)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["C1"], sheet["D1"], sheet["E1"], sheet["F1"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
    )
    workbook.save(target)
    workbook.close()

    with pytest.raises(ReconciliationTargetScopeError, match="BASE_ROLE_MISSING"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )

    _target(target)
    workbook = load_workbook(target)
    workbook["Отчёт 1"]["H1"] = "Единица измерения"
    workbook.save(target)
    workbook.close()
    with pytest.raises(ReconciliationTargetScopeError, match="BASE_ROLE_AMBIGUOUS"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )


def _content_recovery_target(path) -> None:
    _target(path)
    workbook = load_workbook(path)
    sheet = workbook["Отчёт 1"]
    # These labels deliberately carry no global reconciliation alias.  The
    # detail skeleton remains header-bound, so only document/stage may recover.
    sheet["C1"], sheet["D1"] = "Код листа", "Раздел исполнения"
    values = tuple(sheet.cell(3, column).value for column in range(3, 8))
    for column, value in enumerate(values, start=3):
        sheet.cell(3, column).value = None
        sheet.cell(10, column).value = value
    workbook.save(path)
    workbook.close()


def test_preview_recovers_shifted_document_and_stage_roles_from_one_real_pair(tmp_path) -> None:
    target = tmp_path / "content-recovery.xlsx"
    _content_recovery_target(target)

    preview = preview_reconciliation_target(
        target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
    )

    (row,) = preview.rows
    assert row.document_index_normalized == "1214"
    assert row.stage == "13.1"


def test_content_role_recovery_rejects_accidental_numeric_duplicate_pair(tmp_path) -> None:
    target = tmp_path / "content-ambiguous.xlsx"
    _content_recovery_target(target)
    workbook = load_workbook(target)
    sheet = workbook["Отчёт 1"]
    # A second numeric column on the same structural anchor must not be guessed
    # away merely because it looks like an index.
    sheet["A10"] = "9999"
    workbook.save(target)
    workbook.close()

    with pytest.raises(ReconciliationTargetScopeError, match="BASE_ROLE_AMBIGUOUS"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )


def test_content_role_recovery_rejects_multiple_and_formula_only_pairs(tmp_path) -> None:
    target = tmp_path / "content-multiple.xlsx"
    _content_recovery_target(target)
    workbook = load_workbook(target)
    sheet = workbook["Отчёт 1"]
    sheet["A11"], sheet["B11"], sheet["E11"], sheet["F11"], sheet["G11"] = (
        "5678",
        "Этап 13.2",
        "2",
        "Монтаж 2",
        "м",
    )
    workbook.save(target)
    workbook.close()
    with pytest.raises(ReconciliationTargetScopeError, match="BASE_ROLE_AMBIGUOUS"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )

    _content_recovery_target(target)
    workbook = load_workbook(target)
    workbook["Отчёт 1"]["C10"] = '="1214"'
    workbook.save(target)
    workbook.close()
    with pytest.raises(ReconciliationTargetScopeError, match="BASE_ROLE_MISSING"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )

    _content_recovery_target(target)
    workbook = load_workbook(target)
    sheet = workbook["Отчёт 1"]
    sheet["C10"], sheet["D10"] = None, None
    sheet["A10"], sheet["B11"], sheet["E11"], sheet["F11"], sheet["G11"] = (
        "1214",
        "Этап 13.1",
        "1",
        "Монтаж",
        "м",
    )
    workbook.save(target)
    workbook.close()
    with pytest.raises(ReconciliationTargetScopeError, match="BASE_ROLE_MISSING"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )


def test_preview_rejects_target_mutated_during_planning(tmp_path, monkeypatch) -> None:
    target = tmp_path / "historical.xlsx"
    _target(target)
    digest = sha256(target.read_bytes()).hexdigest()
    planner = reconciliation_period_preview.build_period_insertion_plan

    def mutate_after_plan(*args, **kwargs):
        plan = planner(*args, **kwargs)
        target.write_bytes(target.read_bytes() + b"changed")
        return plan

    monkeypatch.setattr(
        reconciliation_period_preview, "build_period_insertion_plan", mutate_after_plan
    )

    with pytest.raises(ValueError, match="RECONCILIATION_TARGET_CHANGED"):
        preview_reconciliation_target(target, digest, "13.1", "2026-08")


def test_preview_rejects_duplicate_raw_merge_inventory(tmp_path) -> None:
    target = tmp_path / "historical.xlsx"
    _target(target)
    with zipfile.ZipFile(target) as archive:
        members = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    root = ET.fromstring(members["xl/worksheets/sheet1.xml"])
    merges = root.find(f"{namespace}mergeCells")
    assert merges is not None
    merges.attrib["count"] = str(len(merges) + 1)
    ET.SubElement(merges, f"{namespace}mergeCell", {"ref": "L1:M1"})
    members["xl/worksheets/sheet1.xml"] = ET.tostring(root)
    with zipfile.ZipFile(target, "w") as archive:
        for name, payload in members.items():
            archive.writestr(name, payload)

    with pytest.raises(ValueError, match="TARGET_HEADER_WINDOW_INVALID"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
        )


def test_current_read_rejects_overlapping_raw_merge_inventory(tmp_path) -> None:
    target = tmp_path / "current.xlsx"
    _target(target, current=True)
    with zipfile.ZipFile(target) as archive:
        members = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    root = ET.fromstring(members["xl/worksheets/sheet1.xml"])
    merges = root.find(f"{namespace}mergeCells")
    assert merges is not None
    merges.attrib["count"] = str(len(merges) + 1)
    ET.SubElement(merges, f"{namespace}mergeCell", {"ref": "M1:N1"})
    members["xl/worksheets/sheet1.xml"] = ET.tostring(root)
    with zipfile.ZipFile(target, "w") as archive:
        for name, payload in members.items():
            archive.writestr(name, payload)

    with pytest.raises(ValueError, match="TARGET_HEADER_WINDOW_INVALID"):
        read_reconciliation_target(target, sha256(target.read_bytes()).hexdigest(), "13.1")


@pytest.mark.parametrize("case", ("count_mismatch", "multiple_containers", "split_competing_pair"))
def test_current_read_rejects_malformed_raw_merge_container(tmp_path, case: str) -> None:
    target = tmp_path / "current.xlsx"
    _target(target, current=True)
    _malform_raw_merge_inventory(target, case)

    with pytest.raises(ValueError, match="TARGET_HEADER_WINDOW_INVALID"):
        read_reconciliation_target(target, sha256(target.read_bytes()).hexdigest(), "13.1")


@pytest.mark.parametrize("case", ("count_mismatch", "multiple_containers", "split_competing_pair"))
@pytest.mark.parametrize(("current", "period"), ((True, None), (False, "2026-08")))
def test_preview_rejects_malformed_raw_merge_container(
    tmp_path, case: str, current: bool, period: str | None
) -> None:
    target = tmp_path / "target.xlsx"
    _target(target, current=current)
    _malform_raw_merge_inventory(target, case)

    with pytest.raises(ValueError, match="TARGET_HEADER_WINDOW_INVALID"):
        preview_reconciliation_target(
            target, sha256(target.read_bytes()).hexdigest(), "13.1", period
        )


def _many_row_target(path, *, current: bool) -> None:
    _target(path, current=current)
    workbook = load_workbook(path)
    sheet = workbook["Отчёт 1"]
    for row in range(3, 104):
        if row > 3:
            sheet.cell(row, 5).value = row - 2
            sheet.cell(row, 6).value = f"Монтаж {row}"
            sheet.cell(row, 7).value = "м"
        if current:
            sheet.cell(row, 12).value = row - 2
            sheet.cell(row, 13).value = f"=L{row}*2"
    sheet["A999999"] = "far dimension"
    workbook.save(path)
    workbook.close()


def _instrument_open(monkeypatch, module):
    original = module.open_dual_workbook
    counts = {"formula": 0, "value": 0, "cell": 0}

    @contextmanager
    def instrumented(request):
        with original(request) as session:
            for view, key in (
                (session.formula_workbook, "formula"),
                (session.value_workbook, "value"),
            ):
                for sheet in view.worksheets:
                    source, cell = sheet._get_source, sheet.cell

                    def counted_source(source=source, key=key):
                        counts[key] += 1
                        return source()

                    def counted_cell(*args, cell=cell, **kwargs):
                        counts["cell"] += 1
                        return cell(*args, **kwargs)

                    sheet._get_source = counted_source
                    sheet.cell = counted_cell
            yield session

    monkeypatch.setattr(module, "open_dual_workbook", instrumented)
    return counts


def _bounded_snapshot_cell_access(monkeypatch):
    original = reconciliation_target_module._SnapshotWorksheet.cell
    calls = 0

    def counted(self, *args, **kwargs):
        nonlocal calls
        calls += 1
        return original(self, *args, **kwargs)

    monkeypatch.setattr(reconciliation_target_module._SnapshotWorksheet, "cell", counted)
    return lambda: calls


def _shifted_header_target(path, *, current: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["C50"], sheet["D50"], sheet["E50"], sheet["F50"], sheet["G50"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
        "Единица измерения",
    )
    sheet.merge_cells("L50:M50")
    sheet["L50"] = "Отчетный период" if current else "Документальная отчетность за весь период"
    if not current:
        sheet["N50"] = "Следующий раздел"
    sheet["L51"], sheet["M51"] = "Количество", "Стоимость"
    sheet["C52"], sheet["D52"], sheet["E52"], sheet["F52"], sheet["G52"] = (
        "1234",
        "Этап 13.1",
        "1",
        "Монтаж",
        "м",
    )
    if current:
        sheet["L52"], sheet["M52"] = 12, 3.5
    sheet["XFD999999"] = "irrelevant"
    workbook.save(path)
    workbook.close()


def test_current_pair_discovery_ignores_far_dimension_column(tmp_path, monkeypatch) -> None:
    target = tmp_path / "shifted-current.xlsx"
    _shifted_header_target(target, current=True)
    access_count = _bounded_snapshot_cell_access(monkeypatch)

    schema, (row,) = read_reconciliation_target(
        target, sha256(target.read_bytes()).hexdigest(), "13.1"
    )

    assert access_count() < 1_000
    assert schema.pair_cardinality == "1"
    assert row.cell_for(LogicalColumn.CURRENT_PERIOD_QUANTITY).coordinate == "L52"
    assert row.selected_quantity.value == 12


def test_preview_discovery_ignores_far_dimension_column(tmp_path, monkeypatch) -> None:
    target = tmp_path / "shifted-historical.xlsx"
    _shifted_header_target(target, current=False)
    original = target.read_bytes()
    access_count = _bounded_snapshot_cell_access(monkeypatch)

    preview = preview_reconciliation_target(target, sha256(original).hexdigest(), "13.1", "2026-08")

    assert access_count() < 1_000
    assert target.read_bytes() == original
    assert preview.schema.status == "OK"
    assert preview.schema.pair_cardinality == "1"
    assert preview.rows[0].cell_for(LogicalColumn.CURRENT_PERIOD_QUANTITY).coordinate == "N52"


def test_public_reconciliation_read_uses_one_snapshot_parse_per_view(tmp_path, monkeypatch) -> None:
    target = tmp_path / "current-many.xlsx"
    _many_row_target(target, current=True)
    counts = _instrument_open(monkeypatch, reconciliation_target_module)

    schema, rows = read_reconciliation_target(
        target, sha256(target.read_bytes()).hexdigest(), "13.1"
    )

    assert counts == {"formula": 1, "value": 1, "cell": 0}
    assert len(rows) == 101
    assert schema.object_blocks[0].start_row == 3
    assert rows[-1].selected_quantity.value == 101
    assert rows[-1].cell_for(LogicalColumn.CURRENT_PERIOD_COST).formula.formula == "=L103*2"


def test_public_reconciliation_preview_uses_one_snapshot_parse_per_view(
    tmp_path, monkeypatch
) -> None:
    target = tmp_path / "historical-many.xlsx"
    _many_row_target(target, current=False)
    counts = _instrument_open(monkeypatch, reconciliation_period_preview)

    preview = preview_reconciliation_target(
        target, sha256(target.read_bytes()).hexdigest(), "13.1", "2026-08"
    )

    assert counts == {"formula": 1, "value": 1, "cell": 0}
    assert len(preview.rows) == 101
    assert preview.schema.object_blocks[0].start_row == 3
    assert preview.rows[-1].selected_quantity.value is None
    assert preview.rows[-1].cell_for(LogicalColumn.CURRENT_PERIOD_COST).coordinate == "O103"

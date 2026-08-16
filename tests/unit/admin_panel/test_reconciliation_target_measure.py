"""Structural current-period target-measure discovery contracts."""

from __future__ import annotations

import zipfile

import pytest
from openpyxl import Workbook

from report_processor.admin_panel import reconciliation_target_measure
from report_processor.admin_panel.reconciliation_target_measure import (
    BoundedHeaderWindow,
    ReconciliationTargetMeasureError,
    discover_historical_target_measures,
    discover_target_measures,
    raw_worksheet_merge_ranges,
)


def _workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["B3"] = "1234"
    sheet["C3"] = "Этап 13.1"
    sheet["D3"] = "1"
    sheet["E3"] = "Монтаж"
    return workbook, sheet


def _pair(
    sheet,
    start: int,
    parent: str | None,
    quantity: str = "Количество",
    cost: str = "Стоимость",
):
    if parent is not None:
        sheet.merge_cells(start_row=1, start_column=start, end_row=1, end_column=start + 1)
        sheet.cell(1, start).value = parent
    sheet.cell(2, start).value = quantity
    sheet.cell(2, start + 1).value = cost


def _raw_merges_payload(declared_count: str, physical_count: int) -> bytes:
    references = "".join(
        f'<mergeCell ref="A{index}:B{index}"/>' for index in range(1, physical_count + 1)
    )
    return (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<mergeCells count="{declared_count}">{references}</mergeCells>'
        "</worksheet>"
    ).encode()


def _raw_merge_workbook(tmp_path, payload: bytes):
    source = tmp_path / "raw-merges.xlsx"
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr("xl/worksheets/sheet1.xml", payload)
    return source


@pytest.mark.parametrize(
    ("declared_count", "physical_count"),
    (("0", 0), ("1", 1), ("64", 64), ("٤٠٩٦", 4_096)),
)
def test_raw_worksheet_merge_count_uses_numeric_limit(
    tmp_path, monkeypatch, declared_count: str, physical_count: int
) -> None:
    source = _raw_merge_workbook(tmp_path, _raw_merges_payload(declared_count, physical_count))
    monkeypatch.setattr(
        reconciliation_target_measure,
        "worksheet_parts",
        lambda _source: {"Отчёт": "xl/worksheets/sheet1.xml"},
    )

    assert len(raw_worksheet_merge_ranges(source, "Отчёт")) == physical_count


def test_raw_worksheet_merge_count_rejects_numeric_value_above_limit(tmp_path, monkeypatch) -> None:
    source = _raw_merge_workbook(tmp_path, _raw_merges_payload("4097", 0))
    monkeypatch.setattr(
        reconciliation_target_measure,
        "worksheet_parts",
        lambda _source: {"Отчёт": "xl/worksheets/sheet1.xml"},
    )

    with pytest.raises(ReconciliationTargetMeasureError, match="TARGET_HEADER_WINDOW_INVALID"):
        raw_worksheet_merge_ranges(source, "Отчёт")


@pytest.mark.parametrize(("declared_count", "physical_count"), (("0", 1), ("1", 0)))
def test_raw_worksheet_merge_count_rejects_declared_physical_mismatch(
    tmp_path, monkeypatch, declared_count: str, physical_count: int
) -> None:
    source = _raw_merge_workbook(tmp_path, _raw_merges_payload(declared_count, physical_count))
    monkeypatch.setattr(
        reconciliation_target_measure,
        "worksheet_parts",
        lambda _source: {"Отчёт": "xl/worksheets/sheet1.xml"},
    )

    with pytest.raises(ReconciliationTargetMeasureError, match="TARGET_HEADER_WINDOW_INVALID"):
        raw_worksheet_merge_ranges(source, "Отчёт")


def test_later_unmerged_same_month_pair_wins_over_historical_documentary_pair() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 10, "Документальная отчетность за весь период")
    sheet["L1"] = "Август 2026"
    sheet["M1"] = "Август 2026"
    _pair(sheet, 12, None)

    pairs = discover_target_measures(workbook, {sheet.title: 3})

    assert [(pair.quantity_letter, pair.cost_letter) for pair in pairs] == [("L", "M")]


@pytest.mark.parametrize("month", ("Март", "Май"))
def test_month_only_current_pair_wins_over_historical_pair(month: str) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 10, "Документальная отчетность за весь период")
    sheet["L1"], sheet["M1"] = month, month
    _pair(sheet, 12, None)

    (pair,) = discover_target_measures(workbook, {sheet.title: 3})

    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")


def test_common_merged_current_parent_is_sufficient_without_calendar_month() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Отчетный период")

    (pair,) = discover_target_measures(workbook, {sheet.title: 3})

    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")


def test_physical_header_window_ignores_far_dimension_column(monkeypatch) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Отчетный период")
    sheet["XFD999999"] = "irrelevant"
    original = sheet.cell
    accesses = 0

    def counted(*args, **kwargs):
        nonlocal accesses
        accesses += 1
        return original(*args, **kwargs)

    monkeypatch.setattr(sheet, "cell", counted)
    (pair,) = discover_target_measures(workbook, {sheet.title: 3})

    assert accesses < 100
    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")


def test_header_work_budget_rejects_wide_intersecting_merge_before_cell_reads(monkeypatch) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Отчетный период")
    sheet.merge_cells("A1:XFD1")
    original = sheet.cell
    accesses = 0

    def counted(*args, **kwargs):
        nonlocal accesses
        accesses += 1
        return original(*args, **kwargs)

    monkeypatch.setattr(sheet, "cell", counted)
    with pytest.raises(ReconciliationTargetMeasureError, match="TARGET_HEADER_WINDOW_INVALID"):
        discover_target_measures(workbook, {sheet.title: 52})

    assert accesses == 0


def test_supplied_header_window_must_equal_actual_competing_pair_evidence() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Отчетный период")
    _pair(sheet, 14, "Отчетный период")
    supplied = BoundedHeaderWindow(1, 2, (12, 13))

    with pytest.raises(ReconciliationTargetMeasureError, match="TARGET_HEADER_WINDOW_INVALID"):
        discover_target_measures(workbook, {sheet.title: 3}, header_windows={sheet.title: supplied})


@pytest.mark.parametrize(
    ("left", "right"),
    (("Август 2026", "Сентябрь 2026"), ("Период", "Период")),
)
def test_missing_or_conflicting_period_identity_fails_closed(left: str, right: str) -> None:
    workbook, sheet = _workbook()
    sheet["L1"], sheet["M1"] = left, right
    _pair(sheet, 12, None)

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


def test_unit_price_and_duplicate_physical_header_evidence_are_not_extra_candidates() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 10, "Текущий отчетный период", cost="Цена за единицу")
    _pair(sheet, 12, "Текущий отчетный период")
    sheet.merge_cells("L2:L3")
    sheet.merge_cells("M2:M3")

    (pair,) = discover_target_measures(workbook, {sheet.title: 4})

    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")


@pytest.mark.parametrize(
    ("quantity_header", "cost_header"),
    (
        ("Март Апрель Количество", "Март Апрель Стоимость"),
        ("Март Количество", "Апрель Стоимость"),
    ),
)
def test_multiple_month_mentions_are_not_an_unambiguous_period(
    quantity_header: str, cost_header: str
) -> None:
    workbook, sheet = _workbook()
    sheet["L1"], sheet["M1"] = quantity_header, cost_header

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 2})


@pytest.mark.parametrize(
    ("quantity_header", "cost_header"),
    (
        ("Март Апрель Количество", "Март Апрель Стоимость"),
        ("Март Количество", "Апрель Стоимость"),
    ),
)
def test_current_parent_does_not_override_conflicting_calendar_evidence(
    quantity_header: str, cost_header: str
) -> None:
    workbook, sheet = _workbook()
    sheet.merge_cells("L1:M1")
    sheet["L1"] = "Текущий отчетный период"
    sheet["L2"], sheet["M2"] = quantity_header, cost_header

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


def test_month_identity_requires_equal_year_presence() -> None:
    workbook, sheet = _workbook()
    sheet["L1"], sheet["M1"] = "Август", "Август 2026"
    _pair(sheet, 12, None)

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


@pytest.mark.parametrize(
    "scope", ("Отчетный период с начала строительства", "Отчетный период итого")
)
def test_cumulative_scope_conflicts_reject_otherwise_current_pair(scope: str) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, scope)

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


def test_total_cost_leaf_with_vsego_is_not_a_historical_conflict() -> None:
    workbook, sheet = _workbook()
    sheet["L1"], sheet["M1"] = "Май", "Май"
    _pair(sheet, 12, None, cost="Стоимость всего")

    (pair,) = discover_target_measures(workbook, {sheet.title: 3})

    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")


@pytest.mark.parametrize(
    "cost",
    ("тыс. руб.", "миллионов рублей", "миллиардах рублей", "млн RUB", "тыс. ₽"),
)
def test_scaled_rub_leaf_without_cost_word_is_a_total_cost(cost: str) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Текущий отчетный период", cost=cost)

    (pair,) = discover_target_measures(workbook, {sheet.title: 3})

    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")


def test_scaled_rub_leaf_under_historical_parent_does_not_become_current() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Документальная отчетность за весь период", cost="млн рублей")
    sheet["N3"] = "suffix"

    (historical,) = discover_historical_target_measures(workbook, {sheet.title: 3})

    assert (historical.quantity_letter, historical.cost_letter) == ("L", "M")
    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


@pytest.mark.parametrize(
    "cost",
    (
        "Цена тыс. руб./м2",
        "Тариф миллионов рублей / шт.",
        "тыс. руб. / пог.м.",
        "млн руб./комплект",
        "млн рублей / маш.-час",
        "млн RUB/кВт·ч",
        "тыс. руб. / 1 м³",
        "млн рублей за машино-час",
        "тыс. руб. на 1 м2",
        "млн руб. на квт ч",
        "млн руб. за квадратный метр",
        "млн руб. за кубический метр",
        "млн руб. за погонный метр",
        "млн руб. за Гкал",
        "млн руб. за человеко-час",
        "Единичная стоимость, млн руб.",
    ),
)
def test_price_or_per_unit_scaled_rub_is_not_a_total_cost(cost: str) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Текущий отчетный период", cost=cost)

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


@pytest.mark.parametrize(
    "price_label",
    ("Цена", "Цены", "Цену", "Ценой", "Цене", "Ценам", "Ценами", "цен."),
)
def test_standalone_price_label_inflections_are_not_total_costs(price_label: str) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Текущий отчетный период", cost=f"{price_label}, млн руб.")

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


@pytest.mark.parametrize(
    "cost",
    (
        "Цена в текущих ценах, млн руб.",
        "Тариф в базисных ценах, млн руб.",
        "Расценка в текущих ценах, млн руб.",
        "Единичная стоимость в базисных ценах, млн руб.",
    ),
)
def test_direct_price_labels_are_not_masked_by_context(cost: str) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Текущий отчетный период", cost=cost)

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


@pytest.mark.parametrize(
    "cost",
    (
        "Стоимость в текущих ценах",
        "Сумма в базисных ценах",
        "Стоимость в ценах 2026",
        "Сумма в базисном уровне цен",
    ),
)
def test_contextual_prices_are_not_unit_prices(cost: str) -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Текущий отчетный период", cost=cost)

    (pair,) = discover_target_measures(workbook, {sheet.title: 3})

    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")


@pytest.mark.parametrize(
    "cost", ("Стоимость, млн руб. за отчетный период", "Стоимость, млн руб. за весь период")
)
def test_currency_over_a_reporting_scope_is_not_a_unit_rate(cost: str) -> None:
    normalized = reconciliation_target_measure._text(cost)

    assert reconciliation_target_measure._total_cost_leaf(normalized)
    assert not reconciliation_target_measure._unit_price(normalized)


@pytest.mark.parametrize(
    "cost",
    (
        "Стоимость, млн руб. за месяц",
        "Стоимость, млн руб. за квартал",
        "Стоимость, млн руб. за год",
        "Стоимость, млн руб. за выполненные работы",
        "Стоимость, млн руб. на дату отчета",
        "Стоимость, млн руб. за все СМР",
        "Стоимость, млн руб. за два дня",
        "Стоимость, млн руб. за шесть дней",
        "Стоимость, млн руб. за десять дней",
        "Стоимость, млн руб. за 1 этап",
        "Стоимость, млн руб. за шесть этапов",
    ),
)
def test_currency_over_a_total_scope_is_not_a_unit_rate(cost: str) -> None:
    normalized = reconciliation_target_measure._text(cost)

    assert reconciliation_target_measure._total_cost_leaf(normalized)
    assert not reconciliation_target_measure._unit_price(normalized)


@pytest.mark.parametrize(
    "cost",
    (
        "Стоимость, млн руб. за м2 за весь период",
        "Стоимость, млн руб. за человеко-час за выполненные работы",
    ),
)
def test_proven_leading_unit_remains_rate_before_later_total_scope(cost: str) -> None:
    normalized = reconciliation_target_measure._text(cost)

    assert reconciliation_target_measure._currency_preposition_scope(normalized) == "rate"
    assert reconciliation_target_measure._unit_price(normalized)
    assert not reconciliation_target_measure._total_cost_leaf(normalized)


def test_unknown_currency_preposition_tail_fails_closed_instead_of_becoming_total_cost() -> None:
    normalized = reconciliation_target_measure._text("Стоимость, млн руб. за оборудование")

    assert reconciliation_target_measure._currency_preposition_scope(normalized) == "unknown"
    assert not reconciliation_target_measure._unit_price(normalized)
    assert not reconciliation_target_measure._total_cost_leaf(normalized)


def test_rubka_is_not_a_ruble_currency_form() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Текущий отчетный период", cost="млн рубка")

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_MISSING"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


def test_competing_scaled_rub_pairs_fail_closed() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Текущий отчетный период", cost="тыс. рублей")
    _pair(sheet, 14, "Текущий отчетный период", cost="млрд руб.")

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_AMBIGUOUS"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


def test_two_distinct_current_period_pairs_are_ambiguous() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Текущий отчетный период")
    _pair(sheet, 14, "Текущий отчетный период")

    with pytest.raises(
        ReconciliationTargetMeasureError, match="TARGET_CURRENT_PERIOD_PAIR_AMBIGUOUS"
    ):
        discover_target_measures(workbook, {sheet.title: 3})


def test_pairs_are_sheet_local_and_can_use_different_columns() -> None:
    workbook, first = _workbook()
    _pair(first, 12, "Текущий отчетный период")
    second = workbook.create_sheet("Дополнение")
    second["B3"], second["C3"], second["D3"], second["E3"] = "1235", "Этап 13.1", "1", "Монтаж"
    _pair(second, 14, "Текущий отчетный период")

    pairs = discover_target_measures(workbook, {first.title: 3, second.title: 3})

    assert [(pair.sheet_name, pair.quantity_letter, pair.cost_letter) for pair in pairs] == [
        ("Дополнение", "N", "O"),
        ("Отчёт", "L", "M"),
    ]


def test_documentary_pair_is_a_separate_fail_closed_insertion_anchor() -> None:
    workbook, sheet = _workbook()
    _pair(sheet, 12, "Документальная отчетность за весь период")
    sheet["N3"] = "suffix"

    (pair,) = discover_historical_target_measures(workbook, {sheet.title: 3})

    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")


def test_nested_historical_parent_is_selected_over_a_broad_outer_ancestor() -> None:
    workbook, sheet = _workbook()
    sheet.merge_cells("J1:N1")
    sheet["J1"] = "Показатели"
    sheet.merge_cells("L2:M2")
    sheet["L2"] = "Документальная отчетность за весь период"
    sheet["L3"], sheet["M3"] = "Количество", "Стоимость"
    sheet["N4"] = "suffix"

    (pair,) = discover_historical_target_measures(workbook, {sheet.title: 4})

    assert (pair.quantity_letter, pair.cost_letter) == ("L", "M")
    assert pair.parent_span == (2, 12, 2, 13)
    assert pair.historical_parent_label == "документальная отчетность за весь период"


def test_competing_historical_ancestors_fail_closed() -> None:
    workbook, sheet = _workbook()
    sheet.merge_cells("J1:N1")
    sheet["J1"] = "Накопленные показатели"
    sheet.merge_cells("L2:M2")
    sheet["L2"] = "Документальная отчетность за весь период"
    sheet["L3"], sheet["M3"] = "Количество", "Стоимость"
    sheet["N4"] = "suffix"

    with pytest.raises(ReconciliationTargetMeasureError, match="TARGET_HISTORICAL_PAIR_MISSING"):
        discover_historical_target_measures(workbook, {sheet.title: 4})

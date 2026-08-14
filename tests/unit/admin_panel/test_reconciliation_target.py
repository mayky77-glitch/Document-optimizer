"""Fail-closed target index and stage selection contracts."""

from __future__ import annotations

from dataclasses import replace
from decimal import Decimal
from hashlib import sha256
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from fixtures.calculation.builders import calculation_rule_set, calculation_source_row, match_result
from report_processor.admin_panel.reconciliation_target import (
    ReconciliationTargetInputError,
    ReconciliationTargetScopeError,
    publish_unchanged_target,
    read_reconciliation_target,
    resolve_reconciliation_stage,
    terminal_index,
    writer_calculations,
)
from report_processor.admin_panel.reconciliation_target_measure import TargetMeasurePair
from report_processor.calculation import calculate_matches
from report_processor.excel import WorkbookOpenRequest, open_dual_workbook
from report_processor.excel_writer import write_target_report
from report_processor.processing.adapters import _materialized
from report_processor.quality_control import WriteDecision
from report_processor.schema import LogicalColumn


class _SparseSheet:
    title = "Отчёт"
    max_row = 1_000_000

    def __init__(self) -> None:
        self.calls = 0
        self._cells = {
            (1, 2): SimpleNamespace(value="Индекс документа"),
            (1, 3): SimpleNamespace(value="Номер этапа"),
            (1, 4): SimpleNamespace(value="Номер п/п"),
            (1, 5): SimpleNamespace(value="Наименование работ"),
            (1, 6): SimpleNamespace(value="Единица измерения"),
            (3, 2): SimpleNamespace(value="1234"),
            (3, 3): SimpleNamespace(value="Этап 13.1"),
            (3, 4): SimpleNamespace(value="1"),
            (3, 5): SimpleNamespace(value="Монтаж"),
            (3, 6): SimpleNamespace(value="м"),
            (4, 2): SimpleNamespace(value="5678"),
            (4, 3): SimpleNamespace(value="Этап 13.2"),
            (4, 4): SimpleNamespace(value="2"),
            (4, 5): SimpleNamespace(value="Монтаж 2"),
            (4, 6): SimpleNamespace(value="м"),
            (1_000_000, 99): SimpleNamespace(value="нерелевантно"),
        }

    def cell(self, row, column):
        self.calls += 1
        return self._cells.get((row, column), SimpleNamespace(value=None))


def test_terminal_index_rejects_year_and_ambiguous_values() -> None:
    assert terminal_index("1234") == "1234"
    assert terminal_index("10.02.0123") == "0123"
    assert terminal_index("1234 2025") is None
    assert terminal_index("2025") is None
    assert terminal_index("10.02.2025") is None
    assert terminal_index("1234 (1) and 5678 (1)") is None


def test_stage_resolution_discovers_exactly_one_stage_only() -> None:
    assert resolve_reconciliation_stage(("13.1",), None) == "13.1"
    assert resolve_reconciliation_stage(("13.1",), "13.1") == "13.1"
    with pytest.raises(ReconciliationTargetScopeError, match="MISSING"):
        resolve_reconciliation_stage(("13.1",), "99.9")
    with pytest.raises(ReconciliationTargetScopeError, match="AMBIGUOUS"):
        resolve_reconciliation_stage(("13.1", "13.2"), None)
    with pytest.raises(ReconciliationTargetScopeError, match="EMPTY"):
        resolve_reconciliation_stage((), None)


def test_stage_discovery_scans_only_sparse_role_rows_and_stops_at_maximum() -> None:
    from report_processor.admin_panel.reconciliation_target import _physical_snapshot, _valid_stages

    sheet = _SparseSheet()
    columns = {
        logical: SimpleNamespace(column_index=index, column_letter=letter, header_text="")
        for logical, index, letter in (
            (LogicalColumn.DOCUMENT_INDEX, 2, "B"),
            (LogicalColumn.STAGE, 3, "C"),
            (LogicalColumn.ROW_NUMBER, 4, "D"),
            (LogicalColumn.WORK_NAME, 5, "E"),
            (LogicalColumn.UNIT, 6, "F"),
        )
    }
    workbook = SimpleNamespace(worksheets=(sheet,))

    assert _valid_stages(
        workbook, {sheet.title: columns}, {sheet.title: _physical_snapshot(sheet)}, maximum=1
    ) == ("13.1", "13.2")
    assert sheet.calls == 0


def test_read_only_sparse_role_scan_ignores_inflated_dimension(tmp_path) -> None:
    from report_processor.admin_panel.reconciliation_target import _physical_snapshot, _role_rows

    target = tmp_path / "sparse.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet["B1"], sheet["C1"], sheet["D1"], sheet["E1"], sheet["F1"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
        "Единица измерения",
    )
    sheet["B3"], sheet["C3"], sheet["D3"], sheet["E3"], sheet["F3"] = (
        "1234",
        "Этап 13.1",
        "1",
        "Монтаж",
        "м",
    )
    sheet["A999999"] = "хвост"
    workbook.save(target)
    workbook.close()
    columns = {
        logical: SimpleNamespace(column_index=index)
        for logical, index in (
            (LogicalColumn.DOCUMENT_INDEX, 2),
            (LogicalColumn.STAGE, 3),
            (LogicalColumn.ROW_NUMBER, 4),
            (LogicalColumn.WORK_NAME, 5),
            (LogicalColumn.UNIT, 6),
        )
    }

    source = _materialized(target, "target:sparse")
    with open_dual_workbook(WorkbookOpenRequest(source)) as session:
        worksheet = session.formula_workbook.active
        snapshot = _physical_snapshot(worksheet)
        assert _role_rows(snapshot, columns) == (1, 3)
        assert snapshot.inspected == (3, 11)


def test_read_only_role_projection_opens_each_view_once_for_many_rows(tmp_path) -> None:
    from report_processor.admin_panel.reconciliation_target import (
        _base_roles,
        _enumerate_stages,
        _first_detail_rows,
        _rows,
        _session_snapshots,
    )
    from report_processor.schema import analyze_workbook_schema

    target = tmp_path / "many.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["B1"], sheet["C1"], sheet["D1"], sheet["E1"], sheet["F1"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
        "Единица измерения",
    )
    sheet.merge_cells("L1:M1")
    sheet["L1"] = "Отчетный период"
    sheet["L2"], sheet["M2"] = "Количество", "Стоимость"
    for row in range(3, 104):
        if row == 3:
            sheet["B3"], sheet["C3"] = "1234", "Этап 13.1"
        sheet.cell(row, 4).value = row - 2
        sheet.cell(row, 5).value = f"Монтаж {row}"
        sheet.cell(row, 6).value = "м"
        sheet.cell(row, 12).value = row - 2
        sheet.cell(row, 13).value = f"=L{row}*2"
    workbook.save(target)
    workbook.close()

    source = _materialized(target, "target:many")
    with open_dual_workbook(WorkbookOpenRequest(source)) as session:
        roles = _base_roles(analyze_workbook_schema(session))
        formula_sheet, value_sheet = session.formula_workbook.active, session.value_workbook.active
        counts = {"formula": 0, "value": 0}
        formula_open, value_open = formula_sheet._get_source, value_sheet._get_source

        def formula_source():
            counts["formula"] += 1
            return formula_open()

        def value_source():
            counts["value"] += 1
            return value_open()

        formula_sheet._get_source = formula_source
        value_sheet._get_source = value_source
        formula_snapshots, value_snapshots = _session_snapshots(session, roles)
        stage = _enumerate_stages(session.formula_workbook, roles, formula_snapshots)[0]
        details = _first_detail_rows(session.formula_workbook, stage, roles, formula_snapshots)
        rows = tuple(
            _rows(
                session,
                stage,
                (TargetMeasurePair("Отчёт", 12, 13, "Количество", "Стоимость"),),
                roles,
                formula_snapshots,
                value_snapshots,
            )
        )

    assert counts == {"formula": 1, "value": 1}
    assert details == {"Отчёт": 3}
    assert len(rows) == 101
    assert rows[0].work_name == "Монтаж 3"
    assert rows[0].selected_quantity.value == Decimal("1")
    assert rows[0].cell_for(LogicalColumn.CURRENT_PERIOD_COST).formula.formula == "=L3*2"


@pytest.mark.parametrize("stage", ("13.1", None), ids=("selected", "no-selected"))
def test_macro_enabled_target_is_rejected_before_reconciliation_review(tmp_path, stage) -> None:
    target = tmp_path / "target.xlsm"
    target.write_bytes(b"not-opened")

    with pytest.raises(ReconciliationTargetInputError, match=r"Целевой отчёт \.xlsm"):
        read_reconciliation_target(target, "digest", stage)


def test_macro_enabled_target_is_rejected_on_no_selected_publish_path(tmp_path) -> None:
    target = tmp_path / "target.xlsm"
    target.write_bytes(b"not-opened")

    with pytest.raises(ReconciliationTargetInputError, match=r"Целевой отчёт \.xlsm"):
        publish_unchanged_target(target, tmp_path / "result.xlsx", "digest")


def test_stage_validation_keeps_broad_printable_labels_without_locations() -> None:
    from report_processor.admin_panel.reconciliation_uploads import validate_stage

    assert validate_stage("Секция (А)") == "Секция (А)"
    assert validate_stage("Секция/Б") == "Секция/Б"
    with pytest.raises(ValueError):
        validate_stage("/private/target.xlsx")
    with pytest.raises(ValueError):
        validate_stage("Лист1!A1")


def test_reader_uses_discovered_cells_for_target_snapshot_provenance(tmp_path) -> None:
    target = tmp_path / "target.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["B1"], sheet["C1"], sheet["D1"], sheet["E1"], sheet["F1"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
        "Единица измерения",
    )
    sheet.merge_cells("L1:M1")
    sheet["L1"] = "Отчетный период"
    sheet["L2"], sheet["M2"] = "Количество", "Стоимость"
    sheet["B3"], sheet["C3"], sheet["D3"], sheet["E3"], sheet["F3"] = (
        "1234",
        "Этап 13.1",
        "1",
        "Монтаж",
        "м",
    )
    sheet["J3"], sheet["K3"], sheet["L3"], sheet["M3"] = 99, 88, 12, 3.5
    workbook.save(target)
    workbook.close()

    schema, (row,) = read_reconciliation_target(
        target, sha256(target.read_bytes()).hexdigest(), "13.1"
    )

    assert [(binding.logical_column, binding.column_letter) for binding in schema.column_bindings][
        -2:
    ] == [
        (LogicalColumn.CURRENT_PERIOD_QUANTITY, "L"),
        (LogicalColumn.CURRENT_PERIOD_COST, "M"),
    ]
    assert row.cell_for(LogicalColumn.CURRENT_PERIOD_QUANTITY).coordinate == "L3"
    assert row.cell_for(LogicalColumn.CURRENT_PERIOD_COST).coordinate == "M3"


def test_writer_updates_only_structurally_discovered_current_measure_cells(tmp_path) -> None:
    target = tmp_path / "target.xlsx"
    output = tmp_path / "result.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчёт"
    sheet["B1"], sheet["C1"], sheet["D1"], sheet["E1"], sheet["F1"] = (
        "Индекс документа",
        "Номер этапа",
        "Номер п/п",
        "Наименование работ",
        "Единица измерения",
    )
    sheet["J1"], sheet["K1"] = "Документальная отчетность", "Документальная отчетность"
    sheet["J2"], sheet["K2"] = "Количество", "Стоимость"
    sheet.merge_cells("L1:M1")
    sheet["L1"] = "Отчетный период"
    sheet["L2"], sheet["M2"] = "Количество", "Стоимость"
    sheet["N1"] = "Нарратив"
    sheet["B3"], sheet["C3"], sheet["D3"], sheet["E3"], sheet["F3"] = (
        "1234",
        "Этап 13.1",
        "1",
        "Монтаж",
        "м",
    )
    sheet["J3"], sheet["K3"], sheet["L3"], sheet["M3"], sheet["N3"] = (99, 88, 0, 0, "text")
    sheet["L3"].number_format = sheet["M3"].number_format = "0.00"
    workbook.save(target)
    workbook.close()

    schema, (target_row,) = read_reconciliation_target(
        target, sha256(target.read_bytes()).hexdigest(), "13.1"
    )
    target_row = replace(target_row, writable=True)
    source = calculation_source_row(quantity=Decimal("12.2"), cost=Decimal("3500000"))
    (calculation,) = calculate_matches((match_result(source),), calculation_rule_set())
    calculation = calculation.__class__(
        calculation.calculation_id,
        calculation.target_row_id,
        calculation.match_result_id,
        target_row,
        calculation.status,
        calculation.quantity,
        calculation.cost_before_coefficient,
        calculation.coefficient,
        calculation.cost,
        calculation.category_totals,
        calculation.trace,
        calculation.warnings,
        calculation.explanation,
    )

    write_target_report(
        target,
        output,
        WriteDecision.ALLOW_WRITE,
        (writer_calculations((calculation,))[0],),
        schema,
    )

    written = load_workbook(output, data_only=True)
    try:
        sheet = written["Отчёт"]
        assert (sheet["J3"].value, sheet["K3"].value, sheet["N3"].value) == (99, 88, "text")
        assert (sheet["L3"].value, sheet["M3"].value) == (12.2, 3.5)
    finally:
        written.close()
